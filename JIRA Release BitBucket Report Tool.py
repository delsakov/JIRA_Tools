from jira import JIRA
from atlassian import Bitbucket
import datetime
from sys import exit
import tkinter as tk
from tkinter import *
from tkinter.filedialog import askopenfilename, asksaveasfilename
from openpyxl.styles import Font, PatternFill
from openpyxl import Workbook, load_workbook
import os
import json
import requests
import traceback
import concurrent.futures
import base64

# Tool properties
current_version = '0.2'
#os.environ['NO_PROXY'] = '<add your site here to skip PROXY>'

# Excel configs
red_font = Font(color='00FF0000', italic=True)
header_font = Font(color='00000000', bold=True)
header_fill = PatternFill(fill_type="solid", fgColor="8db5e2")
hyperlink = Font(underline='single', color='0563C1')
headers = {"Content-type": "application/json", "Accept": "application/json"}
verify = True
JIRA_fields_api = '/rest/api/2/field'
fields_ids_mapping = {}

excel_columns = {'ID': {'index': 0, 'visible': 1, 'name': 'ID', 'field_id': 'key'},
                 'Type': {'index': 1, 'visible': 1, 'name': 'Type', 'field_id': 'issuetype'},
                 'Check Status': {'index': 2, 'visible': 1, 'name': 'Check Status', 'field_id': None},
                 'Summary': {'index': 3, 'visible': 1, 'name': 'Summary', 'field_id': 'summary'},
                 'Components': {'index': 4, 'visible': 1, 'name': 'Component/s', 'field_id': 'components'},
                 'Status': {'index': 6, 'visible': 1, 'name': 'Status', 'field_id': 'status.name'},
                 'Resolution': {'index': 7, 'visible': 1, 'name': 'Resolution', 'field_id': 'resolution'},
                 'fixVersions': {'index': 8, 'visible': 1, 'name': 'Fix Versions', 'field_id': 'fixVersions'},
                 'AD Lead': {'index': 9, 'visible': 1, 'name': 'AD Lead', 'field_id': None},
                 'Reporter': {'index': 10, 'visible': 1, 'name': 'Reporter', 'field_id': 'reporter'},
                 'Assignee': {'index': 11, 'visible': 1, 'name': 'Assignee', 'field_id': 'assignee'},
                 'Epic Link': {'index': 12, 'visible': 1, 'name': 'Epic Link', 'field_id': None},
                 'Epic Name': {'index': 13, 'visible': 1, 'name': 'Epic Name', 'field_id': None},
                 'Team': {'index': 14, 'visible': 1, 'name': 'Team', 'field_id': None},
                 'Labels': {'index': 15, 'visible': 1, 'name': 'Labels', 'field_id': 'labels'},
                 'Sprint': {'index': 16, 'visible': 1, 'name': 'Sprint', 'field_id': None},
                 'Story Points': {'index': 17, 'visible': 1, 'name': 'Story Points', 'field_id': None},
                 'Due Date': {'index': 19, 'visible': 1, 'name': 'Due Date', 'field_id': 'duedate'},
                 'Start Date': {'index': 20, 'visible': 1, 'name': 'Start Date', 'field_id': None},
                 'End Date': {'index': 21, 'visible': 1, 'name': 'End Date', 'field_id': None},
                 'Repository Name': {'index': 22, 'visible': 1, 'name': 'Repository Name(s)', 'field_id': None},
                 }
excel_columns_default = excel_columns

# Program configs
override_checkbox = 0
config_file = 'bitbucket_release_reporting_config.json'
zoom_scale = 90
jira_sheet_title = 'Items from JIRA'
aggregated_sheet = {'name': 'Aggregated',
                    'visible': 1
                    }
JIRA_BASE_URL = '<your JIRA Base URL>'
Bitbucket_URL = '<your BitBucket Base URL>'
fix_version = ''

# Default values
process_aditional_repo_flag = 0
process_mapping_flag = 0
long_title_name = 0
credentials_saved_flag = 0
updated_issues = {}
username = ''
password = ''
time_format = "%Y-%m-%dT%H:%M:%S"
mapping_file = 'JIRA Release Reporting Mappings.xlsx'
report_name = 'JIRA Export.xlsx'
now = datetime.datetime.strptime(str(datetime.datetime.utcnow().isoformat()).split(".", 1)[0], time_format)
default_output_excel = report_name.split('.xls')[0] + '_' + str(now).replace(':', '-').replace(' ', '_') + '_UTC.xlsx'
output_excel = default_output_excel
project_name = ''
release_branch_name = ''
prod_branch = ''
repo_name = ''
jql = ''
repo_names = []
ad_leads = {}
ad_components = {}
ad_teams = {}
ad_labels = {}
ad_lead_check_order = {'ad_leads': {'index': 2, 'type': 'custom'},
                       'ad_teams': {'index': 0, 'type': 'teams'},
                       'ad_components': {'index': 1, 'type': 'components'},
                       'ad_labels': {'index': 3, 'type': 'labels'}
                       }
ad_teams_map = []
ad_components_map = []
ad_leads_map = []
ad_labels_map = []
ad_lead_check_order_map = []
custom_ad_field = ''
additional_repositories = [
    {'project_name': '',
     'repo_name': '',
     'prod_branch': '',
     'release_branch_name': '',
     },
    {'project_name': '',
     'repo_name': '',
     'prod_branch': '',
     'release_branch_name': '',
     }
]

# Checking JIRA connectivity
try:
    jira = JIRA(JIRA_BASE_URL)
except Exception as er:
    print("Exception with JIRA connection: {}".format(er))


# Functions
def get_excel_columns():
    global excel_columns, excel_columns_default
    if not excel_columns or excel_columns == {}:
        excel_columns = excel_columns_default
    else:
        for k in excel_columns_default.keys():
            if k not in excel_columns.keys():
                excel_columns[k] = excel_columns_default[k]
        for i in range(len(excel_columns_default.keys())):
            index_used = 0
            for k in excel_columns_default.keys():
                if excel_columns[k]['index'] == i:
                    if index_used == 0:
                        index_used = 1
                    else:
                        excel_columns[k]['index'] += 1


def threads_processing(function, items):
    global threads, max_retries
    
    items_for_retry = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=threads) as executor:
        futures = {executor.submit(function, i) for i in items}
    for future in futures:
        if future.result()[0] == 1:
            items_for_retry.append(future.result()[1])
    if len(items_for_retry) > 0:
        if max_retries > 0:
            max_retries -= 1
            threads_processing(function, items_for_retry)
        else:
            print("The following items can't be processed: '{}'".format(items_for_retry))
            return


def get_str_from_lst(lst, sep=',', spacing=' ', edging='"', stripping=True):
    """This function returns list as comma separated string - for exporting in excel"""
    if lst is None:
        return None
    elif type(lst) != list:
        return str(lst)
    st = ''
    for i in lst:
        if i != '':
            if stripping is True:
                st += edging + str(i).strip() + edging + sep + spacing
            else:
                st += edging + str(i) + edging + sep + spacing
    if spacing == ' ':
        st = st[0:-2]
    else:
        st = st[0:-1]
    return st


def get_visible_columns():
    visible_columns = []
    for v in excel_columns.values():
        if v['visible'] == 1:
            visible_columns.append(v['index'])
    return visible_columns


def select_output_file():
    global output_excel
    dir_name = os.getcwd()
    output_excel = asksaveasfilename(initialdir=dir_name, title="Select file", filetypes=(("JIRA Report", ".xlsx"), ("all files", "*.*")))
    if not output_excel.endswith('.xlsx'):
        output_excel = output_excel.split('.xl')[0] + '.xlsx'
    out_xls.delete(0, END)
    out_xls.insert(0, output_excel)
    

def select_mapping_file():
    global mapping_file
    dir_name = os.getcwd()
    mapping_file = askopenfilename(initialdir=dir_name, title="Select file", filetypes=(("WPT Release Reporting Mapping File", ".xlsx"), ("all files", "*.*")))
    if not mapping_file.endswith('.xlsx'):
        mapping_file = mapping_file.split('.xl')[0] + '.xlsx'
    mapping.delete(0, END)
    mapping.insert(0, mapping_file)


def generate_mapping_file():
    
    def create_mapping_file():
        global mapping_file
        dir_name = os.getcwd()
        mapping_file = asksaveasfilename(initialdir=dir_name, title="Select file", filetypes=(("WPT Release Reporting Mapping File", ".xlsx"), ("all files", "*.*")))
        if not mapping_file.endswith('.xlsx'):
            mapping_file = mapping_file.split('.xl')[0] + '.xlsx'
        mapping.delete(0, END)
        mapping.insert(0, mapping_file)

    data = {'AD Lead Order': [['Mapping Type (sheet)', 'Processing Mapping Order (starting with "0")', 'CustomField Name'],
                              ['AD Teams', '0', ''],
                              ['AD Components', '1', ''],
                              ['AD CustomField', '2', ''],
                              ['AD Labels', '3', ''],
                              ],
            'AD Teams': [['AD Lead Name', 'Team Name'],
                         ['', '']
                         ],
            'AD Components': [['AD Lead Name', 'Component Name (comma separated parts, contains all of them)'],
                              ['', '']
                              ],
            'AD Labels': [['AD Lead Name', 'Part of Label'],
                          ['', '']
                          ],
            'AD CustomField': [['AD Lead Name', 'CustomField Value'],
                               ['', '']
                               ],
            }
    
    create_mapping_file()
    wb1 = Workbook()

    # Existing sheets - placed after first one
    for k, v in data.items():
        create_excel_sheet(wb1, v, k, template=True)
    
    # Saving Excel file
    save_file(wb1, mapping_file, mapping=True)
    

def get_jira_changelog(project_name, repo_name, ref_from=release_branch_name, ref_to=prod_branch):
    global bitbucket
    jiras =[]
    changelog = bitbucket.get_changelog(project_name, repo_name, ref_from=ref_from, ref_to=ref_to)
    if changelog is not None:
        for i in changelog:
            try:
                if i['properties']['jira-key'] and len(i['parents']) == 1:
                    jiras.extend(i['properties']['jira-key'])
            except Exception as e:
                # No 'properties' found -> merged requests
                # print("Exception for processing {} : {}".format(i, e))
                continue
        jiras = list(set(jiras))
    return jiras


def read_mapping_excel(file_path, excel_sheet_name, columns=2, rows=0, start_row=2):
    global excel_columns, custom_ad_field
    global ad_teams_map, ad_components_map, ad_leads_map, ad_labels_map, ad_lead_check_order_map
    try:
        df = load_workbook(file_path, data_only=True)
        df1 = df.get_sheet_by_name(excel_sheet_name)
        if excel_sheet_name == 'AD Teams':
            ad_teams_map = []
        if excel_sheet_name == 'AD Components':
            ad_components_map = []
        if excel_sheet_name == 'AD CustomField':
            ad_leads_map = []
        if excel_sheet_name == 'AD Labels':
            ad_labels_map = []
        if excel_sheet_name == 'AD Lead Order':
            ad_lead_check_order_map = []
        row_count = rows
        col_count = columns
        
        if rows == 0:
            row_count = df1.max_row
        
        empty_row = ['' for i in range(col_count)]
        
        for row in df1.iter_rows(min_row=start_row, max_row=row_count, max_col=col_count):
            d = []
            for v in row:
                val = v.value
                if val is None:
                    val = ""
                else:
                    val = str(val).strip().replace(u'\xa0', u' ')  # '\xa0' - non-breaking space, replaced by space
                d.append(val)
            if set(d) != set(empty_row):
                if excel_sheet_name == 'AD Teams':
                    ad_teams_map.append(d)
                if excel_sheet_name == 'AD Components':
                    ad_components_map.append(d)
                if excel_sheet_name == 'AD CustomField':
                    ad_leads_map.append(d)
                if excel_sheet_name == 'AD Labels':
                    ad_labels_map.append(d)
                if excel_sheet_name == 'AD Lead Order':
                    if d[0] == 'AD CustomField':
                        custom_ad_field = d[2]
                    ad_lead_check_order_map.append(d[:2])
    
    except:
    
        print(traceback.format_exc())
        
        if excel_sheet_name in ['AD Teams', 'AD Components', 'AD CustomField', 'AD Labels', 'AD Lead Order']:
            print('{} file not found. Default AD Lead calculated will be used...'.format(file_path))


def create_excel_sheet(workbook, sheet_data, title, template=False):
    global long_title_name
    
    if len(title) > 30:
        if long_title_name == 0:
            title = title[0:30]
        else:
            title = title[0:28] + '_' + str(long_title_name)
        long_title_name += 1
    workbook.create_sheet(title)
    ws = workbook.get_sheet_by_name(title)
    
    start_column = 1
    start_row = 1
    visible_cols = get_visible_columns()
    
    # Creating Excel sheet based on data
    for i in range(len(sheet_data)):
        for y in range(len(sheet_data[i])):
            if y in visible_cols:
                if ((y == excel_columns['ID']['index'] or y == excel_columns['Epic Link']['index'])
                        and start_row != 1 and sheet_data[i][y] is not None and sheet_data[i][y] != ''
                        and not template):
                    ws.cell(row=start_row, column=start_column+y).hyperlink = JIRA_BASE_URL + '/browse/' + sheet_data[i][y]
                    ws.cell(row=start_row, column=start_column+y).font = hyperlink
                ws.cell(row=start_row, column=start_column+y).value = sheet_data[i][y]
        start_row += 1
    
    for y in range(1, ws.max_column+1):
        ws.cell(row=1, column=y).fill = header_fill
        ws.cell(row=1, column=y).font = header_font

    # Marking as red
    if not template:
        start_row = 1
        for i in range(len(sheet_data)):
            for y in range(len(sheet_data[i])):
                if y == excel_columns['Check Status']['index']:
                    ws.cell(row=start_row+1, column=start_column+y).font = red_font
            start_row += 1

    # Removing columns
    if not template:
        cols_to_remove = []
        for v in excel_columns.values():
            if v['visible'] == 0:
                cols_to_remove.append(start_column+v['index'])
        cols_to_remove.sort(reverse=True)
        for z in cols_to_remove:
            ws.delete_cols(z)
    
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        if length > 80:
            ws.column_dimensions[column_cells[0].column_letter].width = 80
        else:
            ws.column_dimensions[column_cells[0].column_letter].width = length + 4

    ws.auto_filter.ref = ws.dimensions
    ws.title = title


def save_file(workbook, output_excel, mapping=False):
    # Saving Excel file and removing not required sheets
    try:
        workbook.remove_sheet(workbook['Sheet'])
    except Exception as e:
        print("Exception:", e)
    
    try:
        if mapping:
            set_zoom(workbook, output_excel)
            print("Mapping file \"", output_excel, "\" successfully generated.", sep='')
            print()
        else:
            time_format = "%Y-%m-%dT%H:%M:%S"
            now = datetime.datetime.strptime(str(datetime.datetime.utcnow().isoformat()).split(".", 1)[0], time_format)
            if output_excel == '':
                name = project_name + '_' + get_str_from_lst(repositories).replace(',', '_').replace(' ', '') + '_' + release_branch_name.replace('/', '_')
                if len(name) > 100:
                    name = name[:100]
                output_excel = name + '_' + str(now).replace(':', '-').replace(' ', '_') + '_UTC.xlsx'
            else:
                output_excel = output_excel.split('.xls')[0] + '_' + str(now).replace(':', '-').replace(' ', '_') + '_UTC.xlsx'
            set_zoom(workbook, output_excel)
            print("File \"", output_excel, "\" successfully generated.", sep='')
            print()
            os.system("pause")
            exit()
    except Exception as e:
        print()
        print("ERROR:", e)
        os.system("pause")
        exit()


def set_zoom(workbook, file):
    for ws in workbook.worksheets:
        ws.sheet_view.zoomScale = zoom_scale
    workbook.save(file)


def get_ad_teams():
    """This function returns AD Leads to Teams mapping loaded from the file"""
    global ad_teams_map, ad_teams
    
    if len(ad_teams_map) > 0:
        ad_teams = {}
        for i in range(1, len(ad_teams_map)):
            ad_teams[ad_teams_map[i][0]] = [ad_teams_map[i][1]]
    
    for k in ad_teams.keys():
        ad_teams[k] = list(set(ad_teams[k]))


def get_ad_leads():
    """This function returns AD Leads to Initiative mapping loaded from the file"""
    global ad_leads_map, ad_leads
    
    if len(ad_leads_map) > 0:
        ad_leads = {}
        for i in range(1, len(ad_leads_map)):
            ad_leads[ad_leads_map[i][0]] = [ad_leads_map[i][1]]
    
    for k in ad_leads.keys():
        ad_leads[k] = list(set(ad_leads[k]))


def get_ad_components():
    """This function returns AD Leads to Components mapping loaded from the file"""
    global ad_components_map, ad_components
    
    if len(ad_components_map) > 0:
        ad_components = {}
        for i in range(1, len(ad_components_map)):
            if ad_components_map[i][0] not in ad_components.keys():
                ad_components[ad_components_map[i][0]] = []
            ad_components[ad_components_map[i][0]].append([ad_components_map[i][1]])


def get_ad_labels():
    """This function returns AD Leads to Labels mapping loaded from the file"""
    global ad_labels_map, ad_labels
    
    if len(ad_labels_map) > 0:
        ad_labels = {}
        for i in range(1, len(ad_labels_map)):
            ad_labels[ad_labels_map[i][0]] = [ad_labels_map[i][1]]
    
    for k in ad_labels.keys():
        ad_labels[k] = list(set(ad_labels[k]))


def get_ad_order():
    """This function returns AD Leads order processed loaded from the file"""
    global ad_lead_check_order_map, ad_lead_check_order
    map_type = ''
    type = ''
    
    if len(ad_lead_check_order_map) > 0:
        
        for i in range(1, len(ad_lead_check_order_map)):
            if ad_lead_check_order_map[i][0].strip().upper() == 'AD TEAMS':
                map_type = 'ad_teams'
                type = 'teams'
            elif ad_lead_check_order_map[i][0].strip().upper() == 'AD COMPONENTS':
                map_type = 'ad_components'
                type = 'components'
            elif ad_lead_check_order_map[i][0].strip().upper() == 'AD CUSTOMFIELD':
                map_type = 'ad_leads'
                type = 'custom'
            elif ad_lead_check_order_map[i][0].strip().upper() == 'AD LABELS':
                map_type = 'ad_labels'
                type = 'labels'
            ad_lead_check_order[map_type]['index'] = int(ad_lead_check_order_map[i][1])
            ad_lead_check_order[map_type]['type'] = type


def get_ad_lead(components, initiative, labels, teams):
    global ad_leads, ad_components, ad_labels, ad_lead_check_order, ad_teams
    lead = ''
    equal = 0
    
    def get_ad_initiative(initiative):
        lead = ''
        if initiative:
            for name, inits in ad_leads.items():
                for k in range(len(inits)):
                    inits[k] = inits[k].strip().lower()
                if initiative.strip().lower() in inits:
                    lead = name
                    break
            return lead
    
    def get_ad_team(teams):
        lead = ''
        if teams:
            teams_lst = teams.split(',')
            for t in teams_lst:
                for name, tms in ad_teams.items():
                    for j in range(len(tms)):
                        tms[j] = tms[j].strip().lower()
                    if t.strip().lower() in tms:
                        lead = name
                        break
            return lead
    
    def get_ad_components(components):
        lead = ''
        if components:
            components_lst = components.split(',')
            for component in components_lst:
                for name, comps in ad_components.items():
                    for comp in comps:
                        equal = 0
                        for c in comp:
                            if len(c.split(',')) > 1:
                                for a in c.split(','):
                                    if a.strip().lower() in component.strip().lower():
                                        equal = 1
                                    else:
                                        equal = 0
                                        continue
                            else:
                                if c.strip().lower() in component.strip().lower():
                                    equal = 1
                        if equal == 1:
                            lead = name
                            break
            return lead
    
    def get_ad_labels(labels):
        lead = ''
        if labels:
            labels_lst = labels.split(',')
            for l in labels_lst:
                for name, labs in ad_labels.items():
                    for j in range(len(labs)):
                        labs[j] = labs[j].strip().lower()
                    if l.strip().lower() in labs:
                        lead = name
                        break
            return lead
    
    for k, v in ad_lead_check_order.items():
        for i in range(len(ad_lead_check_order)):
            if v['index'] == i:
                if v['type'] == 'custom':
                    lead = get_ad_initiative(initiative)
                    if lead is not None and lead != '':
                        return lead
                if v['type'] == 'teams':
                    lead = get_ad_team(teams)
                    if lead is not None and lead != '':
                        return lead
                if v['type'] == 'components':
                    lead = get_ad_components(components)
                    if lead is not None and lead != '':
                        return lead
                if v['type'] == 'labels':
                    lead = get_ad_labels(labels)
                    if lead is not None and lead != '':
                        return lead
    return lead


def get_columns():
    global excel_columns
    excel_columns_list = ['' for i in range(len(excel_columns))]
    for v in excel_columns.values():
        excel_columns_list[v['index']] = v['name']
    return excel_columns_list


def get_values_from_ui():
    global project_name, repo_names, prod_branch, release_branch_name, jql, fix_version, report_name, output_excel
    global JIRA_BASE_URL, Bitbucket_URL, mapping_file, excel_columns, username, password

    username = sid.get().strip()
    password = sso_pass.get().strip()
    config_file = conf.get().strip().split('.json')[0] + '.json'
    JIRA_BASE_URL = jira_url.get().strip()
    Bitbucket_URL = repo_url.get().strip()
    
    if len(username) < 6 or len(password) < 3:
        print("Invalid BitBucket credentials were entered!")
        main.destroy()
        os.system("pause")
        exit()
        
    project_name = project.get().strip().replace('/n', '')
    if project_name == '':
        print("Invalid Project '{}' was entered!".format(project_name))
        main.destroy()
        os.system("pause")
        exit()
        
    repo_names = repo.get().strip().replace('/n', '').replace(' ', '').replace('"', '').replace("'", "").split(',')
    if len(repo_names) > 0:
        for i in range(len(repo_names)):
            repo_names[i] = repo_names[i].strip().replace('/n', '')
            
    release_branch_name = branch.get().strip().replace('/n', '')
    if release_branch_name == '':
        release_branch_name = 'develop'
        
    prod_branch = prod.get().strip().replace('/n', '')
    if prod_branch == '':
        prod_branch = 'master'
        
    jql = j_query.get().strip().replace('/n', '')
    fix_version = version.get().strip().replace('/n', '')
    output_excel = out_xls.get().strip().replace('/n', '')
    report_name = output_excel

    mapping_file = mapping.get().strip().replace('/n', '')
    if mapping_file == '':
        excel_columns['AD Lead']['visible'] = 0
    
    if override_checkbox == 1:
        save_config(configfile=config_file)
    main.destroy()

    get_fields_ids()
    
    # Starting main program
    main_program()


def change_override(*args):
    global override_checkbox
    override_checkbox = override.get()


def change_credentials_saved(*args):
    global credentials_saved_flag
    credentials_saved_flag = credentials_saved.get()


def change_aditional_repo(*args):
    global process_aditional_repo_flag
    process_aditional_repo_flag = process_aditional_repo.get()


def change_process_mapping(*args):
    global process_mapping_flag
    process_mapping_flag = process_mapping.get()


def get_issues_by_jql(jql):
    """This function returns list of JIRA keys for provided list of JIRA JQL queries"""
    auth_jira = JIRA(JIRA_BASE_URL)
    issues, items = ([], [])
    start_idx, block_num, block_size = (0, 0, 100)
    while True:
        start_idx = block_num * block_size
        tmp_issues = auth_jira.search_issues(jql_str=jql, startAt=start_idx, maxResults=block_size, fields='key, issuetype')
        if len(tmp_issues) == 0:
            # Retrieve issues until there are no more to come
            break
        issues.extend(tmp_issues)
        block_num += 1
    
    items = list(set([i.key for i in issues]))
    return items
    
    
def main_program():
    global jql, fix_version, updated_issues, excel_columns, mapping_file, ad_lead_check_order, username, password
    global process_aditional_repo_flag, output_excel, process_mapping_flag

    if excel_columns['AD Lead']['visible'] == 1 and process_mapping_flag == 1:
        read_mapping_excel(excel_sheet_name='AD Lead Order', columns=3, file_path=mapping_file)
        get_ad_order()
        if ad_lead_check_order != {}:
            read_mapping_excel(excel_sheet_name='AD Teams', file_path=mapping_file)
            get_ad_teams()
            read_mapping_excel(excel_sheet_name='AD Components', file_path=mapping_file)
            get_ad_components()
            read_mapping_excel(excel_sheet_name='AD CustomField', file_path=mapping_file)
            get_ad_leads()
            read_mapping_excel(excel_sheet_name='AD Labels', file_path=mapping_file)
            get_ad_labels()
            print('AD Leads mapping has been loaded.')
        else:
            print('AD Leads mapping has been skipped - mapping file filled incorrectly.')
    
    # Calculating part - the list of issues associated with Commits in changelog
    print("Bitbucket data is being retrieved...")
    issues = calculate_changelog(project_name, repo_names, prod_branch, release_branch_name, username, password)
    
    if process_aditional_repo_flag == 1 and additional_repositories:
        for i in additional_repositories:
            if i['project_name'] != '' and i['repo_name'] != '' and i['release_branch_name'] != '' and i['prod_branch'] != '':
                opt_issues = calculate_changelog(i['project_name'], i['repo_name'].replace(' ', '').split(','), i['prod_branch'], i['release_branch_name'], username, password)
                new_issues = {}
                for i in [issues, opt_issues]:
                    new_issues.update(i)
                issues = new_issues
    
    # Check if JQL and Fix Versions are not empty
    if fix_version == '' and jql == '':
        excel_columns['Check Status']['visible'] = 0
    else:
        if fix_version != '' and jql == '':
            jql = 'fixVersion = "{}"'.format(str(fix_version).replace('/n', '').strip())
        elif fix_version != '' and jql != '':
            fix_version_jql = 'fixVersion = "{}"'.format(fix_version)
            jql += ' AND ' + fix_version_jql
        try:
            items = get_issues_by_jql(jql)
        except Exception as e:
            items = []
            try:
                error = e.text
            except:
                error = e
            print()
            print("Error while retrieving issues from JIRA due to Exception: {}".format(error))
            print()
        issues[jira_sheet_title] = items
        updated_issues[jira_sheet_title] = []
        updated_issues[jira_sheet_title].append(get_columns())
        
    # Retrieve data from JIRA
    get_jira_metadata(issues)
    
    print("Excel is being generated...")
    print()
    
    wb = Workbook()
    # Add first Aggregated data sheet
    if aggregated_sheet['visible'] == 1:
        create_excel_sheet(wb, all_issues, 'Aggregated list')

    # Existing sheets - placed after first one
    for k, v in updated_issues.items():
        create_excel_sheet(wb, v, k)

    # Saving Excel file
    save_file(wb, output_excel)


def calculate_changelog(project_name, repo_names, prod_branch, release_branch_name, username, password):
    global bitbucket, updated_issues, all_issues, Bitbucket_URL
    
    bitbucket = Bitbucket(url=Bitbucket_URL, username=username, password=password)
    
    print("Issues associated with Commits in {} / {} are being downloaded from BitBucket...".format(project_name, repo_names))
    
    issues = {}
    for repo_name in repo_names:
        updated_issues[repo_name] = []
        issues[repo_name] = []
        updated_issues[repo_name].append(get_columns())
        issue_lst = get_jira_changelog(project_name, repo_name, ref_from=release_branch_name, ref_to=prod_branch)
        if len(issue_lst) == 0:
            issue_lst = get_jira_changelog(project_name, repo_name, ref_to=release_branch_name, ref_from=prod_branch)
        for i in range(len(issue_lst)):
            issues[repo_name].append(issue_lst[i])
    print("Issues for {} / {} have been successfully downloaded from BitBucket.".format(project_name, repo_names))
    
    return issues


def get_jira_metadata(issues):
    global updated_issues, all_issues, repositories, excel_columns, fields_ids_mapping, custom_ad_field
    check_status = ''
    repositories = [k for k in issues.keys()]
    print("Metadata for Issues in {} repositories is being downloaded from JIRA...".format(repositories))
    
    for k, v in issues.items():
        n = 0
        print("Total JIRA issues for '{}' to be processed: {}".format(k, len(v)))
        for i in range(len(v)):
            n += 1
            details = ['' for i in range(len(excel_columns))]
            if v[i] != '':
                try:
                    issue = jira.issue(v[i])
                except:
                    continue
                for field_name in excel_columns.keys():
                    try:
                        if field_name == 'ID':
                            details[excel_columns[field_name]['index']] = v[i]
                        elif field_name == 'Sprint':
                            field_id = fields_ids_mapping[field_name]
                            value = eval('issue.fields.' + field_id)
                            details[excel_columns[field_name]['index']] = '' if value is None else get_str_from_lst([i.split(',name=')[1].split(',')[0] for i in value], edging='')
                        elif field_name == 'Reporter':
                            value = eval('issue.fields.' + excel_columns[field_name]['field_id'])
                            details[excel_columns[field_name]['index']] = 'Anonymous' if value is None else value.displayName if hasattr(value, "displayName") else value
                        elif field_name == 'Assignee':
                            value = eval('issue.fields.' + excel_columns[field_name]['field_id'])
                            details[excel_columns[field_name]['index']] = 'Unassigned' if value is None else value.displayName if hasattr(value, "displayName") else value
                        elif field_name == 'Check Status':
                            check_status = ''
                            if fix_version != '' and fix_version not in [i.name if hasattr(i, "name") else '' for i in issue.fields.fixVersions]:
                                check_status += 'Wrong Fix Version. '
                            if issue.fields.status.name is not None and issue.fields.status.statusCategory.name.upper() != 'DONE':
                                check_status += 'Wrong Status. '
                            details[excel_columns[field_name]['index']] = check_status
                        elif field_name == 'AD Lead':
                            if process_mapping_flag == 1:
                                components = get_str_from_lst([i.name if hasattr(i, "name") else '' for i in issue.fields.components])
                                labels = get_str_from_lst([i for i in issue.fields.labels])
                                team_id = fields_ids_mapping['Team']
                                teams = get_str_from_lst('' if team_id is None else team_id)
                                custom_field_id = fields_ids_mapping[custom_ad_field]
                                custom_leads_value = eval('issue.fields.' + custom_field_id)
                                if custom_leads_value is None:
                                    custom_leads = ''
                                elif type(custom_leads_value) == list:
                                    custom_leads = [item.displayName if hasattr(item, "displayName") else item.value if hasattr(item, "value") else item.name if hasattr(item, "name") else item for item in custom_leads_value][0]
                                else:
                                    custom_leads = custom_leads_value.displayName if hasattr(custom_leads_value, "displayName") else custom_leads_value.value if hasattr(custom_leads_value, "value") else custom_leads_value.name if hasattr(custom_leads_value, "name") else custom_leads_value
                                details[excel_columns[field_name]['index']] = get_ad_lead(components, custom_leads, labels, teams)
                            else:
                                continue
                        elif field_name == 'Repository Name':
                            repository = ''
                            if k != jira_sheet_title:
                                repository = k
                            details[excel_columns[field_name]['index']] = repository
                        elif 'field_id' in excel_columns[field_name].keys() and excel_columns[field_name]['field_id'] is None:
                            try:
                                field_id = fields_ids_mapping[field_name]
                                value = eval('issue.fields.' + field_id)
                                if value is None:
                                    value = ''
                                elif type(value) == list:
                                    value = get_str_from_lst([item.key if hasattr(item, "key") else item.value if hasattr(item, "value") else item.name if hasattr(item, "name") else item for item in value])
                                else:
                                    value = value.key if hasattr(value, "key") else value.value if hasattr(value, "value") else value.name if hasattr(value, "name") else value
                                details[excel_columns[field_name]['index']] = value
                            except:
                                details[excel_columns[field_name]['index']] = ''
                        else:
                            value = eval('issue.fields.' + excel_columns[field_name]['field_id'])
                            if value is None:
                                value = ''
                            elif type(value) == list:
                                value = get_str_from_lst([item.key if hasattr(item, "key") else item.value if hasattr(item, "value") else item.name if hasattr(item, "name") else item for item in value])
                            else:
                                value = value.key if hasattr(value, "key") else value.value if hasattr(value, "value") else value.name if hasattr(value, "name") else value
                            details[excel_columns[field_name]['index']] = value
                    except Exception as e:
                        print(traceback.format_exc())
                        print('field_name:', field_name)
                        print("Exception '{}' for retrieving JIRA details for JIRA_ID: {}".format(e, v[i]))
                # Extend list for Excel export
                updated_issues[k].append(details)

            if n % 100 == 0:
                print("Processed {} issues out of {} so far for {} list.".format(n, len(v), k))

    print("Metadata for Issues in {} repositories has been successfully downloaded from JIRA.".format(repositories))

    if excel_columns['Epic Name']['visible'] == 1:
        print("Epics associated with Issues in {} repositories is being downloaded from JIRA...".format(repositories))
        # Update Epics Names
        epics = set()
        epic_details = []
        for k, v in updated_issues.items():
            for i in range(len(v)):
                if v[i][excel_columns['Epic Link']['index']] != '' and v[i][excel_columns['Epic Link']['index']] is not None and v[i][excel_columns['Epic Link']['index']] != 'Epic Link':
                    epics.add(v[i][excel_columns['Epic Link']['index']])
    
        print("Metadata for Epics in {} repositories is being downloaded from JIRA...".format(repositories))
        
        # Retrieve Epic details from JIRA
        print("Total Epics/ Parents to be processed: {}".format(len(epics)))
        for e in epics:
            epic = jira.issue(e)
            ep = [e, epic.fields.summary]
            epic_details.append(ep)
        print("Epic details has been retrieved successfully.")
        
        # Updating data for Excel reporting with Epic details
        print("Enriching data with Epic Details...")
        for k, v in updated_issues.items():
            for i in range(len(v)):
                if v[i][excel_columns['Epic Link']['index']] != '' and v[i][excel_columns['Epic Link']['index']] is not None and v[i][excel_columns['Epic Link']['index']] != 'Epic Link':
                    for e in epic_details:
                        if v[i][excel_columns['Epic Link']['index']] == e[0]:
                            v[i][excel_columns['Epic Name']['index']] = e[1]
    
        print("Metadata for Epics in {} repositories has been successfully downloaded from JIRA.".format(repositories))

    # First sheet - all data aggregated as well as Duplicates calculated
    all_issues = [get_columns()]
    unique_issues = []
    duplicates = set()
    for k, v in updated_issues.items():
        dd = []
        for i in range(1, len(v)):
            repository_name = k
            if v[i][excel_columns['ID']['index']] not in unique_issues:
                unique_issues.append(v[i][excel_columns['ID']['index']])
                temp = v[i]
                if repository_name in repo_names:
                    temp.append(repository_name)
                dd.append(temp)
            duplicates.add((v[i][excel_columns['ID']['index']], v[i][excel_columns['Check Status']['index']], repository_name))
        all_issues.extend(dd)
    
    # Processing Check Status and Repositories columns
    dup = []
    for j in duplicates:
        dup.append(j[0])
    dup = list(set(dup))
    
    dups = []
    for i in dup:
        temp = [i, '', '']
        for x in duplicates:
            if x[0] == i:
                temp[1] += x[1]
                if x[2] != jira_sheet_title:
                    temp[2] += ', ' + x[2]
        repos = list(set(temp[2].strip().split(',')))
        repos = list(filter(None, repos))
        for r in range(len(repos)):
            repos[r] = repos[r].strip()
        temp[2] = get_str_from_lst(repos)
        if temp[2] == '':
            temp[1] += 'No Commits. '
        stats = list(set(temp[1].split('. ')))
        stats = list(filter(None, stats))
        for j in range(len(stats)):
            stats[j] = stats[j] + '. '
        temp[1] = get_str_from_lst(stats)
        dups.append(temp)
    duplicates = dups
    
    # Data manipulation / cleaning for all sheets
    for k, v in updated_issues.items():
        for i in range(1, len(v)):
            for d in duplicates:
                if v[i][excel_columns['ID']['index']] == d[0]:
                    v[i][excel_columns['Check Status']['index']] = d[1]
                    v[i][excel_columns['Repository Name']['index']] = d[2]

    # Data manipulation / clearing for first aggregated sheet
    for i in range(1, len(all_issues)):
        for d in duplicates:
            if all_issues[i][excel_columns['ID']['index']] == d[0]:
                all_issues[i][excel_columns['Check Status']['index']] = d[1]
                all_issues[i][excel_columns['Repository Name']['index']] = d[2]


def load_config(configfile=config_file):
    global JIRA_BASE_URL, excel_columns, ad_leads, ad_components, ad_labels, ad_lead_check_order, ad_teams
    global aggregated_sheet, jira_sheet_title, report_name, input_excel, project_name, repo_names, prod_branch
    global release_branch_name, jql, fix_version, additional_repositories, output_excel, Bitbucket_URL, mapping_file
    global username, password, process_mapping_flag, process_aditional_repo_flag, credentials_saved_flag
    
    if os.path.exists(configfile) is True:
        try:
            with open(configfile) as json_data_file:
                data = json.load(json_data_file)
            for k, v in data.items():
                if k == 'Bitbucket_URL':
                    Bitbucket_URL = v
                elif k == 'project_name':
                    project_name = v
                elif k == 'JIRA_BASE_URL':
                    JIRA_BASE_URL = v
                elif k == 'repo_names':
                    repo_names = v
                elif k == 'prod_branch':
                    prod_branch = v
                elif k == 'release_branch_name':
                    release_branch_name = v
                elif k == 'jql':
                    jql = v
                elif k == 'fix_version':
                    fix_version = v
                elif k == 'excel_columns':
                    excel_columns = v
                elif k == 'aggregated_sheet':
                    aggregated_sheet = v
                elif k == 'jira_sheet_title':
                    jira_sheet_title = v
                elif k == 'mapping_file':
                    mapping_file = v
                elif k == 'report_name':
                    report_name = v
                    output_excel = report_name
                elif k == 'additional_repositories':
                    additional_repositories = v
                elif k == 'process_aditional_repo_flag':
                    process_aditional_repo_flag = v
                elif k == 'process_mapping_flag':
                    process_mapping_flag = v
                elif k == 'credentials_saved_flag':
                    credentials_saved_flag = v
                elif k == 'auth' and v != '':
                    decoded = base64.b64decode(str.encode(v, 'utf-8')).decode('utf-8', 'ignore')
                    (username, password) = eval(decoded)
            print("Configuration file '{}' has been successfully loaded.".format(configfile))
            get_excel_columns()
        except Exception as er:
            print("Failed to load file '{}', due to Exception: '{}'".format(configfile, er))
            get_excel_columns()
            if override_checkbox == 1:
                print("Configuration file is corrupted. Default '{}' would be created instead.".format(configfile))
                print()
                save_config()
    else:
        print("Config File not found. Default '{}' would be created.".format(configfile))
        get_excel_columns()
        print()
        save_config()


def save_config(configfile=config_file):
    global credentials_saved_flag, username, password, project_name, JIRA_BASE_URL, repo_names, prod_branch
    global release_branch_name, jql, Bitbucket_URL, fix_version, additional_repositories, jira_sheet_title
    global aggregated_sheet, report_name, mapping_file, zoom_scale, excel_columns, process_mapping_flag
    global process_aditional_repo_flag
    
    auth = (username, password)
    encoded = base64.b64encode(bytes(str(auth), 'utf-8'))
    
    if credentials_saved_flag == 1:
        data = {'project_name': project_name,
                'JIRA_BASE_URL': JIRA_BASE_URL,
                'repo_names': repo_names,
                'prod_branch': prod_branch,
                'release_branch_name': release_branch_name,
                'jql': jql,
                'Bitbucket_URL': Bitbucket_URL,
                'fix_version': fix_version,
                'additional_repositories': additional_repositories,
                'jira_sheet_title': jira_sheet_title,
                'aggregated_sheet': aggregated_sheet,
                'report_name': report_name,
                'mapping_file': mapping_file,
                'zoom_scale': zoom_scale,
                'excel_columns': excel_columns,
                'process_mapping_flag': process_mapping_flag,
                'process_aditional_repo_flag': process_aditional_repo_flag,
                'credentials_saved_flag': credentials_saved_flag,
                'auth': encoded.decode('utf-8'),
                }
    else:
        data = {'project_name': project_name,
                'JIRA_BASE_URL': JIRA_BASE_URL,
                'repo_names': repo_names,
                'prod_branch': prod_branch,
                'release_branch_name': release_branch_name,
                'jql': jql,
                'Bitbucket_URL': Bitbucket_URL,
                'fix_version': fix_version,
                'additional_repositories': additional_repositories,
                'jira_sheet_title': jira_sheet_title,
                'aggregated_sheet': aggregated_sheet,
                'report_name': report_name,
                'mapping_file': mapping_file,
                'zoom_scale': zoom_scale,
                'excel_columns': excel_columns,
                'process_mapping_flag': process_mapping_flag,
                'process_aditional_repo_flag': process_aditional_repo_flag,
                'credentials_saved_flag': credentials_saved_flag,
                }
        
    if configfile == '.json':
        time_format = "%Y-%m-%dT%H:%M:%S"
        now = datetime.datetime.strptime(str(datetime.datetime.utcnow().isoformat()).split(".", 1)[0], time_format)
        configfile = 'config' + '_' + str(now).replace(':', '-').replace(' ', '_') + '_UTC.json'
    try:
        with open(configfile, 'w') as outfile:
            json.dump(data, outfile)
    except PermissionError as er:
        print("ERROR: File '{}' has been opened for editing and can't be saved. Exception: {}".format(configfile, er))
        return
    
    print("Config file '{}' has been created.".format(configfile))
    print()


# Open File dialog to open config file in the same location and refresh UI values
def open_file():
    global config_file
    dir_name = os.getcwd()
    filename = askopenfilename(initialdir=dir_name, title="Select file", filetypes=(("Configuration File", "*.json"),
                                                                                    ("all files", "*.*")))
    config_file = filename
    conf.delete(0, END)
    conf.insert(0, config_file.split('/')[-1])
    load_config(configfile=config_file)
    project.delete(0, END)
    project.insert(0, project_name)
    repo.delete(0, END)
    repo.insert(0, get_str_from_lst(repo_names))
    prod.delete(0, END)
    prod.insert(0, prod_branch)
    branch.delete(0, END)
    branch.insert(0, release_branch_name)
    j_query.delete(0, END)
    j_query.insert(0, jql)
    version.delete(0, END)
    version.insert(0, fix_version)
    out_xls.delete(0, END)
    out_xls.insert(0, report_name)


def get_fields_ids():
    global JIRA_BASE_URL, JIRA_fields_api, headers, verify, fields_ids_mapping
    
    try:
        url = JIRA_BASE_URL.strip('/') + JIRA_fields_api
        r = requests.get(url, headers=headers, verify=verify)
        fields_string = r.content.decode('utf-8')
        fields_details = json.loads(fields_string)
        for field in fields_details:
            fields_ids_mapping[field["name"]] = field["id"]
    
    except Exception as e:
        print(traceback.format_exc())


# -----------------Main program------------------------------
load_config()

if release_branch_name == '':
    release_branch_name = 'develop'
if prod_branch == '':
    prod_branch = 'master'

main = tk.Tk()
main.title("WPT - Release Reporting Tool - DIFF Jiras from BitBucket for Candidate Release (v_" + current_version + ")")
tk.Label(main, text="BitBucket Repositori(es) configuration", foreground="black", font=("Helvetica", 10, "bold"), pady=5).grid(row=0, column=0, columnspan=4)

tk.Label(main, text="BitBucket URL:", foreground="black", font=("Helvetica", 10), pady=5).grid(row=1, column=0, columnspan=2)
repo_url = tk.Entry(main, width=50)
repo_url.insert(END, Bitbucket_URL)
repo_url.grid(row=1, column=2, pady=5, columnspan=2)

tk.Label(main, text="BitBucket Project Name:", foreground="black", font=("Helvetica", 10), pady=5).grid(row=2, column=0, columnspan=2)
project = tk.Entry(main, width=50)
project.insert(END, project_name)
project.grid(row=2, column=2, pady=5, columnspan=2)

tk.Label(main, text="Repository Name(s) (comma-separated):", foreground="black", font=("Helvetica", 10), pady=5).grid(row=3, column=0, columnspan=2)
repo = tk.Entry(main, width=50)
repo.insert(END, get_str_from_lst(repo_names))
repo.grid(row=3, column=2, pady=5, columnspan=2)

tk.Label(main, text="Current PROD branch/tag (default = 'master'):", foreground="black", font=("Helvetica", 10), pady=5).grid(row=4, column=0, columnspan=2)
prod = tk.Entry(main, width=50)
prod.insert(END, prod_branch)
prod.grid(row=4, column=2, pady=5, columnspan=2)

tk.Label(main, text="Candidate Release branch/tag (default = 'develop'):", foreground="black", font=("Helvetica", 10), pady=5).grid(row=5, column=0, columnspan=2)
branch = tk.Entry(main, width=50)
branch.insert(END, release_branch_name)
branch.grid(row=5, column=2, pady=5, columnspan=2)

process_aditional_repo = IntVar(value=process_aditional_repo_flag)
Checkbutton(main, text="Add extra repositories (could be configured in configuration .json file as 'additional_repositories' block)", font=("Helvetica", 9, "italic"), variable=process_aditional_repo).grid(row=6, sticky=E, padx=10, columnspan=4, pady=0)
process_aditional_repo.trace('w', change_aditional_repo)

tk.Label(main, text="Read access for BitBucket is required.", foreground="black", font=("Helvetica", 10), padx=5, wraplength=380).grid(row=7, column=0, rowspan=2)
tk.Label(main, text="Username:", foreground="black", font=("Helvetica", 10)).grid(row=7, column=2, pady=5, padx=60, columnspan=2, sticky=W)
sid = tk.Entry(main)
sid.insert(END, username)
sid.grid(row=7, column=3, pady=5)

tk.Label(main, text="Password:", foreground="black", font=("Helvetica", 10)).grid(row=8, column=2, pady=5, padx=60, columnspan=2, sticky=W)
sso_pass = tk.Entry(main, width=20, show="*")
sso_pass.insert(END, password)
sso_pass.grid(row=8, column=3, pady=5)

credentials_saved = IntVar(value=credentials_saved_flag)
Checkbutton(main, text="Save credentials", font=("Helvetica", 9, "italic"), variable=credentials_saved).grid(row=8, column=0, sticky=NW, padx=240, columnspan=4, rowspan=2, pady=10)
credentials_saved.trace('w', change_credentials_saved)

tk.Label(main, text="____________________________________________________________________________________________________________").grid(row=9, columnspan=4)

tk.Label(main, text="JIRA configuration", foreground="black", font=("Helvetica", 10, "bold"), pady=5).grid(row=10, column=0, columnspan=4)

tk.Label(main, text="JIRA Instance URL:", foreground="black", font=("Helvetica", 10), pady=5).grid(row=11, column=0, columnspan=2)
jira_url = tk.Entry(main, width=50)
jira_url.insert(END, JIRA_BASE_URL)
jira_url.grid(row=11, column=2, pady=5, columnspan=2)

tk.Label(main, text="OPTIONAL: JQL (for Team / Component check):", foreground="black", font=("Helvetica", 10), pady=5).grid(row=12, column=0)
j_query = tk.Entry(main, width=50)
j_query.insert(END, jql)
j_query.grid(row=12, column=2, pady=5, columnspan=2)

tk.Label(main, text="OPTIONAL: Fix Version from JIRA (for Release):", foreground="black", font=("Helvetica", 10), pady=5).grid(row=13, column=0)
version = tk.Entry(main, width=50, textvariable=fix_version)
version.insert(END, fix_version)
version.grid(row=13, column=2, columnspan=2, padx=8)

tk.Label(main, text="____________________________________________________________________________________________________________").grid(row=14, columnspan=4)

tk.Label(main, text="Release Reporting Tool configuration", foreground="black", font=("Helvetica", 10, "bold"), pady=5).grid(row=15, column=0, columnspan=4)

tk.Label(main, text="Report File:").grid(row=16, column=0, pady=2, padx=30, sticky=W)
out_xls = tk.Entry(main, width=58)
out_xls.insert(END, report_name)
out_xls.grid(row=16, column=0, padx=110, pady=5, columnspan=4, sticky=W)

tk.Button(main, text='Browse', command=select_output_file, width=15).grid(row=16, column=3, pady=3, padx=8)

process_mapping = IntVar(value=process_mapping_flag)
Checkbutton(main, text="Add AD Lead calculated field in report based on mapping file", font=("Helvetica", 9, "italic"), variable=process_mapping).grid(row=17, sticky=W, padx=80, columnspan=4, pady=0)
process_mapping.trace('w', change_process_mapping)

tk.Button(main, text='Generate Mapping File', command=generate_mapping_file, width=20).grid(row=17, column=3, columnspan=2, pady=3, padx=40, sticky=E)

tk.Label(main, text="Mapping File:").grid(row=18, column=0, pady=2, padx=30, sticky=W)
mapping = tk.Entry(main, width=58)
mapping.insert(END, mapping_file)
mapping.grid(row=18, column=0, padx=110, pady=5, columnspan=4, sticky=W)

tk.Button(main, text='Browse', command=select_mapping_file, width=15).grid(row=18, column=3, pady=3, padx=8)

tk.Label(main, text="____________________________________________________________________________________________________________").grid(row=19, columnspan=4)

override = IntVar()
Checkbutton(main, text="Save current values in file (after execution):", font=("Helvetica", 9, "italic"), variable=override).grid(row=20, sticky=W, padx=30, columnspan=2, pady=0)
override.trace('w', change_override)

conf = tk.Entry(main, width=27)
conf.insert(END, config_file)
conf.grid(row=20, column=0, padx=180, columnspan=4, sticky=E)

tk.Button(main, text='Open saved config file', command=open_file).grid(row=20, column=2, pady=8, padx=32, columnspan=2, stick=E)

tk.Label(main, text="____________________________________________________________________________________________________________").grid(row=21, columnspan=4)

tk.Button(main, text='Quit', font=("Helvetica", 9, "bold"), command=main.quit, width=20, heigh=2).grid(row=22, column=0, pady=10, columnspan=1, padx=50, stick=W)
tk.Button(main, text='Generate JIRA Report', font=("Helvetica", 9, "bold"), state='active', command=get_values_from_ui, width=20, heigh=2).grid(row=22, column=2, pady=10, padx=50, columnspan=2, stick=E)

# Please do not change line below with copyright
tk.Label(main, text="Author: Dmitry Elsakov", foreground="grey", font=("Helvetica", 8, "italic"), pady=0).grid(row=23, column=1, sticky=SE, padx=10, columnspan=3)

tk.mainloop()
