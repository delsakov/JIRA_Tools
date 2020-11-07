from jira import JIRA
from time import sleep
import datetime
from sys import exit
from tkinter.filedialog import askopenfilename, asksaveasfilename
import tkinter as tk
from tkinter import *
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
import json

# This should be your JIRA instance URL, if you don't want to use Proxy
# os.environ['NO_PROXY'] = 'jira.com'

# Excel configs
red_font = Font(color='00FF0000', italic=True)
header_font = Font(color='00000000', bold=True)
header_fill = PatternFill(fill_type="solid", fgColor="8db5e2")
hyperlink = Font(underline='single', color='0563C1')

# Default Excel columns configuration
excel_columns = {'ID': {'index': 0, 'visible': 1, 'name': 'ID'},
                 'Type': {'index': 1, 'visible': 1, 'name': 'Type'},
                 'Summary': {'index': 2, 'visible': 1, 'name': 'Summary'},
                 'Components': {'index': 3, 'visible': 1, 'name': 'Component/s'},
                 'Status': {'index': 4, 'visible': 1, 'name': 'Status'},
                 'fixVersions': {'index': 5, 'visible': 1, 'name': 'Fix Versions'},
                 'Reporter': {'index': 6, 'visible': 1, 'name': 'Reporter'},
                 'Assignee': {'index': 7, 'visible': 1, 'name': 'Assignee'},
                 'Labels': {'index': 8, 'visible': 1, 'name': 'Labels'},
                 'Due Date': {'index': 9, 'visible': 1, 'name': 'Due Date'},
                 'Parent': {'index': 10, 'visible': 1, 'name': 'Parent'},
                 'Priority': {'index': 11, 'visible': 1, 'name': 'Priority'},
                 'Description': {'index': 14, 'visible': 1, 'name': 'Description'},
                 'Created': {'index': 12, 'visible': 1, 'name': 'Created'},
                 'Updated': {'index': 13, 'visible': 1, 'name': 'Updated'},
                 }

# For Aggregated Excel Sheet, if applicable
aggregated_sheet = {'name': 'Aggregated',
                    'visible': 0
                    }

# Program configs - defaults
override_checkbox = 0
config_file = 'config.json'
report_name = 'JIRA Export.xlsx'
zoom_scale = 90
jira_sheet_title = 'Items from JIRA'
JIRA_BASE_URL = 'https://issuetracking.jira.com/jira'
jql = 'issuetype = Story'

# Creation of Excel in-memory
wb = Workbook()

# Name of the reporting Excel with creation time in UTC timezone
time_format = "%Y-%m-%dT%H:%M:%S"
now = datetime.datetime.strptime(str(datetime.datetime.utcnow().isoformat()).split(".", 1)[0], time_format)
default_output_excel = report_name.split('.xls')[0] + '_' + str(now).replace(':', '-').replace(' ', '_') + '_UTC.xlsx'
output_excel = default_output_excel
JIRAs_column = 0

# Formatting functions
def get_str_from_lst(lst, sep=','):
    """This function returns list as comma separated string - for exporting in excel"""
    st = ""
    for l in lst:
        if l != '':
            st += str(l).strip() + sep + ' '
    st = st[0:-2]
    return st


def get_visible_columns():
    visible_columns = []
    for v in excel_columns.values():
        if v['visible'] == 1:
            visible_columns.append(v['index'])
    return visible_columns


# Working with Excel files
def select_output_file():
    global output_excel
    dir_name = os.getcwd()
    output_excel = asksaveasfilename(initialdir=dir_name, title="Select file", filetypes=(("JIRA list with details", ".xlsx"), ("all files", "*.*")))
    if not output_excel.endswith('.xlsx'):
        output_excel += '.xlsx'
    out_xls.delete(0, END)
    out_xls.insert(0, output_excel)


def create_excel_sheet(sheet_data, title):
    global JIRA_BASE_URL
    wb.create_sheet(title)
    ws = wb.get_sheet_by_name(title)
    
    start_column = 1
    start_row = 1
    visible_cols = get_visible_columns()

    # Creating Excel sheet based on data
    for i in range(len(sheet_data)):
        for y in range(len(sheet_data[i])):
            if y in visible_cols:
                if (y == excel_columns['ID']['index'] and start_row != 1 and sheet_data[i][y] is not None and sheet_data[i][y] != ''):
                    ws.cell(row=start_row, column=start_column+y).hyperlink = JIRA_BASE_URL + '/browse/' + sheet_data[i][y]
                    ws.cell(row=start_row, column=start_column+y).font = hyperlink
                ws.cell(row=start_row, column=start_column+y).value = sheet_data[i][y]
        start_row += 1
    
    for y in range(1, ws.max_column+1):
        ws.cell(row=1, column=y).fill = header_fill
        ws.cell(row=1, column=y).font = header_font

    ws.title = title


def remove_columns():
    start_column = 1
    for ws in wb.worksheets:
        # Removing columns
        cols_to_remove = []
        for v in excel_columns.values():
            if v['visible'] == 0:
                cols_to_remove.append(start_column+v['index'])
        cols_to_remove.sort(reverse=True)
        for z in cols_to_remove:
            ws.delete_cols(z)
        # Updating width of columns
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            if length > 80:
                ws.column_dimensions[column_cells[0].column_letter].width = 80
            else:
                ws.column_dimensions[column_cells[0].column_letter].width = length + 4

        ws.auto_filter.ref = ws.dimensions


def save_file():
    global input_excel, output_excel
    # Saving Excel file and removing not required sheets
    sheet_names = wb.sheetnames
    for s in sheet_names:
        ws = wb.get_sheet_by_name(s)
        if ws.dimensions == 'A1:A1':
            wb.remove_sheet(wb[s])
    try:
        if output_excel == '':
            time_format = "%Y-%m-%dT%H:%M:%S"
            now = datetime.datetime.strptime(str(datetime.datetime.utcnow().isoformat()).split(".", 1)[0], time_format)
            output_excel = input_excel.split('.xls')[0] + '_' + str(now).replace(':', '-').replace(' ', '_') + '_UTC.xlsx'
        set_zoom(output_excel)
        print("File \"", output_excel, "\" successfully generated.", sep='')
        print()
        sleep(2)
        exit()
    except Exception as e:
        print()
        print("ERROR:", e)
        os.system("pause")
        exit()


def set_zoom(file, zoom_scale=90):
    for ws in wb.worksheets:
        ws.sheet_view.zoomScale = zoom_scale
    wb.save(file)


def get_columns():
    global excel_columns
    excel_columns_list = ['' for i in range(len(excel_columns))]
    for v in excel_columns.values():
        excel_columns_list[v['index']] = v['name']
    return excel_columns_list


def get_issues_by_jql(jql):
    """This function returns list of JIRA keys for provided list of JIRA JQL queries"""
    auth_jira = JIRA(JIRA_BASE_URL)
    issues, items = ([], [])
    start_idx, block_num, block_size = (0, 0, 100)
    while True:
        start_idx = block_num * block_size
        tmp_issues = auth_jira.search_issues(jql_str=jql, startAt=start_idx, maxResults=block_size, fields='key, issuetype')
        if len(tmp_issues) == 0:
            # Retrieve issues until there are no more to come
            break
        issues.extend(tmp_issues)
        block_num += 1
    
    items = list(set([i.key for i in issues]))
    items = [[i] for i in items]
    return items


def main_program():
    global output_excel, issues, input_excel, report_name, jql, aggregated_sheet, JIRA_BASE_URL
    output_excel = out_xls.get().strip()
    if not output_excel.endswith('.xlsx'):
        output_excel += '.xlsx'
    report_name = output_excel
    JIRA_BASE_URL = jira_instance.get().strip()
    jql = j_query.get().strip()
    config_file = conf.get().strip().split('.json')[0] + '.json'
    if override_checkbox == 1:
        save_config(configfile=config_file)
    master.destroy()
    if JIRA_BASE_URL == '':
        print("JIRA URL has not been entered. Program stopped.")
        os.system("pause")
        exit()
    try:
        jira = JIRA(JIRA_BASE_URL)
    except Exception as er:
        print("Exception with JIRA connection: {}".format(er))
        print("Program stopped.")
        os.system("pause")
        exit()

    issues = {}
    if jql != '':
        issues[jira_sheet_title] = get_issues_by_jql(jql)
    else:
        print("JQL has not been entered. Program stopped.")
        os.system("pause")
        exit()

    updated_issues = {}
    for k, v in issues.items():
        updated_issues[k] = []
        try:
            updated_issues[k].append(get_columns())
        except:
            pass
    
    print("Metadata for Issues is downloading from JIRA...")
    
    for k, v in issues.items():
        n = 0
        print("Total JIRA issues for '{}' sheet to be processed: {}".format(k, len(v)))
        for i in range(len(v)):
            n += 1
            details = ['' for i in range(len(excel_columns))]
            if v[i][JIRAs_column] != '':
                details[excel_columns['ID']['index']] = v[i][JIRAs_column]
                try:
                    issue = jira.issue(v[i][JIRAs_column])
                    details[excel_columns['Type']['index']] = issue.fields.issuetype.name
                    details[excel_columns['Summary']['index']] = issue.fields.summary
                    details[excel_columns['Description']['index']] = str('' if issue.fields.description is None else issue.fields.description.replace('\\\\', '_x000D_'))
                    # Components update
                    components = get_str_from_lst([i.name for i in issue.fields.components])
                    details[excel_columns['Components']['index']] = components
                    # Labels update
                    labels = get_str_from_lst([i for i in issue.fields.labels])
                    details[excel_columns['Labels']['index']] = labels
                    details[excel_columns['Status']['index']] = issue.fields.status.name
                    details[excel_columns['fixVersions']['index']] = get_str_from_lst([i.name for i in issue.fields.fixVersions])
                    details[excel_columns['Reporter']['index']] = get_str_from_lst('' if issue.fields.reporter is None else issue.fields.reporter.displayName for i in range(1))
                    details[excel_columns['Assignee']['index']] = get_str_from_lst('' if issue.fields.assignee is None else issue.fields.assignee.displayName for i in range(1))
                    details[excel_columns['Due Date']['index']] = issue.fields.duedate
                    try:
                        parent = get_str_from_lst('' if issue.fields.parent is None else issue.fields.parent)
                    except:
                        parent = ''
                    details[excel_columns['Parent']['index']] = parent
                    details[excel_columns['Priority']['index']] = issue.fields.priority.name
                    details[excel_columns['Created']['index']] = issue.fields.created.split('T')[0]
                    details[excel_columns['Updated']['index']] = issue.fields.updated.split('T')[0]
                    # Extend list for Excel export
                    updated_issues[k].append(details)
                except Exception as e:
                    print("Exception '{}' for retrieving JIRA details for JIRA_ID: {}".format(e, v[i][JIRAs_column]))
            if n % 100 == 0:
                print("Processed {} issues out of {} so far.".format(n, len(v)))

    print("Metadata for issues was successfully downloaded from JIRA.")
    print()
    
    # First sheet - all data aggregated
    all_issues = [get_columns()]
    unique_issues = []
    duplicates = set()
    for k, v in updated_issues.items():
        dd = []
        for i in range(1, len(v)):
            if v[i][excel_columns['ID']['index']] not in unique_issues:
                unique_issues.append(v[i][excel_columns['ID']['index']])
                temp = v[i]
                temp.append(k)
                dd.append(temp)
            duplicates.add((v[i][excel_columns['ID']['index']], k))
        all_issues.extend(dd)

    if aggregated_sheet['visible'] == 1:
        create_excel_sheet(all_issues, aggregated_sheet['name'])
    
    # Existing sheets - placed after first one
    for k, v in updated_issues.items():
        create_excel_sheet(v, k)
    
    remove_columns()
    
    # Saving Excel file
    save_file()


def load_config(configfile=config_file):
    global JIRA_BASE_URL, excel_columns, aggregated_sheet, jira_sheet_title, report_name, jql, output_excel
    if os.path.exists(configfile) is True:
        try:
            with open(configfile) as json_data_file:
                data = json.load(json_data_file)
            for k, v in data.items():
                if k == 'JIRA_BASE_URL':
                    JIRA_BASE_URL = v
                elif k == 'jql':
                    jql = v
                elif k == 'excel_columns':
                    excel_columns = v
                # elif k == 'aggregated_sheet': #TODO
                #     aggregated_sheet = v
                elif k == 'jira_sheet_title':
                    jira_sheet_title = v
                elif k == 'report_name':
                    report_name = v
                    output_excel = v
            print("Configuration file '{}' has been successfully loaded.".format(configfile))
        except Exception as er:
            print("Failed to load file '{}', due to Exception: '{}'".format(configfile, er))
            if override_checkbox == 1:
                print("Configuration file is corrupted. Default '{}' would be created instead.".format(configfile))
                print()
                save_config()
    else:
        print("Config File not found. Default '{}' would be created.".format(configfile))
        print()
        save_config()


def save_config(configfile=config_file):
    data = {'JIRA_BASE_URL': JIRA_BASE_URL,
            'jql': jql,
            'jira_sheet_title': jira_sheet_title,
            # 'aggregated_sheet': aggregated_sheet,
            'report_name': report_name,
            'zoom_scale': zoom_scale,
            'excel_columns': excel_columns,
            }
    if configfile == '.json':
        time_format = "%Y-%m-%dT%H:%M:%S"
        now = datetime.datetime.strptime(str(datetime.datetime.utcnow().isoformat()).split(".", 1)[0], time_format)
        configfile = 'config' + '_' + str(now).replace(':', '-').replace(' ', '_') + '_UTC.json'
    try:
        with open(configfile, 'w') as outfile:
            json.dump(data, outfile)
    except PermissionError as er:
        print("ERROR: File '{}' has been opened for editing and can't be saved. Exception: {}".format(configfile, er))
        return
    except Exception as ex:
        print("ERROR: '{}' can't be saved. Exception: {}".format(configfile, ex))
        return
    
    print("Config file '{}' has been created.".format(configfile))
    print()


def add_aggrigate_list(*args):
    global aggregated_sheet
    aggregated_sheet['visible'] = add_aggregated.get()


def change_override(*args):
    global override_checkbox
    override_checkbox = override.get()


# Open File dialog to open config file in the same location and refresh UI values
def open_file():
    global config_file, input_excel, report_name, aggregated_sheet
    dir_name = os.getcwd()
    filename = askopenfilename(initialdir=dir_name, title="Select file", filetypes=(("Configuration File", "*.json"),
                                                                                    ("all files", "*.*")))
    if filename != '':
        config_file = filename
        conf.delete(0, END)
        conf.insert(0, config_file.split('/')[-1])
        load_config(configfile=config_file)
        j_query.delete(0, END)
        j_query.insert(0, jql)
        jira_instance.delete(0, END)
        jira_instance.insert(0, JIRA_BASE_URL)
        out_xls.delete(0, END)
        out_xls.insert(0, report_name)
        add_aggregated.set(aggregated_sheet['visible'])
    else:
        print("No config file was found.")


# ------------------ MAIN PROGRAM -----------------------------------
# load_config()
print("Program started. Please DO NOT CLOSE this window...")
print()

master = tk.Tk()
Title = master.title("JIRA Export Tool")
tk.Label(master, text="Please enter JIRA instance URL and required JQL for export. Specify Report File Name.", font=("Helvetica", 10)).grid(row=0, column=0, pady=5, columnspan=3)

tk.Label(master, text="JIRA Instance URL:").grid(row=2, column=0, pady=2, padx=3)
tk.Label(master, text="JQL for Export:", font=("Helvetica", 9)).grid(row=3, column=0, pady=2, padx=3)
tk.Label(master, text="Report File:").grid(row=4, column=0, pady=2, padx=3)

jira_instance = tk.Entry(master, width=70)
jira_instance.insert(END, JIRA_BASE_URL)
jira_instance.grid(row=2, column=1, padx=0, sticky=W, columnspan=2)

j_query = tk.Entry(master, width=70)
j_query.insert(END, jql)
j_query.grid(row=3, column=1, pady=5, columnspan=2, sticky=W)

out_xls = tk.Entry(master, width=50)
out_xls.insert(END, output_excel)
out_xls.grid(row=4, column=1, padx=0, sticky=W)

tk.Button(master, text='Browse', command=select_output_file, width=15).grid(row=4, column=2, pady=3, padx=8)

tk.Label(master, text="________________________________________________________________________________________________").grid(row=5, columnspan=3)

add_aggregated = IntVar(value=aggregated_sheet['visible'])
# Checkbutton(master, text="Add Aggregated list as the first one", font=("Helvetica", 9, "italic"), variable=add_aggregated).grid(row=6, sticky=W, padx=20, columnspan=3, pady=0)
# add_aggregated.trace('w', add_aggrigate_list)

override = IntVar()
Checkbutton(master, text="Save current values in the file for future use:", font=("Helvetica", 9, "italic"), variable=override).grid(row=7, sticky=W, padx=20, columnspan=2, pady=0)
override.trace('w', change_override)

conf = tk.Entry(master, width=20)
conf.insert(END, config_file)
conf.grid(row=7, column=0, padx=135, columnspan=3, sticky=E)

tk.Button(master, text='Reload configs', command=open_file, width=15).grid(row=7, column=2, pady=0, padx=15, columnspan=1, stick=W)

tk.Button(master, text='Quit', font=("Helvetica", 9, "bold"), command=master.quit, width=20, heigh=2).grid(row=8, column=0, pady=5, padx=60, columnspan=2, sticky=W)
tk.Button(master, text='Generate Report', font=("Helvetica", 9, "bold"), state='active', command=main_program, width=20, heigh=2).grid(row=8, column=1, pady=10, padx=60, columnspan=2, sticky=E)

tk.mainloop()
