# This Migration Tool has been created by Dmitry Elsakov
# The main source code has been created over weekends and distributed over GPL-3.0 License
# The license details could be found here: https://github.com/delsakov/JIRA_Tools/
# Please do not change notice above and copyright

from jira import JIRA
from atlassian import jira
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Protection
from sys import exit
import logging
import io
import traceback
from tkinter import *
from tkinter.filedialog import askopenfilename, askdirectory
import tkinter as tk
import os
import datetime
import isodate
import time
from time import sleep
import requests
import urllib.parse
from requests.auth import HTTPBasicAuth
import urllib.request
import urllib3
from bs4 import BeautifulSoup
import json
import shutil
import threading
import concurrent.futures
from itertools import zip_longest

# Migration Tool properties
current_version = '5.3'
config_file = 'config.json'

# JIRA Default configuration
JIRA_BASE_URL_OLD = ''
project_old = ''
JIRA_BASE_URL_NEW = ''
project_new = ''
template_project = ''
new_project_name = ''
team_project_prefix = ''
dummy_parent = ''
read_only_scheme_name = 'Read-Only'
protection_password = 'dmitry*'
excel_locked = 1
verify = True
max_number_for_dummy_parent_search = 20000
retry_number_allowed = 12

# JIRA API configs
JIRA_sprint_api = '/rest/agile/1.0/sprint/'
JIRA_core_api = '/rest/api/2/issue/'
JIRA_team_api = '/rest/teams-api/1.0/team'
JIRA_board_api = '/rest/agile/1.0/board/'
JIRA_status_api = '/rest/api/2/status'
JIRA_fields_api = '/rest/api/2/field'
JIRA_attachment_api = '/rest/api/2/attachment/'
JIRA_users_api = '/rest/api/2/user?username={}'
JIRA_versions_api = '/rest/api/2/version/{}'
JIRA_components_api = '/rest/api/2/component/{}'
JIRA_component_api = '/rest/api/2/component'
JIRA_create_users_api = '/rest/api/2/user'
JIRA_imported_api = '/rest/jira-importers-plugin/1.0/importer/json'
JIRA_labelit_api = '/rest/labelit/1.0/items'
JIRA_workflowscheme_api = '/rest/projectconfig/1/workflowscheme/{}'
JIRA_workflow_api = '/rest/projectconfig/1/workflow?workflowName={}&projectKey={}'
JIRA_create_project_api = '/rest/scriptrunner/latest/custom/createProject'
JIRA_update_team_api = '/rest/scriptrunner/latest/custom/updateTeam'
JIRA_update_parent_link_api = '/rest/scriptrunner/latest/custom/updateParentLink'
JIRA_get_permissions_scheme_api = '/rest/api/2/permissionscheme'
JIRA_assign_permission_scheme_api = '/rest/api/2/project/{}/permissionscheme'
headers = {"Content-type": "application/json", "Accept": "application/json"}

# Disabling WARNINGS - hide them from console. All warnings will be covered by program itself.
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
urllib3.disable_warnings(urllib3.exceptions.HTTPWarning)
urllib3.disable_warnings(urllib3.exceptions.ConnectionError)
urllib3.disable_warnings()

# Sprints configs
old_sprints = {}
new_sprints = {}
old_board_id = 0
new_board_id = 0
default_board_name = 'Shared Sprints'

# Excel configs
header_font = Font(color='00000000', bold=True)
header_fill = PatternFill(fill_type="solid", fgColor="8db5e2")
hyperlink = Font(underline='single', color='0563C1')
project_tab_color = '32CD32'  # Green
mandatory_tab_color = 'FA8072'  # Red
optional_tab_color = 'F4A460'  # Amber
mandatory_template_tabs = ['Project', 'Issuetypes', 'Fields', 'Statuses', 'Priority', 'Links']
hide_tabs = False
zoom_scale = 100
wb = Workbook()
default_validation = {}
excel_columns_validation_ranges = {'0': 'A2:A1048576',
                                   '1': 'B2:B1048576',
                                   '2': 'C2:C1048576',
                                   '3': 'D2:D1048576',
                                   '4': 'E2:E1048576',
                                   '5': 'F2:F1048576',
                                   '6': 'G2:G1048576',
                                   '7': 'H2:H1048576',
                                   '8': 'I2:I1048576',
                                   '9': 'J2:J1048576',
                                   '10': 'K2:K1048576',
                                   '11': 'L2:L1048576',
                                   '12': 'M2:M1048576',
                                   '13': 'N2:N1048576',
                                   '14': 'O2:O1048576',
                                   '15': 'P2:P1048576',
                                   '16': 'Q2:Q1048576',
                                   '17': 'R2:R1048576',
                                   '18': 'S2:S1048576',
                                   '19': 'T2:T1048576',
                                   '20': 'U2:U1048576',
                                   '21': 'V2:V1048576',
                                   '22': 'W2:W1048576',
                                   '23': 'X2:X1048576',
                                   '24': 'Y2:Y1048576',
                                   '25': 'Z2:Z1048576',
                                   '26': 'AA2:AA1048576',
                                   '27': 'AB2:AB1048576',
                                   '28': 'AC2:AC1048576',
                                   '29': 'AD2:AD1048576',
                                   '30': 'AE2:AE1048576',
                                   '31': 'AF2:AF1048576',
                                   '32': 'AG2:AG1048576',
                                   }
process_complete_folder = 'Complete'
process_partially_complete_folder = 'Partially_Complete'
validation_template_error_folder = 'Validation_Failed'
processing_error_folder = 'Other_Failed'

# Migration configs
temp_dir_name = 'Attachments_Temp/'
logs_folder = 'Migration_Tool_Logs/'
log_file = 'MIGRATION_TOOL_OUT.txt'
mapping_file = ''
default_configuration_file = ''
jira_system_fields = ['Summary', 'Sprint', 'Epic Link', 'Epic Name', 'Story Points', 'Parent Link', 'Flagged', 'Target start', 'Target end']
jira_system_skip_fields = ['Issue Type', 'Project', 'Linked Issues', 'Attachment', 'Parent', 'worklog', 'Time Tracking']
additional_mapping_fields = ['Description', 'Labels', 'Due Date', 'Target start', 'Target end']
skipping_tabs_for_default_mapping = ['Project', 'Issuetypes', 'Fields', 'Statuses', 'Priority']
old_fields_ids_mapping = {}
limit_migration_data = 0  # 0 if all
save_validation_details = 0  # to save validation details in the .json file
start_jira_key = 1
dummy_process = 1
create_remote_link_for_old_issue = 0
username, password = ('', '')
auth = (username, password)
items_lst = {}
sub_tasks = {}
old_sub_tasks = {}
new_issues_ids = {}
teams = {}
total_data = {}
total_data["projects"] = []
total_data["users"] = []
total_data["links"] = []
issues_lst = []
issues_set = set()
users_set = set()
users = []
already_migrated_set = set()
last_updated_date = 'YYYY-MM-DD'
refresh_issuetypes = 'ALL'
processing_jira_jql = ''
updated_issues_num = 0
threads = 1
migrated_text = 'Migrated to'
verbose_logging = 0
recently_updated_days = 365
reconciliation_updated_days = 365
shifted_key_val = 100
shifted_by = 1000
migrate_fixversions_check = 1
migrate_components_check = 1
migrate_sprints_check = 1
migrate_attachments_check = 1
migrate_comments_check = 1
migrate_statuses_check = 1
migrate_links_check = 1
migrate_teams_check = 1
migrate_metadata_check = 1
change_configuration_flag = 1
bulk_processing_flag = 0
processing_error = 0
validation_template_error = 0
validation_error = 0
process_complete = 0
process_partially_complete = 0
credentials_saved_flag = 0
force_update_flag = 0
delete_dummy_flag = 0
skip_migrated_flag = 1
last_updated_days_check = 1
including_dependencies_flag = 1
force_sprints_update_flag = 0
merge_projects_flag = 0
merge_projects_start_flag = 0
set_source_project_read_only = 0
json_importer_flag = 1
including_users_flag = 1
process_only_last_updated_date_flag = 0
refresh_already_migrated_flag = 0
replace_complete_statuses_flag = 1
check_template_flag = 1
skip_existing_issuetypes_validation_flag = 1
clear_additional_configuration_flag = 0
process_reconciliation_flag = 0
process_reconciliation_excel_flag = 0
control_logic_flag = 0
retry_logic_flag = 1
override_template_flag = 0
remaining_previous = 0
previous_JIRA_BASE_URL_NEW = ''
supported_issuetypes = ''
recently_updated = ''

# Required for creation JSON file - total_data have to be dumped in JSON file for processing from UI.
json_thread_lock = threading.Lock()
json_current_size = 0
sleep_count = 0
default_sleep_time = 10
multiple_json_data_processing = 0
previous_multiple_json_data_processing = 0
json_files_autoupload = 0
json_file_part_num = 1
failed_issues = set()
migrated_issues_lst = []
already_processed_json_importer_issues = set()
already_processed_users = set()
processed_issues_set = set()
teams_to_be_added_set = set()
skipped_issuetypes = []
processed_issuetypes = []
total_processed = 0
max_json_file_size = 10
pool_size = 1

# Concurrent processing configs
teams_thread_lock = threading.Lock()
default_max_retries = 5
max_retries = default_max_retries

# Mappings
issuetypes_mappings = {}
fields_mappings = {}
status_mappings = {}
field_value_mappings = {}
link_mappings = {}

# Transitions mapping - for status changes
old_transitions = {}
new_transitions = {}
old_statuses = {}
new_statuses = {}


# Functions list
def read_excel(file_path=mapping_file, columns=0, rows=0, start_row=2):
    """Function for reading Mapping Excel file and saves all mappings for further processing."""
    global issuetypes_mappings, fields_mappings, status_mappings, field_value_mappings, verbose_logging, project_old
    global JIRA_BASE_URL_OLD, JIRA_BASE_URL_NEW,  project_new, link_mappings, skipped_issuetypes, processed_issuetypes
    global override_template_flag
    
    print("[START] Mapping file '{}' is opened for processing.".format(file_path))
    mapping_type = 1
    try:
        with open(file_path, 'rb') as f:
            in_mem_file = io.BytesIO(f.read())
        df = load_workbook(in_mem_file, read_only=True, data_only=True, keep_vba=True, keep_links=True)
        excel_sheet_names = df.get_sheet_names()
        for excel_sheet_name in excel_sheet_names:
            value_mappings = {}
            df1 = df.get_sheet_by_name(excel_sheet_name)
            row_count = rows
            col_count = columns
            if rows == 0:
                row_count = df1.max_row
            if columns == 0:
                col_count = df1.max_column
            empty_row = ['' for i in range(col_count)]
            for row in df1.iter_rows(min_row=start_row, max_row=row_count, max_col=col_count):
                d = []
                for v in row:
                    val = v.value
                    if val is None:
                        val = ""
                    else:
                        val = str(val)
                    d.append(val)
                if set(d) != set(empty_row) and d != []:
                    if excel_sheet_name == 'Project':
                        JIRA_BASE_URL_OLD = d[0].strip('/').strip()
                        JIRA_BASE_URL_NEW = d[2].strip('/').strip()
                        project_old = d[1].strip()
                        project_new = d[3].strip()
                        if d[4] == 'Target -> Source':
                            mapping_type = 0
                        break
                    elif excel_sheet_name == 'Issuetypes':
                        if mapping_type == 0:
                            issuetypes_mappings[d[2]] = {"hierarchy": d[1], "issuetypes": d[3].split(',')}
                        else:
                            if d[0] == '':
                                break
                            elif d[1] in issuetypes_mappings.keys():
                                issuetypes_mappings[d[1]]["issuetypes"].append(d[0])
                            else:
                                issuetypes_mappings[d[1]] = {"hierarchy": '2', "issuetypes": [d[0]]}
                            if override_template_flag == 0:
                                if d[1] == '':
                                    skipped_issuetypes.append(d[0])
                                else:
                                    processed_issuetypes.append(d[0])
                    elif excel_sheet_name == 'Links':
                        if mapping_type == 0:
                            link_mappings[d[0]] = [d[2]]
                        else:
                            if d[0] == '':
                                break
                            elif d[2] in link_mappings.keys():
                                link_mappings[d[2]].append(d[0])
                            else:
                                link_mappings[d[2]] = [d[0]]
                    elif excel_sheet_name == 'Statuses':
                        if mapping_type == 0:
                            for issuetype in d[0].split(','):
                                if issuetype.strip() not in status_mappings.keys():
                                    status_mappings[issuetype] = {d[1]: d[2].split(',')}
                                else:
                                    status_mappings[issuetype][d[1]] = d[2].split(',')
                        else:
                            if d[0] == '':
                                break
                            elif d[0] in status_mappings.keys():
                                if d[2] in status_mappings[d[0]].keys():
                                    status_mappings[d[0]][d[2]].append(d[1])
                                else:
                                    status_mappings[d[0]][d[2]] = [d[1]]
                            else:
                                status_mappings[d[0]] = {d[2]: [d[1]]}
                            if d[2] == '' and verbose_logging == 1:
                                print("[WARNING] The mapping of '{}' status for '{}' Issuetype not found. Default status would be used.".format(d[1], d[0]))
                    elif excel_sheet_name == 'Fields':
                        if mapping_type == 0:
                            for issuetype in d[0].split(','):
                                if issuetype not in fields_mappings.keys():
                                    fields_mappings[issuetype] = {d[1]: d[2].split(',')}
                                else:
                                    fields_mappings[issuetype][d[1]] = d[2].split(',')
                        else:
                            if d[0] == '':
                                break
                            elif d[0] in fields_mappings.keys():
                                if d[2] in fields_mappings[d[0]].keys():
                                    fields_mappings[d[0]][d[2]].append(d[1])
                                else:
                                    fields_mappings[d[0]][d[2]] = [d[1]]
                            else:
                                fields_mappings[d[0]] = {d[2]: [d[1]]}
                            if d[2] == '' and verbose_logging == 1:
                                print("[WARNING] The mapping of '{}' field for '{}' Issuetype not found. Field values will be dropped.".format(d[1], d[0]))
                    elif excel_sheet_name == 'Priority':
                        try:
                            if mapping_type == 0:
                                value_mappings[d[0]] = d[1].split(',')
                            else:
                                if d[1] not in value_mappings.keys():
                                    value_mappings[d[1]] = d[0].split(',')
                                else:
                                    value_mappings[d[1]].extend(d[0].split(','))
                        except:
                            print("[ERROR] Data on the sheet '{}' is invalid, Skipping...".format(excel_sheet_name))
                            continue
                    else:
                        if len(d) <= 2:
                            try:
                                if mapping_type == 0:
                                    value_mappings[d[0]] = d[1].split(';')
                                else:
                                    if d[1] not in value_mappings.keys():
                                        value_mappings[d[1]] = d[0].split(';')
                                    else:
                                        value_mappings[d[1]].extend(d[0].split(';'))
                            except:
                                print("[ERROR] Data on the sheet '{}' is invalid, Skipping...".format(excel_sheet_name))
                                continue
                        else:
                            try:
                                if mapping_type == 1:
                                    if d[1] + ' --> ' + d[2] not in value_mappings.keys():
                                        value_mappings[d[1] + ' --> ' + d[2]] = d[0].split(';')
                                    else:
                                        value_mappings[d[1] + ' --> ' + d[2]].extend(d[0].split(';'))
                            except:
                                print("[ERROR] Data on the sheet '{}' is invalid, Skipping...".format(excel_sheet_name))
                                continue
            
            if excel_sheet_name not in ['Project', 'Issuetypes', 'Statuses', 'Fields']:
                field_value_mappings[excel_sheet_name] = value_mappings
    except Exception as e:
        print("[ERROR] '{}' file not found. Mappings can't be processed. Error: '{}'".format(file_path, e))
        if verbose_logging == 1:
            print(traceback.format_exc())
        os.system("pause")
        exit()
    print("[END] Mapping data has been successfully processed.")
    print("")


def read_default_mappings_excel(file_path=default_configuration_file, columns=0, rows=0, start_row=2):
    """Function for reading Mapping Excel file and saves all mappings for further processing."""
    global field_value_mappings, link_mappings, verbose_logging, skipping_tabs_for_default_mapping, status_mappings
    global default_configuration_file, override_template_flag, issuetypes_mappings, fields_mappings
    global processed_issuetypes, skipped_issuetypes
    
    print("[START] Default Mapping file '{}' is opened for processing.".format(default_configuration_file))
    if override_template_flag == 1:
        issuetypes_mappings = {}
        fields_mappings = {}
        status_mappings = {}
        field_value_mappings['Priority'] = {}
    try:
        df = load_workbook(file_path, read_only=True, data_only=True, keep_vba=True, keep_links=True)
        excel_sheet_names = df.get_sheet_names()
        for excel_sheet_name in excel_sheet_names:
            value_mappings = {}
            df1 = df.get_sheet_by_name(excel_sheet_name)
            row_count = rows
            col_count = columns
            if rows == 0:
                row_count = df1.max_row
            if columns == 0:
                col_count = df1.max_column
            empty_row = ['' for i in range(col_count)]
            for row in df1.iter_rows(min_row=start_row, max_row=row_count, max_col=col_count):
                d = []
                for v in row:
                    val = v.value
                    if val is None:
                        val = ""
                    else:
                        val = str(val)
                    d.append(val)
                if set(d) != set(empty_row) and d != [] and excel_sheet_name not in skipping_tabs_for_default_mapping:
                    if excel_sheet_name == 'Links':
                        if d[0] == '':
                            break
                        elif d[2] in link_mappings.keys():
                            link_mappings[d[2]].append(d[0])
                        else:
                            link_mappings[d[2]] = [d[0]]
                    else:
                        if len(d) <= 2:
                            try:
                                if d[1] not in value_mappings.keys():
                                    value_mappings[d[1]] = d[0].split(';')
                                else:
                                    value_mappings[d[1]].extend(d[0].split(';'))
                            except:
                                print("[ERROR] Data on the sheet '{}' is invalid, Skipping...".format(excel_sheet_name))
                                continue
                        else:
                            try:
                                if d[1] + ' --> ' + d[2] not in value_mappings.keys():
                                    value_mappings[d[1] + ' --> ' + d[2]] = d[0].split(';')
                                else:
                                    value_mappings[d[1] + ' --> ' + d[2]].extend(d[0].split(';'))
                            except:
                                print("[ERROR] Data on the sheet '{}' is invalid, Skipping...".format(excel_sheet_name))
                                continue
                if excel_sheet_name == 'Links':
                    for n_field, o_fields in link_mappings.items():
                        link_mappings[n_field] = o_fields
                elif override_template_flag == 1 and excel_sheet_name in ['Issuetypes', 'Fields', 'Statuses', 'Priority']:
                    if excel_sheet_name == 'Issuetypes':
                        if d[0] == '':
                            break
                        elif d[1] in issuetypes_mappings.keys():
                            issuetypes_mappings[d[1]]["issuetypes"].append(d[0])
                        else:
                            issuetypes_mappings[d[1]] = {"hierarchy": '2', "issuetypes": [d[0]]}
                        if d[1] == '':
                            skipped_issuetypes.append(d[0])
                        else:
                            processed_issuetypes.append(d[0])
                    elif excel_sheet_name == 'Fields':
                        if d[0] == '':
                            break
                        elif d[0] in fields_mappings.keys():
                            if d[2] in fields_mappings[d[0]].keys():
                                fields_mappings[d[0]][d[2]].append(d[1])
                            else:
                                fields_mappings[d[0]][d[2]] = [d[1]]
                        else:
                            fields_mappings[d[0]] = {d[2]: [d[1]]}
                    elif excel_sheet_name == 'Statuses':
                        if d[0] == '':
                            break
                        elif d[0] in status_mappings.keys():
                            if d[2] in status_mappings[d[0]].keys():
                                status_mappings[d[0]][d[2]].append(d[1])
                            else:
                                status_mappings[d[0]][d[2]] = [d[1]]
                        else:
                            status_mappings[d[0]] = {d[2]: [d[1]]}
                    elif excel_sheet_name == 'Priority':
                        if d[1] not in field_value_mappings['Priority'].keys():
                            field_value_mappings['Priority'][d[1]] = d[0].split(',')
                        else:
                            field_value_mappings['Priority'][d[1]].extend(d[0].split(','))
                elif excel_sheet_name in field_value_mappings.keys():
                    for n_field, o_fields in value_mappings.items():
                        if n_field in field_value_mappings[excel_sheet_name].keys():
                            field_value_mappings[excel_sheet_name][n_field].extend(o_fields)
                        else:
                            field_value_mappings[excel_sheet_name][n_field] = o_fields
        print("[END] Default Mapping data has been successfully processed.")
    except Exception as e:
        print("[ERROR] '{}' file not found. Default Mappings can't be processed. Skipping... Error: '{}'".format(file_path, e))
        if verbose_logging == 1:
            print(traceback.format_exc())
    print("")


def get_transitions_per_issuetype(params):
    global JIRA_workflow_api, old_transitions, new_transitions, auth, headers, verify
    
    try:
        statuses_lst = []
        jira_url, workflow_name, issuetype, project, new = params
        url0 = jira_url + JIRA_workflow_api.format(urllib.parse.quote_plus(workflow_name), project)
        url1 = jira_url + '/' + JIRA_workflow_api.format(urllib.parse.quote_plus(workflow_name), project)
        r = requests.get(url0, auth=auth, headers=headers, verify=verify)
        if r.status_code == 200:
            workflow_string = r.content.decode('utf-8')
        else:
            r = requests.get(url1, auth=auth, headers=headers, verify=verify)
            workflow_string = r.content.decode('utf-8')
        workflow_data = json.loads(workflow_string)
        transition_details = []
        for status in workflow_data["sources"]:
            if len(status["targets"]) > 0:
                for target in status["targets"]:
                    transition_details.append([status["fromStatus"]["name"], target['transitionName'], target['toStatus']['name']])
            else:
                statuses_lst.append(status["fromStatus"]["name"])
                transition_details.append([status["fromStatus"]["name"], status["fromStatus"]["name"], ''])
        temp_transitions = []
        for transition in transition_details:
            if transition[2] == '':
                for missing_status in statuses_lst:
                    if transition[2] == '':
                        transition[2] = missing_status
                    else:
                        temp_transitions.append([missing_status, missing_status, missing_status])
        transition_details.extend(temp_transitions)
        if new is False:
            old_transitions[issuetype] = transition_details
        else:
            new_transitions[issuetype] = transition_details
        return (0, params)
    except:
        return (1, params)


def get_transitions(project, jira_url, new=False):
    global old_transitions, new_transitions, auth, migrate_statuses_check, headers, verify, verbose_logging
    global replace_complete_statuses_flag, JIRA_workflow_api, default_max_retries, max_retries
    
    print("[START] Retrieving Transitions and Statuses for {} '{}' project from JIRA.".format('Target' if new is True else 'Source', project))
    
    def get_workflows(project, jira_url, new):
        global sub_tasks, auth, old_sub_tasks, new_issues_ids, JIRA_workflowscheme_api
        url = jira_url + JIRA_workflowscheme_api.format(project)
        r = requests.get(url, auth=auth, headers=headers, verify=verify)
        workflow_schema_string = r.content.decode('utf-8')
        workflow_schema_details = json.loads(workflow_schema_string)
        workflows = {}
        issuetypes = {}
        for issuetype in workflow_schema_details['issueTypes']:
            issuetypes[issuetype['id']] = issuetype['name']
            if new is True and issuetype['subTask'] is True:
                sub_tasks[issuetype['name']] = issuetype['id']
            elif issuetype['subTask'] is True:
                old_sub_tasks[issuetype['name']] = issuetype['id']
            if new is True and issuetype['subTask'] is False:
                new_issues_ids[issuetype['name']] = issuetype['id']
        for workflow in workflow_schema_details['mappings']:
            workflows[workflow['name']] = [issuetypes[i] for i in workflow['issueTypes']]
        return workflows
    
    try:
        params = []
        for workflow_name, workflow_details in get_workflows(project, jira_url, new).items():
            for issuetype in workflow_details:
                params.append((jira_url, workflow_name, issuetype, project, new))
        max_retries = default_max_retries
        threads_processing(get_transitions_per_issuetype, params)
        
        if replace_complete_statuses_flag == 1:
            get_statuses(jira_url=jira_url, new=new)
        
        print("[END] Transitions and Statuses for {} '{}' project has been successfully retrieved.".format('Target' if new is True else 'Source', project))
        print("")
    except Exception as e:
        migrate_statuses_check = 0
        print("[WARNING] No PROJECT ADMIN right for the {} '{}' project. Statuses WILL NOT be updated / migrated.".format('Target' if new is True else 'Source', project))
        print("[ERROR] Transitions and Statuses can't be retrieved due to '{}'".format(e))
        if verbose_logging == 1:
            print(traceback.format_exc())


def get_hierarchy_config():
    global sub_tasks, issuetypes_mappings, issue_details_new
    
    for issuetype, details in issuetypes_mappings.items():
        try:
            if issuetype in sub_tasks.keys():
                issuetypes_mappings[issuetype]['hierarchy'] = '3'
            elif 'Epic Link' in issue_details_new[issuetype].keys():
                issuetypes_mappings[issuetype]['hierarchy'] = '2'
            elif 'Epic Name' in issue_details_new[issuetype].keys():
                issuetypes_mappings[issuetype]['hierarchy'] = '1'
            else:
                issuetypes_mappings[issuetype]['hierarchy'] = '0'
        except:
            print("[WARNING] '{}' Issue Type(s) mapped in mapping file to '{}'. Skipping...".format(details['issuetypes'], issuetype))
            print("")
    
    # Removing non-mapped items
    issuetypes_mappings.pop("", None)


def calculate_statuses(transitions):
    issuetypes_lst = []
    issuetype_statuses = {}
    for k, v in transitions.items():
        statuses_lst = []
        issuetypes_lst.append(k)
        for l in v:
            statuses_lst.append(l[0])
            statuses_lst.append(l[2])
        issuetype_statuses[k] = list(set(statuses_lst))
    statuses_lst = []
    for k, v in issuetype_statuses.items():
        for status in v:
            statuses_lst.append([k, status, ''])
    return statuses_lst, list(set(issuetypes_lst))


def set_project_as_read_only(jira_url, project):
    global read_only_scheme_name, JIRA_get_permissions_scheme_api, JIRA_assign_permission_scheme_api, headers, auth, verify
    
    read_only_scheme_id = None
    url = jira_url + JIRA_get_permissions_scheme_api
    r = requests.get(url, auth=auth, headers=headers, verify=verify)
    permission_schemes_string = r.content.decode('utf-8')
    permission_schemes_details = json.loads(permission_schemes_string)
    for p in permission_schemes_details["permissionSchemes"]:
        if read_only_scheme_name in p["name"]:
            read_only_scheme_id = p["id"]
            break
    
    url1 = jira_url + JIRA_assign_permission_scheme_api.format(project)
    body = json.dumps({"id": read_only_scheme_id})
    r = requests.put(url1, data=body, auth=auth, headers=headers, verify=verify)
    return r.status_code


def prepare_template_data():
    global old_transitions, new_transitions, issue_details_old, default_validation, jira_system_fields
    global JIRA_BASE_URL_OLD, project_old, JIRA_BASE_URL_NEW, project_new, migrate_statuses_check, json_importer_flag
    global additional_mapping_fields, jira_old, jira_new
    
    template_excel = {}
    old_statuses, new_statuses, old_issuetypes, new_issuetypes = ([], [], [], [])
    
    # Project details
    project_details = [['Source Project JIRA URL', 'Source Project JIRA Key', 'Target Project JIRA URL', 'Target Project JIRA Key', 'Template type']]
    project_details.append([JIRA_BASE_URL_OLD, project_old, JIRA_BASE_URL_NEW, project_new, 'Source -> Target'])
    
    # IssueTypes
    if migrate_statuses_check == 1:
        old_statuses, old_issuetypes = calculate_statuses(old_transitions)
        new_statuses, new_issuetypes = calculate_statuses(new_transitions)
    else:
        for issuetype in issue_details_old.keys():
            old_issuetypes.append(issuetype)
        old_issuetypes = list(set(old_issuetypes))
        for issuetype in issue_details_new.keys():
            new_issuetypes.append(issuetype)
        new_issuetypes = list(set(new_issuetypes))
    issue_types_map_lst = [['Source Issue type', 'Target Issue Type']]
    for o_it in old_issuetypes:
        issue_types_map_lst.append([o_it, ''])
    
    # Fields
    fields_map_lst = [['Source Issue Type', 'Source Field Name', 'Target Field Name']]
    for issuetype, fields in issue_details_old.items():
        for field, details in fields.items():
            if details['custom'] is True and field not in jira_system_fields:
                fields_map_lst.append([issuetype, field, ''])
        # Add IssueType Name for mapping
        fields_map_lst.append([issuetype, issuetype + ' issuetype.name', ''])
        # Add IssueType Name for mapping
        fields_map_lst.append([issuetype, issuetype + ' issuetype.status', ''])
    
    new_fields_val = additional_mapping_fields[:]
    for issuetype, fields in issue_details_new.items():
        for field, details in fields.items():
            if details['custom'] is True and field not in jira_system_fields:
                new_fields_val.append(field.title())
    try:
        f_val = list(set(new_fields_val[:]))
        f_val.sort()
    except:
        f_val = list(new_fields_val[:])
    
    # Statuses
    statuses_map_lst = []
    if migrate_statuses_check == 1:
        statuses_map_lst = [['Source Issue Type', 'Source Status', 'Target Status']]
        statuses_map_lst.extend(old_statuses)
    
    # Priority
    priority_map_lst = [['Source Priority', 'Target Priority']]
    priority_old_lst = []
    priority_new_lst = []
    for field_values in issue_details_old.values():
        if 'Priority' in field_values.keys():
            for p in field_values['Priority']['allowed values']:
                priority_old_lst.append(p)
    priority_old_lst = list(set(priority_old_lst))
    for priority in priority_old_lst:
        priority_map_lst.append([priority, ''])
    for field_values in issue_details_new.values():
        if 'Priority' in field_values.keys():
            for p in field_values['Priority']['allowed values']:
                priority_new_lst.append(p)
    try:
        pr_val = list(set(priority_new_lst[:]))
        pr_val.sort()
    except:
        pass
    
    # Issue Linkage Types
    links_map_lst = [['Source Link Name', 'Sourse Link Details', 'Target Link Name']]
    links_new_lst = []
    for link in jira_new.issue_link_types():
        links_new_lst.append(link.name)
    try:
        l_val = list(set(links_new_lst[:]))
        l_val.sort()
    except:
        pass
    for link in jira_old.issue_link_types():
        links_map_lst.append([link.name, link.inward + ' / ' + link.outward, link.name if link.name in links_new_lst else ''])
    
    # Combine all data under one dictionary
    template_excel['Project'] = project_details
    template_excel['Issuetypes'] = issue_types_map_lst
    template_excel['Fields'] = fields_map_lst
    if migrate_statuses_check == 1:
        template_excel['Statuses'] = statuses_map_lst
    template_excel['Priority'] = priority_map_lst
    template_excel['Links'] = links_map_lst
    
    # Other fields
    for field_values in issue_details_new.values():
        for field_name, field_data in field_values.items():
            if field_name not in template_excel.keys() and field_name not in jira_system_fields:
                if field_data['type'] != 'option-with-child':
                    new_field_map_lst = [["Source '{}'".format(field_name), "Target '{}'".format(field_name)]]
                else:
                    new_field_map_lst = [["Source '{}'".format(field_name), "Target Level 1 of '{}'".format(field_name), "Target Level 2 of '{}'".format(field_name)]]
                if field_data['custom'] is True and field_data['allowed values'] is not None:
                    for value in field_data['allowed values']:
                        if field_data['type'] != 'option-with-child':
                            new_field_map_lst.append(['', value])
                        else:
                            new_field_map_lst.append(['', value[0], value[1]])
                    
                    template_excel[field_name] = new_field_map_lst
    
    # Calculate and update validation
    if migrate_statuses_check == 1:
        new_statuses_val = []
        for i in new_statuses:
            new_statuses_val.append(i[1])
        try:
            st_val = list(set(new_statuses_val[:]))
            st_val.sort()
        except:
            pass
        default_validation['Statuses'] = '"' + get_str_from_lst(st_val, spacing='') + '"'
    default_validation['Fields'] = '"' + get_str_from_lst(f_val, spacing='', stripping=False) + '"'
    new_issuetypes_val = []
    for i in new_issuetypes:
        new_issuetypes_val.append(i)
    try:
        i_val = list(set(new_issuetypes_val[:]))
        i_val.sort()
    except:
        pass
    default_validation['Issuetypes'] = '"' + get_str_from_lst(i_val, spacing='') + '"'
    default_validation['Priority'] = '"' + get_str_from_lst(pr_val, spacing='') + '"'
    default_validation['Links'] = '"' + get_str_from_lst(l_val, spacing='') + '"'
    
    return template_excel


def get_new_issuetype(old_issuetype):
    global issuetypes_mappings
    
    for issuetype, details in issuetypes_mappings.items():
        if old_issuetype in details['issuetypes']:
            return issuetype


def get_issues_by_jql(jira, jql, types=None, sprint=None, migrated=None, non_migrated=False, control=False, max_result=limit_migration_data):
    """This function returns list of JIRA keys for provided list of JIRA JQL queries"""
    global items_lst, limit_migration_data, verbose_logging, max_retries, default_max_retries, already_migrated_set
    global issues_lst, issues_set, issuetypes_mappings
    
    def sprint_update(param):
        global items_lst, old_sprints, issue_details_old, already_migrated_set, JIRA_board_api, headers
        global old_fields_ids_mapping
        
        jira, types, non_migrated, jql, control, start_idx, max_res = param
        try:
            sprint_field_id = issue_details_old['Story']['Sprint']['id']
        except:
            sprint_field_id = old_fields_ids_mapping['Sprint']
        issues = jira.search_issues(jql_str=jql, startAt=start_idx, maxResults=max_res, json_result=False, fields=eval("'issuetype," + sprint_field_id + "'"))
        try:
            for issue in issues:
                issue_sprints = eval('issue.fields.' + sprint_field_id)
                if issue_sprints is not None:
                    for sprint in issue_sprints:
                        sprint_id, name, state, start_date, end_date, board_id, board_name = ('', '', '', '', '', '', '')
                        for attr in sprint[sprint.find('[')+1:-1].split(','):
                            if 'id=' in attr:
                                sprint_id = attr.split('id=')[1]
                            if 'name=' in attr:
                                name = attr.split('name=')[1]
                            if 'state=' in attr:
                                state = attr.split('state=')[1]
                            if 'startDate=' in attr:
                                start_date = '' if attr.split('startDate=')[1] == '<null>' else attr.split('startDate=')[1]
                            if 'endDate=' in attr:
                                end_date = '' if attr.split('endDate=')[1] == '<null>' else attr.split('endDate=')[1]
                            if 'rapidViewId=' in attr:
                                board_id = '' if attr.split('rapidViewId=')[1] == '<null>' else attr.split('rapidViewId=')[1]
                                url = JIRA_BASE_URL_NEW + JIRA_board_api + str(board_id)
                                try:
                                    r = requests.get(url, auth=auth, headers=headers)
                                    board_details = r.content.decode('utf-8')
                                    board_data = json.loads(board_details)
                                    board_name = board_data["name"]
                                except:
                                    pass
                        if name not in old_sprints.keys():
                            old_sprints[name] = {"id": sprint_id, "startDate": start_date, "endDate": end_date, "state": state.upper(), "originBoardName": board_name}
                if (control is False and (types is None or issue.key in already_migrated_set)
                    or (control is True and issue.key not in already_migrated_set)):
                    continue
                elif issue.fields.issuetype.name not in items_lst.keys():
                    items_lst[issue.fields.issuetype.name] = set()
                    items_lst[issue.fields.issuetype.name].add(issue.key)
                else:
                    items_lst[issue.fields.issuetype.name].add(issue.key)
            return (0, param)
        except:
            return (1, param)
    
    def issue_list_update(param):
        global items_lst, already_migrated_set, jira_old, project_new, project_old, verbose_logging
        global migrate_metadata_check, json_importer_flag
        
        jira, types, non_migrated, jql, control, start_idx, max_res = param
        issues = jira.search_issues(jql_str=jql, startAt=start_idx, maxResults=max_res, json_result=False, fields='issuetype')
        
        try:
            for issue in issues:
                if non_migrated is True:
                    non_migrated_key = get_shifted_key(issue.key.replace(project_new + '-', project_old + '-'), reversed=True)
                    try:
                        non_migrated_issue = jira_old.issue(non_migrated_key)
                        if non_migrated_issue.fields.issuetype.name not in items_lst.keys():
                            items_lst[non_migrated_issue.fields.issuetype.name] = set()
                        items_lst[non_migrated_issue.fields.issuetype.name].add(non_migrated_key)
                    except:
                        if migrate_metadata_check == 1 and json_importer_flag == 1:
                            print("[WARNING] Issue '{}' has been removed from Source '{}' project. Removing from Target...".format(non_migrated_key, project_old))
                            delete_issue(issue.key)
                elif (control is False and issue.key in already_migrated_set) or (control is True and issue.key not in already_migrated_set):
                    continue
                elif issue.fields.issuetype.name not in items_lst.keys():
                    items_lst[issue.fields.issuetype.name] = set()
                    items_lst[issue.fields.issuetype.name].add(issue.key)
                else:
                    items_lst[issue.fields.issuetype.name].add(issue.key)
            return (0, param)
        except Exception as e:
            if verbose_logging == 1:
                print("[ERROR] JIRA can't perform the search for '{}' due to: '{}'".format(jql, e))
                print(traceback.format_exc())
            return (1, param)
    
    def migrated_update(param):
        global already_migrated_set, project_old, project_new
        
        jira, types, non_migrated, jql, control, start_idx, max_res = param
        issues = jira.search_issues(jql_str=jql, startAt=start_idx, maxResults=max_res, json_result=False)
        
        try:
            for issue in issues:
                already_migrated_set.add(get_shifted_key(issue.key.replace(project_new + '-', project_old + '-'), reversed=True))
            return (0, param)
        except:
            return (1, param)
    
    def issue_list_upload(param):
        global issues_lst, issues_set, already_migrated_set
        
        jira, types, non_migrated, jql, control, start_idx, max_res = param
        try:
            issues = jira.search_issues(jql_str=jql, startAt=start_idx, maxResults=max_res, json_result=False, fields='summary,issuetype,priority,status,assignee,reporter,created,duedate,parent,labels,resolution')
            for issue in issues:
                if issue.key not in issues_set:
                    issues_set.add(issue.key)
                    if non_migrated is False:
                        if issue.key in already_migrated_set:
                            continue
                        issues_lst.append(issue.key)
                    else:
                        issue_type = issue.fields.issuetype.name
                        issue_summary = issue.fields.summary.replace('\n', ' ').replace('\t', ' ')
                        issue_priority = '' if issue.fields.priority is None else issue.fields.priority.name
                        issue_status = issue.fields.status.name
                        issue_resolution = '' if issue.fields.resolution is None else issue.fields.resolution.name
                        issue_assignee = 'Unassigned' if issue.fields.assignee is None else issue.fields.assignee.displayName
                        issue_reporter = 'Anonymous' if issue.fields.reporter is None else issue.fields.reporter.displayName
                        issue_created = issue.fields.created.split('.')[0].replace("T", " ")
                        issue_due_date = '' if issue.fields.duedate is None else issue.fields.duedate
                        issue_parent = '' if not hasattr(issue.fields, 'parent') else '' if issue.fields.parent is None else issue.fields.parent.key
                        issue_labels = '' if issue.fields.labels is None else get_str_from_lst([i for i in issue.fields.labels])
                        issues_lst.append([issue.key, issue_type, issue_summary, issue_priority, issue_status, issue_resolution, issue_assignee, issue_reporter, issue_created, issue_due_date, issue_parent, issue_labels])
            return (0, param)
        except:
            if max_res == 1:
                return Exception
            try:
                for i in range(100):
                    try:
                        param = (jira, types, non_migrated, jql, control, start_idx + i, 1)
                        migrated_update(param)
                    except:
                        continue
                return (0, param)
            except:
                return (1, param)
    
    start_idx, block_num, block_size = (0, 0, 100)
    if max_result != 0 and block_size > max_result:
        block_size = max_result
    
    try:
        total = jira.search_issues(jql_str=jql, json_result=True, maxResults=1)['total']
    except:
        total = 0
    
    if total == 0:
        return None
    
    params = [(jira, types, non_migrated, jql, control, block_num * block_size, block_size) for block_num in range(0, total // block_size + 1)]
    
    if types is not None and sprint is None:
        max_retries = default_max_retries
        threads_processing(issue_list_update, params)
    elif sprint is not None:
        max_retries = default_max_retries
        threads_processing(sprint_update, params)
    elif migrated is not None:
        max_retries = default_max_retries
        threads_processing(migrated_update, params)
    else:
        issues_lst = []
        issues_set = set()
        max_retries = default_max_retries
        threads_processing(issue_list_upload, params)
        return issues_lst


def get_str_from_lst(lst, sep=',', spacing=' ', stripping=True):
    """This function returns list as comma separated string - for exporting in excel"""
    if lst is None:
        return None
    elif type(lst) != list:
        return str(lst)
    st = ''
    for i in lst:
        if i != '':
            if stripping is True:
                st += str(i).strip() + sep + spacing
            else:
                st += str(i) + sep + spacing
    if spacing == ' ':
        st = st[0:-2]
    else:
        st = st[0:-1]
    return st


def grouper(iterable, n, fill_value=None):
    """Collect data into fixed chunks or blocks"""
    if fill_value is None:
        if len(iterable) > n:
            return [list(iterable)[i:i+n] for i in range(0, len(iterable), n)]
        else:
            return [list(iterable)]
    else:
        args = [iter(iterable)] * n
        return zip_longest(*args, fillvalue=fill_value)


def create_temp_folder(folder, clean=True):
    """Create temp local folder for temporarily store attachments"""
    global verbose_logging
    
    local_folder = folder
    if clean and os.path.exists(local_folder):
        for filename in os.listdir(local_folder):
            file_path = os.path.join(local_folder, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
                elif os.path.isdir(file_path):
                    shutil.rmtree(file_path)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (file_path, e))
                if verbose_logging == 1:
                    print(traceback.format_exc())
        print("[INFO] Folder '{}' has been cleaned up.".format(local_folder))
        print("")
    elif not os.path.exists(local_folder):
        os.mkdir(local_folder)
        print("[INFO] Folder '{}' has been created".format(local_folder))
        print("")


def clean_temp_folder(folder):
    """Clean the folder tree"""
    shutil.rmtree(folder)


def get_jira_connection():
    global auth, threads, verify, create_remote_link_for_old_issue, JIRA_BASE_URL_OLD, JIRA_BASE_URL_NEW
    global jira_old, jira_new, atlassian_jira_old, verbose_logging
    
    # Check SSL certification and use unsecured connection if not available
    try:
        jira1 = JIRA(JIRA_BASE_URL_OLD, max_retries=1)
    except:
        try:
            jira1 = JIRA(JIRA_BASE_URL_OLD, max_retries=1, options={'verify': False})
            verify = False
            print("")
            print("[WARNING] SSL verification failed. Further processing would be with skipping SSL verification -> insecure connection processing.")
            print("")
        except Exception as e:
            try:
                if e.status_code == 503:
                    print("[ERROR] JIRA '{}' not available. Please check connectivity and try again later.".format(JIRA_BASE_URL_OLD))
                    print("")
                    os.system("pause")
                    exit()
            except:
                pass
    
    try:
        jira2 = JIRA(JIRA_BASE_URL_NEW, max_retries=1)
    except:
        try:
            jira2 = JIRA(JIRA_BASE_URL_NEW, max_retries=1, options={'verify': False})
            verify = False
            print("")
            print("[WARNING] SSL verification failed. Further processing would be with skipping SSL verification -> insecure connection processing.")
            print("")
        except Exception as e:
            try:
                if e.status_code == 503:
                    print("[ERROR] JIRA '{}' not available. Please check connectivity and try again later.".format(JIRA_BASE_URL_NEW))
                    print("")
                    os.system("pause")
                    exit()
            except:
                pass
    
    try:
        try:
            jira_old = JIRA(JIRA_BASE_URL_OLD, auth=auth, logging=False, async_=True, async_workers=threads, max_retries=0, options={'verify': verify})
            if JIRA_BASE_URL_OLD == JIRA_BASE_URL_NEW:
                jira_new = jira_old
            else:
                jira_new = JIRA(JIRA_BASE_URL_NEW, auth=auth, logging=False, async_=True, async_workers=threads, max_retries=0, options={'verify': verify})
        except Exception as e:
            print("[ERROR] Login to JIRA failed. JIRA is unavailable or credentials are invalid. Exception: '{}'".format(e))
            if verbose_logging == 1:
                print(traceback.format_exc())
            os.system("pause")
            exit()
        if create_remote_link_for_old_issue == 1 or migrate_attachments_check == 1:
            atlassian_jira_old = jira.Jira(JIRA_BASE_URL_OLD, username=username, password=password)
    except Exception as e:
        print("[ERROR] Login to JIRA failed. JIRA is unavailable or credentials are invalid. Exception: '{}'".format(e))
        if verbose_logging == 1:
            print(traceback.format_exc())
        os.system("pause")
        exit()
    jira_old = JIRA(JIRA_BASE_URL_OLD, auth=auth, logging=False, async_=True, async_workers=threads, max_retries=3, options={'verify': verify})
    if JIRA_BASE_URL_OLD == JIRA_BASE_URL_NEW:
        jira_new = jira_old
    else:
        jira_new = JIRA(JIRA_BASE_URL_NEW, auth=auth, logging=False, async_=True, async_workers=threads, max_retries=3, options={'verify': verify})

def get_total_teams():
    global auth, headers, verify, JIRA_BASE_URL_NEW, JIRA_team_api
    
    url_retrieve = JIRA_BASE_URL_NEW + JIRA_team_api + '/count'
    r = requests.get(url=url_retrieve, auth=auth, headers=headers, verify=verify)
    if str(r.status_code) == '200':
        teams_string = r.content.decode('utf-8')
        teams_lst = json.loads(teams_string)
        return teams_lst
    else:
        print("[ERROR] Portfolio / Advanced Roadmaps Plug-in(s) not available. Teams migration will be skipping...")
        return ''


def get_all_shared_teams():
    global teams, verbose_logging, auth, headers, migrate_teams_check, threads, max_retries
    
    def retrieve_teams(i):
        global teams, auth, headers, migrate_teams_check, verify, JIRA_BASE_URL_NEW, JIRA_team_api
        
        try:
            url_retrieve = JIRA_BASE_URL_NEW + JIRA_team_api + '?size=100&page=' + str(i)
            r = requests.get(url=url_retrieve, auth=auth, headers=headers, verify=verify)
            teams_string = r.content.decode('utf-8')
            try:
                teams_lst = json.loads(teams_string)
            except:
                print("[ERROR] Portfolio Add on not available for Target JIRA project. Teams will not be migrated.")
                migrate_teams_check = 0
                return (0, i)
            try:
                for team in teams_lst:
                    if team['shareable'] is True:
                        teams[team['title'].upper().strip()] = team['id']
            except:
                pass
            if verbose_logging == 1:
                print("[INFO] Teams retrieved from JIRA so far: {}".format(len(teams)))
            return (0, i)
        except Exception as e:
            if verbose_logging == 1:
                print("[ERROR] Error while processing: '{}'".format(e))
            return (1, i)
    
    print("[START] Reading ALL available shared teams.")
    try:
        pages = [i for i in range(1, get_total_teams() // 100 + 2)]
        max_retries = default_max_retries
        threads_processing(retrieve_teams, pages)
        print("[END] All teams has been loaded for further items processing.")
    except:
        migrate_teams_check = 0


def get_team_id(team_name):
    global teams, team_project_prefix, auth, verbose_logging
    
    def create_new_team(team_name):
        global teams, auth, headers, verify
        url_create = JIRA_BASE_URL_NEW + JIRA_team_api
        body = eval('{"title": team_name, "shareable": "true"}')
        r = requests.post(url_create, json=body, auth=auth, headers=headers, verify=verify)
        team_id = int(r.content.decode('utf-8'))
        teams[team_name.upper().strip()] = team_id
        return str(team_id)
    
    if teams == {}:
        get_all_shared_teams()
    
    if team_project_prefix.strip() != '':
        team_name_to_check = team_project_prefix + team_name.strip()
    else:
        team_name_to_check = team_name.strip()
    
    try:
        return str(teams[team_name_to_check.upper().strip()])
    except Exception as e:
        print("Creating NEW Team: '{}'".format(team_name_to_check))
        if verbose_logging == 1:
            print(e)
        return create_new_team(team_name_to_check.strip())


def get_team_name(team_id, jira_url=None):
    global JIRA_BASE_URL_OLD, auth, headers, verify
    
    if team_id == '':
        return None
    if jira_url is None:
        jira_url = JIRA_BASE_URL_OLD
    url_team_name = jira_url + JIRA_team_api + '/{}'.format(str(team_id))
    try:
        r = requests.get(url=url_team_name, auth=auth, headers=headers, verify=verify)
        team_details_json = r.content.decode('utf-8')
        team_details = json.loads(team_details_json)
        team_name = team_details['title']
    except:
        team_name = None
    
    return team_name


def get_lm_field_values(field_name, type):
    global JIRA_BASE_URL_NEW, JIRA_labelit_api, headers, verify, auth, project_new, issue_details_new
    
    url = JIRA_BASE_URL_NEW + JIRA_labelit_api
    proj = jira_new.project(project_new).id
    field_id = issue_details_new[type][field_name]['id']
    lm_data = ['labels']
    offset = 0
    labels = []
    while len(lm_data) > 0:
        data = {"customFieldId": field_id,
                "projectId": proj,
                "offset": offset}
        r = requests.get(url, params=data, auth=auth, headers=headers, verify=verify)
        lm_data = eval(r.content.decode('utf-8'))
        offset += 1000
        for l in lm_data:
            labels.append(l["name"])
    return labels


def add_lm_field_value(value, field_name, type):
    global JIRA_BASE_URL_NEW, JIRA_labelit_api, headers, verify, auth, project_new, issue_details_new
    
    url = JIRA_BASE_URL_NEW + JIRA_labelit_api
    proj = jira_new.project(project_new).id
    field_id = issue_details_new[type][field_name]['id']
    name = str(value).strip().replace(' ', '_').replace(',', '_').replace('?', '_')
    body = {"name": name,
            "customFieldId": field_id,
            "projectId": proj}
    r = requests.post(url, json=body, auth=auth, headers=headers, verify=verify)
    if str(r.status_code) != '201':
        print("[WARNING] Value '{}' can't be added into Label Manager field.".format(name))


def create_sprint(data):
    global auth, headers, JIRA_BASE_URL_NEW, JIRA_sprint_api, new_sprints, verify, jira_old, project_old
    global verbose_logging
    
    url = JIRA_BASE_URL_NEW + JIRA_sprint_api
    try:
        if data["startDate"] == '':
            data.pop("startDate", None)
        if data["endDate"] == '':
            data.pop("endDate", None)
        try:
            jql_count = "project = '{}' and issueFunction in completeInSprint('{}', '{}')".format(project_old, data["originBoardName"], data["name"])
            issues_cnt = jira_old.search_issues(jql_count, startAt=0, maxResults=1, json_result=True)['total']
        except:
            issues_cnt = 0
        if issues_cnt == 0:
            return (0, data)
        r = requests.post(url, json=data, auth=auth, headers=headers, verify=verify)
        new_sprint_details = r.content.decode('utf-8')
        new_sprint = json.loads(new_sprint_details)
        new_sprints[new_sprint['name']] = {"id": new_sprint['id'], "state": new_sprint['state']}
        return (0, data)
    except Exception as e:
        print('Exception: {}'.format(e))
        if verbose_logging == 1:
            print(traceback.format_exc())
        return (1, data)

def update_sprint(data):
    global auth, headers, JIRA_BASE_URL_NEW, JIRA_sprint_api, new_sprints, verify, verbose_logging
    
    sprint_id, body, closed = data
    url = JIRA_BASE_URL_NEW + JIRA_sprint_api + str(sprint_id)
    try:
        r = requests.post(url, json=body, auth=auth, headers=headers, verify=verify)
        if r.status_code == 400:
            body["startDate"] = datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.%f%z')
            body["endDate"] = (datetime.datetime.utcnow() + datetime.timedelta(days=14)).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
            r = requests.post(url, json=body, auth=auth, headers=headers, verify=verify)
        new_sprint_details = r.content.decode('utf-8')
        new_sprint = json.loads(new_sprint_details)
        new_sprints[new_sprint['name']] = {"state": new_sprint['state']}
        if closed == 1:
            body = {"state": "CLOSED"}
            rr = requests.post(url, json=body, auth=auth, headers=headers, verify=verify)
            new_sprint_details = rr.content.decode('utf-8')
            new_sprint = json.loads(new_sprint_details)
            new_sprints[new_sprint['name']] = {"state": new_sprint['state']}
        return (0, data)
    except Exception as e:
        if verbose_logging == 1:
            print(traceback.format_exc())
        return (0, data)


def delete_sprint(sprint_id):
    global auth, headers, JIRA_BASE_URL_NEW, JIRA_sprint_api, new_sprints, verify, verbose_logging
    
    url = JIRA_BASE_URL_NEW + JIRA_sprint_api + str(sprint_id)
    try:
        r = requests.delete(url, auth=auth, headers=headers, verify=verify)
        if r.status_code == 400:
            param = (sprint_id, {"state": "CLOSED"}, 0)
            update_sprint(param)
            r = requests.delete(url, auth=auth, headers=headers, verify=verify)
    except Exception as e:
        if verbose_logging == 1:
            print(traceback.format_exc())


def refresh_sprints():
    global old_sprints, new_sprints, jira_old, jira_new, new_board_id, project_old, process_only_last_updated_date_flag
    global last_updated_date, default_max_retries, max_retries
    
    sprint_names_for_delete = set()
    sprint_ids_for_delete = set()
    # New Sprints processing
    print("[INFO] Retrieving Sprints data from Target '{}' project.".format(project_new))
    if len(new_sprints) < 1 and len(jira_new.sprints(board_id=new_board_id)) > 0:
        for n_sprint in jira_new.sprints(board_id=new_board_id):
            if n_sprint.name in new_sprints.keys() or n_sprint.name in sprint_names_for_delete:
                sprint_names_for_delete.add(n_sprint.name)
                sprint_ids_for_delete.add(n_sprint.id)
            new_sprints[n_sprint.name] = {"id": n_sprint.id, "state": n_sprint.state}
    
    # Old Sprint processing
    print("[INFO] Retrieving Sprints data from Source '{}' project.".format(project_old))
    if len(old_sprints) < 1:
        if process_only_last_updated_date_flag == 1 and last_updated_date not in ['YYYY-MM-DD', '']:
            print("[INFO] Sprints would be calculated as changed from the '{}' date.".format(last_updated_date))
            jql_sprints = "project = {} AND updated >= {} AND Sprint is not EMPTY".format(project_old, last_updated_date)
        elif last_updated_days_check == 1 and recently_updated_days not in ['']:
            print("[INFO] Sprints would be calculated as changed within last '{}' days.".format(recently_updated_days))
            jql_sprints = "project = {} AND updated >= startOfDay(-{}) AND Sprint is not EMPTY".format(project_old, recently_updated_days)
        else:
            print("[INFO] ALL Sprints would be retrieved from Source project. It could take some time, please wait...".format(last_updated_date))
            jql_sprints = "project = {} AND Sprint is not EMPTY".format(project_old)
        get_issues_by_jql(jira_old, jql=jql_sprints, sprint=True)
    print("")
    
    # Calculation different Sprints for re-upload
    print("[INFO] Calculating the Sprints difference...")
    for n_sprint, n_values in new_sprints.items():
        if n_sprint in old_sprints.keys() and old_sprints[n_sprint]["state"] != n_values["state"]:
            sprint_ids_for_delete.add(n_values["id"])
            sprint_names_for_delete.add(n_sprint)
    if process_only_last_updated_date_flag != 1:
        for n_sprint, n_values in new_sprints.items():
            if n_sprint not in old_sprints.keys():
                sprint_ids_for_delete.add(n_values["id"])
                sprint_names_for_delete.add(n_sprint)
    
    # Removing changed Sprints and searching for issues to re-upload
    print("[INFO] The following Sprints would be deleted for further re-linkage for Issues: '{}'".format(sprint_names_for_delete))
    print("[INFO] Issues with changed Sprints calculating...")
    for sprint in sprint_names_for_delete:
        if sprint in old_sprints.keys():
            jql = "project = {} AND Sprint = {}".format(project_old, old_sprints[sprint]["id"])
            get_issues_by_jql(jira_old, jql=jql, types=True)
    print("")
    
    print("[INFO] Changed Sprints would be deleted...")
    for sprint in sprint_ids_for_delete:
        delete_sprint(sprint)


def migrate_sprints(board_id=old_board_id, proj_old=None, project=project_new, name=default_board_name, param='FUTURE'):
    global old_sprints, new_sprints, jira_old, jira_new, limit_migration_data, limit_migration_data, auth, max_retries
    global max_id, start_jira_key, headers, recently_updated, JIRA_BASE_URL_NEW, JIRA_sprint_api, default_max_retries
    global JIRA_board_api, new_board_id, supported_issuetypes
    
    start_time = time.time()
    if param == 'FUTURE':
        print("[START] Sprints and Issues processing has been started. All relevant Sprints and Issues are retrieving from Source JIRA.")
        if new_board_id == 0:
            new_board, n = (0, 0)
            for board in jira_new.boards():
                if board.name == name and project in board.filter.query:
                    new_board = board.id
                    new_board_id = new_board
                    break
            if new_board == 0:
                board = jira_new.create_board(name, project, location_type='project')
                new_board = board.id
                new_board_id = new_board
        else:
            new_board = new_board_id
        
        if len(jira_new.sprints(board_id=new_board)) > 0:
            for n_sprint in jira_new.sprints(board_id=new_board):
                new_sprints[n_sprint.name] = {"id": n_sprint.id, "state": n_sprint.state}
        if proj_old is None:
            print("[INFO] Sprints to be migrated from board '{}'.".format(board_id))
            url = JIRA_BASE_URL_NEW + JIRA_board_api + str(board_id)
            try:
                r = requests.get(url, auth=auth, headers=headers)
                board_details = r.content.decode('utf-8')
                board_data = json.loads(board_details)
                board_name = board_data["name"]
            except:
                board_name = ''
            for sprint in jira_old.sprints(board_id=board_id):
                if sprint.name not in new_sprints.keys():
                    try:
                        old_sprint = jira_old.sprint(sprint.id)
                        old_sprints[sprint.name] = {"id": sprint.id, "startDate": old_sprint.startDate, "endDate": old_sprint.endDate, "state": old_sprint.state.upper(), "originBoardName": board_name}
                    except:
                        old_sprints[sprint.name] = {"id": sprint.id, "startDate": '', "endDate": '', "state": sprint.state.upper(), "originBoardName": board_name}
                n += 1
                if (n % 20) == 0:
                    print("[INFO] Downloaded metadata for {} out of {} Sprints so far...".format(n, len(jira_old.sprints(board_id=board_id))))
        else:
            print("[INFO] All Sprints to be migrated from old '{}' project and will be added into new '{}' project, '{}' board.".format(proj_old, project, name))
            if limit_migration_data != 0:
                if start_jira_key != max_id:
                    jql_sprints = 'project = {} AND key >= {} AND key < {} {} {} order by key ASC'.format(project_old, start_jira_key, max_id, recently_updated, supported_issuetypes)
                else:
                    jql_sprints = 'project = {} AND key >= {} AND key <= {} {} {} order by key ASC'.format(project_old, start_jira_key, max_id, recently_updated, supported_issuetypes)
            else:
                jql_sprints = 'project = {} AND key >= {} {} {} order by key ASC'.format(proj_old, start_jira_key, recently_updated, supported_issuetypes)
            get_issues_by_jql(jira_old, jql=jql_sprints, types=True, sprint=True)
            print("[END] Sprints and Issues has been retrieved from Source JIRA.")
            print("[INFO] Sprints / Issues has been retrieved in '{}' seconds".format(time.time() - start_time))
            print("")
        
        sprint_start_time = time.time()
        print("[START] Missing Sprints to be created...")
        sprint_details = []
        for o_sprint_name, o_sprint_details in old_sprints.items():
            if o_sprint_name not in new_sprints.keys():
                sprint_details.append({"originBoardId": new_board, "name": o_sprint_name, "startDate": old_sprints[o_sprint_name]['startDate'], "endDate": old_sprints[o_sprint_name]['endDate'], "originBoardName": old_sprints[o_sprint_name]['originBoardName']})
        max_retries = default_max_retries
        threads_processing(create_sprint, sprint_details)
        print("[END] Sprints have been created with '{}' states.".format(param))
        print("[INFO] Sprints have been created in '{}' seconds".format(time.time() - sprint_start_time))
        print("")
    else:
        print("[START] Sprint statuses to be updated to '{}'.".format(param))
        sprint_details = []
        for o_sprint_name, o_sprint_details in old_sprints.items():
            if o_sprint_name in new_sprints.keys():
                body = {}
                closed = 0
                if param == 'ACTIVE' and new_sprints[o_sprint_name]["state"] == 'FUTURE' and old_sprints[o_sprint_name]['state'] != 'FUTURE':
                    body = {"state": "ACTIVE"}
                    if old_sprints[o_sprint_name]['state'] == 'CLOSED':
                        closed = 1
                if param == 'CLOSED' and new_sprints[o_sprint_name]["state"] == 'ACTIVE' and old_sprints[o_sprint_name]['state'] == 'CLOSED':
                    body = {"state": "CLOSED"}
                elif param == 'CLOSED' and new_sprints[o_sprint_name]["state"] == 'FUTURE' and old_sprints[o_sprint_name]['state'] == 'CLOSED':
                    body = {"state": "ACTIVE"}
                    closed = 1
                if body != {}:
                    sprint_details.append((new_sprints[o_sprint_name]["id"], body, closed))
        max_retries = default_max_retries
        threads_processing(update_sprint, sprint_details)
        print("[END] Sprint statuses have been updated to '{}'.".format(param))


def migrate_components():
    global jira_new, jira_old, max_retries, default_max_retries, project_old, project_new, verbose_logging
    
    print("[START] Components migration has been started.")
    old_components = jira_old.project_components(project_old)
    new_components = jira_new.project_components(project_new)
    
    def update_component(data):
        global jira_new, verbose_logging, max_retries, default_max_retries, auth, JIRA_BASE_URL_NEW, JIRA_components_api
        global headers, verify, JIRA_component_api
        
        try:
            id, name, project, description, lead_name, assignee_type, assignee_valid, archived = data
            try:
                body = json.dumps({"project": project,
                                   "name": name.strip(),
                                   "description": description,
                                   "leadUserName": lead_name,
                                   "assigneeType": assignee_type,
                                   "isAssigneeTypeValid": assignee_valid,
                                   "archived": archived})
                if id is not None:
                    url = JIRA_BASE_URL_NEW + JIRA_components_api.format(id)
                    r = requests.put(url, data=body, auth=auth, headers=headers, verify=verify)
                else:
                    url = JIRA_BASE_URL_NEW + JIRA_component_api
                    r = requests.post(url, data=body, auth=auth, headers=headers, verify=verify)
            except Exception as e:
                print('Exception: {}'.format(e.text))
                if verbose_logging == 1:
                    print(traceback.format_exc())
            return (0, data)
        except:
            return (1, data)
    
    new_components_dict = {}
    components_data = []
    for new_component in new_components:
        id, description, assignee_type, lead_name, assignee_valid, archived = (None, None, None, None, None, None)
        if hasattr(new_component, 'id'):
            id = new_component.id
        if hasattr(new_component, 'description'):
            description = new_component.description
        if hasattr(new_component, 'assigneeType'):
            assignee_type = new_component.assigneeType
        if hasattr(new_component, 'lead') and hasattr(new_component.lead, 'name'):
            lead_name = new_component.lead.name
        if hasattr(new_component, 'isAssigneeTypeValid'):
            assignee_valid = new_component.isAssigneeTypeValid
        if hasattr(new_component, 'archived'):
            archived = new_component.archived
        new_components_dict[new_component.name.strip().upper()] = {"id": id,
                                                                   "assigneeType": assignee_type,
                                                                   "leadUserName": lead_name,
                                                                   "isAssigneeTypeValid": assignee_valid,
                                                                   "archived": archived,
                                                                   "description": description,
                                                                   }
    
    for component in old_components:
        description, assignee_type, lead_name, assignee_valid, archived = (None, None, None, None, None)
        name = component.name.strip().upper()
        if hasattr(component, 'description'):
            description = component.description
        if hasattr(component, 'assigneeType'):
            assignee_type = component.assigneeType
        if hasattr(component, 'lead') and hasattr(component.lead, 'name'):
            lead_name = component.lead.name
        if hasattr(component, 'isAssigneeTypeValid'):
            assignee_valid = component.isAssigneeTypeValid
        if hasattr(component, 'archived'):
            archived = component.archived
        
        if name not in [n_component.upper() for n_component in new_components_dict.keys()]:
            components_data.append((None, component.name, project_new, description, lead_name, assignee_type, assignee_valid, archived))
        elif description != new_components_dict[name]["description"] or archived != new_components_dict[name]["archived"] or assignee_type != new_components_dict[name]["assigneeType"] or lead_name != new_components_dict[name]["leadUserName"]:
            components_data.append((new_components_dict[name]["id"], component.name.strip(), project_new, description, lead_name, assignee_type, assignee_valid, archived))
    
    max_retries = default_max_retries
    threads_processing(update_component, components_data)
    print("[END] All components have been succsessfully migrated.")


def migrate_versions():
    global jira_new, jira_old, verbose_logging
    
    print("[START] FixVersions (Releases) migration has been started.")
    old_versions = jira_old.project_versions(project_old)
    new_versions = jira_new.project_versions(project_new)
    
    def update_version(data):
        global jira_new, verbose_logging, max_retries, default_max_retries, auth, JIRA_BASE_URL_NEW, JIRA_versions_api
        global headers, verify
        
        id, name, project, description, release_date, start_date, archived, released = data
        try:
            if id is not None:
                project_id = jira_new.project(project_new).id
                url = JIRA_BASE_URL_NEW + JIRA_versions_api.format(id)
                body = json.dumps({"projectId": project_id,
                                   "name": name.strip(),
                                   "description": description,
                                   "releaseDate": release_date,
                                   "startDate": start_date,
                                   "released": released,
                                   "archived": archived})
                r = requests.put(url, data=body, auth=auth, headers=headers, verify=verify)
            else:
                jira_new.create_version(name, project, description=description, releaseDate=release_date, startDate=start_date, archived=archived, released=released)
            return (0, data)
        except Exception as e:
            print('Exception: {}'.format(e.text))
            if verbose_logging == 1:
                print(traceback.format_exc())
            return (1, data)
    
    versions = []
    new_versions_dict = {}
    for new_version in new_versions:
        id, description, release_date, start_date, archived, released = (None, None, None, None, None, None)
        if hasattr(new_version, 'id'):
            id = new_version.id
        if hasattr(new_version, 'description'):
            description = new_version.description
        if hasattr(new_version, 'releaseDate'):
            release_date = new_version.releaseDate
        if hasattr(new_version, 'startDate'):
            start_date = new_version.startDate
        if hasattr(new_version, 'archived'):
            archived = new_version.archived
        if hasattr(new_version, 'released'):
            released = new_version.released
        new_versions_dict[new_version.name.strip().upper()] = {"id": id,
                                                               "released": released,
                                                               "archived": archived,
                                                               "startDate": start_date,
                                                               "releaseDate": release_date,
                                                               "description": description,
                                                               }
    for version in old_versions:
        description, release_date, start_date, archived, released = (None, None, None, None, None)
        name = version.name.strip().upper()
        if hasattr(version, 'description'):
            description = version.description
        if hasattr(version, 'releaseDate'):
            release_date = version.releaseDate
        if hasattr(version, 'startDate'):
            start_date = version.startDate
        if hasattr(version, 'archived'):
            archived = version.archived
        if hasattr(version, 'released'):
            released = version.released
        
        if name not in [n_version.upper() for n_version in new_versions_dict.keys()]:
            versions.append((None, version.name.strip(), project_new, description, release_date, start_date, archived, released))
        elif released != new_versions_dict[name]["released"] or archived != new_versions_dict[name]["archived"] or release_date != new_versions_dict[name]["releaseDate"]:
            versions.append((new_versions_dict[name]["id"], version.name.strip(), project_new, description, release_date, start_date, archived, released))
    
    max_retries = default_max_retries
    threads_processing(update_version, versions)
    print("[END] All FixVersions (Releases) have been succsessfully migrated.")


def migrate_comments(old_issue, new_issue):
    for comment in jira_old.comments(old_issue):
        comment_match = 0
        try:
            new_data = eval("'*[' + comment.author.displayName + '|~' + comment.author.name + ']* added on *_' + comment.created[:10] + ' ' + comment.created[11:19] + '_*: \\\\\\ '")
        except:
            new_data = eval("'*Anonymous* added on *_' + comment.created[:10] + ' ' + comment.created[11:19] + '_*: \\\\\\ '")
        len_new_data = len(new_data)
        for new_comment in jira_new.comments(new_issue):
            if comment.body == new_comment.body[len_new_data:] or comment.body == new_comment.body:
                comment_match = 1
        if comment_match == 0:
            data = eval("new_data + comment.body")
            jira_new.add_comment(new_issue, body=str(data))


def get_new_link_type(old_link):
    global link_mappings
    
    new_link = old_link
    try:
        for k, v in link_mappings.items():
            if old_link in v:
                new_link = k
                return new_link
    except:
        pass
    return new_link


def migrate_links(old_issue, new_issue):
    global project_old, project_new, jira_new
    
    outward_issue_links = {}
    inward_issue_links = {}
    outward_issue_links_old = {}
    inward_issue_links_old = {}
    outward_issue_links_new = {}
    inward_issue_links_new = {}
    links_for_del = []
    
    try:
        old_links = old_issue.fields.issuelinks
    except:
        old_links = []
    try:
        new_links = new_issue.fields.issuelinks
    except:
        new_links = []
    
    for link in new_links:
        if hasattr(link, "outwardIssue"):
            if link.outwardIssue.key not in outward_issue_links:
                outward_issue_links[link.outwardIssue.key] = {}
            outward_issue_links[link.outwardIssue.key][link.type.name] = str(link)
        if hasattr(link, "inwardIssue"):
            if link.inwardIssue.key not in inward_issue_links:
                inward_issue_links[link.inwardIssue.key] = {}
            inward_issue_links[link.inwardIssue.key][link.type.name] = str(link)
    
    for link in old_links:
        new_link = get_new_link_type(link.type.name)
        if hasattr(link, "outwardIssue"):
            new_id = get_shifted_key(link.outwardIssue.key).replace(project_old + '-', project_new + '-')
            if new_id not in outward_issue_links_old.keys():
                outward_issue_links_old[new_id] = []
            outward_issue_links_old[new_id].append(new_link)
            if new_id not in outward_issue_links.keys() or (new_id in outward_issue_links.keys()
                                                            and new_link not in outward_issue_links[new_id].keys()):
                try:
                    jira_new.create_issue_link(new_link, new_issue.key, new_id)
                except:
                    pass
        
        if hasattr(link, "inwardIssue"):
            new_id = get_shifted_key(link.inwardIssue.key).replace(project_old + '-', project_new + '-')
            if new_id not in inward_issue_links_old:
                inward_issue_links_old[new_id] = []
            inward_issue_links_old[new_id].append(new_link)
            if new_id not in inward_issue_links.keys() or (new_id in inward_issue_links.keys()
                                                           and new_link not in inward_issue_links[new_id].keys()):
                try:
                    jira_new.create_issue_link(new_link, new_id, new_issue.key)
                except:
                    pass
    
    try:
        new_links = new_issue.fields.issuelinks
    except:
        new_links = []
    
    for link in new_links:
        if hasattr(link, "outwardIssue"):
            if link.outwardIssue.key not in outward_issue_links_new:
                outward_issue_links_new[link.outwardIssue.key] = {}
            outward_issue_links_new[link.outwardIssue.key][link.type.name] = str(link)
        if hasattr(link, "inwardIssue"):
            if link.inwardIssue.key not in inward_issue_links_new:
                inward_issue_links_new[link.inwardIssue.key] = {}
            inward_issue_links_new[link.inwardIssue.key][link.type.name] = str(link)
    
    for k, v in outward_issue_links_new.items():
        for link_type, id in v.items():
            if k not in outward_issue_links_old.keys() or (k in outward_issue_links_old.keys() and link_type not in outward_issue_links_old[k]):
                links_for_del.append(str(id))
    
    for k, v in inward_issue_links_new.items():
        for link_type, id in v.items():
            if k not in inward_issue_links_old.keys() or (k in inward_issue_links_old.keys() and link_type not in inward_issue_links_old[k]):
                links_for_del.append(str(id))
    
    for link in links_for_del:
        try:
            jira_new.delete_issue_link(link)
        except:
            pass


def migrate_attachments(old_issue, new_issue, retry=True):
    global temp_dir_name, jira_old, JIRA_attachment_api, atlassian_jira_old, verbose_logging, jira_new
    
    new_attachments = {}
    old_attachments = {}
    
    if new_issue is not None and new_issue.fields.attachment:
        for new_attachment in new_issue.fields.attachment:
            try:
                new_attachments[new_attachment.id] = new_attachment.filename
            except:
                pass
    
    if not os.path.exists(temp_dir_name):
        create_temp_folder(temp_dir_name)
    
    try:
        if old_issue.fields.attachment:
            for attachment in old_issue.fields.attachment:
                old_attachments[attachment.id] = attachment.filename
                if attachment.filename not in new_attachments.values():
                    file = attachment.get()
                    filename = attachment.filename
                    temp_name = 'temp_' + attachment.id
                    full_name = os.path.join(temp_dir_name, temp_name)
                    with open(full_name, 'wb') as f:
                        f.write(file)
                    with open(full_name, 'rb') as file_new:
                        jira_new.add_attachment(new_issue.key, file_new, filename)
                    if os.path.exists(full_name):
                        os.remove(full_name)
    except:
        try:
            attachments = {}
            old_issue_full = jira_old.issue(old_issue.key, expand='changelog')
            for log in old_issue_full.raw['changelog']['histories']:
                for item in log['items']:
                    if item["field"] == "Attachment":
                        if item["to"] is not None:
                            if item["to"] in attachments.keys() and (("added" in attachments[item["to"]] and log["created"] > attachments[item["to"]]["added"]) or "added" not in attachments[item["to"]]):
                                attachments[item["to"]]["added"] = log["created"]
                            else:
                                attachments[item["to"]] = {}
                                attachments[item["to"]]["added"] = log["created"]
                                attachments[item["to"]]["name"] = item["toString"]
                        else:
                            if item["from"] in attachments.keys() and (("removed" in attachments[item["from"]] and log["created"] > attachments[item["from"]]["removed"]) or "removed" not in attachments[item["from"]]):
                                attachments[item["from"]]["removed"] = log["created"]
                            else:
                                attachments[item["from"]] = {}
                                attachments[item["from"]]["removed"] = log["created"]
                                attachments[item["from"]]["name"] = item["fromString"]
            
            for k, v in attachments.items():
                old_attachments[k] = v["name"]
                if "added" in v.keys() and v["name"] not in new_attachments.values() and ("removed" not in v.keys() or ("removed" in v.keys() and v["added"] > v["removed"])):
                    attachment = atlassian_jira_old.get_attachment(k)
                    if 'content' in attachment and 'filename' in attachment and v["name"] == attachment["filename"]:
                        file_url = attachment["content"]
                        r = requests.get(file_url, allow_redirects=True)
                        filename = attachment["filename"]
                        temp_name = 'temp_' + k
                        full_name = os.path.join(temp_dir_name, temp_name)
                        with open(full_name, 'wb') as f:
                            f.write(r.content)
                        with open(full_name, 'rb') as file_new:
                            jira_new.add_attachment(new_issue.key, file_new, filename)
                        if os.path.exists(full_name):
                            os.remove(full_name)
        
        except Exception as e:
            if retry is True:
                migrate_attachments(old_issue, new_issue, retry=False)
            else:
                print("[ERROR] Attachments from '{}' issue can't be loaded due to: '{}'.".format(old_issue.key, e))
                if verbose_logging == 1:
                    print(traceback.format_exc())
    
    for id, new_name in new_attachments.items():
        if new_name not in old_attachments.values():
            jira_new.delete_attachment(id)



def migrate_status(new_issue, old_issue):
    global new_transitions, verbose_logging
    
    def find_shortest_path(graph, start, end, path):
        path = path + [start]
        if start == end:
            return path
        if start not in graph.keys():
            return None
        shortest = None
        for node in graph[start]:
            if node not in path:
                new_path = find_shortest_path(graph, node, end, path)
                if new_path and (not shortest or len(new_path) < len(shortest)):
                    shortest = new_path
        return shortest
    
    resolution = None
    new_issue_type = None
    issue_type = old_issue.fields.issuetype.name
    new_status = get_new_status(old_issue.fields.status.name, issue_type)
    old_status = new_issue.fields.status.name
    if old_issue.fields.resolution:
        resolution = old_issue.fields.resolution.name
    
    if new_status == '':
        return
    graph = {}
    
    new_issue_type = get_new_issuetype(issue_type)
    
    for t in new_transitions[new_issue_type]:
        if t[0].upper() in graph.keys():
            graph[t[0].upper()].append(t[2].upper())
        else:
            graph[t[0].upper()] = [t[2].upper()]
    if old_status is None:
        old_status = graph[new_issue_type][0][0]
    transition_path = find_shortest_path(graph, old_status.upper(), new_status.upper(), [])
    if transition_path is None:
        return
    
    status_transitions = []
    for i in range(1, len(transition_path)):
        for t in new_transitions[new_issue_type]:
            if t[0].upper() == transition_path[i-1] and t[2].upper() == transition_path[i]:
                status_transitions.append(t[1])
    
    for s in status_transitions:
        if resolution is None:
            jira_new.transition_issue(new_issue, transition=s)
        else:
            try:
                jira_new.transition_issue(new_issue, transition=s, fields={"resolution": {"name": resolution}})
            except:
                try:
                    jira_new.transition_issue(new_issue, transition=s)
                except Exception as e:
                    print("[ERROR] Status can't be changed due to '{}'".format(e.text))
                    if verbose_logging == 1:
                        print(traceback.format_exc())
                    return


def update_issue_json(old_issue, new_issue_type, new_status, new=False, new_issue=None, subtask=None, retry_number=None):
    global retry_number_allowed, verbose_logging, max_retries
    
    if retry_number is None:
        retry_number = retry_number_allowed
    try:
        status_code, status = migrate_change_history(old_issue, new_issue_type, new_status, new=new, new_issue=new_issue, subtask=subtask)
        retry_number -= 1
        if str(status_code) == '202':
            return status
        elif retry_number <= 0:
            status = 'SKIP'
            if max_retries == 1:
                print("[ERROR] JSON Importer can't process '{}' issue. Will be re-tried later.".format(old_issue.key))
        else:
            sleep(retry_number_allowed - retry_number)
            status = update_issue_json(old_issue, new_issue_type, new_status, new=new, new_issue=new_issue, subtask=subtask, retry_number=retry_number)
        return status
    except Exception as er:
        print("[ERROR] JIRA Importer Plugin Failed. Issue '{}' can't be processed due to: '{}'. Skipped.".format(old_issue.key, er))
        print(traceback.format_exc())
        return 'SKIP'


def get_new_issue_after_json(key, retry_number=None):
    global retry_number_allowed, verbose_logging
    
    if retry_number is None:
        retry_number = retry_number_allowed
    try:
        new_issue = jira_new.issue(key, expand="changelog")
        retry_number -= 1
        if new_issue is None and retry_number > 0:
            sleep(retry_number_allowed - retry_number)
            new_issue = get_new_issue_after_json(key, retry_number=retry_number)
        return new_issue
    except Exception as er:
        retry_number -= 1
        if retry_number > 0:
            sleep(retry_number_allowed - retry_number)
            new_issue = get_new_issue_after_json(key, retry_number=retry_number)
            if new_issue is not None:
                return new_issue
        if verbose_logging == 1:
            print(traceback.format_exc())
    return None


def get_new_status(old_status, old_issue_type, new_issue_type=None):
    global status_mappings
    
    def get_status(type, status):
        global new_transitions
        
        new_statuses, new_issuetypes = ([], [])
        new_statuses, new_issuetypes = calculate_statuses(new_transitions)
        for l in new_statuses:
            if l[0] == type and l[1] == status:
                return status
        default_status = new_transitions[type][0][0]
        print("[WARNING] Mapping of '{}' Source Status for '{}' Target Issuetype Status hasn't been found! Default '{}' Status would be used instead.".format(status, type, default_status))
        return default_status
    
    try:
        for n_status, o_statuses in status_mappings[old_issue_type].items():
            for o_status in o_statuses:
                if old_status.upper() == o_status.upper() and n_status != '':
                    return n_status
        if new_issue_type is not None:
            return get_status(new_issue_type, old_status)
    except:
        return get_status(new_issue_type, old_status)
    return None


def get_parent_for_subtask(issue, issue_type, reprocess=False):
    global issue_details_old, dummy_parent, dummy_process, project_old, project_new
    
    parent = None
    try:
        parent = issue.fields.parent.key
        if project_old in parent:
            return get_shifted_key(parent.replace(project_old + '-', project_new + '-'))
    except:
        pass
    try:
        parent = eval('issue.fields.' + issue_details_old[issue_type]['Epic Link']['id'])
        if project_old in parent:
            return get_shifted_key(parent.replace(project_old + '-', project_new + '-'))
    except:
        pass
    try:
        parent = eval('issue.fields.' + issue_details_old[issue_type]['Parent Link']['id'])
        if project_old in parent:
            return get_shifted_key(parent.replace(project_old + '-', project_new + '-'))
    except:
        pass
    if dummy_parent != '':
        parent = dummy_parent
        dummy_process = 1
    elif reprocess is False:
        get_dummy_parent()
        get_parent_for_subtask(issue, issue_type, reprocess=True)
    return parent


def process_issue(key, reprocess=False):
    """ Main migration Function - issue migration processing here.
    :param key: JIRA Issue key
    :param reprocess: Flag for re-processing (if issue was deleted beforehand)
    :return: '0' if successful and '1' if failed plus input key, for reprocess
    """
    global items_lst, jira_new, project_new, jira_old, migrate_comments_check, migrate_links_check, migrated_text
    global migrate_attachments_check, migrate_statuses_check, migrate_metadata_check, create_remote_link_for_old_issue
    global max_id, json_importer_flag, issuetypes_mappings, sub_tasks, failed_issues, issue_details_old, dummy_parent
    global multiple_json_data_processing, verbose_logging, force_update_flag, max_retries, default_max_retries
    global including_dependencies_flag, dummy_process, force_update_flag
    
    if verbose_logging == 1:
        print("[INFO] Processing '{}' issue".format(key))
    try:
        new_issue_type = ''
        parent = None
        issue_type = None
        new_status = None
        new_issue = None
        existent_parent = None
        if reprocess is False:
            new_issue_key = get_shifted_key(project_new + '-' + str(key.split('-')[1]))
        else:
            new_issue_key = project_new + '-' + str(key.split('-')[1])
        try:
            old_issue = jira_old.issue(key, expand="changelog")
            issue_type = old_issue.fields.issuetype.name
            new_issue_type = get_new_issuetype(issue_type)
            new_status = get_new_status(old_issue.fields.status.name, issue_type, new_issue_type)
        except:
            delete_issue(new_issue_key)
            return (0, key)
        try:
            new_issue = jira_new.issue(new_issue_key, expand="changelog")
            existent_new_type = new_issue.fields.issuetype.name
            if new_issue_type in sub_tasks.keys() and json_importer_flag == 1:
                # Checking Parent within same Project
                existent_parent = None if new_issue.fields.parent is None else new_issue.fields.parent.key
                if existent_parent is not None and existent_parent != dummy_parent and str(project_new + '-') not in str(existent_parent):
                    delete_issue(new_issue_key)
                    return process_issue(key)
                if new_issue_type != existent_new_type:
                    parent_calculated = get_parent_for_subtask(old_issue, issue_type)
                    parent = None if parent_calculated is None else parent_calculated
                    try:
                        parent_issue = jira_new.issue(parent)
                        if parent_issue.fields.issuetype.name in sub_tasks.keys():
                            parent = dummy_parent
                            dummy_process = 1
                    except:
                        try:
                            if including_dependencies_flag == 1 or force_update_flag == 1:
                                parent_issue_old = jira_old.issue(parent.replace(project_new + '-', project_old + '-'))
                                for k, v in issuetypes_mappings.items():
                                    if parent_issue_old.fields.issuetype.name in v and project_old in str(parent):
                                        process_issue(parent.replace(project_new + '-', project_old + '-'), reprocess=True)
                                        break
                            parent_issue = jira_new.issue(parent)
                            if parent_issue.fields.issuetype.name in sub_tasks.keys():
                                parent = dummy_parent
                                dummy_process = 1
                        except:
                            try:
                                if dummy_parent != '':
                                    parent = dummy_parent
                                    dummy_process = 1
                            except:
                                print("[ERROR] Parent '{}' for '{}' has not been mapped in Mapping file or can't be found in Source project. Sub-Task '{}' would not be created. Skipped.".format(parent, new_issue_type, new_issue_key))
                                delete_issue(new_issue_key)
                                return (0, key)
                    convert_to_subtask(parent, new_issue, sub_tasks[new_issue_type])
                    status = update_issue_json(old_issue, new_issue_type, new_status, new=False, new_issue=new_issue)
                    if status == 'SKIP' and key not in failed_issues:
                        failed_issues.add(key)
                        return (1, key)
                else:
                    parent_calculated = get_parent_for_subtask(old_issue, issue_type)
                    parent = None if parent_calculated is None else parent_calculated
                    if (hasattr(new_issue.fields, 'parent') and parent != new_issue.fields.parent.key) or not hasattr(new_issue.fields, 'parent'):
                        delete_issue(new_issue_key)
                        return process_issue(key)
                    try:
                        parent_issue = jira_new.issue(parent)
                        if parent_issue.fields.issuetype.name in sub_tasks.keys():
                            delete_issue(new_issue_key)
                            return process_issue(key)
                    except:
                        pass
                    if force_update_flag == 1 or new_issue.fields.status.name.upper() != new_status.upper():
                        status = update_issue_json(old_issue, new_issue_type, new_status, new=False, new_issue=new_issue, subtask=True)
                        if status == 'SKIP' and key not in failed_issues:
                            failed_issues.add(key)
                            return (1, key)
            elif new_issue_type not in sub_tasks.keys() and hasattr(new_issue.fields, 'parent') and new_issue.fields.parent is not None and json_importer_flag == 1:
                delete_issue(new_issue_key)
                return process_issue(key)
            elif force_update_flag == 1 and json_importer_flag == 1:
                status = update_issue_json(old_issue, new_issue_type, new_status, new=False, new_issue=new_issue)
                if status == 'SKIP' and key not in failed_issues:
                    failed_issues.add(key)
                    return (1, key)
        except Exception as e:
            if json_importer_flag == 0:
                print("[ERROR] Missing issue key '{}' in Target project. Exception: '{}'".format(old_issue.key, e))
                if verbose_logging == 1:
                    print(traceback.format_exc())
                return(0, key)
            else:
                if new_issue_type in sub_tasks.keys():
                    status = update_issue_json(old_issue, new_issue_type, new_status, new=True, subtask=True)
                    if status == 'SKIP' and key not in failed_issues:
                        failed_issues.add(key)
                        return (1, key)
                    new_issue = get_new_issue_after_json(new_issue_key)
                    if new_issue is None:
                        failed_issues.add(key)
                        return (1, key)
                    parent_calculated = get_parent_for_subtask(old_issue, issue_type)
                    parent = None if parent_calculated is None else parent_calculated
                    try:
                        parent_issue = jira_new.issue(parent)
                        if parent_issue.fields.issuetype.name in sub_tasks.keys():
                            parent = dummy_parent
                            dummy_process = 1
                    except:
                        try:
                            if including_dependencies_flag == 1 or force_update_flag == 1:
                                parent_issue_old = jira_old.issue(parent.replace(project_new + '-', project_old + '-'))
                                for k, v in issuetypes_mappings.items():
                                    if parent_issue_old.fields.issuetype.name in v and project_old in str(parent):
                                        process_issue(parent.replace(project_new + '-', project_old + '-'), reprocess=True)
                                        break
                            parent_issue = jira_new.issue(parent)
                            if parent_issue.fields.issuetype.name in sub_tasks.keys():
                                parent = dummy_parent
                                dummy_process = 1
                        except:
                            try:
                                if dummy_parent != '':
                                    parent = dummy_parent
                                    dummy_process = 1
                            except:
                                print("[ERROR] Parent '{}' for '{}' has not been mapped in Mapping file or can't be found in Source project. Sub-Task '{}' would not be created. Skipped.".format(parent, new_issue_type, new_issue_key))
                                delete_issue(new_issue_key)
                                return (0, key)
                    convert_to_subtask(parent, new_issue, sub_tasks[new_issue_type])
                    status = update_issue_json(old_issue, new_issue_type, new_status, new=False, new_issue=new_issue)
                    if status == 'SKIP' and key not in failed_issues:
                        failed_issues.add(key)
                        return (1, key)
                    new_issue = get_new_issue_after_json(new_issue_key)
                    if new_issue is None:
                        failed_issues.add(key)
                        return (1, key)
                else:
                    status = update_issue_json(old_issue, new_issue_type, new_status, new=True)
                    if status == 'SKIP' and key not in failed_issues:
                        failed_issues.add(key)
                        return (1, key)
                    new_issue = get_new_issue_after_json(new_issue_key)
                    if new_issue is None:
                        failed_issues.add(key)
                        print("[ERROR] Issue '{}' can't be created. Details: '{}'".format(new_issue_key, e))
                        return(1, key)
                    status = update_issue_json(old_issue, new_issue_type, new_status, new=False, new_issue=new_issue)
                    if status == 'SKIP' and key not in failed_issues:
                        failed_issues.add(key)
                        return (1, key)
        if migrate_comments_check == 1 and json_importer_flag == 0 and multiple_json_data_processing == 0 and new_issue is not None:
            migrate_comments(old_issue, new_issue)
        if migrate_links_check == 1 and new_issue is not None:
            migrate_links(old_issue, new_issue)
        if migrate_attachments_check == 1 and new_issue is not None:
            migrate_attachments(old_issue, new_issue)
        if migrate_statuses_check == 1 and json_importer_flag == 0 and multiple_json_data_processing == 0 and new_issue is not None:
            try:
                migrate_status(new_issue, old_issue)
            except Exception as e:
                print("[ERROR] Status can't be migrated due to: '{}'".format(e))
                if verbose_logging == 1:
                    print(traceback.format_exc())
        if create_remote_link_for_old_issue == 1:
            remote_link_exist = 0
            try:
                for r_link in jira_old.remote_links(old_issue.key):
                    if r_link.object.title == new_issue.key and r_link.relationship == migrated_text:
                        remote_link_exist = 1
            except:
                pass
            if remote_link_exist == 0:
                atlassian_jira_old.create_or_update_issue_remote_links(old_issue.key, JIRA_BASE_URL_NEW + '/browse/' + new_issue.key, title=new_issue.key, relationship='Migrated to')
        if migrate_metadata_check == 1 and new_issue is not None:
            update_new_issue_type(old_issue, new_issue, new_issue_type)
        return (0, key)
    except Exception as e:
        if max_retries == default_max_retries:
            print("[ERROR] Exception while processing '{}' issue: '{}'.".format(key, e))
            if verbose_logging == 1:
                print(traceback.format_exc())
        return (1, key)


def migrate_issues(issuetype):
    global items_lst, threads, max_retries, default_max_retries, pool_size, failed_issues, issuetypes_mappings
    global skipped_issuetypes
    
    for type in issuetypes_mappings[issuetype]['issuetypes']:
        if type in items_lst.keys() and type not in skipped_issuetypes:
            print("[INFO] The total number of '{}' issuetype: {}".format(type, len(items_lst[type])))
            print("[START] Copying from old '{}' Issuetype to new '{}' Issuetype...".format(type, issuetype))
            max_retries = default_max_retries
            if pool_size > 1:
                processes_processing(process_issue, items_lst[type])
            else:
                threads_processing(process_issue, items_lst[type])
        else:
            print("[INFO] No issues under '{}' issuetype were found. Skipping...".format(type))
            continue
        print("[END] '{}' issuetype has been migrated to '{}' Issuetype.".format(type, issuetype))
        print("")


def get_fields_ids():
    global JIRA_BASE_URL_OLD, JIRA_fields_api, auth, headers, verify, old_fields_ids_mapping
    
    try:
        url = JIRA_BASE_URL_OLD + JIRA_fields_api
        r = requests.get(url, auth=auth, headers=headers, verify=verify)
        fields_string = r.content.decode('utf-8')
        fields_details = json.loads(fields_string)
        for field in fields_details:
            old_fields_ids_mapping[field["name"]] = field["id"]
    
    except Exception as e:
        if verbose_logging == 1:
            print(traceback.format_exc())


def get_fields_list_by_project(jira, project, old=False):
    auth_jira = jira
    allfields = auth_jira.fields()
    
    def retrieve_custom_field(field_id):
        for field in allfields:
            if field['id'] == field_id:
                return field['custom']
    
    if old:
        get_fields_ids()
    
    proj = auth_jira.project(project)
    if proj.archived is True:
        print("[ERROR] Project '{}' is ARCHIVED.".format(project))
        return {}
    
    try:
        project_fields = auth_jira.createmeta(projectKeys=proj, expand='projects.issuetypes.fields')
        is_types = project_fields['projects'][0]['issuetypes']
    except:
        print("[ERROR] NO ACCESS to the '{}' project.".format(proj))
        return {}
    issuetype_fields = {}
    for issuetype in is_types:
        issuetype_name = issuetype['name']
        issuetype_fields[issuetype_name] = {}
        for field_id in issuetype['fields']:
            field_name = issuetype['fields'][field_id]['name']
            allowed_values = []
            if 'allowedValues' in issuetype['fields'][field_id]:
                for i in issuetype['fields'][field_id]['allowedValues']:
                    if 'children' in i:
                        for ch in i['children']:
                            if 'value' in ch.keys():
                                allowed_values.append([i['value'], ch['value']])
                    elif 'name' in i:
                        allowed_values.append(i['name'])
                    elif 'value' in i:
                        allowed_values.append(i['value'])
            
            default_val = None
            if issuetype['fields'][field_id]['hasDefaultValue'] is not False:
                if type(issuetype['fields'][field_id]['defaultValue']) == float:
                    default_val = str(issuetype['fields'][field_id]['defaultValue'])
                elif 'name' in issuetype['fields'][field_id]['defaultValue']:
                    default_val = issuetype['fields'][field_id]['defaultValue']['name']
                elif type(issuetype['fields'][field_id]['defaultValue']) == dict:
                    default_val = issuetype['fields'][field_id]['defaultValue']['value']
                elif type(issuetype['fields'][field_id]['defaultValue']) == list:
                    try:
                        default_val = issuetype['fields'][field_id]['defaultValue'][0]['value']
                    except:
                        default_val = issuetype['fields'][field_id]['defaultValue'][0]
                else:
                    default_val = issuetype['fields'][field_id]['defaultValue']
            
            field_attributes = {'id': field_id,
                                'required': issuetype['fields'][field_id]['required'],
                                'custom': retrieve_custom_field(field_id),
                                'type': issuetype['fields'][field_id]['schema']['type'],
                                'custom type': None if 'custom' not in issuetype['fields'][field_id]['schema'] else issuetype['fields'][field_id]['schema']['custom'].replace('com.atlassian.jira.plugin.system.customfieldtypes:', ''),
                                'allowed values': None if allowed_values == [] else allowed_values,
                                'default value': default_val,
                                'validated': True if 'allowedValues' in issuetype['fields'][field_id] else False}
            issuetype_fields[issuetype_name][field_name] = field_attributes
    return issuetype_fields


def load_file():
    global mapping_file, bulk_processing_flag
    dir_name = os.getcwd()
    if bulk_processing_flag == 0:
        mapping_file = askopenfilename(initialdir=dir_name, title="Select file", filetypes=(("Migration JIRA Template", ".xlsx .xls"), ("all files", "*.*")))
    else:
        mapping_file = askdirectory(initialdir=dir_name, title="Select Folder with Templates")
    file.delete(0, END)
    file.insert(0, mapping_file)


def load_default_file():
    global default_configuration_file
    dir_name = os.getcwd()
    default_configuration_file = askopenfilename(initialdir=dir_name, title="Select file", filetypes=(("Migration JIRA Template", ".xlsx .xls"), ("all files", "*.*")))
    default_file.delete(0, END)
    default_file.insert(0, default_configuration_file)


def create_excel_sheet(sheet_data, title):
    global JIRA_BASE_URL, header, output_excel, default_validation, issue_details_new, issue_details_old
    global jira_system_fields, additional_mapping_fields, new_transitions, excel_locked
    global project_tab_color, mandatory_tab_color, optional_tab_color, mandatory_template_tabs, hide_tabs
    
    try:
        wb.create_sheet(title)
    except:
        converted_value = ''
        for letter in title:
            if letter.isalpha() or letter.isnumeric() or letter in [' ']:
                converted_value += letter
            else:
                converted_value += '_'
        title = converted_value
        wb.create_sheet(title)
    
    ws = wb.get_sheet_by_name(title)

    if excel_locked == 1:
        ws.protection.password = protection_password
        ws.protection.sheet = True
        ws.protection.enable()
        ws.protection.sort = False
        ws.protection.formatCells = False
    
    start_column = 1
    start_row = 1
    
    # Creating Excel sheet based on data
    for i in range(len(sheet_data)):
        for y in range(len(sheet_data[i])):
            try:
                ws.cell(row=start_row, column=start_column+y).value = sheet_data[i][y]
            except:
                converted_value = ''
                for letter in sheet_data[i][y]:
                    if letter.isalpha() or letter.isnumeric() or letter in [',', '.', ';', ':', '&', '"', "'", ' ']:
                        converted_value += letter
                    else:
                        converted_value += '?'
                ws.cell(row=start_row, column=start_column+y).value = converted_value
        start_row += 1
    
    for y in range(1, ws.max_column+1):
        ws.cell(row=1, column=y).fill = header_fill
        ws.cell(row=1, column=y).font = header_font

    # Unlocking cells which could be updated
    if excel_locked == 1:
        start_row = 1
        if title == 'Project':
            pass
        elif title in ['Issuetypes', 'Priority']:
            for i in range(1, ws.max_row+1):
                ws.cell(row=start_row+i, column=2).protection = Protection(locked=False)
        elif title in ['Fields', 'Statuses', 'Links']:
            for i in range(1, ws.max_row+1):
                ws.cell(row=start_row+i, column=3).protection = Protection(locked=False)
        else:
            for i in range(1, ws.max_row+1):
                ws.cell(row=start_row+i, column=1).protection = Protection(locked=False)
            if 'Source' in sheet_data[0][1]:
                for i in range(1, ws.max_row+1):
                    ws.cell(row=start_row+i, column=2).protection = Protection(locked=False)

    # Column width formatting
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        if length > 80:
            ws.column_dimensions[column_cells[0].column_letter].width = 80
        else:
            ws.column_dimensions[column_cells[0].column_letter].width = length + 4
    
    # Data Validation added - for short lists in the beginning
    if title == 'Project':
        fields_val = {}
        for issuetype, fields in issue_details_new.items():
            fields_val[issuetype] = additional_mapping_fields[:]
            for field in fields.keys():
                if issue_details_new[issuetype][field]['custom'] is True and field not in jira_system_fields:
                    fields_val[issuetype].append(field)
        
        for issuetype, fields in fields_val.items():
            issuetype_fields = []
            for field in fields:
                issuetype_fields.append(field)
            start_row = 2
            start_column = ws.max_column + 1
            
            for f in issuetype_fields:
                ws.cell(row=start_row+1, column=start_column).value = f
                start_row += 1
            col_letter = ws.cell(row=start_row, column=start_column).column_letter
            issuetype_value = DefinedName(issuetype.replace('-', '_').replace(' ', '__'), attr_text='Project!$' + col_letter + '$3:$' + col_letter + '$' + str(len(issuetype_fields) + 2))
            wb.defined_names.append(issuetype_value)
            ws.column_dimensions[col_letter].hidden = True
        
        statuses_val = {}
        for issuetype, statuses in new_transitions.items():
            statuses_lst = []
            for status in statuses:
                statuses_lst.append(status[0])
                statuses_lst.append(status[2])
            statuses_lst = list(set(statuses_lst))
            statuses_val[issuetype] = statuses_lst
        
        for issuetype, statuses in statuses_val.items():
            issuetype_statuses = []
            for status in statuses:
                issuetype_statuses.append(status)
            start_row = 2
            start_column = ws.max_column + 1
            
            for s in issuetype_statuses:
                ws.cell(row=start_row+1, column=start_column).value = s
                start_row += 1
            col_letter = ws.cell(row=start_row, column=start_column).column_letter
            issuetype_value = DefinedName(issuetype.replace('-', '_').replace(' ', '__') + 'STATUS', attr_text='Project!$' + col_letter + '$3:$' + col_letter + '$' + str(len(issuetype_statuses) + 2))
            wb.defined_names.append(issuetype_value)
            ws.column_dimensions[col_letter].hidden = True
    
    if title == 'Issuetypes':
        start_row = 1
        start_column = ws.max_column + 1
        for i in default_validation['Issuetypes'].split(','):
            ws.cell(row=start_row+1, column=start_column).value = i.replace('"', '')
            start_row += 1
        col_letter = ws.cell(row=start_row, column=start_column).column_letter
        formula1 = '$' + col_letter + '$2:$' + col_letter + '$' + str(len(default_validation['Issuetypes'].split(',')) + 1)
        issuetypes_val = DataValidation(type="list", formula1=formula1, allow_blank=True)
        ws.add_data_validation(issuetypes_val)
        issuetypes_val.add(excel_columns_validation_ranges['1'])
        ws.column_dimensions[col_letter].hidden = True
    
    if title == 'Fields':
        issuetypes_val = DataValidation(type="list", formula1='INDIRECT(SUBSTITUTE(SUBSTITUTE(VLOOKUP($A2,Issuetypes!$A$2:$B$' + str(len(issue_details_old.keys()) + 1) + ',2,FALSE)," ","__"),"-","_"))', allow_blank=False)
        ws.add_data_validation(issuetypes_val)
        issuetypes_val.add(excel_columns_validation_ranges['2'])
    
    if title == 'Statuses':
        issuetypes_val = DataValidation(type="list", formula1='INDIRECT(CONCATENATE(SUBSTITUTE(SUBSTITUTE(VLOOKUP($A2,Issuetypes!$A$2:$B$' + str(len(issue_details_old.keys()) + 1) + ',2,FALSE)," ","__"),"-","_"),"STATUS"))', allow_blank=False)
        ws.add_data_validation(issuetypes_val)
        issuetypes_val.add(excel_columns_validation_ranges['2'])
    
    if title == 'Priority':
        priority_val = DataValidation(type="list", formula1=default_validation['Priority'], allow_blank=False)
        ws.add_data_validation(priority_val)
        priority_val.add(excel_columns_validation_ranges['1'])
    
    if title == 'Links' and len(default_validation['Links']) > 0:
        start_row = 1
        start_column = ws.max_column + 1
        for i in default_validation['Links'].split(','):
            ws.cell(row=start_row+1, column=start_column).value = i.replace('"', '')
            start_row += 1
        col_letter = ws.cell(row=start_row, column=start_column).column_letter
        formula1 = '$' + col_letter + '$2:$' + col_letter + '$' + str(len(default_validation['Links'].split(',')) + 1)
        links_val = DataValidation(type="list", formula1=formula1, allow_blank=True)
        ws.add_data_validation(links_val)
        links_val.add(excel_columns_validation_ranges['2'])
        ws.column_dimensions[col_letter].hidden = True
    
    ws.title = title
    
    # Coloring sheets
    if title == 'Project':
        ws.sheet_properties.tabColor = project_tab_color
    elif title in ['Issuetypes', 'Fields', 'Statuses', 'Priority']:
        ws.sheet_properties.tabColor = mandatory_tab_color
    elif title == 'Links':
        ws.sheet_properties.tabColor = optional_tab_color
    
    sheet_names = wb.sheetnames
    for s in sheet_names:
        ws = wb.get_sheet_by_name(s)
        if ws.dimensions == 'A1:A1':
            wb.remove_sheet(wb[s])
    
    # Hiding all non-mandatory sheets
    if hide_tabs is True and title not in mandatory_template_tabs:
        ws.sheet_state = 'hidden'


def save_excel():
    """Saving prepared Excel File. Applying zooming / scaling upon saving."""
    global zoom_scale, mapping_file, project_old, project_new, verbose_logging
    try:
        if mapping_file == '':
            mapping_file = "Mappings '{}'-'{}'.xlsx".format(project_old, project_new)
        
        if os.path.exists(mapping_file) is True:
            overwrite_popup()
        
        for ws in wb.worksheets:
            ws.sheet_view.zoomScale = zoom_scale
            ws.auto_filter.ref = ws.dimensions
        wb.save(mapping_file)
        print("[END] Mapping file '{}' successfully generated.".format(mapping_file))
        print('')
        sleep(2)
        exit()
    except Exception as e:
        if verbose_logging == 1:
            print(traceback.format_exc())
        print('')
        print("[ERROR] ", e)
        os.system("pause")
        exit()


def get_minfields_issuetype(issue_details, all=0):
    """Function for find out the issue type with minimal mandatory fields for Dummy issue creation."""
    min = 999
    i_types = {}
    min_type = ''
    min_fields = []
    
    for issuetype, fields in issue_details.items():
        mandatory_fields = set([field['id'] if field['required'] is True and field['id'] not in ['project', 'issuetype', 'summary'] else '' for field in fields.values()])
        if all == 0 and len(mandatory_fields) < min:
            min = len(mandatory_fields)
            min_type = issuetype
            min_fields = mandatory_fields
        else:
            mandatory_fields.remove('')
            i_types[issuetype] = mandatory_fields
    if all == 0:
        min_fields.remove('')
        return (min_type, min_fields)
    else:
        return i_types


def get_dummy_parent(retry=False, retry_number=None):
    global project_old, jira_new, jira_old, issue_details_new, issuetypes_mappings, dummy_parent
    global project_new, auth, headers, verify, json_importer_flag, max_number_for_dummy_parent_search, dummy_process
    global start_jira_key, verbose_logging, retry_number_allowed, multiple_json_data_processing
    
    if dummy_parent != '':
        return
    
    if retry_number is None:
        retry_number = retry_number_allowed
    if retry is False:
        print("[START] Searching / creating Dummy Parent for orphan Sub-Tasks.")
    jql_new = "project = {} AND summary ~ DUMMY_PARENT".format(project_new)
    parent = get_issues_by_jql(jira_new, jql_new, max_result=1)
    
    if parent is not None and parent != []:
        dummy_parent = parent[0]
        dummy_process = 1
        print("[END] Dummy Parent was found. Dummy Parent Key is '{}'".format(dummy_parent))
        print("")
        return
    
    if json_importer_flag == 1 or multiple_json_data_processing == 1:
        parent_key = None
        jql = "project = {}".format(project_old)
        total_old = jira_old.search_issues(jql, startAt=0, maxResults=1, json_result=True)['total']
        
        if int(total_old) > max_number_for_dummy_parent_search:
            try:
                max_key = find_max_id(project_old + '-' + str(max_number_for_dummy_parent_search), jira_old, project_old)
                jql = "project = {} AND key >= {} AND key <= {}".format(project_old, start_jira_key, max_key)
                total_old = jira_old.search_issues(jql, startAt=0, maxResults=1, json_result=True)['total']
            except:
                pass
        
        issues_for_parent = get_issues_by_jql(jira_old, jql, max_result=0)
        for i in range(1, len(issues_for_parent)):
            temp_key = str(project_old + '-' + str(i))
            if temp_key not in issues_for_parent:
                try:
                    parent_issue = jira_old.issue(temp_key)
                except:
                    parent_key = temp_key
                    break
        
        if parent_key is None:
            jql_max = 'project = {} order by key DESC'.format(project_old)
            max_processing_key = jira_new.search_issues(jql_str=jql_max, maxResults=1, json_result=False)[0].key
            parent_key = str(project_old + '-' + str(int(max_processing_key.split('-')[1]) + 1000))
        
        # Calculating Issuetype for Dummy Parent. Default would be 'Story'
        issuetype = 'Story'
        for k, v in issuetypes_mappings.items():
            if v['hierarchy'] in ['2', '3']:
                issuetype = k
                break
        
        parent_key = get_shifted_key(parent_key.replace(project_old + '-', project_new + '-'))
        url = JIRA_BASE_URL_NEW + JIRA_imported_api
        data = {}
        data["projects"] = []
        project_issue = {}
        project_details = {}
        project_details["key"] = project_new
        project_issue["key"] = parent_key
        project_issue["externalId"] = parent_key
        project_issue["issueType"] = issuetype
        project_issue["summary"] = "DUMMY_PARENT"
        project_issue["description"] = "DUMMY_PARENT (for migrated orphan Sub-Tasks)"
        project_details["issues"] = [project_issue]
        data["projects"].append(project_details)
        try:
            params = {"notifyUsers": "false"}
            r = requests.post(url, json=data, auth=auth, headers=headers, verify=verify, params=params)
            retry_number -= 1
            if str(r.status_code) == '202':
                dummy_parent = parent_key
                dummy_process = 0
            elif str(r.status_code) == '409' and retry_number > 0:
                sleep(retry_number_allowed - retry_number)
                get_dummy_parent(retry=True, retry_number=retry_number)
                return
            else:
                print("[ERROR] JSON Importer can't process '{}' issue. Dummy Parent can't be created.".format(parent_key))
        except Exception as e:
            if verbose_logging == 1:
                print(traceback.format_exc())
            print("[ERROR] JSON Importer error: '{}'".format(e))
            dummy_parent = None
        
        print("[END] Dummy Parent was created. Dummy Parent Key is '{}'".format(dummy_parent))
        print("")
        return


def delete_issue(key, retry_number=None):
    global auth, headers, JIRA_BASE_URL_NEW, JIRA_core_api, verify, verbose_logging, retry_number_allowed
    global max_retries
    
    if retry_number is None:
        retry_number = retry_number_allowed
    
    url = JIRA_BASE_URL_NEW + JIRA_core_api + str(key)
    try:
        retry_number -= 1
        r = requests.delete(url, auth=auth, headers=headers, verify=verify)
        if str(r.status_code) == '400':
            params = {"deleteSubtasks": "true"}
            r = requests.delete(url, auth=auth, headers=headers, verify=verify, params=params)
        if str(r.status_code) in ['204', '404']:
            return (0, key)
        elif retry_number > 0:
            sleep(retry_number_allowed - retry_number)
            return delete_issue(key, retry_number=retry_number)
        else:
            if max_retries == 1:
                print("[ERROR] Issue '{}' can't be deleted. Skipped.".format(key))
            return (1, key)
    except Exception as e:
        if max_retries == 1:
            print("[ERROR] Issue '{}' can't be deleted due to: '{}'".format(key, e))
        if verbose_logging == 1:
            print(traceback.format_exc())
        return (0, key)


def delete_extra_issues(max_id):
    """Function for removal extra Dummy Issues created via Migration Process (to have same ids while migration)"""
    global start_jira_key, jira_old, jira_new, project_new, project_old, verbose_logging, delete_dummy_flag, threads
    global recently_updated, max_retries, default_max_retries, supported_issuetypes
    
    # Check if that Issue available in the Source JIRA Project
    max_id = find_max_id(max_id, project=project_old, jira=jira_old)
    
    # Calculating total Number of Issues in OLD JIRA Project
    jql_total_old = "project = '{}' AND key >= {} AND key <= {} {} {}".format(project_old, start_jira_key, max_id, recently_updated, supported_issuetypes)
    total_old = jira_old.search_issues(jql_total_old, startAt=0, maxResults=1, json_result=True)['total']
    
    # Calculating total Number of Migrated Issues to NEW JIRA Project
    jql_total_new = "project = '{}' AND (labels not in ('MIGRATION_NOT_COMPLETE') OR labels is EMPTY) AND key >= {} AND key <= {}".format(project_new, get_shifted_key(start_jira_key.replace(project_old + '-', project_new + '-')), get_shifted_key(max_id.replace(project_old + '-', project_new + '-')))
    total_new = jira_new.search_issues(jql_total_new, startAt=0, maxResults=1, json_result=True)['total']
    
    print("[INFO] Total issues in Source Project: '{}' and total migrated issues: '{}'.".format(total_old, total_new))
    
    jql_total_new_for_deletion = "project = '{}' AND labels in ('MIGRATION_NOT_COMPLETE') AND key >= {} AND key <= {}".format(project_new, get_shifted_key(start_jira_key.replace(project_old + '-', project_new + '-')), get_shifted_key(max_id.replace(project_old + '-', project_new + '-')))
    total_new_for_deletion = jira_new.search_issues(jql_total_new_for_deletion, startAt=0, maxResults=1, json_result=True)['total']
    
    if delete_dummy_flag == 0:
        if total_old == total_new:
            print("[INFO] Total 'dummy' issues to be deleted in new project: '{}'.".format(total_new_for_deletion))
            if total_new_for_deletion > 0:
                print("[START] 'Dummy' issue deletion is started. Please wait...")
                issues_for_delete = get_issues_by_jql(jira_new, jql_total_new_for_deletion, max_result=0)
                if issues_for_delete is not None:
                    max_retries = default_max_retries
                    threads_processing(delete_issue, issues_for_delete)
                print("[END] 'Dummy' issues has been successfuly removed from target '{}' JIRA Project.".format(project_new))
        else:
            print("[ERROR] Not ALL issues have been migrated. 'Dummy' issues will not be removed to avoid any mapping issues.")
    else:
        print("[INFO] 'Dummy' issues will not be deleted due to 'Skip dummy deletion' flag was set.")


def create_dummy_issues(total_number, batch_size=100):
    """Creating Dummy Issue with all defaulted mandatory fields + specific Summary for further processing."""
    global issue_details_new, max_retries, default_max_retries, project_new, verbose_logging
    
    def create_issues(data_lst):
        global jira_new, verbose_logging
        
        try:
            issues = jira_new.create_issues(field_list=data_lst)
            if verbose_logging == 1:
                print("[INFO] Created dummy issues: '{}'".format(issues))
            return (0, data_lst)
        except Exception as e:
            print("[ERROR] Issues can't be created due to '{}'".format(e))
            if verbose_logging == 1:
                print(traceback.format_exc())
            return (1, data_lst)
    
    if total_number == 0:
        return
    
    print("[START] Dummy issues will be created. Total dummy issues to be created: '{}'.".format(total_number))
    issuetype = get_minfields_issuetype(issue_details_new)[0]
    fields = get_minfields_issuetype(issue_details_new)[1]
    
    new_data = {}
    new_data['project'] = project_new
    new_data['issuetype'] = eval('{"name": "' + issuetype + '"}')
    new_data['summary'] = "Dummy issue - for migration"
    new_data['labels'] = ['MIGRATION_NOT_COMPLETE']
    
    for field in fields:
        for f in issue_details_new[issuetype]:
            if issue_details_new[issuetype][f]['id'] == field:
                default_value = issue_details_new[issuetype][f]['default value']
                allowed = issue_details_new[issuetype][f]['allowed values']
                type = issue_details_new[issuetype][f]['type']
                custom_type = issue_details_new[issuetype][f]['custom type']
                if type == 'option':
                    value = allowed[0] if default_value is None else default_value
                    new_data[field] = eval('{"value": "' + value + '"}')
                elif field in ['components', 'versions', 'fixVersions'] or custom_type == 'multiversion':
                    value = allowed[0] if default_value is None else default_value
                    new_data[field] = eval('[{"name": "' + value + '"}]')
                elif type == 'option-with-child':
                    new_data[field] = eval('{"value": "' + allowed[0][0] + '", "child": {"value": "' + allowed[0][1] + '"}}')
                elif type == 'string':
                    new_data[field] = 'Dummy' if default_value is None else default_value
                elif type == 'number':
                    new_data[field] = 0 if default_value is None else default_value
                elif type == 'array' and issue_details_new[issuetype][f]['id'] != 'labels':
                    new_data[field] = ['Dummy'] if default_value is None else [default_value]
                else:
                    new_data[field] = default_value
    
    if total_number > batch_size:
        batch_count = total_number // batch_size + 1
        numbers = [[new_data for i in range(batch_size)] for j in range(batch_count - 1)]
        numbers.append([new_data for k in range(total_number - batch_size * (batch_count - 1))])
    else:
        numbers = [[new_data for i in range(total_number)]]
    
    max_retries = default_max_retries
    threads_processing(create_issues, numbers)
    print("[END] Dummy Issues have been created.")


def convert_to_subtask(parent, new_issue, sub_task_id):
    """ This function will convert issue to sub-task via parsing HTML page and apply emulation of conversion via UI. """
    global auth, verify, json_importer_flag, verbose_logging, dummy_parent
    
    if parent is None or parent == '':
        parent = dummy_parent
    
    session = requests.Session()
    
    url0 = JIRA_BASE_URL_NEW + '/secure/ConvertIssueSetIssueType.jspa?id=' + new_issue.id
    r = session.get(url=url0, auth=auth, verify=verify)
    soup = BeautifulSoup(r.text, features="lxml")
    try:
        guid = soup.find_all("input", type="hidden", id="guid")[0]['value']
    except Exception as e:
        if json_importer_flag == 0:
            print("[ERROR] Issue '{}' can't be converted to Sub-Task. Details: '{}'.".format(new_issue.key, e))
            if verbose_logging == 1:
                print(traceback.format_exc())
        return
    
    url_11 = JIRA_BASE_URL_NEW + '/secure/ConvertIssueSetIssueType.jspa'
    payload_11 = {
        "parentIssueKey": parent,
        "issuetype": sub_task_id,
        "id": new_issue.id,
        "guid": guid,
        "Next >>": "Next >>",
        "atl_token": session.cookies.get('atlassian.xsrf.token'),
    }
    
    r = session.post(url=url_11, data=payload_11, headers={"Referer": url0}, verify=verify)
    r.raise_for_status()
    
    url_12 = JIRA_BASE_URL_NEW + '/secure/ConvertIssueUpdateFields.jspa'
    payload_12 = {
        "id": new_issue.id,
        "guid": guid,
        "Next >>": "Next >>",
        "atl_token": session.cookies.get('atlassian.xsrf.token'),
    }
    
    r = session.post(url=url_12, data=payload_12, verify=verify)
    r.raise_for_status()
    
    url_13 = JIRA_BASE_URL_NEW + '/secure/ConvertIssueConvert.jspa'
    payload_13 = {
        "id": new_issue.id,
        "guid": guid,
        "Finish": "Finish",
        "atl_token": session.cookies.get('atlassian.xsrf.token'),
    }
    
    r = session.post(url=url_13, data=payload_13, verify=verify)
    r.raise_for_status()


def convert_to_issue(new_issue, issuetype):
    """ This function will convert sub-task to issue via parsing HTML page and apply emulation of conversion via UI. """
    global auth, verify, new_issues_ids, verbose_logging
    
    session = requests.Session()
    
    url0 = JIRA_BASE_URL_NEW + '/secure/ConvertSubTask.jspa?id=' + new_issue.id
    r = session.get(url=url0, auth=auth, verify=verify)
    soup = BeautifulSoup(r.text, features="lxml")
    try:
        guid = soup.find_all("input", type="hidden", id="guid")[0]['value']
    except Exception as e:
        print("[ERROR] Issue '{}' can't be converted to Issue from Sub-Task. Details: '{}'.".format(new_issue.key, e))
        if verbose_logging == 1:
            print(traceback.format_exc())
        return
    
    url_11 = JIRA_BASE_URL_NEW + '/secure/ConvertSubTaskSetIssueType.jspa'
    payload_11 = {
        "issuetype": new_issues_ids[issuetype],
        "id": new_issue.id,
        "guid": guid,
        "Next >>": "Next >>",
        "atl_token": session.cookies.get('atlassian.xsrf.token'),
    }
    
    r = session.post(url=url_11, data=payload_11, headers={"Referer": url0}, verify=verify)
    r.raise_for_status()
    
    url_12 = JIRA_BASE_URL_NEW + '/secure/ConvertSubTaskUpdateFields.jspa'
    payload_12 = {
        "id": new_issue.id,
        "guid": guid,
        "Next >>": "Next >>",
        "atl_token": session.cookies.get('atlassian.xsrf.token'),
    }
    
    r = session.post(url=url_12, data=payload_12, verify=verify)
    r.raise_for_status()
    
    url_13 = JIRA_BASE_URL_NEW + '/secure/ConvertSubTaskConvert.jspa'
    payload_13 = {
        "id": new_issue.id,
        "guid": guid,
        "Finish": "Finish",
        "atl_token": session.cookies.get('atlassian.xsrf.token'),
    }
    
    r = session.post(url=url_13, data=payload_13, verify=verify)
    r.raise_for_status()


def load_config(message=True):
    """Loading pre-saved values from 'config.json' file."""
    global mapping_file, JIRA_BASE_URL_OLD, JIRA_BASE_URL_NEW, project_old, project_new, last_updated_date, start_jira_key
    global team_project_prefix, old_board_id, default_board_name, temp_dir_name, limit_migration_data, threads, pool_size
    global template_project, new_project_name, verbose_logging, auth, username, password, credentials_saved_flag
    global default_configuration_file, bulk_processing_flag, max_json_file_size, protection_password
    
    if os.path.exists(config_file) is True:
        try:
            with open(config_file) as json_data_file:
                data = json.load(json_data_file)
            for k, v in data.items():
                if k == 'mapping_file':
                    mapping_file = v
                elif k == 'default_configuration_file':
                    default_configuration_file = v
                elif k == 'JIRA_BASE_URL_OLD':
                    JIRA_BASE_URL_OLD = v
                elif k == 'JIRA_BASE_URL_NEW':
                    JIRA_BASE_URL_NEW = v
                elif k == 'project_old':
                    project_old = v.strip()
                elif k == 'project_new':
                    project_new = v.strip()
                elif k == 'team_project_prefix':
                    team_project_prefix = v
                elif k == 'max_json_file_size':
                    max_json_file_size = v
                elif k == 'old_board_id':
                    old_board_id = v
                elif k == 'default_board_name':
                    default_board_name = v.strip()
                elif k == 'temp_dir_name':
                    temp_dir_name = v.strip()
                elif k == 'limit_migration_data':
                    limit_migration_data = v
                elif k == 'start_jira_key':
                    start_jira_key = v
                elif k == 'last_updated_date':
                    last_updated_date = v
                elif k == 'threads':
                    threads = v
                elif k == 'pool_size':
                    pool_size = v
                elif k == 'template_project':
                    template_project = v
                elif k == 'auth':
                    auth = v
                elif k == 'username':
                    username = v
                elif k == 'password':
                    password = v
                elif k == 'credentials_saved_flag':
                    credentials_saved_flag = v
                elif k == 'new_project_name':
                    new_project_name = v.strip()
                elif k == 'bulk_processing_flag':
                    bulk_processing_flag = v
                elif k == 'protection_password':
                    protection_password = v
            if message is True:
                print("[INFO] Configuration has been successfully loaded from '{}' file.".format(config_file))
        except Exception as er:
            print("[ERROR] Configuration file is corrupted. Default '{}' would be created instead.".format(config_file))
            print('')
            if verbose_logging == 1:
                print(traceback.format_exc())
            save_config()
    else:
        print("[INFO] Config File not found. Default '{}' would be created.".format(config_file))
        print("[INFO] Migration configuration default values will be load from that file.")
        print('')
        save_config()


def save_config(message=True):
    global credentials_saved_flag
    
    if credentials_saved_flag == 1:
        data = {'mapping_file': mapping_file,
                'default_configuration_file': default_configuration_file,
                'JIRA_BASE_URL_OLD': JIRA_BASE_URL_OLD,
                'JIRA_BASE_URL_NEW': JIRA_BASE_URL_NEW,
                'project_old': project_old,
                'project_new': project_new,
                'team_project_prefix': team_project_prefix,
                'max_json_file_size': max_json_file_size,
                'old_board_id': old_board_id,
                'default_board_name': default_board_name,
                'new_transitions': new_transitions,
                'temp_dir_name': temp_dir_name,
                'limit_migration_data': limit_migration_data,
                'start_jira_key': start_jira_key,
                'last_updated_date': last_updated_date,
                'threads': threads,
                'pool_size': pool_size,
                'template_project': template_project,
                'new_project_name': new_project_name,
                'auth': auth,
                'username': username,
                'password': password,
                'credentials_saved_flag': credentials_saved_flag,
                'bulk_processing_flag': bulk_processing_flag,
                'protection_password': protection_password,
                }
    else:
        data = {'mapping_file': mapping_file,
                'default_configuration_file': default_configuration_file,
                'JIRA_BASE_URL_OLD': JIRA_BASE_URL_OLD,
                'JIRA_BASE_URL_NEW': JIRA_BASE_URL_NEW,
                'project_old': project_old,
                'project_new': project_new,
                'team_project_prefix': team_project_prefix,
                'max_json_file_size': max_json_file_size,
                'old_board_id': old_board_id,
                'default_board_name': default_board_name,
                'new_transitions': new_transitions,
                'temp_dir_name': temp_dir_name,
                'limit_migration_data': limit_migration_data,
                'start_jira_key': start_jira_key,
                'last_updated_date': last_updated_date,
                'threads': threads,
                'pool_size': pool_size,
                'template_project': template_project,
                'new_project_name': new_project_name,
                'credentials_saved_flag': credentials_saved_flag,
                'bulk_processing_flag': bulk_processing_flag,
                'protection_password': protection_password,
                }
    
    try:
        with open(config_file, 'w') as outfile:
            json.dump(data, outfile)
    except PermissionError as er:
        print("[ERROR] File '{}' has been opened for editing and can't be saved. Exception: {}".format(config_file, er))
        return
    
    if message is True:
        print("[INFO] Config file '{}' has been created.".format(config_file))
    else:
        print("[INFO] Config file '{}' has been updated.".format(config_file))
    print("")


def get_statuses(jira_url, new=False):
    global headers, verify, JIRA_status_api, auth, old_statuses, new_statuses, verbose_logging
    
    statuses = {}
    try:
        url_statuses = jira_url + JIRA_status_api
        r = requests.get(url_statuses, auth=auth, headers=headers, verify=verify)
        statuses_string = r.content.decode('utf-8')
        statuses_details = json.loads(statuses_string)
        for status in statuses_details:
            if status["name"].upper() not in statuses.keys():
                statuses[status["name"].upper()] = {}
            statuses[status["name"].upper()]["key"] = status["statusCategory"]["key"]
            statuses[status["name"].upper()]["id"] = status["id"]
    except Exception as e:
        print("[ERROR] Statuses Categories can't be processed due to: '{}'.".format(e))
        if verbose_logging == 1:
            print(traceback.format_exc())
    
    if new is False:
        old_statuses = statuses
    else:
        new_statuses = statuses


def get_priority(new_issue_type, old_issue, message=False):
    global field_value_mappings, issue_details_new, issuetypes_mappings, project_new, max_retries, default_max_retries
    
    if new_issue_type is None:
        for k, v in issuetypes_mappings.items():
            if v['hierarchy'] in ['2', '3']:
                new_issue_type = k
                break
    default_priority = issue_details_new[new_issue_type]['Priority']['default value']
    proposed_priority = default_priority
    try:
        old_priority = old_issue.fields.priority.name
    except:
        old_priority = ''
    try:
        for new_value, old_values in field_value_mappings['Priority'].items():
            if str(old_priority.strip()) in old_values and new_value != '':
                proposed_priority = new_value
                if proposed_priority in issue_details_new[new_issue_type]['Priority']['allowed values']:
                    return proposed_priority
        if message is True and max_retries == default_max_retries:
            print("[WARNING] Priority '{}' hasn't been found in the Target '{}' project for '{}' issue. Default '{}' Priority would be used instead.".format(old_priority, project_new, old_issue.key, default_priority.upper()))
    except Exception as e:
        if message is True and max_retries == default_max_retries:
            print("[WARNING] Priority hasn't been found in the Target '{}' project for '{}' issue. Default '{}' Priority would be used instead. ERROR: '{}'.".format(project_new, old_issue.key, default_priority.upper(), e))
    return default_priority


def update_issues_json(data):
    global project_new, max_json_file_size, total_data, json_files_autoupload
    
    def create_json_file(data):
        global project_new, json_file_part_num
        filename = 'JSON_Importer_' + project_new + '_PART_' + str(json_file_part_num) + '.json'
        print("[INFO] Approximate size of the {} file to be created: '{}'".format(filename, str(float(len(str(data)) / 1024 / 1024)) + " Mb"))
        json_file_part_num += 1
        try:
            with open(filename, 'w') as outfile:
                json.dump(data, outfile)
            print("[INFO] File '{}' has been created.".format(filename))
        except:
            print("[ERROR] JSON File can't be created.")
    
    def upload_json_file(data, retry_number=None):
        global JIRA_BASE_URL_NEW, JIRA_imported_api, auth, headers, verify, retry_number_allowed, max_retries
        global json_files_autoupload, verbose_logging, json_file_part_num
        
        def check_json_load_status(url):
            global username, password, verify, json_file_part_num, sleep_count, default_sleep_time
            
            r1 = requests.get(url=json.loads(r.text)['job']['log'], auth=HTTPBasicAuth(username, password), verify=verify)
            if 'INFO - Finished Importing : Issue Links & Subtasks' in r1.text:
                print("")
                print("[INFO] The Part number '{}' was successfully processed by JIRA.".format(str(json_file_part_num)))
                sleep_count = 0
                return
            else:
                sleep(default_sleep_time)
                sleep_count += default_sleep_time
                if sleep_count % 100 == 0:
                    print("...")
                return check_json_load_status(url)
        
        if retry_number is None:
            retry_number = retry_number_allowed
        try:
            url = JIRA_BASE_URL_NEW + JIRA_imported_api
            params = {"notifyUsers": "false"}
            r = requests.post(url, json=data, auth=auth, headers=headers, verify=verify, params=params)
            if str(r.status_code) == '202':
                print("[INFO] Processing. Please wait. The log details available here: {}".format(json.loads(r.text)['job']['log']))
                check_json_load_status(json.loads(r.text)['job']['log'])
                json_file_part_num += 1
            elif retry_number <= 0:
                if max_retries == 1:
                    print("[ERROR] JSON Importer can't process generated data. Files will ve saved instead.")
                    create_json_file(data)
                    json_files_autoupload = 0
            else:
                print("[ERROR] JSON Importer: '{}'".format(r.content))
                retry_number -= 1
                sleep(retry_number_allowed - retry_number)
                upload_json_file(data, retry_number=retry_number)
        except Exception as e:
            print("[ERROR] JSON Importer error: '{}'".format(e))
            if verbose_logging == 1:
                print(traceback.format_exc())
    
    if json_files_autoupload == 0:
        print("[INFO] File would be created...")
    else:
        print("[INFO] Data would be uploaded into Target JIRA...")
    temp_data = {}
    temp_data["users"] = []
    temp_data["links"] = []
    temp_data["projects"] = [{"key": project_new, "issues": []}]
    number_processed = 0
    processed = 0
    current_size = 0
    max_size = float(max_json_file_size / 1.1)
    
    if "users" in data.keys():
        temp_size = float(len(str(data["users"])) / 1024 / 1024)
        if temp_size <= max_size:
            temp_size = float(len(str(temp_data)) / 1024 / 1024)
            temp_data["users"].extend(data["users"])
            total_data["users"] = []
            current_size += temp_size
        else:
            temp_data["users"] = []
            for count, user in enumerate(data["users"]):
                temp_data["users"].append(user)
                current_size += float(len(str(user)) / 1024 / 1024)
                if current_size > max_size:
                    if json_files_autoupload == 0:
                        create_json_file(temp_data)
                    else:
                        upload_json_file(temp_data)
                    processed = 1
                    current_size = 0
                    temp_data = {}
                    temp_data["users"] = []
                    temp_data["links"] = []
                    temp_data["projects"] = [{"key": project_new, "issues": []}]
                    number_processed = count
                    break
            if processed == 1 and number_processed > 0:
                total_data["users"] = total_data["users"][number_processed:]
            else:
                total_data["users"] = []
    
    number_processed = 0
    if processed == 0 and "projects" in data.keys() and "issues" in data["projects"][0].keys():
        if (current_size + float(len(str(data["projects"][0]["issues"])) / 1024 / 1024)) <= max_size:
            temp_data["projects"][0]["issues"].extend(data["projects"][0]["issues"])
            temp_size = float(len(str(temp_data)) / 1024 / 1024)
            total_data["projects"] = [{"key": project_new, "issues": []}]
            current_size += temp_size
        else:
            temp_data["projects"] = [{"key": project_new, "issues": []}]
            for count, issue in enumerate(data["projects"][0]["issues"]):
                temp_data["projects"][0]["issues"].append(issue)
                current_size += float(len(str(issue)) / 1024 / 1024)
                if current_size > max_size:
                    if json_files_autoupload == 0:
                        create_json_file(temp_data)
                    else:
                        upload_json_file(temp_data)
                    processed = 1
                    current_size = 0
                    temp_data = {}
                    temp_data["users"] = []
                    temp_data["links"] = []
                    temp_data["projects"] = [{"key": project_new, "issues": []}]
                    number_processed = count
                    break
            if processed == 1 and number_processed > 0:
                total_data["projects"][0]["issues"] = total_data["projects"][0]["issues"][number_processed:]
            else:
                total_data["projects"] = [{"key": project_new, "issues": []}]
    
    number_processed = 0
    if processed == 0 and "links" in data.keys():
        if (current_size + float(len(str(data["links"])) / 1024 / 1024)) <= max_size:
            temp_data["links"].extend(data["links"])
            temp_size = float(len(str(temp_data)) / 1024 / 1024)
            total_data["users"] = []
            current_size += temp_size
        else:
            temp_data["links"] = []
            for count, link in enumerate(data["links"]):
                temp_data["links"].append(link)
                current_size += float(len(str(link)) / 1024 / 1024)
                if current_size > max_size:
                    if json_files_autoupload == 0:
                        create_json_file(temp_data)
                    else:
                        upload_json_file(temp_data)
                    processed = 1
                    temp_data = {}
                    temp_data["users"] = []
                    temp_data["projects"] = [{"key": project_new, "issues": []}]
                    temp_data["links"] = []
                    number_processed = count
                    break
            if processed == 1 and number_processed > 0:
                total_data["links"] = total_data["links"][number_processed:]
            else:
                total_data["users"] = []
    
    if processed == 0:
        if json_files_autoupload == 0:
            create_json_file(temp_data)
        else:
            upload_json_file(temp_data)
        temp_data = {}
        temp_data["users"] = []
        temp_data["projects"] = [{"key": project_new, "issues": []}]
        temp_data["links"] = []
    else:
        try:
            rest_size = float(len(str(total_data)) / 1024 / 1024)
            if rest_size > max_size:
                update_issues_json(total_data)
        except:
            pass


def check_new_team(old_issue, new_issuetype):
    global teams, teams_to_be_added_set, fields_mappings, verbose_logging
    
    old_issuetype = old_issue.fields.issuetype.name
    if 'Team' in fields_mappings[old_issuetype].keys():
        for o_field in fields_mappings[old_issuetype]['Team']:
            o_field_value = get_value(old_field=o_field, new_field='Team', old_issue=old_issue, new_issuetype=new_issuetype, preprocess=True)
            if o_field == 'Team':
                o_field_value = get_team_name(o_field_value)
            if o_field_value is not None and o_field_value:
                o_field_value = o_field_value.strip()
                if o_field_value.upper() not in teams.keys() and o_field_value not in teams_to_be_added_set:
                    teams_to_be_added_set.add(o_field_value)
                    if verbose_logging == 1:
                        print("[INFO] New Team '{}' to be added in JIRA. From '{}' issue.".format(o_field_value, old_issue.key))
                break


def migrate_change_history(old_issue, new_issue_type, new_status, new=False, new_issue=None, subtask=None):
    global auth, verify, project_old, project_new, headers, JIRA_BASE_URL_NEW, JIRA_imported_api, new_board_id
    global issuetypes_mappings, issue_details_old, migrate_sprints_check, migrate_comments_check, including_users_flag
    global migrate_statuses_check, migrate_metadata_check, already_processed_json_importer_issues, max_json_file_size
    global multiple_json_data_processing, total_data, already_processed_users, total_processed, jira_old, dummy_parent
    global replace_complete_statuses_flag, verbose_logging, old_fields_ids_mapping, migrate_teams_check, teams
    global json_thread_lock, json_current_size, users_set, users, jira_new
    
    def check_status(new_status, new_issue_type):
        global new_transitions, max_retries, default_max_retries
        
        default_status = new_transitions[new_issue_type][0][0]
        for statuses in new_transitions[new_issue_type]:
            for status in statuses:
                if new_status.upper() == status.upper():
                    return status
        return default_status
    
    def get_new_complete_status(old_status_id, old_status_name, old_issuetype):
        global old_statuses, new_statuses, status_mappings
        
        try:
            if old_statuses[old_status_name.upper()]["key"] == 'done':
                for new_status, old_status_lst in status_mappings[old_issuetype].items():
                    if old_status_name in old_status_lst:
                        return (new_statuses[new_status.upper()]["id"], new_status)
            else:
                return (old_status_id, old_status_name)
        except:
            return (old_status_id, old_status_name)
    
    def get_watchers(jira, key):
        global already_processed_users, users_set, users
        
        user = {}
        watchers = []
        try:
            watcher = jira.watchers(key)
            if watcher.watchers != []:
                for w in watcher.watchers:
                    user_name = w.name
                    watchers.append(user_name)
                    if user_name not in users_set and user_name not in already_processed_users:
                        user["name"] = user_name
                        user["fullname"] = w.displayName
                        user["email"] = w.emailAddress
                        user["active"] = w.active
                        user["groups"] = ["jira-users"]
                        users_set.add(user_name)
                        users.append(user)
        except:
            pass
        return watchers
    
    def get_duration(jira_duration):
        weeks, days, hours, minutes, seconds = (0, 0, 0, 0, 0)
        if type(jira_duration) == int:
            remaining = jira_duration
            seconds = remaining
        else:
            time_lst = str(jira_duration).split(' ')
            for t in time_lst:
                if 'w' in t:
                    weeks = float(t.split('w')[0])
                elif 'd' in t:
                    days = float(t.split('d')[0])
                elif 'h' in t:
                    hours = float(t.split('h')[0])
                elif 'm' in t:
                    minutes = float(t.split('m')[0])
                elif 's' in t:
                    seconds = float(t.split('s')[0])
                else:
                    seconds = 0
        duration = isodate.duration_isoformat(datetime.timedelta(weeks=weeks, days=days, hours=hours, minutes=minutes, seconds=seconds))
        return duration

    if migrate_statuses_check == 0:
        new_status = None
    # Checking Portfolio Teams
    if multiple_json_data_processing == 1 and migrate_teams_check == 1 and teams != {}:
        try:
            check_new_team(old_issue, new_issue_type)
        except:
            pass
    
    existed_histories = []
    existed_worklogs = []
    existed_comments = []
    data = {}
    data["projects"] = []
    project_issue = {}
    project_details = {}
    project_details["key"] = project_new
    if total_data["projects"] == []:
        total_data["projects"] = [{"key": project_new, "issues": []}]
    project_details["issues"] = []
    histories = []
    worklogs = []
    comments = []
    sprints = []
    users = []
    users_set = set()
    if multiple_json_data_processing == 0:
        already_processed_users = set()
    
    # Checking the already existed data
    if new is False:
        for log in new_issue.raw['changelog']['histories']:
            created = datetime.datetime.strptime(log['created'], '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
            existed_histories.append(created)
        try:
            for log in new_issue.raw['fields']['worklog']['worklogs']:
                created = datetime.datetime.strptime(log['started'], '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
                existed_worklogs.append(created)
        except:
            pass
        try:
            for log in new_issue.raw['fields']['comment']['comments']:
                created = datetime.datetime.strptime(log['created'], '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
                existed_comments.append(created)
        except:
            pass
    
    # Processing Issue History here
    for log in old_issue.raw['changelog']['histories']:
        history = {}
        user = {}
        created = datetime.datetime.strptime(log['created'], '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
        if created not in existed_histories:
            if 'author' in log:
                user_name = log['author']['name'].upper()
                if user_name not in users_set and user_name not in already_processed_users:
                    user["name"] = user_name
                    user["fullname"] = log['author']['displayName']
                    user["email"] = log['author']['emailAddress']
                    user["active"] = log['author']['active']
                    user["groups"] = ["jira-users"]
                    users_set.add(user_name)
                    already_processed_users.add(user_name)
                    users.append(user)
                history["author"] = user_name
            created = datetime.datetime.strptime(log['created'], '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
            history["created"] = created
            history["items"] = []
            for item in log['items']:
                new_item = {}
                new_item["fieldType"] = item['fieldtype']
                new_item["field"] = item['field']
                new_item["from"] = item['from']
                new_item["fromString"] = item['fromString']
                if replace_complete_statuses_flag == 1 and item['field'] == 'status':
                    try:
                        (new_item["to"], new_item["toString"]) = get_new_complete_status(item['to'], item['toString'], old_issue.fields.issuetype.name)
                    except:
                        new_item["to"] = item['to']
                        new_item["toString"] = item['toString']
                else:
                    new_item["to"] = item['to']
                    new_item["toString"] = item['toString']
                history["items"].append(new_item)
            histories.append(history)
    
    # Processing Issue Worklogs here
    for log in old_issue.raw['fields']['worklog']['worklogs']:
        worklog = {}
        user = {}
        created = datetime.datetime.strptime(log['started'], '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
        if created not in existed_worklogs:
            if 'author' in log:
                user_name = log['author']['name'].upper()
                if user_name not in users_set and user_name not in already_processed_users:
                    user["name"] = user_name
                    user["fullname"] = log['author']['displayName']
                    user["email"] = log['author']['emailAddress']
                    user["active"] = log['author']['active']
                    user["groups"] = ["jira-users"]
                    users_set.add(user_name)
                    already_processed_users.add(user_name)
                    users.append(user)
                worklog["author"] = user_name
            worklog["startDate"] = datetime.datetime.strptime(log['started'], '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
            worklog["timeSpent"] = get_duration(log["timeSpent"])
            try:
                worklog["comment"] = log["comment"]
            except:
                pass
            worklogs.append(worklog)
    
    # Processing Comments here
    for log in old_issue.raw['fields']['comment']['comments']:
        comment = {}
        created = datetime.datetime.strptime(log['created'], '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
        if created not in existed_comments:
            user = {}
            try:
                user_name = log['author']['name'].upper()
                if user_name not in users_set and user_name not in already_processed_users:
                    user["name"] = user_name
                    user["fullname"] = log['author']['displayName']
                    user["email"] = log['author']['emailAddress']
                    user["active"] = log['author']['active']
                    user["groups"] = ["jira-users"]
                    users_set.add(user_name)
                    already_processed_users.add(user_name)
                    users.append(user)
                comment["author"] = user_name
            except:
                user_name = 'Anonymous'
                if user_name not in users_set:
                    user["name"] = user_name
                    user["fullname"] = 'Anonymous'
                    users.append(user)
                    users_set.add(user_name)
                comment["author"] = user_name
            comment["created"] = datetime.datetime.strptime(log['created'], '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
            comment["body"] = log["body"]
            comments.append(comment)
    
    # Sprints
    if subtask is None and migrate_sprints_check == 1:
        try:
            try:
                sprint_field_id = issue_details_old[old_issue.fields.issuetype.name]['Sprint']['id']
            except:
                sprint_field_id = old_fields_ids_mapping['Sprint']
            issue_sprints = eval('old_issue.fields.' + sprint_field_id)
            if issue_sprints is not None:
                for sprint in issue_sprints:
                    sprint_detail = {}
                    name, state, start_date, end_date, complete_date = ('', '', None, None, None)
                    for attr in sprint[sprint.find('[')+1:-1].split(','):
                        if 'name=' in attr:
                            name = attr.split('name=')[1]
                        if 'state=' in attr:
                            state = attr.split('state=')[1]
                        if 'startDate=' in attr:
                            start_date = None if attr.split('startDate=')[1] == '<null>' else datetime.datetime.strptime(attr.split('startDate=')[1][:-3]+'00', '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
                        if 'endDate=' in attr:
                            end_date = None if attr.split('endDate=')[1] == '<null>' else datetime.datetime.strptime(attr.split('endDate=')[1][:-3]+'00', '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
                        if 'completeDate=' in attr:
                            complete_date = None if attr.split('completeDate=')[1] == '<null>' else datetime.datetime.strptime(attr.split('completeDate=')[1][:-3]+'00', '%Y-%m-%dT%H:%M:%S.%f%z').astimezone(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.%f%z')
                    sprint_detail["rapidViewId"] = new_board_id
                    sprint_detail["state"] = state
                    sprint_detail["startDate"] = start_date
                    sprint_detail["endDate"] = end_date
                    sprint_detail["completeDate"] = complete_date
                    sprint_detail["name"] = name
                    sprints.append(sprint_detail)
                project_issue["customFieldValues"] = [{"fieldName": "Sprint", "fieldType": "com.pyxis.greenhopper.jira:gh-sprint", "value": sprints}]
        except:
            pass
    
    project_issue["key"] = get_shifted_key(old_issue.key.replace(project_old + '-', project_new + '-'))
    if new is True:
        project_issue["externalId"] = get_shifted_key(old_issue.key.replace(project_old + '-', project_new + '-'))
    project_issue["issueType"] = new_issue_type
    
    if subtask is not None and new is True:
        try:
            parent_id = get_shifted_key(old_issue.fields.parent.key.replace(project_old + '-', project_new + '-'))
            try:
                parent = jira_new.issue(get_shifted_key(old_issue.fields.parent.key.replace(project_old + '-', project_new + '-')))
                if parent.fields.issuetype.name in sub_tasks:
                    parent_id = dummy_parent
            except:
                pass
            link = {"name": "sub-task-link",
                    "sourceId": get_shifted_key(old_issue.key.replace(project_old + '-', project_new + '-')),
                    "destinationId": parent_id
                    }
            data["links"] = [link]
        except:
            if dummy_parent != '':
                link = {"name": "sub-task-link",
                        "sourceId": get_shifted_key(old_issue.key.replace(project_old + '-', project_new + '-')),
                        "destinationId": dummy_parent
                        }
                data["links"] = [link]
    
    try:
        user = {}
        user_name = old_issue.fields.reporter.name.upper()
        if user_name not in users_set and user_name not in already_processed_users:
            user["name"] = user_name
            user["fullname"] = old_issue.fields.reporter.displayName
            user["email"] = old_issue.fields.reporter.emailAddress
            user["active"] = old_issue.fields.reporter.active
            user["groups"] = ["jira-users"]
            users_set.add(user_name)
            already_processed_users.add(user_name)
            users.append(user)
        if migrate_metadata_check == 1:
            project_issue["reporter"] = user_name
    except:
        pass
    try:
        user = {}
        user_name = old_issue.fields.assignee.name.upper()
        if user_name not in users_set and user_name not in already_processed_users:
            user["name"] = user_name
            user["fullname"] = old_issue.fields.assignee.displayName
            user["email"] = old_issue.fields.assignee.emailAddress
            user["active"] = old_issue.fields.assignee.active
            user["groups"] = ["jira-users"]
            users_set.add(user_name)
            already_processed_users.add(user_name)
            users.append(user)
        if migrate_metadata_check == 1:
            project_issue["assignee"] = user_name
    except:
        pass
    if migrate_comments_check == 1:
        project_issue["comments"] = comments
    if migrate_metadata_check == 1:
        try:
            project_issue["originalEstimate"] = get_duration(old_issue.fields.timeoriginalestimate)
        except:
            project_issue["originalEstimate"] = None
        try:
            project_issue["timeSpent"] = get_duration(old_issue.fields.timeSpent)
        except:
            project_issue["timeSpent"] = None
        try:
            project_issue["estimate"] = get_duration(old_issue.fields.timeestimate)
        except:
            project_issue["estimate"] = None
        if new_status is not None:
            project_issue["status"] = check_status(new_status, new_issue_type)
        project_issue["resolutionDate"] = old_issue.fields.resolutiondate
        project_issue["resolution"] = None if old_issue.fields.resolution is None else old_issue.fields.resolution.name
        project_issue["priority"] = get_priority(new_issue_type, old_issue, message=False)
        project_issue["created"] = old_issue.fields.created
        project_issue["summary"] = old_issue.fields.summary.replace('\n', ' ').replace('\t', ' ')
        project_issue["updated"] = old_issue.fields.updated
        project_issue["watchers"] = get_watchers(jira_old, old_issue.key)
        if new is True:
            project_issue["labels"] = ['MIGRATION_NOT_COMPLETE']
    project_issue["history"] = histories
    project_issue["worklogs"] = worklogs
    project_details["issues"].append(project_issue)
    data["projects"].append(project_details)
    if including_users_flag == 1:
        data["users"] = users
    
    if multiple_json_data_processing == 1 and old_issue.key not in already_processed_json_importer_issues:
        already_processed_json_importer_issues.add(old_issue.key)
        total_data["projects"][0]["issues"].append(project_issue)
        json_current_size += float(len(str(project_issue))) / 1024 / 1024
        if 'users' in data.keys():
            total_data["users"].extend(data["users"])
            json_current_size += float(len(str(data["users"]))) / 1024 / 1024
        if 'links' in data.keys():
            total_data["links"].extend(data["links"])
            json_current_size += float(len(str(data["links"]))) / 1024 / 1024
        if len(already_processed_json_importer_issues) > 0 and len(already_processed_json_importer_issues) % 500 == 0:
            print("[INFO] Processed '{}' out of '{}' issues so far.".format(len(already_processed_json_importer_issues), total_processed))
        json_thread_lock.acquire()
        if json_current_size > float(max_json_file_size / 1.1):
            update_issues_json(total_data)
            json_current_size = float(len(str(total_data))) / 1024 / 1024
        json_thread_lock.release()
        return '202'
    elif json_importer_flag == 1 and multiple_json_data_processing == 0:
        already_processed_json_importer_issues.add(old_issue.key)
        try:
            url = JIRA_BASE_URL_NEW + JIRA_imported_api
            params = {"notifyUsers": "false"}
            r = requests.post(url, json=data, auth=auth, headers=headers, verify=verify, params=params)
            return (r.status_code, r.content)
        except Exception as e:
            print("[ERROR] JSON Importer error: '{}'".format(e))
            if verbose_logging == 1:
                print(traceback.format_exc())
            return (0, e)


def get_correct_components(components):
    global project_new, jira_new
    
    new_components = jira_new.project_components(project_new)
    new_components_lst = []
    for new_component in new_components:
        new_components_lst.append(new_component.name.strip())
    
    new_components_detail = []
    for component in components:
        if component['name'] in new_components_lst:
            new_components_detail.append({"name": component['name'].strip()})
    return new_components_detail


def get_correct_versions(versions):
    global project_new, jira_new
    
    new_versions = jira_new.project_versions(project_new)
    new_versions_lst = []
    for new_version in new_versions:
        new_versions_lst.append(new_version.name.strip())
    
    new_versions_detail = []
    for version in versions:
        if version['name'] in new_versions_lst:
            new_versions_detail.append({"name": version['name'].strip()})
    return new_versions_detail


def check_user(user):
    global JIRA_BASE_URL_NEW, JIRA_users_api, auth, headers, verify
    
    def create_user(user):
        global JIRA_BASE_URL_NEW, JIRA_create_users_api, auth, headers, verify, including_users_flag
        if including_users_flag == 1:
            try:
                user_name = user.name
                user_display_name = user.displayName
                user_email = user.emailAddress
            except:
                user_name = user
                user_display_name = user
                user_email = user
            data = {"emailAddress": user_email,
                    "displayName": user_display_name,
                    "name": user_name
                    }
            payload = json.dumps(data)
            url = JIRA_BASE_URL_NEW + JIRA_create_users_api
            r = requests.post(url, data=payload, auth=auth, headers=headers, verify=verify)
            if str(r.status_code) == '201':
                return True
        return False
    
    if user is None:
        return False
    try:
        user_name = user.name
    except:
        user_name = user
    try:
        url = JIRA_BASE_URL_NEW + JIRA_users_api.format(user_name)
        r = requests.get(url, auth=auth, headers=headers, verify=verify)
        users_string = r.content.decode('utf-8')
        user_details = json.loads(users_string)
        if 'errorMessages' in user_details:
            try:
                return create_user(user)
            except:
                return False
        else:
            return True
    except:
        return False


def get_new_value_from_mapping(old_value, field_name):
    global field_value_mappings
    try:
        for new_value, old_values in field_value_mappings[field_name].items():
            if type(old_value) == list:
                for o_val in old_value:
                    if str(o_val.strip()) in old_values or str(o_val.strip()) == new_value:
                        return new_value
            else:
                if str(old_value.strip()) in old_values or str(old_value.strip()) == new_value:
                    return new_value
        return old_value
    except:
        return old_value


def get_value(old_field, new_field, old_issue, new_issuetype, preprocess=False):
    global issue_details_old, issue_details_new, teams_thread_lock
    
    old_value = None
    old_issuetype = old_issue.fields.issuetype.name
    
    try:
        value = eval('old_issue.fields.' + issue_details_old[old_issuetype][old_field]['id'])
    except:
        try:
            value = eval('old_issue.fields.' + old_field.strip())
        except:
            value = None
    if value is None and old_field not in ['Source Status', 'Source Issuetype']:
        try:
            temp = issue_details_old[old_issuetype][old_field]['type']
        except:
            if max_retries == default_max_retries:
                print("[ERROR] Field '{}' for '{}' Issue Type in Mapping Template can't be found in Source Project. Please check for extra spaces missing.".format(old_field, old_issuetype))
                print("[INFO] Available fields for the Source Project's '{}' issuetype are: '{}'".format(old_issuetype, issue_details_old[old_issuetype].keys()))
        try:
            temp = issue_details_new[new_issuetype][new_field]['type']
        except:
            if max_retries == default_max_retries:
                print("[ERROR] Field '{}' for '{}' Issue Type in Mapping Template can't be found in Target Project. Please check for extra spaces missing.".format(new_field, new_issuetype))
                print("[INFO] Available fields for the Target Project's '{}' issuetype are: '{}'".format(new_issuetype, issue_details_new[new_issuetype].keys()))
    if issue_details_old[old_issuetype][old_field]['type'] == 'string' and issue_details_old[old_issuetype][old_field]['custom type'] == 'textfield' and issue_details_old[old_issuetype][old_field]['custom type'] != 'com.atlassian.teams:rm-teams-custom-field-team':
        try:
            value = value.replace('\n', '').replace('\t', ' ')
        except:
            pass
    elif issue_details_old[old_issuetype][old_field]['type'] == 'number' and issue_details_old[old_issuetype][old_field]['custom type'] != 'com.atlassian.teams:rm-teams-custom-field-team':
        try:
            value = int(float(str(value).replace('\n', '').replace('\t', ' ')))
        except:
            pass
    elif issue_details_old[old_issuetype][old_field]['custom type'] == 'labels' and issue_details_old[old_issuetype][old_field]['custom type'] != 'com.atlassian.teams:rm-teams-custom-field-team':
        value = get_str_from_lst(value)
    elif issue_details_old[old_issuetype][old_field]['type'] == 'option-with-child' and value is not None and issue_details_old[old_issuetype][old_field]['custom type'] != 'com.atlassian.teams:rm-teams-custom-field-team':
        value_value = value.value
        try:
            value_child = value.child.value
            mapped_value = get_new_value_from_mapping(value_value + ' --> ' + value_child, new_field)
        except:
            value_child = None
            mapped_value = value_value
        if mapped_value is not None and value_child is not None:
            try:
                mapped_value_value = value.split(' --> ')[0]
                mapped_value_child = value.split(' --> ')[1]
            except:
                mapped_value_value = value_value
                mapped_value_child = value_child
        else:
            mapped_value_value = value_value
            mapped_value_child = value_child
        if issue_details_new[new_issuetype][new_field]['type'] == 'option-with-child':
            if issue_details_new[new_issuetype][new_field]['validated'] is True:
                for values in issue_details_new[new_issuetype][new_field]['allowed values']:
                    if mapped_value_value == values[0] and mapped_value_child == values[1]:
                        old_value = {"value": mapped_value_value, "child": {"value": mapped_value_child}}
                        return old_value
                    else:
                        old_value = None
            else:
                old_value = {"value": value_value, "child": {"value": value_child}}
        elif issue_details_new[new_issuetype][new_field]['type'] in ['option']:
            if issue_details_new[new_issuetype][new_field]['validated'] is True:
                for values in issue_details_new[new_issuetype][new_field]['allowed values']:
                    if mapped_value_value == values:
                        old_value = mapped_value_value
                        return old_value
                    if mapped_value_child == values:
                        old_value = mapped_value_child
                        return old_value
                old_value = None
                return old_value
        else:
            if value_value is not None and value_child is not None:
                old_value = value_value + ' --> ' + value_child
            elif value_child is not None:
                old_value = value_value
            else:
                old_value = value_child
    elif issue_details_old[old_issuetype][old_field]['custom type'] == 'com.atlassian.teams:rm-teams-custom-field-team':
        value = get_team_name(value)
    elif issue_details_old[old_issuetype][old_field]['custom type'] in ['multiversion', 'multiuserpicker', 'multiselect', 'multicheckboxes'] and value is not None:
        value = [item.value if hasattr(item, "value") else item.name if hasattr(item, "name") else item for item in value]
    old_value = value
    
    if issue_details_old[old_issuetype][old_field]['type'] in ['string', 'number', 'array'] and issue_details_new[new_issuetype][new_field]['custom type'] != 'com.atlassian.teams:rm-teams-custom-field-team':
        old_value = get_str_from_lst(value)
        if issue_details_new[new_issuetype][new_field]['type'] == 'option-with-child':
            value = get_new_value_from_mapping(value[0] if type(value) == list else value, new_field)
            try:
                value_value = value.split(' --> ')[0]
                value_child = value.split(' --> ')[1]
                old_value = {"value": value_value, "child": {"value": value_child}}
            except:
                old_value = value
        elif issue_details_new[new_issuetype][new_field]['custom type'] == 'labels' or new_field == 'Labels':
            old_value = str(old_value).replace(' ', '_').replace('\n', '_').replace('\t', '_')
    elif issue_details_old[old_issuetype][old_field]['type'] in ['option', 'user'] and issue_details_new[new_issuetype][new_field]['custom type'] != 'com.atlassian.teams:rm-teams-custom-field-team':
        if issue_details_old[old_issuetype][old_field]['custom type'] in ['userpicker'] and old_value is not None:
            try:
                old_value = [item.value if hasattr(item, "value") else item.name if hasattr(item, "name") else item for item in old_value]
            except:
                old_value = [old_value]
        else:
            old_value = value.value if hasattr(value, "value") else value.name if hasattr(value, "name") else value
    elif issue_details_new[new_issuetype][new_field]['custom type'] == 'com.atlassian.teams:rm-teams-custom-field-team':
        if new_issuetype in sub_tasks.keys():
            return None
        else:
            team_value = None
            try:
                team_value_obj = old_value[0] if type(old_value) == list else old_value
                team_value = team_value_obj.value if hasattr(team_value_obj, "value") else team_value_obj.name if hasattr(team_value_obj, "name") else team_value_obj
                if type(team_value) == list:
                    team_value = team_value[0]
            except:
                old_value = None
            if preprocess is False:
                teams_thread_lock.acquire()
                team = '' if old_value is None else get_team_id(team_value)
                teams_thread_lock.release()
            else:
                team = '' if old_value is None else team_value
            return team
    else:
        return get_new_value_from_mapping(old_value, new_field)
    
    return get_new_value_from_mapping(old_value, new_field)


def get_label_value(value):
    if value is None:
        return ''
    elif hasattr(value, 'displayName'):
        return value.displayName
    elif hasattr(value, 'name'):
        return value.name
    elif hasattr(value, 'value'):
        return value.value
    else:
        return value


def update_parent_link(key, parent, field_id):
    global JIRA_BASE_URL_NEW, JIRA_update_parent_link_api, headers, verify, auth, verbose_logging
    
    url_parent_link_update = JIRA_BASE_URL_NEW + JIRA_update_parent_link_api
    if parent is None:
        params = {'key': key,
                  'fieldId': field_id}
    else:
        params = {'key': key,
                  'fieldId': field_id,
                  'parent': parent}
    r = requests.post(url_parent_link_update, auth=auth, headers=headers, verify=verify, params=params)
    if verbose_logging == 1:
        print(r.content)


def update_team(key, new_team_id, field_id):
    global JIRA_BASE_URL_NEW, JIRA_update_team_api, headers, verify, auth, verbose_logging
    
    url_parent_link_update = JIRA_BASE_URL_NEW + JIRA_update_team_api
    if new_team_id is None:
        params = {'key': key,
                  'fieldId': field_id}
    else:
        params = {'key': key,
                  'fieldId': field_id,
                  'newId': new_team_id}
    r = requests.post(url_parent_link_update, auth=auth, headers=headers, verify=verify, params=params)
    if verbose_logging == 1:
        print(r.content)


def update_new_issue_type(old_issue, new_issue, issuetype):
    """Function for Issue Metadata Update - the most complicated part of the migration"""
    global issue_details_old, issuetypes_mappings, sub_tasks, issue_details_new, create_remote_link_for_old_issue
    global jira_new, items_lst, json_importer_flag, migrate_teams_check, including_dependencies_flag, dummy_parent
    global jira_system_skip_fields, old_fields_ids_mapping, force_update_flag, last_updated_days_check
    global processed_issues_set, process_only_last_updated_date
    
    old_issuetype = old_issue.fields.issuetype.name
    
    def get_old_system_field(new_field, old_issue=old_issue, new_issuetype=issuetype):
        global issue_details_old, new_sprints, issuetypes_mappings
        
        old_issuetype = old_issue.fields.issuetype.name
        
        if new_field == 'Sprint':
            if issuetype in sub_tasks.keys() or issuetypes_mappings[issuetype]['hierarchy'] in ['0', '1']:
                return None
            else:
                sprint_field = get_old_field('Sprint')
                issue_sprints = None if sprint_field is None else sprint_field
                if issue_sprints is not None:
                    new_issue_sprints = []
                    for sprint in issue_sprints:
                        name = ''
                        for attr in sprint[sprint.find('[')+1:-1].split(','):
                            if 'name=' in attr:
                                name = attr.split('name=')[1]
                            if name in new_sprints.keys():
                                new_issue_sprints.append(new_sprints[name]['id'])
                                break
                    if len(new_issue_sprints) == 0:
                        new_issue_sprints = None
                    else:
                        # Only one LAST Sprint will be assigned to the issue ## TO DO
                        new_issue_sprints = new_issue_sprints[-1]
                else:
                    new_issue_sprints = None
            return new_issue_sprints
        try:
            value = eval('old_issue.fields.' + issue_details_old[old_issuetype][new_field]['id'])
            try:
                if value == []:
                    return value
            except:
                pass
        except:
            return None
        if type(value) == list:
            cont_value = []
            for v in value:
                if hasattr(v, 'name'):
                    if ((issue_details_old[old_issuetype][new_field]['type'] == 'user' and check_user(v))
                        or issue_details_old[old_issuetype][new_field]['type'] != 'user'):
                        cont_value.append({"name": get_new_value_from_mapping(v.name, new_field)})
                elif hasattr(v, 'value'):
                    if issue_details_new[issuetype][new_field]['custom type'] == 'multiselect':
                        if issue_details_new[new_issuetype][new_field]['validated'] is True and v is not None:
                            found = 0
                            for values in issue_details_new[new_issuetype][new_field]['allowed values']:
                                if str(v.value) == str(values):
                                    cont_value.append({"value": str(v.value)})
                                    found = 1
                                    break
                            if found == 0:
                                cont_value.append({"value": None})
                        else:
                            cont_value.append({"value": get_new_value_from_mapping(v.value, new_field)})
                else:
                    cont_value.append(get_new_value_from_mapping(v, new_field))
            return cont_value
        else:
            if hasattr(value, 'name'):
                if issue_details_old[old_issuetype][new_field]['type'] == 'user' and not check_user(value):
                    return None
                elif new_field == 'Priority' and get_new_value_from_mapping(value.name, new_field) == '':
                    return {"name": issue_details_new[issuetype]['Priority']['default value']}
                else:
                    return {"name": get_new_value_from_mapping(value.name, new_field)}
            elif hasattr(value, 'value'):
                return {"value": get_new_value_from_mapping(value.value, new_field)}
            else:
                if issue_details_new[issuetype][new_field]['type'] == 'array':
                    if value is None:
                        value = []
                    else:
                        try:
                            value = value.split(',')
                        except:
                            value = [value.replace(' ', '_').replace('\n', '_').replace('\t', '_')]
                elif value is None and issue_details_new[issuetype][new_field]['type'] == 'string':
                    value = ''
                elif new_field in ['Epic Link', 'Parent Link']:
                    if value is None:
                        return None
                    else:
                        value = get_shifted_key(value.replace(project_old + '-', project_new + '-'))
                elif new_field in ['Summary']:
                    value = value.replace('\n', ' ').replace('\t', ' ')
                elif issue_details_new[issuetype][new_field]['custom type'] == 'float':
                    try:
                        if type(value) == str:
                            value = value.strip()
                        value = float(value)
                    except:
                        value = None
                return value
    
    def get_old_field(new_field, old_issue=old_issue, new_issuetype=issuetype, data_val={}):
        global fields_mappings, issue_details_old, issue_details_new, max_retries, default_max_retries
        value = None
        concatenated_value = None
        processed = False
        old_issuetype = old_issue.fields.issuetype.name
        
        old_field = ''
        o_field_val = ''
        try:
            if new_field in fields_mappings[old_issuetype].keys():
                old_field = fields_mappings[old_issuetype][new_field]
        except:
            return old_field
        if old_field == '':
            try:
                if new_field in issue_details_old[old_issuetype].keys():
                    old_field = [new_field]
            except:
                if new_field == 'Sprint' and new_field in issue_details_old[old_issuetype].keys():
                    val = eval('old_issue.fields.' + issue_details_old[old_issuetype][new_field]['id'])
                    return val
                return value
        for o_field in old_field:
            if 'issuetype.name' in o_field or 'issuetype.status' in o_field:
                if 'issuetype.name' in o_field:
                    if issue_details_new[new_issuetype][new_field]['type'] != 'string' and issue_details_new[new_issuetype][new_field]['type'] != 'array':
                        try:
                            value_type = get_str_from_lst(old_issue.fields.issuetype.name)
                        except:
                            value_type = None
                        return get_new_value_from_mapping(value_type, new_field)
                    o_field = 'Source Issuetype'
                    try:
                        o_field_val = get_str_from_lst(old_issue.fields.issuetype.name)
                    except:
                        o_field_val = get_new_value_from_mapping(old_issuetype, new_field)
                elif 'issuetype.status' in o_field:
                    if issue_details_new[new_issuetype][new_field]['type'] != 'string' and issue_details_new[new_issuetype][new_field]['type'] != 'array':
                        try:
                            value_status = get_str_from_lst(old_issue.fields.status.name)
                        except:
                            value_status = None
                        return get_new_value_from_mapping(value_status, new_field)
                    o_field = 'Source Status'
                    try:
                        o_field_val = get_str_from_lst(old_issue.fields.status.name)
                    except:
                        o_field_val = ''
                processed = True
            if issue_details_new[new_issuetype][new_field]['type'] == 'string':
                if concatenated_value is None:
                    if new_field != 'Description':
                        concatenated_value = ''
                    else:
                        try:
                            concatenated_value = data_val['description'] + '\r\n----\r\n'
                        except:
                            concatenated_value = '----\r\n'
                if processed is True:
                    added_value = '' if o_field_val is None else o_field_val
                else:
                    try:
                        try:
                            calculated_value = [i.displayName for i in get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype)]
                        except:
                            try:
                                calculated_value = [i.name for i in get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype)]
                            except:
                                calculated_value = [i.value for i in get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype)]
                    except:
                        calculated_value = get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype)
                    added_value = '' if calculated_value is None else get_str_from_lst(calculated_value)
                if new_field == 'Description':
                    concatenated_value += '' if added_value == '' else '\r\n *[' + o_field + ']:* ' + added_value
                elif issue_details_new[new_issuetype][new_field]['custom type'] == 'textarea':
                    concatenated_value += '' if get_str_from_lst(added_value) == '' else '[' + o_field + ']: ' + get_str_from_lst(added_value) + ' \r\n \r\n'
                else:
                    concatenated_value += '' if get_str_from_lst(added_value) == '' else '[' + o_field + ']: ' + get_str_from_lst(added_value) + ' '
            elif issue_details_new[new_issuetype][new_field]['type'] == 'number':
                if concatenated_value is None:
                    concatenated_value = 0
                try:
                    added_value = int(get_str_from_lst(re.findall(r"\d*", get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype))))
                except:
                    added_value = 0
                concatenated_value += 0 if get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype) is None else added_value
            elif issue_details_new[new_issuetype][new_field]['type'] == 'array':
                if concatenated_value is None:
                    if new_field != 'Labels' or (new_field == 'Labels' and 'labels' in data_val.keys() and data_val['labels'] is None):
                        concatenated_value = []
                    else:
                        if 'labels' not in data_val.keys():
                            data_val['labels'] = []
                        concatenated_value = data_val['labels']
                if issue_details_new[new_issuetype][new_field]['custom type'] == 'labels' or new_field == 'Labels':
                    if processed is True:
                        label_add_value = o_field_val
                    else:
                        label_add_value = get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype)
                    if label_add_value is not None and type(label_add_value) == list:
                        concatenated_value.extend([str(get_label_value(i)).strip('_').replace(' ', '_').replace('\n', '_').replace('\t', '_').strip() for i in label_add_value])
                    else:
                        try:
                            values = label_add_value.split(',')
                            for value in values:
                                concatenated_value.append('' if value is None else str(value).strip('_').replace(' ', '_').replace('\n', '_').replace('\t', '_').strip())
                        except:
                            if label_add_value is not None:
                                concatenated_value.append('' if label_add_value is None else str(label_add_value).strip('_').replace(' ', '_').replace('\n', '_').replace('\t', '_').strip())
                elif issue_details_new[new_issuetype][new_field]['custom type'] == 'rs.codecentric.label-manager-project:labelManagerCustomField':
                    if processed is True:
                        value = str(o_field_val).strip('_').replace(' ', '_').replace('\n', '_').replace('\t', '_').strip()
                    else:
                        try:
                            values = get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype).split(',')
                            for val in values:
                                value = str(get_label_value(val)).strip('_').strip().replace(' ', '_').replace('\n', '_').replace('\t', '_')
                                if value not in get_lm_field_values(new_field, new_issuetype):
                                    add_lm_field_value(value, new_field, new_issuetype)
                                concatenated_value.append(value)
                        except:
                            value = get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype)
                            if value is not None:
                                concatenated_value.append(value)
                elif issue_details_new[new_issuetype][new_field]['custom type'] == 'multiuserpicker':
                    if get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype) is not None:
                        if type(get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype)) == list:
                            concatenated_value.extend(get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype))
                        else:
                            concatenated_value.append(get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype))
                else:
                    concatenated_value.append('' if get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype) is None else str(get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype)))
            elif issue_details_new[new_issuetype][new_field]['type'] == 'option':
                value = str(get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype))
                if issue_details_new[new_issuetype][new_field]['validated'] is True:
                    for values in issue_details_new[new_issuetype][new_field]['allowed values']:
                        if value == values:
                            existing_value = value
                            return existing_value
                concatenated_value = value
            else:
                value = get_value(o_field, new_field=new_field, old_issue=old_issue, new_issuetype=issuetype)
                if str(value) != '':
                    return value
        value = concatenated_value
        return value
    
    def update_issuetype(issuetype, old_issuetype, new_issue=new_issue):
        global issue_details_new, verbose_logging
        data = {}
        mandatory_fields = get_minfields_issuetype(issue_details_new, all=1)
        data['issuetype'] = {'name': issuetype}
        for field in mandatory_fields[old_issuetype]:
            for f in issue_details_new[old_issuetype]:
                if issue_details_new[old_issuetype][f]['id'] == field:
                    default_value = issue_details_new[old_issuetype][f]['default value']
                    allowed = issue_details_new[old_issuetype][f]['allowed values']
                    type = issue_details_new[old_issuetype][f]['type']
                    custom_type = issue_details_new[old_issuetype][f]['custom type']
                    if type == 'option':
                        value = allowed[0] if default_value is None else default_value
                        data[field] = eval('{"value": "' + value + '"}')
                    elif field in ['components', 'versions', 'fixVersions'] or custom_type == 'multiversion':
                        value = allowed[0] if default_value is None else default_value
                        data[field] = eval('[{"name": "' + value + '"}]')
                    elif type == 'option-with-child':
                        data[field] = eval('{"value": "' + allowed[0][0] + '", "child": {"value": "' + allowed[0][1] + '"}}')
                    elif type == 'string':
                        data[field] = 'Dummy' if default_value is None else default_value
                    elif type == 'number':
                        data[field] = 0 if default_value is None else default_value
                    elif type == 'array':
                        data[field] = ['Dummy'] if default_value is None else [default_value]
                    else:
                        data[field] = default_value
        try:
            new_issue.update(notify=False, fields=data)
        except Exception as e:
            if verbose_logging == 1:
                print("[ERROR] Exception for updating Issuetype for '{}' issue from '{}' to '{}'.".format(new_issue.key, old_issuetype, issuetype))
                print("[ERROR] Exception: '{}'".format(e))
            new_issue.update(notify=True, fields=data)
    
    data_val = {}
    data_value = None
    diff_issuetypes = 0
    try:
        parent = new_issue.fields.parent.key
    except:
        parent = None
    new_issuetype = new_issue.fields.issuetype.name
    if new_issuetype != issuetype:
        diff_issuetypes = 1
    # Checking for Sub-Task and convert to Sub-Task if necessary
    if issuetype in sub_tasks.keys():
        parent_calculated = get_parent_for_subtask(old_issue, old_issuetype)
        try:
            existent_parent = None if new_issue.fields.parent is None else new_issue.fields.parent.key
        except:
            existent_parent = None
        parent = None if parent_calculated is None else parent_calculated
        if parent is not None and new_issuetype != issuetype:
            convert_to_subtask(parent, new_issue, sub_tasks[issuetype])
            diff_issuetypes = 0
        elif existent_parent is None and new_issuetype == issuetype:
            delete_issue(new_issue.key)
            return process_issue(old_issue.key)
    elif (new_issuetype not in sub_tasks.keys() and parent is not None) or (new_issuetype in sub_tasks.keys() and issuetype not in sub_tasks.keys()):
        convert_to_issue(new_issue, issuetype)
        diff_issuetypes = 0
    
    data_val['summary'] = old_issue.fields.summary
    data_val['issuetype'] = {'name': issuetype}
    if diff_issuetypes == 1:
        update_issuetype(issuetype, new_issuetype)
    
    # Non-linked Custom fields - clear them out
    for new_field in issue_details_new[issuetype].keys():
        if issuetype in sub_tasks.keys() and new_field in ['Sprint', 'Parent Link', 'Team']:
            continue
        if new_field not in jira_system_skip_fields and new_field not in jira_system_fields and old_issuetype in fields_mappings.keys() and process_only_last_updated_date_flag == 0:
            if issue_details_new[issuetype][new_field]['type'] in ['string']:
                data_value = ''
            elif issue_details_new[issuetype][new_field]['type'] in ['array']:
                data_value = []
            else:
                data_value = None
            data_val[issue_details_new[issuetype][new_field]['id']] = data_value
    
    # System fields
    for n_field, n_values in issue_details_new[issuetype].items():
        if issuetype in sub_tasks.keys() and n_field in ['Sprint', 'Parent Link', 'Team']:
            continue
        elif n_field not in issue_details_old[old_issuetype].keys() and n_field in jira_system_fields:
            try:
                orig_data = eval('old_issue.fields.' + old_fields_ids_mapping[n_field])
                if type(orig_data) == list:
                    data_val[issue_details_new[issuetype][n_field]['id']] = [{"value": i.value} if hasattr(i, 'value') else {"name": i.name} if hasattr(i, 'name') else i for i in orig_data]
                else:
                    data_val[issue_details_new[issuetype][n_field]['id']] = {"value": orig_data.value} if hasattr(orig_data, 'value') else {'name': orig_data.name} if hasattr(orig_data, 'name') else orig_data
            except:
                continue
        elif (n_values['custom type'] is None and n_field not in jira_system_skip_fields) or (n_field in jira_system_fields):
            try:
                data_val[n_values['id']] = get_old_system_field(n_field)
            except:
                continue
        # Check for mandatory fields
        if n_field not in jira_system_skip_fields and issue_details_new[issuetype][n_field]['required']is True and (data_val[issue_details_new[issuetype][n_field]['id']] is None
                                                                         or data_val[issue_details_new[issuetype][n_field]['id']] == ''):
            if issue_details_new[issuetype][n_field]['type'] in ['string']:
                data_value = 'TBD'
            elif issue_details_new[issuetype][n_field]['type'] in ['array']:
                if issue_details_new[issuetype][n_field]['custom type'] == 'labels':
                    data_value = ['TBD']
                else:
                    data_value = [{'value': issue_details_new[issuetype][n_field]['allowed values'][0] if issue_details_new[issuetype][n_field]['allowed values'] is not None else {'value': 'TBD'}}]
            elif issue_details_new[issuetype][n_field]['type'] in ['option']:
                data_value = {'value': issue_details_new[issuetype][n_field]['allowed values'][0]}
            data_val[issue_details_new[issuetype][n_field]['id']] = data_value

    # Post_fix for labels
    if 'labels' in data_val.keys() and data_val['labels'] is None:
        try:
            data_val['labels'] = old_issue.fields.labels
        except:
            data_val['labels'] = []
    
    # Custom fields
    if old_issuetype in fields_mappings.keys():
        for n_field in fields_mappings[old_issuetype].keys():
            if ((issuetype in sub_tasks.keys() and n_field in ['Sprint', 'Parent Link', 'Team'])
                or n_field == ''
                or n_field not in issue_details_new[issuetype].keys()
                or n_field in jira_system_skip_fields
                or (n_field in jira_system_fields and n_field not in additional_mapping_fields)):
                continue
            data_value = None
            o_field_value = get_old_field(n_field, data_val=data_val)
            n_field_value = '' if (o_field_value is None or o_field_value == 'None') else o_field_value
            if issue_details_new[issuetype][n_field]['type'] in ['number', 'date']:
                data_value = None if n_field_value == '' else n_field_value
            elif issue_details_new[issuetype][n_field]['type'] in ['string']:
                data_value = '' if n_field_value == '' else get_str_from_lst(n_field_value)
            elif issue_details_new[issuetype][n_field]['type'] in ['user', 'array']:
                if issue_details_new[issuetype][n_field]['custom type'] == 'multiuserpicker':
                    if type(n_field_value) == list and n_field_value != '':
                        data_value = []
                        for i in n_field_value:
                            if i is not None and check_user(i):
                                try:
                                    try:
                                        users = jira_new.search_users(i.name)
                                    except:
                                        users = jira_new.search_users(i)
                                    if users == []:
                                        data_value.append({"name": None})
                                    else:
                                        try:
                                            data_value.append({"name": users[0].name})
                                        except:
                                            try:
                                                data_value.append({"name": i.name})
                                            except:
                                                data_value.append({"name": i})
                                except:
                                    data_value.append({"name": None})
                        if data_value == []:
                            data_value.append({"name": None})
                    else:
                        if not check_user(n_field_value) and n_field_value != '':
                            print("[WARNING] No '{}' User found on Target JIRA instance for '{}' issuetype. Field: '{}'.".format(n_field_value.name, old_issue.key, n_field))
                            n_field_value = ''
                        data_value = None if n_field_value == '' else [{"name": n_field_value.name}]
                if issue_details_new[issuetype][n_field]['custom type'] == 'userpicker':
                    if type(n_field_value) == list and n_field_value != '':
                        data_value = None
                        for i in n_field_value:
                            if i is not None and check_user(i):
                                try:
                                    try:
                                        users = jira_new.search_users(i.name)
                                    except:
                                        users = jira_new.search_users(i)
                                    if users == []:
                                        data_value = {"name": None}
                                    else:
                                        try:
                                            data_value = {"name": users[0].name}
                                        except:
                                            try:
                                                data_value = {"name": i.name}
                                            except:
                                                data_value = {"name": i}
                                except:
                                    data_value = {"name": None}
                    else:
                        if not check_user(n_field_value) and n_field_value != '':
                            print("[WARNING] No '{}' User found on Target JIRA instance for '{}' issuetype. Field: '{}'.".format(n_field_value, old_issue.key, n_field))
                            n_field_value = ''
                        data_value = None if n_field_value == '' else {"name": n_field_value.name} if hasattr(n_field_value, 'name') else {"name": n_field_value}
                elif issue_details_new[issuetype][n_field]['custom type'] in ['labels', 'rs.codecentric.label-manager-project:labelManagerCustomField'] or n_field == 'Labels':
                    if type(n_field_value) == list and n_field_value != '':
                        data_value = []
                        for i in n_field_value:
                            added_val = str(i.replace(' ', '_').replace('\n', '_').replace('\t', '_'))
                            if len(added_val) >= 255 and i is not None and i != 'None' and i != ' ':
                                print("[WARNING] Label '{}' is longer than allowed 255 characters. It would be trimmed.".format(added_val))
                                data_value.append(added_val[:255])
                            elif i is not None and i != 'None' and i != ' ':
                                data_value.append(added_val)
                    else:
                        added_val = str(n_field_value.replace(' ', '_').replace('\n', '_').replace('\t', '_'))
                        if len(added_val) >= 255:
                            print("[WARNING] Label '{}' is longer than allowed 255 characters. It would be trimmed.".format(added_val))
                            data_value = [added_val[:255]]
                        else:
                            data_value = [added_val]
                    data_value = list(set(data_value))
                    if '' in data_value:
                        data_value.remove('')
                elif issue_details_new[issuetype][n_field]['custom type'] == 'multiselect':
                    if issue_details_new[new_issuetype][n_field]['validated'] is True and n_field_value is not None:
                        data_value = []
                        for val in n_field_value:
                            for values in issue_details_new[new_issuetype][n_field]['allowed values']:
                                if str(val) == str(values):
                                    data_value.append({"value": str(val)})
                                    break
                    else:
                        data_value = [{"name": get_str_from_lst(n_field_value)}]
                elif issue_details_new[issuetype][n_field]['custom type'] == 'multiuserpicker':
                    data_value = None if data_value == [] else data_value
                elif issue_details_new[issuetype][n_field]['type'] in ['array']:
                    data_value = None if n_field_value == [''] else n_field_value
                else:
                    data_value = None if n_field_value == '' else {"name":  str(n_field_value)}
            elif issue_details_new[issuetype][n_field]['type'] in ['option'] and issue_details_new[issuetype][n_field]['validated'] is True:
                data_value = None
                for value in issue_details_new[new_issuetype][n_field]['allowed values']:
                    if str(n_field_value) == str(value):
                        data_value = {"value":  str(n_field_value)}
                        break
            elif issue_details_new[issuetype][n_field]['type'] == 'option-with-child':
                if n_field_value == '':
                    data_value = '[!DROP]'
                elif type(n_field_value) == str:
                    try:
                        value_value = n_field_value.split(' --> ')[0]
                        value_child = n_field_value.split(' --> ')[1]
                        if issue_details_new[new_issuetype][n_field]['validated'] is True:
                            for values in issue_details_new[new_issuetype][n_field]['allowed values']:
                                if value_value == values[0] and value_child == values[1]:
                                    data_value = {"value": value_value, "child": {"value": value_child}}
                                    break
                    except:
                        data_value = None
                else:
                    data_value = n_field_value
            else:
                data_value = n_field_value
            
            # Cheking the field MAX lenght and trimming all the extra info
            if issue_details_new[issuetype][n_field]['custom type'] == 'textfield' and len(data_value) > 255:
                print("[WARNING] The value in '{}' field would be trimmed. It exceeds the allowed limit of 255 characters.".format(n_field))
                print("[INFO] Removed part: '{}'".format(data_value[254:]))
                data_value = data_value[:254]
            
            if data_value != '[!DROP]':
                data_val[issue_details_new[issuetype][n_field]['id']] = data_value
            
            # Check for mandatory fields
            if issue_details_new[issuetype][n_field]['required']is True and (data_val[issue_details_new[issuetype][n_field]['id']] is None
                                                                             or data_val[issue_details_new[issuetype][n_field]['id']] == ''):
                if issue_details_new[issuetype][n_field]['type'] in ['string']:
                    data_value = 'TBD'
                elif issue_details_new[issuetype][n_field]['type'] in ['array']:
                    if issue_details_new[issuetype][n_field]['custom type'] == 'labels':
                        data_value = ['TBD']
                    else:
                        data_value = [{'value': issue_details_new[issuetype][n_field]['allowed values'][0] if issue_details_new[issuetype][n_field]['allowed values'] is not None else {'value': 'TBD'}}]
                elif issue_details_new[issuetype][n_field]['type'] in ['option']:
                    data_value = {'value': issue_details_new[issuetype][n_field]['allowed values'][0]}
                data_val[issue_details_new[issuetype][n_field]['id']] = data_value
    
    # Fix for Team management JIRA Portfolio Team field - JPOSERVER-2322
    if migrate_teams_check == 1 and issuetype not in sub_tasks.keys():
        try:
            new_existent_team = eval('new_issue.fields.' + issue_details_new[issuetype]['Team']['id'])
            if type(new_existent_team) == list:
                new_existent_team = new_existent_team[0]
            try:
                new_existent_team = new_existent_team.value.upper()
            except:
                try:
                    new_existent_team = new_existent_team.name.upper()
                except:
                    new_existent_team = new_existent_team.upper()
            if issue_details_new[issuetype]['Team']['custom type'] == 'com.atlassian.teams:rm-teams-custom-field-team':
                new_existent_team = get_team_name(new_existent_team.strip(), jira_url=JIRA_BASE_URL_NEW).upper().strip()
        except:
            new_existent_team = None
        try:
            new_team = data_val[issue_details_new[issuetype]['Team']['id']].upper()
            if type(new_team) == list:
                new_team = new_team[0]
            if issue_details_new[issuetype]['Team']['custom type'] == 'com.atlassian.teams:rm-teams-custom-field-team':
                new_team = get_team_name(new_team.strip(), jira_url=JIRA_BASE_URL_NEW).upper().strip()
        except:
            new_team = None
        if new_team == '':
            new_team = None
            try:
                update_team(new_issue.key, None, issue_details_new[issuetype]['Team']['id'])
            except:
                data_val.pop(issue_details_new[issuetype]['Team']['id'], None)
        if new_team != new_existent_team and new_existent_team is None:
            if verbose_logging == 1:
                print("[INFO] Team would be updated to '{}'".format(new_team))
        elif json_importer_flag == 1 and new_team != new_existent_team and new_existent_team is not None:
            try:
                update_team(new_issue.key, get_team_id(new_team), issue_details_new[issuetype]['Team']['id'])
            except:
                delete_issue(new_issue.key)
                return process_issue(old_issue.key)
        else:
            try:
                data_val.pop(issue_details_new[issuetype]['Team']['id'], None)
            except:
                pass
    elif issuetype in sub_tasks.keys():
        try:
            data_val.pop(issue_details_new[issuetype]['Team']['id'], None)
        except:
            pass
    # Due to JIRA issue JPOSERVER-2322, Teams can't be updated / cleared from Issue via update. Fixing:
    try:
        if data_val[issue_details_new[issuetype]['Team']['id']] is None:
            data_val.pop(issue_details_new[issuetype]['Team']['id'], None)
    except:
        pass
    
    # Post-processing for Assignee, if issue was converted from Sub-Task
    new_assignee = None
    new_assignee_key = None
    old_assignee = None
    old_assignee_key = None
    if new_issue.fields.assignee is not None and check_user(new_issue.fields.assignee):
        new_assignee = new_issue.fields.assignee.name
        new_assignee_key = new_issue.fields.assignee.key
    if old_issue.fields.assignee is not None:
        old_assignee = old_issue.fields.assignee.name
        old_assignee_key = old_issue.fields.assignee.key
    if (new_assignee != old_assignee or new_assignee_key != old_assignee_key) and old_assignee is not None and check_user(old_issue.fields.assignee.name):
        data_val['assignee'] = {"name": old_issue.fields.assignee.name}
    elif new_assignee != old_assignee and old_assignee is None:
        data_val['assignee'] = None
    
    # Post-processing for Reporter for JSON importer case
    try:
        new_assignee_name = None if new_issue.fields.assignee is None else new_issue.fields.assignee.name
        new_assignee_key = None if new_issue.fields.assignee is None else new_issue.fields.assignee.key
        old_assignee_name = None if old_issue.fields.assignee is None else old_issue.fields.assignee.name
        old_assignee_key = None if old_issue.fields.assignee is None else old_issue.fields.assignee.key
        if diff_issuetypes != 0 and (new_assignee_name == old_assignee_name or new_assignee_key == old_assignee_key or not check_user(old_assignee_name)):
            data_val.pop('assignee', None)
    except:
        pass
    try:
        new_reporter_name = None if new_issue.fields.reporter is None else new_issue.fields.reporter.name
        new_reporter_key = None if new_issue.fields.reporter is None else new_issue.fields.reporter.key
        old_reporter_name = None if old_issue.fields.reporter is None else old_issue.fields.reporter.name
        old_reporter_key = None if old_issue.fields.reporter is None else old_issue.fields.reporter.key
        if new_reporter_name == old_reporter_name or new_reporter_key == old_reporter_key or not check_user(old_reporter_name):
            data_val.pop('reporter', None)
    except:
        pass
    try:
        if new_issue.fields.priority.name == get_priority(issuetype, old_issue, message=True):
            data_val.pop('priority', None)
    except:
        pass
    if json_importer_flag == 1 or multiple_json_data_processing == 1:
        try:
            data_val.pop(issue_details_new[issuetype]['Sprint']['id'], None)
        except:
            pass
    
    # Post-processing fix for Assignee, Reporter
    if 'assignee' in data_val.keys() and data_val['assignee'] == []:
        data_val['assignee'] = None
    if 'reporter' in data_val.keys() and data_val['reporter'] == []:
        # Reporter is mandatory, can't be set to 'None' or ''
        data_val.pop('reporter', None)
    
    # Post-processing fix for Components, versions
    if 'components' in data_val.keys() and data_val['components'] is None:
        data_val['components'] = []
    
    # Post-processing for OLD Components / Versions with spaces in the very beginning
    fix_version_name_new, version_name_new, components_name_new = (None, None, None)
    for name, values in issue_details_new[issuetype].items():
        if 'fixVersions' in values["id"]:
            fix_version_name_new = name
        elif 'versions' in values["id"]:
            version_name_new = name
        elif 'components' in values["id"]:
            components_name_new = name
    
    if (('versions' not in data_val.keys()
         or ('versions' in data_val.keys() and data_val['versions'] == [])
         or ('versions' in data_val.keys() and data_val['versions'] is None))
        and version_name_new is not None):
        try:
            temp_versions = [{'name': i.name.strip()} for i in old_issue.fields.versions]
            data_val['versions'] = get_correct_versions(temp_versions)
        except:
            pass
    elif 'versions' in data_val.keys() and data_val['versions'] != [] and data_val['versions'] is not None:
        temp_versions = []
        for version in data_val['versions']:
            temp_versions.append({'name': version['name'].strip()})
        data_val['versions'] = get_correct_versions(temp_versions)
    
    if (('fixVersions' not in data_val.keys()
         or ('fixVersions' in data_val.keys() and data_val['fixVersions'] == [])
         or ('fixVersions' in data_val.keys() and data_val['fixVersions'] is None))
        and fix_version_name_new is not None):
        try:
            temp_versions = [{'name': i.name.strip()} for i in old_issue.fields.fixVersions]
            data_val['fixVersions'] = get_correct_versions(temp_versions)
        except:
            pass
    elif 'fixVersions' in data_val.keys() and data_val['fixVersions'] != [] and data_val['fixVersions'] is not None:
        temp_versions = []
        for version in data_val['fixVersions']:
            temp_versions.append({'name': version['name'].strip()})
        data_val['fixVersions'] = get_correct_versions(temp_versions)
    
    if (('components' not in data_val.keys()
         or ('components' in data_val.keys() and data_val['components'] == [])
         or ('components' in data_val.keys() and data_val['components'] is None))
        and components_name_new is not None):
        try:
            temp_components = [{'name': i.name.strip()} for i in old_issue.fields.components]
            data_val['components'] = get_correct_components(temp_components)
        except:
            pass
    elif 'components' in data_val.keys() and data_val['components'] != [] and data_val['components'] is not None:
        temp_components = []
        for component in data_val['components']:
            temp_components.append({'name': component['name'].strip()})
        data_val['components'] = get_correct_components(temp_components)
    
    # Post-processing fix for Parent Links (which is not part of migration)
    try:
        parent_link_id_to_add = data_val[issue_details_new[issuetype]['Parent Link']['id']]
        try:
            existent_parent = eval('new_issue.fields.' + issue_details_new[issuetype]['Parent Link']['id'])
        except:
            existent_parent = None
        if issue_details_new[issuetype]['Parent Link']['id'] in data_val.keys() and parent_link_id_to_add is not None:
            try:
                parent_issue = jira_old.issue(get_shifted_key(parent_link_id_to_add.replace(project_new + '-', project_old + '-'), reversed=True))
                if existent_parent is not None and parent_issue.key == get_shifted_key(existent_parent.replace(project_new + '-', project_old + '-'), reversed=True):
                    data_val.pop(issue_details_new[issuetype]['Parent Link']['id'], None)
                elif parent_link_id_to_add != existent_parent and existent_parent is not None:
                    if json_importer_flag == 1:
                        try:
                            update_parent_link(new_issue.key, parent_link_id_to_add, issue_details_new[issuetype]['Parent Link']['id'])
                        except:
                            delete_issue(new_issue.key)
                            return process_issue(old_issue.key)
                    else:
                        data_val.pop(issue_details_new[issuetype]['Parent Link']['id'], None)
                elif parent_link_id_to_add is not None and existent_parent is None and json_importer_flag == 1:
                    for k, v in issuetypes_mappings.items():
                        if parent_issue.fields.issuetype.name in v['issuetypes'] and project_old in str(parent_link_id_to_add):
                            process_issue(parent_link_id_to_add.replace(project_new + '-', project_old + '-'), reprocess=True)
                else:
                    print("[WARNING] Parent '{}' can't be processed for '{}'. Dropped.".format(parent_link_id_to_add, new_issue.key))
                    data_val.pop(issue_details_new[issuetype]['Parent Link']['id'], None)
            except:
                print("[WARNING] Parent Link '{}' for '{}' can't be found in the same JIRA instance / project.".format(parent_link_id_to_add, new_issue.key))
                data_val.pop(issue_details_new[issuetype]['Parent Link']['id'], None)
        elif issue_details_new[issuetype]['Parent Link']['id'] in data_val.keys() and parent_link_id_to_add is None and existent_parent is not None:
            try:
                update_parent_link(new_issue.key, None, issue_details_new[issuetype]['Parent Link']['id'])
            except:
                delete_issue(new_issue.key)
                return process_issue(old_issue.key)
    except:
        pass
    
    # Post-processing fix for Epic Links (which is not part of migration)
    try:
        if issue_details_new[issuetype]['Epic Link']['id'] in data_val.keys():
            try:
                epic_new_issue = jira_new.issue(data_val[issue_details_new[issuetype]['Epic Link']['id']])
            except:
                try:
                    epic_issue = jira_old.issue(data_val[issue_details_new[issuetype]['Epic Link']['id']].replace(project_new + '-', project_old + '-'))
                    for k, v in issuetypes_mappings.items():
                        if epic_issue.fields.issuetype.name in v['issuetypes'] and project_old in str(epic_issue.key):
                            process_issue(data_val[issue_details_new[issuetype]['Epic Link']['id']], reprocess=True)
                            break
                    else:
                        data_val.pop(issue_details_new[issuetype]['Epic Link']['id'], None)
                except:
                    data_val.pop(issue_details_new[issuetype]['Epic Link']['id'], None)
    except:
        pass
    
    # Post-processing for Description (if empty and no mappings from other fields)
    if 'description' in data_val.keys():
        if data_val['description'] == '\r\n----\r\n' or data_val['description'] is None:
            data_val['description'] = ' '
        if len(data_val['description']) > 32767:
            trimmed_data = data_val['description'][32767:]
            data_val['description'] = str(data_val['description'].encode('utf-8', errors='ignore'))[:32767]
            print("[WARNING] '{}' - 'Description' field value is too long. The trimmed data: '{}'".format(new_issue.key, trimmed_data))
    
    # Post-processing for Labels
    if (json_importer_flag == 1 or multiple_json_data_processing == 1) and 'labels' not in data_val.keys():
        data_val['labels'] = []
    
    # Post-processing for Epic Name
    try:
        if issue_details_new[issuetype]['Epic Name']['id'] in data_val.keys() and (data_val[issue_details_new[issuetype]['Epic Name']['id']] is None or data_val[issue_details_new[issuetype]['Epic Name']['id']] == ''):
            data_val[issue_details_new[issuetype]['Epic Name']['id']] = data_val['summary']
    except:
        pass
    
    if verbose_logging == 1:
        print("[INFO] The currently processing: '{}'".format(old_issue.key))
        print("[INFO] The details for update: '{}'".format(data_val))
        print("")
    
    # Post-processing for other system-related fields:
    if "worklog" in data_val.keys():
        data_val.pop("worklog", None)
    
    try:
        new_issue.update(notify=False, fields=data_val)
        processed_issues_set.add(old_issue.key)
    except Exception as er:
        try:
            if "epic.error.not.found" in er.text:
                data_val.pop(issue_details_new[issuetype]['Epic Link']['id'], None)
            new_issue.update(notify=False, fields=data_val)
            processed_issues_set.add(old_issue.key)
        except Exception as e:
            try:
                if "User" in e.text and 'does not exist' in e.text:
                    user_name = e.text.split('\'')[1]
                    if "assignee" in data_val.keys() and data_val['assignee'] is not None and data_val['assignee']['name'] == user_name:
                        data_val.pop('assignee', None)
                    if "reporter" in data_val.keys() and data_val['reporter'] is not None and data_val['reporter']['name'] == user_name:
                        data_val.pop('reporter', None)
                elif "The reporter specified is not a user" in e.text:
                    data_val.pop('reporter', None)
                elif "cannot be assigned issues" in e.text:
                    data_val.pop('assignee', None)
                elif "does not exist for the field 'project'." in e.text:
                    try:
                        new_issue = jira_new.issue(new_issue.key)
                    except Exception as er:
                        print("[ERROR] Session was killed by JIRA. Exception: '{}'".format(er.text))
                elif "You do not have permission to edit issue" in e.text:
                    try:
                        data_val.pop(issue_details_new[issuetype]['Parent Link']['id'], None)
                        new_issue.update(notify=False, fields=data_val)
                    except:
                        data_val.pop(issue_details_new[issuetype]['Epic Link']['id'], None)
                elif "Component name" in e.text and "is not valid" in e.text:
                    try:
                        missing_component = e.text.replace("Component name '", "").replace("' is not valid", "")
                        new_components = []
                        for component in data_val["components"]:
                            if component['name'] not in missing_component:
                                new_components.append({"name": component['name']})
                        data_val["components"] = new_components
                    except:
                        data_val.pop('components', None)
                elif "Issue type is a sub-task but parent issue key or id not specified." in e.text and json_importer_flag == 1 and dummy_parent != '':
                    data_val["parent"] = dummy_parent
                try:
                    new_issue.update(notify=False, fields=data_val)
                    processed_issues_set.add(old_issue.key)
                except:
                    print("[ERROR] Exception for '{}' is '{}'".format(new_issue.key, e))
                    print("[INFO] The details for update: '{}'".format(data_val))
            except:
                print("[ERROR] Exception for '{}' is '{}'".format(new_issue.key, e))
                print("[INFO] The details for update: '{}'".format(data_val))
                failed_issues.add(old_issue.key)
                if verbose_logging == 1:
                    print(traceback.format_exc())


def check_target_project():
    global JIRA_create_project_api, auth, project_new, template_project, jira_new, JIRA_BASE_URL_NEW, headers, verify
    global new_project_name
    
    try:
        proj = jira_new.project(project_new)
        return
    except:
        if new_project_name == '':
            new_project_name = project_new
        data = {"key": project_new,
                "parent": template_project,
                "name": new_project_name
                }
        url = JIRA_BASE_URL_NEW + JIRA_create_project_api
        r = requests.post(url=url, params=data, auth=auth, headers=headers, verify=verify)
        if str(r.status_code) != '200':
            print("[ERROR] Project can't be created due to: '{}'".format(r.content))
        else:
            print("[INFO] New Target project '{}' has been successfully created.".format(project_new))
            print("")


def generate_template():
    """Function for Excel Mapping Template Generation - saving user configuration and processing data"""
    global jira_old, jira_new, auth, username, password, project_old, project_new, mapping_file, JIRA_BASE_URL_NEW
    global JIRA_BASE_URL_OLD, issue_details_old, issue_details_new, migrate_statuses_check, threads, verify
    global json_importer_flag, verbose_logging
    
    username = user.get()
    password = passwd.get()
    mapping_file = file.get()
    if mapping_file == '':
        mapping_file = "Migration Template for {} project to {} project.xlsx".format(project_old, project_new)
    else:
        mapping_file = mapping_file.split('.xls')[0] + '.xlsx'
    main.destroy()
    change_mappings_configs()
    if len(username) < 3 or len(password) < 3:
        print("[INFO] Please enter JIRA Credentials on new window.")
        print("")
        jira_authorization_popup()
    else:
        auth = (username, password)
        if verbose_logging == 1:
            print("[INFO] A connection attempt to JIRA server is started.")
        get_jira_connection()
    
    if project_old == '' or project_new == '' or JIRA_BASE_URL_NEW == '' or JIRA_BASE_URL_OLD == '':
        print("[ERROR] Configuration parameters are not set. Exiting...")
        os.system("pause")
        exit()
    
    print("[START] Template is being generated. Please wait...")
    print("")
    print("[START] Fields configuration downloading from Source '{}' and Target '{}' projects".format(project_old, project_new))
    
    check_global_admin_rights()
    
    if json_importer_flag == 1:
        check_target_project()
    
    try:
        issue_details_old = get_fields_list_by_project(jira_old, project_old)
        issue_details_new = get_fields_list_by_project(jira_new, project_new)
    except Exception as e:
        print("[ERROR] Issue Details can't be processed due to '{}'.".format(e))
        if verbose_logging == 1:
            print(traceback.format_exc())
        os.system("pause")
        exit()
    
    print("[END] Fields configuration successfully processed.")
    print("")
    if migrate_statuses_check == 1:
        get_transitions(project_new, JIRA_BASE_URL_NEW, new=True)
        get_transitions(project_old, JIRA_BASE_URL_OLD)
    get_hierarchy_config()
    
    for k, v in prepare_template_data().items():
        create_excel_sheet(v, k)
    save_excel()


def threads_processing(function, items):
    global threads, max_retries
    
    items_for_retry = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=threads) as executor:
        futures = {executor.submit(function, i) for i in items}
    for future in futures:
        if future.result()[0] == 1:
            items_for_retry.append(future.result()[1])
    if len(items_for_retry) > 0:
        if max_retries > 0:
            max_retries -= 1
            threads_processing(function, items_for_retry)
        else:
            print("The following items can't be processed: '{}'".format(items_for_retry))
            return


def processes_processing(function, items):
    global pool_size, threads
    
    chunck_size = len(items) // pool_size
    if len(items) % pool_size > 0:
        chunck_size += 1
    with concurrent.futures.ProcessPoolExecutor(max_workers=pool_size) as executor:
        futures = {executor.submit(threads_processing(function, i), i) for i in grouper(items, chunck_size)}


def check_global_admin_rights():
    global project_new, JIRA_BASE_URL_NEW, JIRA_imported_api, headers, verify, auth, json_importer_flag
    global multiple_json_data_processing
    
    data = {"projects": [{"key": project_new}]}
    url = JIRA_BASE_URL_NEW + JIRA_imported_api
    try:
        r = requests.post(url, json=data, auth=auth, headers=headers, verify=verify)
        if str(r.status_code) not in ['202', '409']:
            json_importer_flag = 0
            multiple_json_data_processing = 1
            print("[WARNING] No Global Admin Rights for Target Project. JSON files for Change History migration will be created.")
            print("")
        else:
            json_importer_flag = 1
            print("[INFO] Global Admin access check for Target Project has been successfully validated.")
            print("")
    except:
        json_importer_flag = 0
        print("[WARNING] No Global Admin Rights for Target Project. JSON files for Change History migration will be created.")
        print("")


def validate_template():
    global issue_details_new, issuetypes_mappings, fields_mappings, status_mappings, new_transitions, save_validation_details
    global issue_details_old, old_transitions, jira_old, project_old, skip_existing_issuetypes_validation_flag, field_value_mappings
    
    def try_to_validate(field, field_lst):
        for avail_field in field_lst:
            if field.strip().upper() == avail_field.strip().upper():
                return avail_field
        return None
    
    if save_validation_details == 1:
        data = {'issue_details_new': issue_details_new,
                'new_transitions': new_transitions,
                }
        validation_file = 'validation_data.json'
        try:
            with open(validation_file, 'w') as outfile:
                json.dump(data, outfile)
        except:
            pass
    
    template_error = 0
    error_processed = 0
    default_priority = ''
    old_issuetypes_totals = {}
    for issuetype, old_issuetypes in issuetypes_mappings.items():
        for old_issuetype in old_issuetypes['issuetypes']:
            if skip_existing_issuetypes_validation_flag == 1:
                try:
                    jql = "project = {} AND issuetype = {}".format(project_old, old_issuetype)
                    total = jira_old.search_issues(jql, startAt=0, maxResults=1, json_result=True)['total']
                except:
                    total = 0
            else:
                total = 1
            old_issuetypes_totals[old_issuetype] = total
    
    # Checking Target issuetype mappings
    if ('' in issuetypes_mappings.keys() and len(issuetypes_mappings.keys()) <= 1) or ('' not in issuetypes_mappings.keys() and len(issuetypes_mappings.keys()) < 1):
        print("")
        print("[ERROR] [ISSUETYPES] Template wasn't populated. Issietypes mapping is EMPTY.")
        template_error = 1
    for issuetype, old_issuetypes in issuetypes_mappings.items():
        if issuetype != '':
            type_not_found = 1
            for type in issue_details_new.keys():
                if type == issuetype:
                    type_not_found = 0
                    break
            if type_not_found == 1:
                print("[ERROR] Issuetype '{}' is not available in Target project. Mapped to '{}'".format(issuetype, old_issuetypes['issuetypes']))
                template_error = 1
                error_processed = 1
    
    # Checking Source Issue Types
    for issuetype, old_issuetypes in issuetypes_mappings.items():
        for old_issuetype in old_issuetypes['issuetypes']:
            if old_issuetypes_totals[old_issuetype] > 0:
                type_not_found = 1
                for o_type in issue_details_old.keys():
                    if o_type == old_issuetype:
                        type_not_found = 0
                        break
                if type_not_found == 1 and issuetype != '':
                    print("[ERROR] Issuetype '{}' is not available in Source project. Mapped to '{}'".format(old_issuetype, issuetype))
                    proposed_value = try_to_validate(old_issuetype, list(issue_details_old.keys()))
                    if proposed_value is None:
                        print("[INFO] Available issuetypes are: '{}'".format(issue_details_old.keys()))
                    else:
                        print("[PROPOSED CHANGE] Issuetype '{}' could be renamed to '{}'".format(old_issuetype, proposed_value))
                    template_error = 1
                    error_processed = 1
    
    if template_error == 1 and error_processed == 1:
        print("")
        error_processed = 0
    
    # Checking Target field mappings
    total_fields_mapped = 0
    for o_it in fields_mappings.keys():
        if ('' in fields_mappings[o_it].keys() and len(fields_mappings[o_it].keys()) > 1) or ('' not in fields_mappings[o_it].keys() and len(fields_mappings[o_it].keys()) > 0):
            total_fields_mapped += 1
    if total_fields_mapped < 1:
        print("[WARNING] [FIELDS] Fields mapping hasn't been populated. ONLY System fields would be migrated.")
    for issuetype, values in issuetypes_mappings.items():
        for old_issuetype in values['issuetypes']:
            if old_issuetypes_totals[old_issuetype] > 0:
                try:
                    for new_field, old_fields in fields_mappings[old_issuetype].items():
                        if new_field != '':
                            field_not_found = 1
                            try:
                                for field in issue_details_new[issuetype].keys():
                                    if new_field == field:
                                        field_not_found = 0
                                        break
                            except:
                                field_not_found = 1
                            if field_not_found == 1:
                                print("[ERROR] Target Field '{}' is incorrect. Details: Source Issuetype: '{}', Target Issuetype: '{}'. Mapped to Source fields: '{}'".format(new_field, old_issuetype, issuetype, old_fields))
                                template_error = 1
                                error_processed = 1
                except:
                    if issuetype != '':
                        print("[WARNING] Please check the '{}' Source Issuetype value. Looks like name is incorrect. It would be skipped.".format(old_issuetype))
    
    # Checking Source field mappings
    for issuetype, values in issuetypes_mappings.items():
        for old_issuetype in values['issuetypes']:
            if old_issuetypes_totals[old_issuetype] > 0:
                try:
                    for new_field, old_fields in fields_mappings[old_issuetype].items():
                        if new_field != '':
                            for old_field in old_fields:
                                field_not_found = 1
                                try:
                                    for o_field in issue_details_old[old_issuetype].keys():
                                        if old_field == o_field or 'issuetype.name' in old_field or 'issuetype.status' in old_field:
                                            field_not_found = 0
                                            break
                                except:
                                    field_not_found = 1
                                if field_not_found == 1 and issuetype != '':
                                    proposed_value = try_to_validate(old_field, list(issue_details_old[old_issuetype].keys()))
                                    print("[ERROR] Source Field '{}' is incorrect. Details: Source Issuetype: '{}', Target Issuetype: '{}'. Mapped to Target fields: '{}'".format(old_field, old_issuetype, issuetype, new_field))
                                    if proposed_value is None:
                                        print("[INFO] Available Fields are: '{}'".format(list(issue_details_old[old_issuetype].keys())))
                                    else:
                                        print("[PROPOSED CHANGE] Source Field '{}' could be renamed to '{}'".format(old_field, proposed_value))
                                    template_error = 1
                                    error_processed = 1
                except:
                    if issuetype != '':
                        print("[WARNING] Please check the '{}' Source Issuetype value. Looks like name is incorrect. It would be skipped.".format(old_issuetype))
    
    if template_error == 1 and error_processed == 1:
        print("")
    
    # Checking Target statuses mappings
    new_issuetype_statuses = {}
    for k, v in new_transitions.items():
        statuses_lst = []
        for l in v:
            statuses_lst.append(l[0])
            statuses_lst.append(l[2])
        new_issuetype_statuses[k] = list(set(statuses_lst))
    
    total_statuses_mapped = 0
    for o_it in status_mappings.keys():
        if ('' in status_mappings[o_it].keys() and len(status_mappings[o_it].keys()) > 1) or ('' not in status_mappings[o_it].keys() and len(status_mappings[o_it].keys()) > 0):
            total_statuses_mapped += 1
    if total_statuses_mapped < 1:
        print("[WARNING] [STATUSES] Statuses mapping hasn't been populated. DEFAULT Status would be used for ALL statuses.")
    
    for issuetype, values in issuetypes_mappings.items():
        for old_issuetype in values['issuetypes']:
            if old_issuetypes_totals[old_issuetype] > 0:
                try:
                    for n_status, old_statuses in status_mappings[old_issuetype].items():
                        if n_status != '':
                            status_not_found = 1
                            try:
                                for new_status in new_issuetype_statuses[issuetype]:
                                    if new_status.upper() == n_status.upper():
                                        status_not_found = 0
                                        break
                                if status_not_found == 1:
                                    print("[ERROR] Target Status '{}' is incorrect. Details: Source Issuetype: '{}', Target Issuetype: '{}'. Mapped to Source statuses: '{}'".format(n_status, old_issuetype, issuetype, old_statuses))
                                    print("[INFO] Available Statuses for Target '{}' issuetype '{}': ".format(n_status, new_issuetype_statuses[issuetype]))
                                    template_error = 1
                            except:
                                pass
                        else:
                            for n_issuetype, old_issuetypes in issuetypes_mappings.items():
                                if n_issuetype != '' and old_issuetype in old_issuetypes['issuetypes']:
                                    print("[WARNING] Old Statuses '{}' have not been mapped for '{}' Issuetype in Source project. DEFAULT Status would be used for them.".format(old_statuses, old_issuetype))
                except:
                    if issuetype != '':
                        print("[WARNING] Please check the '{}' Source Issuetype value in Statuses. Looks like name is incorrect. It would be skipped.".format(old_issuetype))
    
    # Checking Source statuses mappings
    old_issuetype_statuses = {}
    for k, v in old_transitions.items():
        statuses_lst = []
        for l in v:
            statuses_lst.append(l[0])
            statuses_lst.append(l[2])
        old_issuetype_statuses[k] = list(set(statuses_lst))
    
    for issuetype, values in issuetypes_mappings.items():
        for old_issuetype in values['issuetypes']:
            if old_issuetypes_totals[old_issuetype] > 0:
                try:
                    for n_status, old_statuses in status_mappings[old_issuetype].items():
                        if n_status != '':
                            for old_status in old_statuses:
                                status_not_found = 1
                                try:
                                    for o_status in old_issuetype_statuses[old_issuetype]:
                                        if old_status.upper() == o_status.upper():
                                            status_not_found = 0
                                            break
                                    if status_not_found == 1 and issuetype != '':
                                        proposed_value = try_to_validate(old_status, list(old_issuetype_statuses[old_issuetype]))
                                        print("[ERROR] Source Status '{}' is incorrect. Details: Source Issuetype: '{}', Target Issuetype: '{}'. Mapped to Target statuses: '{}'".format(old_status, old_issuetype, issuetype, n_status))
                                        if proposed_value is None:
                                            print("[INFO] Available Statuses are: '{}'".format(list(old_issuetype_statuses[old_issuetype])))
                                        else:
                                            print("[PROPOSED CHANGE] Status '{}' could be renamed to '{}'".format(old_status, proposed_value))
                                        template_error = 1
                                except:
                                    pass
                except:
                    if issuetype != '':
                        print("[WARNING] Please check the '{}' Source Issuetype value in Statuses. Looks like name is incorrect. It would be skipped.".format(old_issuetype))
    
    # Validate Target Priority
    priority_new_lst = []
    for field_values in issue_details_new.values():
        if 'Priority' in field_values.keys():
            for p in field_values['Priority']['allowed values']:
                priority_new_lst.append(p)
    
    if len(priority_new_lst) > 0:
        for issuetype in issuetypes_mappings.keys():
            if issuetype != '':
                default_priority = issue_details_new[issuetype]['Priority']['default value']
                break
        
        if len(field_value_mappings['Priority'].keys()) == 0:
            print("[WARNING] [PRIORITIES] Priorities mapping hasn't been populated. DEFAULT '{}' Priority would be used for ALL issues.".format(default_priority))
        else:
            for new_value, old_values in field_value_mappings['Priority'].items():
                if new_value not in priority_new_lst and new_value != '':
                    print("[WARNING] Priority '{}' does not exist in Target '{}' project. DEFAULT '{}' Priority would be used instead.".format(new_value, project_new, default_priority))
                elif new_value == '':
                    print("[WARNING] Source Priorities '{}' have not been mapped. DEFAULT '{}' Priority would be used for them.".format(old_values, default_priority))
    
    # Validate Source Priority
    priority_old_lst = []
    for field_values in issue_details_old.values():
        if 'Priority' in field_values.keys():
            for p in field_values['Priority']['allowed values']:
                priority_old_lst.append(p)
    
    if len(priority_old_lst) > 0:
        for new_value, old_values in field_value_mappings['Priority'].items():
            for o_val in old_values:
                if o_val not in priority_old_lst and o_val != '':
                    print("[WARNING] Priority '{}' does not exist in Source '{}' project.".format(o_val, project_old))
    
    if template_error == 1:
        print("")
        print("[FATAL ERROR] Template filled incorectly. Please fix issues and try again.")
        print("")
        return 1
    else:
        return 0


def find_max_id(key, jira, project):
    jql_max = 'project = {} order by key DESC'.format(project)
    max_processing_key = jira.search_issues(jql_str=jql_max, maxResults=1, json_result=False)[0].key
    jql_min = 'project = {} order by key ASC'.format(project)
    min_processing_key = jira.search_issues(jql_str=jql_min, maxResults=1, json_result=False)[0].key
    
    if int(key.split('-')[1]) > int(max_processing_key.split('-')[1]):
        key = max_processing_key
    
    try:
        max_issue = jira.issue(key)
        if project in max_issue.key:
            return key
        else:
            key = key.split('-')[0] + '-' + str(int(key.split('-')[1]) - 1)
            if int(key.split('-')[1]) < int(min_processing_key.split('-')[1]):
                print("")
                print("[ERROR] Min issue available is: '{}'. End issue id is lower than min key.".format(min_processing_key))
                print("")
                os.system("pause")
                exit()
            key = find_max_id(key, jira, project)
            return key
    except:
        key = key.split('-')[0] + '-' + str(int(key.split('-')[1]) - 1)
        if int(key.split('-')[1]) < int(min_processing_key.split('-')[1]):
            print("")
            print("[ERROR] Min issue available is: '{}'. End issue id is lower than min key.".format(min_processing_key))
            print("")
            os.system("pause")
            exit()
        key = find_max_id(key, jira, project)
        return key


def find_min_id(key, jira, project):
    jql_max = 'project = {} order by key DESC'.format(project)
    max_processing_key = jira.search_issues(jql_str=jql_max, maxResults=1, json_result=False)[0].key
    jql_min = 'project = {} order by key ASC'.format(project)
    min_processing_key = jira.search_issues(jql_str=jql_min, maxResults=1, json_result=False)[0].key
    
    if int(key.split('-')[1]) < int(min_processing_key.split('-')[1]):
        key = min_processing_key
    
    try:
        min_issue = jira.issue(key)
        if project in min_issue.key:
            return key
        else:
            key = key.split('-')[0] + '-' + str(int(key.split('-')[1]) + 1)
            if int(key.split('-')[1]) > int(max_processing_key.split('-')[1]):
                print("")
                print("[ERROR] Max issue available is: '{}'. Start issue id is higher than max key.".format(max_processing_key))
                print("")
                os.system("pause")
                exit()
            key = find_min_id(key, jira, project)
            return key
    except:
        key = key.split('-')[0] + '-' + str(int(key.split('-')[1]) + 1)
        if int(key.split('-')[1]) > int(max_processing_key.split('-')[1]):
            print("")
            print("[ERROR] Max issue available is: '{}'. Start issue id is higher than max key.".format(max_processing_key))
            print("")
            os.system("pause")
            exit()
        key = find_min_id(key, jira, project)
        return key


def split_processing_jql(jql):
    global project_old, recently_updated, supported_issuetypes
    
    max_size = len(str(recently_updated)) + len(str(supported_issuetypes)) + len('project =   ') + len(str(project_old)) + 4
    remaining_size = 2000 - max_size
    
    if jql in ['', 'key in ()']:
        return None
    elif ('=' in jql
          or '~' in jql
          or '!' in jql
          or ' IN ' in jql.upper()
          or ' AND ' in jql.upper()
          or ' OR ' in jql.upper()
          or ' NOT ' in jql.upper()
          or ' IS ' in jql.upper()):
        if len(str(jql)) > remaining_size:
            print("[ERROR] The JQL provided: '{}' with extra checks is longer then 2000 characters. Processing will be skipped.".format(jql))
            return None
        return [jql]
    else:
        jql_issues_list_tmp = jql.split(',')
        jql_issues_list = [i.replace("'", "").strip() for i in jql_issues_list_tmp]
        jql_issues = []
        if len(str(jql_issues_list)) > remaining_size:
            new_jql = "key in ("
            for issue_key in jql_issues_list:
                new_jql += issue_key + ","
                if len(str(new_jql)) > remaining_size - len(str(issue_key)) - 7:
                    new_jql = new_jql[:-1] + ")"
                    jql_issues.append(new_jql)
                    new_jql = "key in ("
        else:
            jql_issues.append(str("key in (" + jql + ")"))
        return jql_issues


def migration_process(start_jira_key, max_processing_key, max_id, reprocess=False):
    global skip_migrated_flag, process_only_last_updated_date_flag, project_old, project_new, recently_updated_days
    global processing_jira_jql, issuetypes_mappings, recently_updated, last_updated_days_check, total_processed
    global including_dependencies_flag, jira_old, jira_new, already_migrated_set, skipped_issuetypes, issues_lst
    global migrated_issues_lst, multiple_json_data_processing, processed_issuetypes, items_lst, failed_issues
    global processed_issues_set, total_data, max_retries, override_template_flag, teams_to_be_added_set
    global already_processed_json_importer_issues, max_json_file_size, control_logic_flag, refresh_already_migrated_flag
    global refresh_issuetypes, teams_thread_lock, supported_issuetypes, previous_multiple_json_data_processing
    
    # Check if already migrated or partially migrated should be reprocessed
    if refresh_already_migrated_flag == 1 and reprocess is False:
        if refresh_issuetypes in ['', 'ALL']:
            print("[INFO] ALL Migrated Issues in Target '{}' project will be updated. Please wait...".format(project_new))
            jql_refresh_migrated = "project = {} AND summary !~ 'DUMMY_PARENT'".format(project_new)
            get_issues_by_jql(jira_new, jql=jql_refresh_migrated, types=True, non_migrated=True)
        else:
            print("[INFO] Migrated Issues with {} issuetypes in Target '{}' project will be updated. Please wait...".format(refresh_issuetypes, project_new))
            jql_refresh_migrated = "project = {} AND issuetype in ({})".format(project_new, refresh_issuetypes)
            try:
                get_issues_by_jql(jira_new, jql=jql_refresh_migrated, types=True, non_migrated=True)
            except:
                print("[ERROR] The '{}' list of issuetypes can't be found in Target JIRA. All issuetypes would be used for refresh data instead.".format(refresh_issuetypes))
                jql_refresh_migrated = "project = {} ".format(project_new)
                get_issues_by_jql(jira_new, jql=jql_refresh_migrated, types=True, non_migrated=True)
    elif skip_migrated_flag == 1 or force_update_flag == 1 or including_dependencies_flag == 1 or process_only_last_updated_date_flag == 1 and reprocess is False:
        print("[START] Checking for previously NON-COMPLETED issues. They would be re-processed.")
        jql_non_migrated = "project = {} AND labels = MIGRATION_NOT_COMPLETE".format(project_new)
        non_migrated_count = jira_new.search_issues(jql_non_migrated, startAt=0, maxResults=1, json_result=True)['total']
        print("[END] Total '{}' previously NON-COMPLETED issues will be added based on JQL: '{}'.".format(non_migrated_count, jql_non_migrated))
        get_issues_by_jql(jira_new, jql=jql_non_migrated, types=True, non_migrated=True)
        print("")
    
    # Check already migrated issues
    if skip_migrated_flag == 1 and process_only_last_updated_date_flag == 0 and reprocess is False:
        print("[START] Checking for already migrated issues. They will be skipped.")
        start_already_migrated_time = time.time()
        already_migrated_set = set()
        try:
            jql_last_migrated = "project = {} AND (labels not in ('MIGRATION_NOT_COMPLETE') OR labels is EMPTY)".format(project_new)
            get_issues_by_jql(jira_new, jql_last_migrated, migrated=True, max_result=0)
        except Exception as e:
            print("[WARNING] Already migrated issues can't be processed due to: '{}'".format(e))
        print("[END] Already migrated issues have been calculated. Number: '{}'.".format(len(already_migrated_set)))
        print("[INFO] Already migrated issues retrieved in '{}' seconds.".format(time.time() - start_already_migrated_time))
        print("")
    
    if reprocess is True:
        already_migrated_set -= failed_issues
        already_migrated_set |= processed_issues_set
    
    # Check issues updated within the last number of days
    if last_updated_days_check == 1 and process_only_last_updated_date_flag == 0 and reprocess is False:
        try:
            jql_recently_updated = "project = '{}' AND updated >= startOfDay(-{}) order by key ASC".format(project_old, recently_updated_days)
            new_start_jira_key = jira_old.search_issues(jql_str=jql_recently_updated, maxResults=1, json_result=False)[0].key
        except:
            jql_recently_updated = "project = '{}' order by key ASC".format(project_old)
            new_start_jira_key = jira_old.search_issues(jql_str=jql_recently_updated, maxResults=1, json_result=False)[0].key
        if int(start_jira_key.split('-')[1]) < int(new_start_jira_key.split('-')[1]):
            start_jira_key = new_start_jira_key
        if int(start_jira_key.split('-')[1]) > int(max_processing_key.split('-')[1]):
            start_jira_key = max_processing_key
        recently_updated = "AND updated >= startOfDay(-{})".format(recently_updated_days)
    
    # Check for only mapped issuetypes
    if len(skipped_issuetypes) >= len(processed_issuetypes) or override_template_flag == 1:
        supported_issuetypes = "AND issuetype in ({})".format(str(processed_issuetypes)[1:-1])
    elif len(skipped_issuetypes) == 0:
        supported_issuetypes = ''
    else:
        supported_issuetypes = "AND issuetype not in ({})".format(str(skipped_issuetypes)[1:-1])
    
    # Add last updated issues to migration / update process
    if process_only_last_updated_date_flag == 1 and last_updated_date not in ['YYYY-MM-DD', ''] and reprocess is False:
        print("[START] Recently updated Issues loading was started from Source project. It could take some time... Please wait...")
        try:
            if control_logic_flag == 0:
                jql_latest = "project = '{}' AND updated >= {} {}".format(project_old, last_updated_date, supported_issuetypes)
                if json_importer_flag == 1 or multiple_json_data_processing == 1:
                    get_issues_by_jql(jira_old, jql_latest, types=True, max_result=0)
                else:
                    get_issues_by_jql(jira_old, jql_latest, types=True, sprint=True, max_result=0)
                print("[INFO] Updated issues after '{}' date would be processed, total number is '{}'".format(last_updated_date, sum([len(items_lst[i]) for i in items_lst.keys()])))
            else:
                print("[INFO] Control logic would be applied - only recently created issues would be updated.")
                jql_latest = "project = '{}' AND created >= {} {}".format(project_old, last_updated_date, supported_issuetypes)
                if json_importer_flag == 1 or multiple_json_data_processing == 1:
                    get_issues_by_jql(jira_old, jql_latest, types=True, max_result=0)
                else:
                    get_issues_by_jql(jira_old, jql_latest, types=True, sprint=True, max_result=0)
                print("[INFO] Created issues after '{}' date would be processed, total number is '{}'".format(last_updated_date, sum([len(items_lst[i]) for i in items_lst.keys()])))
                jql_last_migrated = "project = '{}' AND (labels not in ('MIGRATION_NOT_COMPLETE') OR labels is EMPTY)".format(project_new)
                get_issues_by_jql(jira_new, jql_last_migrated, migrated=True, max_result=0)
                print("[INFO] Already migrated '{}' issues would be checked for any updates after '{}' date.".format(len(already_migrated_set), last_updated_date))
                jql_latest_control = "project = '{}' AND updated >= {} {}".format(project_old, last_updated_date, supported_issuetypes)
                if json_importer_flag == 1 or multiple_json_data_processing == 1:
                    get_issues_by_jql(jira_old, jql_latest_control, types=True, control=True, max_result=0)
                else:
                    get_issues_by_jql(jira_old, jql_latest_control, types=True, sprint=True, control=True, max_result=0)
            print("[END] Recently updated Issues have been successfully loaded for processing: '{}'.".format(sum([len(items_lst[i]) for i in items_lst.keys()])))
            print("")
        except:
            print("[ERROR] The value for Last Updated '{}' not in correct 'YYYY-MM-DD' format.".format(last_updated_date))
    
    # Sprints / issues migration check
    start_issues_time = time.time()
    print("[START] Issues loading from Source project was started. It could take some time... Please wait...")
    
    if migrate_sprints_check == 1 and json_importer_flag == 0 and multiple_json_data_processing == 0 and reprocess is False:
        start_sprints_time = time.time()
        if old_board_id == 0:
            migrate_sprints(proj_old=project_old, project=project_new, name=default_board_name)
        else:
            migrate_sprints(proj_old=project_old, board_id=old_board_id, project=project_new, name=default_board_name)
        if verbose_logging == 1:
            print("[INFO] Sprints migrated in '{}' seconds.".format(time.time() - start_sprints_time))
            print("")
    
    if process_only_last_updated_date_flag == 1 and reprocess is False:
        print("[INFO] Only last updated issues will be processed. Other options will be ignored.")
        print("")
    elif refresh_already_migrated_flag == 1 and reprocess is False:
        print("[INFO] Only already processed issues will be processed. Other options will be ignored.")
        print("")
    elif reprocess is True:
        processing_jql_lst = split_processing_jql(get_str_from_lst(list(failed_issues)))
        if processing_jql_lst is not None:
            for processing_jql in processing_jql_lst:
                processing_jql = "project = {} AND {} {} {}".format(project_old, processing_jql, recently_updated, supported_issuetypes)
                print("[INFO] The issues would be processed based on JQL: '{}'".format(processing_jql))
                get_issues_by_jql(jira_old, jql=processing_jql, types=True)
    else:
        if limit_migration_data != 0:
            if start_jira_key == max_id:
                recently_updated = "AND key >= {} AND key < {} {}".format(start_jira_key, max_id, recently_updated)
            else:
                recently_updated = "AND key >= {} AND key <= {} {}".format(start_jira_key, max_id, recently_updated)
        processing_jql_lst = split_processing_jql(processing_jira_jql)
        if processing_jql_lst is not None:
            for processing_jql in processing_jql_lst:
                processing_jql = "project = {} AND {} {} {}".format(project_old, processing_jql, recently_updated, supported_issuetypes)
                print("[INFO] The issues would be processed based on JQL: '{}'".format(processing_jql))
                get_issues_by_jql(jira_old, jql=processing_jql, types=True)
        if including_dependencies_flag == 1:
            if processing_jql_lst is None:
                processing_jql_lst = ['']
            for processing_jql in processing_jql_lst:
                if processing_jql == '':
                    dependencies_jql = "project = {} {} {}".format(project_old, recently_updated, processing_jql)
                    main_jql = "project = {} {} {} {}".format(project_old, recently_updated, processing_jql, supported_issuetypes)
                else:
                    dependencies_jql = "project = {} {} AND {}".format(project_old, recently_updated, processing_jql)
                    main_jql = "project = {} {} AND {} {}".format(project_old, recently_updated, processing_jql, supported_issuetypes)
                main_count = jira_old.search_issues(main_jql, startAt=0, maxResults=1, json_result=True)['total']
                print("[INFO] Total '{}' issues would be added based on JQL: '{}'.".format(main_count, main_jql))
                get_issues_by_jql(jira_old, jql=main_jql, types=True)
                print("")
                print("[INFO] Dependencies would be processed, as Epics, SubTasks, Parents and Linked issues...")
                if skip_migrated_flag == 1:
                    print("[INFO] NOTE: The total numbers below could be reduced based on already migrated issues - they would be skipped.")
                epics_jql = "project = {} {} AND issueFunction in epicsOf(\"{}\")".format(project_old, supported_issuetypes, dependencies_jql)
                epics_count = jira_old.search_issues(epics_jql, startAt=0, maxResults=1, json_result=True)['total']
                print("[INFO] Total '{}' Dependency Epics would be added based on JQL: '{}'.".format(epics_count, epics_jql))
                get_issues_by_jql(jira_old, jql=epics_jql, types=True)
                subtasks_jql = "project = {} {} AND issueFunction in subtasksOf(\"{}\")".format(project_old, supported_issuetypes, dependencies_jql)
                subtasks_count = jira_old.search_issues(subtasks_jql, startAt=0, maxResults=1, json_result=True)['total']
                print("[INFO] Total '{}' Dependency SubTasks would be added based on JQL: '{}'.".format(subtasks_count, subtasks_jql))
                get_issues_by_jql(jira_old, jql=subtasks_jql, types=True)
                parents_jql = "project = {} {} AND issueFunction in parentsOf(\"{}\")".format(project_old, supported_issuetypes, dependencies_jql)
                parents_count = jira_old.search_issues(parents_jql, startAt=0, maxResults=1, json_result=True)['total']
                print("[INFO] Total '{}' Dependency Parents would be added based on JQL: '{}'.".format(parents_count, parents_jql))
                get_issues_by_jql(jira_old, jql=parents_jql, types=True)
                links_jql = "project = {} {} AND issueFunction in linkedIssuesOf(\"{}\")".format(project_old, supported_issuetypes, dependencies_jql)
                links_count = jira_old.search_issues(links_jql, startAt=0, maxResults=1, json_result=True)['total']
                print("[INFO] Total '{}' Dependency Linked Issues would be added based on JQL: '{}'.".format(links_count, links_jql))
                get_issues_by_jql(jira_old, jql=links_jql, types=True)
        if last_updated_days_check == 1 and including_dependencies_flag == 0 and processing_jql_lst is None:
            recently_updated_jql = "project = {} {} {}".format(project_old, recently_updated, supported_issuetypes)
            print("[INFO] The Recently Updated issues would be processed based on JQL: '{}'".format(recently_updated_jql))
            get_issues_by_jql(jira_old, jql=recently_updated_jql, types=True)
    print("[END] Issues have been loaded from Target '{}' project.".format(project_old))
    print("")
    
    # Calculating Minimal and Maximal issues to be migrated
    min_issue, max_issue = (0, 0)
    failed_issues = set()
    migrated_issues_lst = []
    try:
        for k, v in items_lst.items():
            new_list = list(set(v) - set(processed_issues_set))
            if len(new_list) > 0:
                items_lst[k] = new_list
            else:
                items_lst.pop(k, None)
    except:
        pass
    for k, v in items_lst.items():
        if k not in skipped_issuetypes:
            if min_issue == 0 or min([int(i.split('-')[1]) for i in v]) < min_issue:
                min_issue = min([int(i.split('-')[1]) for i in v])
            if max_issue == 0 or max([int(i.split('-')[1]) for i in v]) > max_issue:
                max_issue = max([int(i.split('-')[1]) for i in v])
            migrated_issues_lst.extend(v)
    min_issue_key = project_old + '-' + str(min_issue)
    max_issue_key = project_old + '-' + str(max_issue)
    migrated_issues_lst = list(set(migrated_issues_lst))
    print("[INFO] The Number of Unique issues to be migrated: {}".format(len(migrated_issues_lst)))
    if len(migrated_issues_lst) > 0:
        start_jira_key = min_issue_key
        max_id = max_issue_key
    print("[INFO] The first issue to be migrated: {}".format(start_jira_key))
    print("[INFO] The last issue to be migrated: {}".format(max_id))
    print("[INFO] Issues loaded in '{}' seconds.".format(time.time() - start_issues_time))
    print("")
    
    # Extra Logging
    if verbose_logging == 1:
        print('[INFO] The list of migrated issues by type:', str(items_lst))
    
    # Creating dummy parent
    if len(sub_tasks) > 0:
        for k in sub_tasks.keys():
            try:
                for task in issuetypes_mappings[k]['issuetypes']:
                    if task in items_lst.keys():
                        get_dummy_parent()
                        break
                if dummy_parent != '':
                    break
            except:
                pass
    
    # Restore file-processing for pre-updates processing
    if previous_multiple_json_data_processing == 1 and multiple_json_data_processing == 0:
        multiple_json_data_processing = 1
    
    # -----Metadata Migration-------
    # Creating JSON file for importing data
    if reprocess is False and multiple_json_data_processing == 1:
        start_placeholders_time = time.time()
        print("[START] JSON Importer file(s) will be created.")
        print("[INFO] Data would be loaded from Source JIRA. JSON Files will be created with maximum size up to ~{} Mb.".format(max_json_file_size))
        total_processed = len(migrated_issues_lst)
        for i in range(4):
            for k, v in issuetypes_mappings.items():
                if v['hierarchy'] == str(i):
                    for i_type in issuetypes_mappings[k]['issuetypes']:
                        if i_type in items_lst.keys():
                            max_retries = default_max_retries
                            threads_processing(json_file_process_issue, items_lst[i_type])
        # Saving latest file if all issues processed
        if total_data["projects"] != [{"key": project_new, "issues": []}]:
            update_issues_json(total_data)
            total_data = {}
            total_data["projects"] = [{"key": project_new, "issues": []}]
            total_data["users"] = []
            total_data["links"] = []
        
        print("[END] JSON Importer file(s) have been successfully processed.")
        print("[INFO] JSON Importer file(s) have been created/checked in '{}' seconds.".format(time.time() - start_placeholders_time))
        print("")
        if json_files_autoupload == 0:
            print("[INFO] Please process JSON files - incrementally all parts in JIRA 'Projects -> Import External Project -> JSON' and continue migration process.")
            print("")
        
        if len(teams_to_be_added_set) > 0:
            print("[START] Teams would be checked for existence in Target JIRA instance.")
            if verbose_logging == 1:
                print("[INFO] Teams to be added: '{}'".format(list(teams_to_be_added_set)))
            for team_name in teams_to_be_added_set:
                teams_thread_lock.acquire()
                team_id = get_team_id(team_name)
                teams_thread_lock.release()
            print("[END] Teams checks have been completed. Please continue migration process after all JSON files are processed.")
        if json_files_autoupload == 0:
            os.system("pause")
        print("[INFO] Migration process will be continued.")
        print("")
    
    # Removing file-processing for post-updates processing
    if multiple_json_data_processing == 1 and json_importer_flag == 1:
        previous_multiple_json_data_processing = multiple_json_data_processing
        multiple_json_data_processing = 0
    
    # Main migration process starts here
    for i in range(4):
        for k, v in issuetypes_mappings.items():
            if v['hierarchy'] == str(i):
                migrate_issues(issuetype=k)


def json_file_process_issue(key):
    global jira_new, project_new, jira_old, issuetypes_mappings, sub_tasks, already_processed_json_importer_issues
    global verbose_logging
    
    if key in already_processed_json_importer_issues:
        return (0, key)
    new_issue_key = project_new + '-' + key.split('-')[1]
    try:
        old_issue = jira_old.issue(key, expand="changelog")
        issue_type = old_issue.fields.issuetype.name
        new_issue_type = get_new_issuetype(issue_type)
        new_status = None
        
        try:
            new_status = get_new_status(old_issue.fields.status.name, issue_type)
        except:
            print("[ERROR] Status '{}' can't be mapped for '{}' - check Mapping file. Default status would be used.".format(old_issue.fields.status.name, issue_type))
        
        try:
            new_issue = jira_new.issue(new_issue_key, expand="changelog")
            if new_issue_type in sub_tasks.keys():
                status = migrate_change_history(old_issue, new_issue_type, new_status, new=False, new_issue=new_issue, subtask=True)
            else:
                status = migrate_change_history(old_issue, new_issue_type, new_status, new=False, new_issue=new_issue)
        except Exception as e:
            if new_issue_type in sub_tasks.keys():
                status = migrate_change_history(old_issue, new_issue_type, new_status, new=True, subtask=True)
            else:
                status = migrate_change_history(old_issue, new_issue_type, new_status, new=True)
        
        if str(status) == '202':
            return (0, key)
        else:
            return (1, key)
    except Exception as e:
        print("[ERROR] Exception: '{}'.".format(e))
        print(traceback.format_exc())
        return (1, key)


def get_new_board_id():
    global new_board_id, jira_new, project_new, default_board_name
    
    if new_board_id == 0:
        if len(jira_new.boards()) == 0:
            board = jira_new.create_board(default_board_name, project_new, location_type='project')
            new_board_id = board.id
            return
        for board in jira_new.boards():
            if board.name == default_board_name and project_new in board.filter.query:
                new_board_id = board.id
                return
        if new_board_id == 0:
            board = jira_new.create_board(default_board_name, project_new, location_type='project')
            new_board_id = board.id


def process_one_template(mapping_file):
    global default_configuration_file, username, password, auth, issue_details_old, issue_details_new, verbose_logging
    global jira_old, jira_new, project_old, project_new, merge_projects_flag, shifted_by, migrate_statuses_check
    global json_importer_flag, multiple_json_data_processing, bulk_processing_flag, process_only_last_updated_date_flag
    global JIRA_BASE_URL_NEW, JIRA_BASE_URL_OLD, check_template_flag, processing_error, validation_template_error
    global start_jira_key, limit_migration_data, new_board_id, migrate_sprints_check, force_sprints_update_flag
    global migrate_teams_check, fields_mappings, migrate_metadata_check, sub_tasks, issuetypes_mappings
    global recently_updated_days, items_lst, max_processing_key, recently_updated, process_complete, remaining_previous
    global process_partially_complete, dummy_parent, set_source_project_read_only, verify, already_migrated_set
    global previous_JIRA_BASE_URL_NEW, skip_new_process, old_transitions, old_statuses, status_mappings, teams
    global field_value_mappings, link_mappings, old_sub_tasks, new_transitions, new_statuses, new_issues_ids
    global old_sprints, new_sprints, old_fields_ids_mapping, total_data, issues_lst, failed_issues, processed_issuetypes
    global already_processed_json_importer_issues, already_processed_users, migrated_issues_lst, processing_jql_lst
    global skipped_issuetypes, retry_logic_flag, override_template_flag, refresh_already_migrated_flag
    global supported_issuetypes, reconciliation_updated_days, process_reconciliation_flag
    
    start_time = time.time()
    
    # Set default values for each project
    verify = True
    max_processing_key, recently_updated, supported_issuetypes = ('', '', '')
    process_partially_complete, process_complete, processing_error, validation_template_error = (0, 0, 0, 0)
    dummy_parent, JIRA_BASE_URL_NEW, JIRA_BASE_URL_OLD, project_old, project_new = ('', '', '', '', '')
    same_new_project, remaining_previous, skip_new_process, new_board_id, remaining = (0, 0, 0, 0, 0)
    issue_details_old, old_transitions, old_statuses, issuetypes_mappings, fields_mappings = ({}, {}, {}, {}, {})
    status_mappings, field_value_mappings, link_mappings, sub_tasks, old_sub_tasks = ({}, {}, {}, {}, {})
    old_sprints, new_sprints, old_fields_ids_mapping, total_data, items_lst = ({}, {}, {}, {}, {})
    issue_details_new, new_transitions, new_statuses, new_issues_ids, skipped_issuetypes = ({}, {}, {}, {}, [])
    issues_lst, already_migrated_set, already_processed_json_importer_issues = (set(), set(), set())
    already_processed_users, failed_issues, processed_issuetypes = (set(), set(), [])
    
    # Loading data from Excel
    read_excel(file_path=mapping_file.strip())
    if JIRA_BASE_URL_NEW == '' or JIRA_BASE_URL_OLD == '' or project_old == '':
        print("[ERROR] File '{}' has incorrect format. Skipped.".format(mapping_file))
        print("")
        processing_error = 1
        return
    if project_new == '':
        project_new = project_old
    
    if JIRA_BASE_URL_NEW != previous_JIRA_BASE_URL_NEW:
        teams = {}
    else:
        same_new_project = 1
    
    previous_JIRA_BASE_URL_NEW = JIRA_BASE_URL_NEW
    total_data["projects"] = [{"key": project_new, "issues": []}]
    total_data["users"] = []
    total_data["links"] = []
    
    # Loading data from default configuration file
    if os.path.exists(default_configuration_file) is True or default_configuration_file != '':
        read_default_mappings_excel(file_path=default_configuration_file.strip())
    
    # Create jira connection details
    print("[START] JIRA connection is checking...")
    get_jira_connection()
    print("[END] Connection with JIRA has been successfully established.")
    print("")
    
    # Check Global Admin Access
    if (json_importer_flag == 1 or multiple_json_data_processing == 1) and same_new_project == 0:
        check_global_admin_rights()
    if json_importer_flag == 1 and bulk_processing_flag == 0:
        check_target_project()
    
    print("[START] Fields configuration downloading from Source '{}' and Target '{}' projects.".format(project_old, project_new))
    
    try:
        issue_details_old = get_fields_list_by_project(jira_old, project_old, old=True)
        issue_details_new = get_fields_list_by_project(jira_new, project_new)
    except Exception as e:
        print("[ERROR] Issue Details can't be processed due to '{}'.".format(e))
        print("")
        if verbose_logging == 1:
            print(traceback.format_exc())
        processing_error = 1
        return
    
    if issue_details_old == {} or issue_details_new == {}:
        print("[ERROR] No access to the projects. Migration skipped.")
        print("")
        processing_error = 1
        return
    
    print("[END] Fields configuration successfully processed.")
    print("")
    
    # Check if Target Project should not be re-written by Source Project
    if merge_projects_flag == 1:
        print("[START] Calculating difference in Issue Keys for Target Project.")
        get_shifted_val()
        print("[END] The difference of Issue Keys for Target Project would be: '{}'".format(shifted_by))
        print("")
    
    if migrate_statuses_check == 1 or json_importer_flag == 1 or multiple_json_data_processing == 1:
        get_transitions(project_new, JIRA_BASE_URL_NEW, new=True)
        try:
            get_transitions(project_old, JIRA_BASE_URL_OLD, new=False)
        except:
            print("[WARNING] No PROJECT ADMIN rigts available for Source '{}' project. Sub-Tasks can't be converted into Issues.".format(project_old))
    
    # Validate values in template
    if check_template_flag == 1:
        print("[START] Template validation started.")
        try:
            validation_template_error = validate_template()
            if validation_template_error == 1:
                return
            print("[END] Template validation has been completed. No critical issues were found.")
        except Exception as e:
            print("[ERROR] Exception while processing validation: '{}'.".format(e))
            if verbose_logging == 1:
                print(traceback.format_exc())
        print("")
    
    get_hierarchy_config()
    
    # Calculating the highest level of available Key in OLD project
    if process_only_last_updated_date_flag == 1 or bulk_processing_flag == 1:
        limit_migration_data = 0
        start_jira_key = 1
    start_jira_key = project_old + '-' + str(start_jira_key)
    jql_max = 'project = {} order by key DESC'.format(project_old)
    if limit_migration_data != 0:
        try:
            max_processing_key = project_old + '-' + str(int(limit_migration_data) + int(start_jira_key.split('-')[1]))
        except:
            max_processing_key = jira_old.search_issues(jql_str=jql_max, maxResults=1, json_result=False)[0].key
    else:
        try:
            max_processing_key = jira_old.search_issues(jql_str=jql_max, maxResults=1, json_result=False)[0].key
        except:
            print("[ERROR] No issues found in Source project. Exiting...")
            print("")
            processing_error = 1
            return
    start_jira_key = find_min_id(start_jira_key, jira_old, project_old)
    
    # Creating Agile board for the Project for further Sprints migration - if there are no yet one
    if new_board_id == 0 and migrate_sprints_check == 1:
        print("[START] Agile Board processing for Sprints.")
        get_new_board_id()
        print("[END] Agile Board has been found / created.")
        print("")
    
    # Checking for Sprint changed between Source and Target
    if migrate_sprints_check == 1 and (json_importer_flag == 1 or multiple_json_data_processing == 1) and force_sprints_update_flag == 1:
        print("[START] Changed Sprints / Issues would be refreshed in Target '{}' project.".format(project_new))
        refresh_sprints()
        print("[END] Changed Sprints has been removed. They will be re-created with latest data available.")
        print("")
    
    # Calculating Max ID for the project
    try:
        max_id = find_max_id(max_processing_key, jira_old, project_old)
    except:
        print("[ERROR] There no issues below '{}'. Exiting...".format(max_processing_key))
        print("")
        processing_error = 1
        return
    
    # Teams Migration (skipping if no mapping to Portfolio Teams)
    if migrate_teams_check == 1 and same_new_project == 0:
        teams_processed = 0
        start_teams_time = time.time()
        for f_mappings in fields_mappings.values():
            if 'Team' in f_mappings.keys():
                get_all_shared_teams()
                print("[INFO] Teams loaded in '{}' seconds.".format(time.time() - start_teams_time))
                print("")
                teams_processed = 1
                break
        if teams_processed == 0:
            print("[WARNING] Teams mapping hasn't been found - Teams would not be processed.")
            print("")
            migrate_teams_check = 0
    elif migrate_teams_check == 0:
        for issuestype, fields in fields_mappings.items():
            if 'Team' in fields.keys():
                fields_mappings[issuestype].pop('Team', None)
    
    # Components Migration
    if migrate_components_check == 1:
        start_components_time = time.time()
        migrate_components()
        print("[INFO] Components migrated in '{}' seconds.".format(time.time() - start_components_time))
        print("")
    
    # FixVersions Migration
    if migrate_fixversions_check == 1:
        start_versions_time = time.time()
        migrate_versions()
        print("[INFO] FixVersions migrated in '{}' seconds.".format(time.time() - start_versions_time))
        print("")
    
    # Creating missing Dummy issues
    if json_importer_flag == 0 and multiple_json_data_processing == 0 and migrate_metadata_check == 1:
        start_dummy_time = time.time()
        jql_max_new = 'project = {} order by key desc'.format(project_new)
        try:
            max_new_id = jira_new.search_issues(jql_str=jql_max_new, maxResults=1, json_result=False)[0].key
        except:
            max_new_id = None
        if max_new_id is not None and max_id is not None and int(max_new_id.split('-')[1]) < int(max_id.split('-')[1]):
            issues_for_creation = int(max_id.split('-')[1]) - int(max_new_id.split('-')[1])
        elif max_new_id is None and max_id is not None:
            issues_for_creation = int(max_id.split('-')[1])
        else:
            issues_for_creation = 0
        if len(migrated_issues_lst) > 0:
            create_dummy_issues(issues_for_creation, batch_size=100)
        print("[INFO] Dummy Issues created in '{}' seconds.".format(time.time() - start_dummy_time))
        print("")
    
    # Main Migration block
    start_processing_time = time.time()
    migration_process(start_jira_key, max_processing_key, max_id)
    print("[INFO] Issues have been migrated in '{}' seconds.".format(time.time() - start_processing_time))
    print("")
    
    # Re-try missed items
    start_retry_time = time.time()
    
    while True:
        start_update_sprints = time.time()
        
        # Reconciliation logic
        if process_reconciliation_flag == 1:
            diff = set()
            old_issues_keys_set = set()
            new_issues_keys_set = set()
            new_issues_lst = [['Target JIRA Key', 'Issue Type', 'Summary', 'Priority', 'Status', 'Resolution', 'Assignee', 'Reporter', 'Created', 'Due Date', 'Parent', 'Labels']]
            old_issues_lst = [['Source JIRA Key', 'Issue Type', 'Summary', 'Priority', 'Status', 'Resolution', 'Assignee', 'Reporter', 'Created', 'Due Date', 'Parent', 'Labels']]
            missing_issues_lst = [['Source JIRA Key', 'Issue Type', 'Summary', 'Priority', 'Status', 'Resolution', 'Assignee', 'Reporter', 'Created', 'Due Date', 'Parent', 'Labels']]
            header_set = {'Target JIRA Key', 'Source JIRA Key'}
            
            if reconciliation_updated_days != '0':
                print("[START] Reconciliation process is started. The difference between Source and Target projects for the last '{}' days  will be calculated.".format(reconciliation_updated_days))
                jql_old = 'project = {} {} AND updated >= startOfDay(-{})'.format(project_old, supported_issuetypes, reconciliation_updated_days)
            else:
                print("[START] Reconciliation process is started. The difference between Source and Target projects will be calculated.")
                jql_old = 'project = {} {}'.format(project_old, supported_issuetypes)
            jql_new = "project = {} AND (labels not in ('MIGRATION_NOT_COMPLETE') OR labels is EMPTY)".format(project_new)
            
            print("[INFO] Source Issues for Reconciliation logic is being loaded with JQL: '{}'...".format(jql_old))
            old_issues_lst.extend(get_issues_by_jql(jira_old, jql=jql_old, non_migrated=True))
            old_issues_keys_set = set([i[0] for i in old_issues_lst])
            print("[INFO] Source issues count: '{}'".format(len(old_issues_keys_set) - 1))
            print("[INFO] Target Issues for Reconciliation logic is being loaded with JQL: '{}'...".format(jql_new))
            new_issues_lst.extend(get_issues_by_jql(jira_new, jql=jql_new, non_migrated=True))
            new_issues_keys_set = set([get_shifted_key(i[0], reversed=True).replace(project_new + '-', project_old + '-') for i in new_issues_lst])
            print("[INFO] Target issues count: '{}'".format(len(new_issues_keys_set) - 1))
            diff = old_issues_keys_set - new_issues_keys_set - header_set
            failed_issues |= diff
            for issue in old_issues_lst:
                if issue[0] in diff:
                    missing_issues_lst.append(issue)
            if len(diff) > 0:
                print("[WARNING] Not all issues were migrated. The number of missing issues: '{}'".format(len(diff)))
                if verbose_logging == 1:
                    print("Issues missing in the Target Instance: '{}'".format(list(diff)))
            else:
                print("[INFO] No missing issues for Reconciliation logic were found.")
            print("[END] Reconciliation process has been completed.")
            print("")
        
        # Calculating total Number of Issues in OLD JIRA Project
        try:
            jql_total_old = "project = {} {} {}".format(project_old, recently_updated, supported_issuetypes)
            total_old = jira_old.search_issues(jql_total_old, startAt=0, maxResults=1, json_result=True)['total']
        except:
            total_old = len(set(migrated_issues_lst))
        
        # Calculating total Number of Migrated Issues to NEW JIRA Project
        jql_total_new = "project = '{}' AND (labels not in ('MIGRATION_NOT_COMPLETE') OR labels is EMPTY) AND summary !~ 'DUMMY_PARENT'".format(project_new)
        total_new = jira_new.search_issues(jql_total_new, startAt=0, maxResults=1, json_result=True)['total']
        
        jql_non_completed_new = "project = '{}' AND labels in ('MIGRATION_NOT_COMPLETE') ".format(project_new)
        non_completed_new = jira_new.search_issues(jql_non_completed_new, startAt=0, maxResults=1, json_result=True)['total']
        
        failed_issues = set(migrated_issues_lst) | failed_issues
        failed_issues = failed_issues - processed_issues_set
        
        if int(total_old) <= int(total_new) and int(non_completed_new) == 0 and len(failed_issues) == 0 and refresh_already_migrated_flag == 0:
            # Update and Close Sprints - after migration of issues are done
            if migrate_sprints_check == 1 and json_importer_flag == 0 and multiple_json_data_processing == 0:
                migrate_sprints(proj_old=project_old, param='CLOSED')
                migrate_sprints(proj_old=project_old, param='ACTIVE')
            print("[INFO] ALL Issues have been updated.")
            print("[INFO] Issues in Source Project: '{}'".format(total_old))
            print("[INFO] Issues in Target Project: '{}'".format(total_new))
            print("")
            break
        elif ((len(set(migrated_issues_lst)) == len(processed_issues_set) and len(failed_issues) == 0)
              or (refresh_already_migrated_flag == 1 and len(failed_issues) == 0)):
            print("[INFO] ALL Requested Issues have been migrated.")
            print("[INFO] Issues in Source Project: '{}'".format(total_old))
            print("[INFO] Issues in Target Project: '{}'".format(total_new))
            break
        else:
            remaining = int(non_completed_new) + len(failed_issues)
            remaining = remaining if remaining > 0 else (int(total_old) - int(total_new))
            if migrate_sprints_check == 1 and json_importer_flag == 0 and multiple_json_data_processing == 0:
                print("[WARNING] Not ALL issues have been migrated from '{}' project. Remaining Issues: '{}'. Sprints will not be CLOSED until ALL issues migrated.".format(project_old, remaining if remaining > 0 else 0))
                print("[INFO] Sprints have been updated in '{}' seconds.".format(time.time() - start_update_sprints))
            else:
                print("[WARNING] Not ALL issues have been migrated from '{}' project. Remaining Issues: '{}'.".format(project_old, remaining if remaining > 0 else 0))
            print("[ERROR] Not processed issues so far: '{}'".format(list(failed_issues)))
            print("")
        
        if (retry_logic_flag == 1 or process_reconciliation_flag == 1) and remaining > 0 and (remaining_previous > remaining or remaining_previous == 0):
            remaining_previous = remaining
            print("")
            print("[INFO] Re-try logic for skipped issues started.")
            
            # Reset Default lists
            items_lst, issues_lst = ({}, {})
            already_migrated_set = set()
            
            migration_process(start_jira_key, max_processing_key, max_id, reprocess=True)
            print("[INFO] Issues have been migrated in '{}' seconds.".format(time.time() - start_retry_time))
            print("")
        else:
            print("[WARNING] Not All issues were migrated from '{}' project. Remaining Issues: '{}'.".format(project_old, remaining if remaining > 0 else 0))
            break
    
    # Delete issues with Summary = 'Dummy Issue'
    if json_importer_flag == 0 and multiple_json_data_processing == 0 and migrate_metadata_check == 1:
        start_delete_time = time.time()
        delete_extra_issues(max_id)
        print("[INFO] Dummy issues have been deleted/skipped in '{}' seconds.".format(time.time() - start_delete_time))
        print("")
    if dummy_process == 0:
        delete_issue(dummy_parent)
    
    if process_reconciliation_excel_flag == 1:
        create_recon_excel(old_issues_lst, new_issues_lst, missing_issues_lst)
    
    # Update Source Project as Read-Only after migration
    if set_source_project_read_only == 1:
        status = set_project_as_read_only(JIRA_BASE_URL_OLD, project_old)
        if str(status) != str(200):
            print("[ERROR] Source '{}' project can't be set to Read-Only.".format(project_old))
            print("")
        else:
            print("[INFO] Source Project has been updated to Read-Only after migration.")
            print("")
    
    if remaining == 0:
        print("[COMPLETE] Source Project: '{}'; Target Project: '{}'".format(project_old, project_new))
        process_complete = 1
    else:
        print("[PARTIALLY COMPLETE] Source Project: '{}'; Target Project: '{}'".format(project_old, project_new))
        process_partially_complete = 1
    print("[INFO] TOTAL processing time: '{}' seconds.".format(time.time() - start_time))
    print("")


def create_recon_excel(old_issues_lst, new_issues_lst, missing_issues_lst):
    
    def save_excel():
        # Saving Excel file and removing not required sheets
        global project_old, verbose_logging
        
        sheet_names = wb.sheetnames
        for s in sheet_names:
            ws = wb.get_sheet_by_name(s)
            if ws.dimensions == 'A1:A1':
                wb.remove_sheet(wb[s])
        
        report_name = '{} Source vs Target Issue.xlsx'.format(project_old)
        try:
            set_zoom(report_name)
            print("[END] Reconciliation File '{}' has been successfully generated.".format(report_name))
            print("")
        except Exception as e:
            print("[ERROR] Reconciliation Excel can't be created due to: {}".format(e))
            if verbose_logging == 1:
                print(traceback.format_exc())
    
    def set_zoom(file, zoom_scale=90):
        for ws in wb.worksheets:
            ws.sheet_view.zoomScale = zoom_scale
        wb.save(file)
    
    def create_excel_sheet(sheet_data, title, new=False):
        global JIRA_BASE_URL_OLD, JIRA_BASE_URL_NEW, hyperlink
        
        wb.create_sheet(title)
        ws = wb.get_sheet_by_name(title)
        
        start_column = 1
        start_row = 1
        
        if new is True:
            JIRA_BASE_URL = JIRA_BASE_URL_NEW
        else:
            JIRA_BASE_URL = JIRA_BASE_URL_OLD
        
        # Creating Excel sheet based on data
        for i in range(len(sheet_data)):
            for y in range(len(sheet_data[i])):
                try:
                    if start_row != 1 and sheet_data[i][y] != '' and sheet_data[i][y] != 'n/a' and y == 0:
                        ws.cell(row=start_row, column=start_column+y).hyperlink = JIRA_BASE_URL + '/browse/' + sheet_data[i][y]
                        ws.cell(row=start_row, column=start_column+y).font = hyperlink
                    ws.cell(row=start_row, column=start_column+y).value = sheet_data[i][y]
                except:
                    converted_value = ''
                    for letter in sheet_data[i][y]:
                        if letter.isalpha() or letter.isnumeric() or letter in [',', '.', ';', ':', '&', '"', "'", ' ', '-', '_']:
                            converted_value += letter
                        else:
                            converted_value += '?'
                    ws.cell(row=start_row, column=start_column+y).value = converted_value
            start_row += 1
        
        for y in range(1, ws.max_column+1):
            ws.cell(row=1, column=y).fill = header_fill
            ws.cell(row=1, column=y).font = header_font
        
        # Column width formatting
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            if length > 80:
                ws.column_dimensions[column_cells[0].column_letter].width = 80
            else:
                ws.column_dimensions[column_cells[0].column_letter].width = length + 4
        
        ws.title = title
        ws.auto_filter.ref = ws.dimensions
    
    print("[START] Reconciliation Excel will be generated.")
    create_excel_sheet(old_issues_lst, title='Source Issues')
    create_excel_sheet(new_issues_lst, title='Target Issues', new=True)
    create_excel_sheet(missing_issues_lst, title='Missing Issues')
    save_excel()


def move_processed_template(folder, filename):
    global process_complete, process_partially_complete, validation_template_error, processing_error
    global process_complete_folder, process_partially_complete_folder, validation_template_error_folder
    global processing_error_folder
    
    old_file_path = os.path.join(folder, filename)
    if process_complete == 1:
        folder_path = os.path.join(folder, process_complete_folder)
        if not os.path.exists(folder_path):
            os.mkdir(folder_path)
        new_file_path = os.path.join(folder_path, filename)
        shutil.move(old_file_path, new_file_path)
    elif process_partially_complete == 1:
        folder_path = os.path.join(folder, process_partially_complete_folder)
        if not os.path.exists(folder_path):
            os.mkdir(folder_path)
        new_file_path = os.path.join(folder_path, filename)
        shutil.move(old_file_path, new_file_path)
    elif validation_template_error == 1:
        folder_path = os.path.join(folder, validation_template_error_folder)
        if not os.path.exists(folder_path):
            os.mkdir(folder_path)
        new_file_path = os.path.join(folder_path, filename)
        shutil.move(old_file_path, new_file_path)
    elif processing_error == 1:
        folder_path = os.path.join(folder, processing_error_folder)
        if not os.path.exists(folder_path):
            os.mkdir(folder_path)
        new_file_path = os.path.join(folder_path, filename)
        shutil.move(old_file_path, new_file_path)
    return


def main_program():
    """Migration Processing Main Function - covering 'End to End' process."""
    global auth, username, password, mapping_file, temp_dir_name, validation_error, last_updated_date
    global shifted_by, read_only_scheme_name, shifted_key_val, recently_updated_days, bulk_processing_flag
    global credentials_saved_flag, refresh_issuetypes, reconciliation_updated_days
    
    username = user.get()
    password = passwd.get()
    credentials_saved_flag = credentials_saved.get()
    read_only_scheme_name = permission_scheme.get().strip()
    last_updated_date = last_updated_main.get().strip()
    refresh_issuetypes = refresh_issuetypes_field.get().strip()
    
    if refresh_issuetypes not in ['', 'ALL']:
        new_refresh_issuetypes = []
        for issuetype in refresh_issuetypes.split(','):
            new_refresh_issuetypes.append("'" + str(issuetype.strip()) + "'")
        refresh_issuetypes = get_str_from_lst(new_refresh_issuetypes)
    
    try:
        shifted_by = int(start_num.get().strip())
    except:
        shifted_by = 1000
    try:
        shifted_key_val = int(shift_num.get().strip())
    except:
        shifted_key_val = 1000
    if bulk_processing_flag == 0:
        mapping_file = file.get().strip().split('.xls')[0] + '.xlsx'
    
    recently_updated_days = days.get().strip()
    try:
        recently_updated_days = str(int(recently_updated_days))
    except:
        print("[ERROR] The number of Days for Last Updated period should be a Number. Default value '365' will be used.")
        recently_updated_days = '365'
    
    reconciliation_updated_days = recon_days.get().strip()
    try:
        reconciliation_updated_days = str(int(reconciliation_updated_days))
    except:
        print("[ERROR] The number of Days for Reconciliation logic should be a Number. Default value '365' will be used.")
        reconciliation_updated_days = '365'
    
    # Checking the all mandatory fields are populated on Config page
    if validation_error == 1:
        change_configs()
    
    # Checking the Mapping File available
    if os.path.exists(mapping_file) is False or (bulk_processing_flag == 0 and mapping_file == '.xlsx'):
        load_file()
    main.destroy()
    
    if os.path.exists(mapping_file) is False or mapping_file == '.xlsx':
        print("[ERROR] Mapping File not found. Migration failed.")
        os.system("pause")
        exit()
    
    # Checking the JIRA credentials
    if len(username) < 3 or len(password) < 3:
        print('[ERROR] JIRA credentials are required. Please enter them on new window.')
        jira_authorization_popup()
    else:
        auth = (username, password)
    
    # Save some data in config file
    save_config(message=False)
    
    # Starting Program
    print("[START] Migration process has been started. Please wait...")
    print("")
    
    # Creating / Cleaning Folder for Attachments migration
    if migrate_attachments_check == 1:
        create_temp_folder(temp_dir_name)
    
    if bulk_processing_flag == 1:
        for filename in os.listdir(mapping_file):
            if '.xlsx' in filename:
                file_path = os.path.join(mapping_file, filename)
                process_one_template(file_path)
                move_processed_template(mapping_file, filename)
                print("___________________________________________________________________________")
                print("")
    else:
        process_one_template(mapping_file)
    
    # Cleaning Folder for Attachments migration
    if migrate_attachments_check == 1:
        clean_temp_folder(temp_dir_name)
    
    print("[INFO] Migration complete.")
    print("")
    os.system("pause")
    exit()


def overwrite_popup():
    """Function which shows Pop-Up window with question about overriding Excel file, if it already exists"""
    global mapping_file
    
    def create_new():
        global mapping_file
        popup.destroy()
        popup.quit()
        time_format = "%Y-%m-%dT%H:%M:%S"
        now = datetime.datetime.strptime(str(datetime.datetime.utcnow().isoformat()).split(".", 1)[0], time_format)
        mapping_file = mapping_file.split('.xls')[0] + '_' + str(now).replace(':', '-').replace(' ', '_') + '_UTC.xlsx'
    
    def override():
        popup.destroy()
        popup.quit()
    
    popup = tk.Tk()
    popup.title("Override File?")
    
    l1 = tk.Label(popup, text="File '{}' already exist.".format(mapping_file), foreground="black", font=("Helvetica", 10), pady=4, padx=8)
    l1.grid(row=0, column=0, columnspan=2)
    l2 = tk.Label(popup, text="Do you want to override existing file OR create a new one?", foreground="black", font=("Helvetica", 10), pady=4, padx=8)
    l2.grid(row=1, column=0, columnspan=2)
    
    b1 = tk.Button(popup, text="Override", font=("Helvetica", 9, "bold"), command=override, width=20, heigh=2)
    b1.grid(row=2, column=0, pady=10, padx=8)
    b2 = tk.Button(popup, text="Create New", font=("Helvetica", 9, "bold"), command=create_new, width=20, heigh=2)
    b2.grid(row=2, column=1, pady=10, padx=8)
    
    tk.mainloop()


def get_shifted_val():
    global shifted_key_val, shifted_by, jira_new, project_new, merge_projects_flag
    
    if merge_projects_flag == 1:
        jql_max = 'project = {} order by key DESC'.format(project_new)
        max_processing_key = jira_new.search_issues(jql_str=jql_max, maxResults=1, json_result=False)[0].key
        shifted_by = int(shifted_key_val) + int(max_processing_key.split('-')[1])


def get_shifted_key(key, reversed=False):
    global shifted_by, merge_projects_flag, merge_projects_start_flag
    
    new_key = key
    if merge_projects_flag == 0 and merge_projects_start_flag == 0:
        return new_key
    if reversed is False:
        new_id = int(key.split('-')[1]) + int(shifted_by)
    else:
        new_id = int(key.split('-')[1]) - int(shifted_by)
    new_project = str(key.split('-')[0])
    new_key = str(new_project) + '-' + str(new_id)
    return new_key


def change_configs():
    """Function which shows Pop-Up window with question about JIRA credentials, if not entered"""
    global start_jira_key, limit_migration_data, template_project, new_project_name, processing_jira_jql
    global default_board_name, old_board_id, team_project_prefix, last_updated_date, threads, pool_size, project_old
    global max_json_file_size
    
    def config_save():
        global start_jira_key, limit_migration_data, pool_size, template_project, new_project_name, processing_jira_jql
        global default_board_name, old_board_id, team_project_prefix, validation_error, last_updated_date, threads
        global project_old, max_json_file_size
        
        validation_error = 0
        
        start_jira_key = first_issue.get()
        limit_migration_data = migrated_number.get()
        default_board_name = new_board.get()
        old_board_id = old_board.get()
        team_project_prefix = new_teams.get()
        threads = threads_num.get().strip()
        pool_size = process_num.get().strip()
        last_updated_date = last_updated.get().strip()
        template_project = template_proj.get()
        new_project_name = name_proj.get()
        processing_jira_jql = jql.get().strip()
        max_json_file_size = max_size.get().strip()
        
        if last_updated_date == 'YYYY-MM-DD':
            last_updated_date = ''
        
        config_popup.destroy()
        
        if start_jira_key == '':
            start_jira_key = 1
        try:
            start_jira_key = int(start_jira_key.strip())
            if start_jira_key < 1:
                start_jira_key = 1
        except:
            try:
                start_jira_key = str(start_jira_key.strip()).split('-')[1]
            except:
                print("[ERROR] Start Issue Key is invalid.")
        
        if limit_migration_data in ['', 'ALL']:
            limit_migration_data = 0
        try:
            limit_migration_data = int(limit_migration_data.strip())
            if limit_migration_data < 0:
                limit_migration_data = 0
                print("[ERROR] Number of Total migrated issues can't be NEGATIVE. Defaulted to '0' - 'ALL'.")
        except:
            print("[ERROR] Number of Total migrated issues is invalid. Default value for ALL would be used.")
            limit_migration_data = 0
        
        try:
            default_board_name = str(default_board_name.strip())
            if default_board_name == '' and migrate_sprints_check == 1:
                default_board_name = 'Shared Sprints'
                print("[ERROR] Board Name for migrated Sprints can't be empty. Default Name '{}' will be used instead.".format(default_board_name))
        except:
            print("[ERROR] New Board name for migrated Sprints is invalid.")
            if migrate_sprints_check == 1:
                validation_error = 1
        
        try:
            if old_board_id.strip() != '':
                old_board_id = int(old_board_id.strip())
        except:
            if migrate_sprints_check == 1:
                print("[ERROR] Board ID for Sprints migration from Source JIRA in invalid. By default ALL Sprints will be migrated.")
            old_board_id = 0
        
        try:
            team_project_prefix = str(team_project_prefix)
        except:
            print("[ERROR] Prefix for Team names is invalid. Default '[{}] ' will be used.".format(project_old))
            team_project_prefix = '[' + project_old + '] '
        
        try:
            threads = int(threads.strip())
        except:
            print("[ERROR] Parallel Threads number is invalid. Default '1' will be used.")
            threads = 1
        
        try:
            max_json_file_size = int(max_json_file_size.strip())
        except:
            print("[ERROR] File Size is invalid. Default '10' will be used.")
            max_json_file_size = 10
        
        try:
            pool_size = int(pool_size.strip())
        except:
            print("[ERROR] Number of processes is invalid. Default '1' will be used.")
            pool_size = 1
        
        try:
            template_project = str(template_project).strip()
        except:
            template_project = ''
        
        try:
            new_project_name = str(new_project_name).strip()
        except:
            new_project_name = ''
        
        if processing_jira_jql == "key in ()":
            processing_jira_jql = ''
        
        if validation_error == 1:
            print("[WARNING] Mandatory Config data is invalid or empty. Please check the Config data again.")
        
        last_updated_main.delete(0, END)
        last_updated_main.insert(0, last_updated_date)
        
        save_config()
        config_popup.quit()
    
    def config_popup_close():
        config_popup.destroy()
        config_popup.quit()
    
    def check_similar(field, value):
        """ This function required for fixing same valu duplication issue for second Tk window """
        global start_jira_key, limit_migration_data, pool_size, default_board_name, old_board_id, new_project_name
        global team_project_prefix, validation_error, last_updated_date, threads, template_project, processing_jira_jql
        global max_json_file_size
        
        fields = {"start_jira_key": start_jira_key,
                  "limit_migration_data": limit_migration_data,
                  "default_board_name": default_board_name,
                  "old_board_id": old_board_id,
                  "team_project_prefix": team_project_prefix,
                  "max_json_file_size": max_json_file_size,
                  "validation_error": validation_error,
                  "last_updated_date": last_updated_date,
                  "threads": threads,
                  "pool_size": pool_size,
                  "template_project": template_project,
                  "new_project_name": new_project_name,
                  "processing_jira_jql": processing_jira_jql,
                  }
        for f, v in fields.items():
            if str(value) == str(v) and field != f:
                return check_similar(field, ' ' + str(value))
        else:
            return value
    
    config_popup = tk.Tk()
    config_popup.title("JIRA Migration Tool - Configuration")
    
    tk.Label(config_popup, text="Detailed Configuration for migration. Defaults are '0' or empty for ALL Sprints / Issues:", foreground="black", font=("Helvetica", 11, "italic"), padx=10, pady=10, wraplength=600).grid(row=3, column=0, columnspan=5)
    
    start_jira_key = check_similar("start_jira_key", start_jira_key)
    
    tk.Label(config_popup, text="Start migration from (Issue Key or Number):", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=300).grid(row=4, column=0, columnspan=2)
    first_issue = tk.Entry(config_popup, width=20, textvariable=start_jira_key)
    first_issue.insert(END, start_jira_key)
    first_issue.grid(row=4, column=2, columnspan=1, padx=8)
    
    if limit_migration_data == 0:
        limit_migration_data = 'ALL'
    limit_migration_data = check_similar("limit_migration_data", limit_migration_data)
    
    tk.Label(config_popup, text="Number for migration:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=200).grid(row=4, column=3)
    migrated_number = tk.Entry(config_popup, width=10, textvariable=limit_migration_data)
    migrated_number.delete(0, END)
    migrated_number.insert(0, limit_migration_data)
    migrated_number.grid(row=4, column=4, columnspan=1, padx=8)
    
    if processing_jira_jql == '':
        processing_jira_jql = "key in ()"
    processing_jira_jql = check_similar("processing_jira_jql", processing_jira_jql)
    
    tk.Label(config_popup, text="OR (instead of above) JQL/comma-separated Issues:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=300).grid(row=5, column=0, columnspan=2)
    jql = tk.Entry(config_popup, width=62, textvariable=processing_jira_jql)
    jql.insert(END, processing_jira_jql)
    jql.grid(row=5, column=2, columnspan=3, padx=8)
    
    default_board_name = check_similar("default_board_name", default_board_name)
    
    tk.Label(config_popup, text="New Board name for migrated Sprints:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=250).grid(row=6, column=0, columnspan=2)
    new_board = tk.Entry(config_popup, width=20, textvariable=default_board_name)
    new_board.insert(END, default_board_name)
    new_board.grid(row=6, column=2, columnspan=1, padx=8)
    
    if old_board_id == 0:
        old_board_id = ''
    old_board_id = check_similar("old_board_id", old_board_id)
    
    tk.Label(config_popup, text="Sprints from Board ID only:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=200).grid(row=6, column=3)
    old_board = tk.Entry(config_popup, width=10, textvariable=old_board_id)
    old_board.delete(0, END)
    old_board.insert(0, old_board_id)
    old_board.grid(row=6, column=4, columnspan=1, padx=8)
    
    team_project_prefix = check_similar("team_project_prefix", team_project_prefix)
    
    tk.Label(config_popup, text="Prefix for Team names, if migrated:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=250).grid(row=7, column=0, columnspan=2)
    new_teams = tk.Entry(config_popup, width=20, textvariable=team_project_prefix)
    new_teams.insert(END, team_project_prefix)
    new_teams.grid(row=7, column=2, columnspan=1, padx=8)
    
    max_json_file_size = check_similar("max_json_file_size", max_json_file_size)
    
    tk.Label(config_popup, text="Approximate JSON file size, Mb:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=250).grid(row=8, column=0, columnspan=2)
    max_size = tk.Entry(config_popup, width=20, textvariable=max_json_file_size)
    max_size.insert(END, max_json_file_size)
    max_size.grid(row=8, column=2, columnspan=1, padx=8)
    
    threads = check_similar("threads", threads)
    
    tk.Label(config_popup, text="Parallel Threads:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=200).grid(row=7, column=3)
    threads_num = tk.Entry(config_popup, width=10, textvariable=threads)
    threads_num.delete(0, END)
    threads_num.insert(0, threads)
    threads_num.grid(row=7, column=4, columnspan=1, padx=8)
    
    pool_size = check_similar("pool_size", pool_size)
    
    tk.Label(config_popup, text="Number Processes:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=200).grid(row=8, column=3)
    process_num = tk.Entry(config_popup, width=10, textvariable=pool_size)
    process_num.delete(0, END)
    process_num.insert(0, pool_size)
    process_num.grid(row=8, column=4, columnspan=1, padx=8)
    
    if last_updated_date == '':
        last_updated_date = 'YYYY-MM-DD'
    
    last_updated_date = check_similar("last_updated_date", last_updated_date)
    
    tk.Label(config_popup, text="Force update issues changed after that date, i.e. 'last updated >=  :", foreground="black", font=("Helvetica", 10), pady=7, padx=8, wraplength=500).grid(row=9, column=0, columnspan=4)
    last_updated = tk.Entry(config_popup, width=15, textvariable=last_updated_date)
    last_updated.insert(END, last_updated_date)
    last_updated.grid(row=9, column=3, columnspan=2, padx=70, stick=W)
    
    tk.Label(config_popup, text="____________________________________________________________________________________________________________").grid(row=10, columnspan=5)
    
    tk.Label(config_popup, text="If Source Project doesn't exist, it could be created as copy of Template Project Key:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=550).grid(row=11, column=0, columnspan=4, stick=W)
    template_proj = tk.Entry(config_popup, width=20, textvariable=template_project)
    template_proj.insert(END, template_project)
    template_proj.grid(row=11, column=3, columnspan=2, padx=30, stick=E)
    
    tk.Label(config_popup, text="and Target Project Name would be:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=550).grid(row=12, column=1, columnspan=3, stick=W)
    name_proj = tk.Entry(config_popup, width=45, textvariable=new_project_name)
    name_proj.insert(END, new_project_name)
    name_proj.grid(row=12, column=2, columnspan=3, padx=30, stick=E)
    
    tk.Label(config_popup, text="____________________________________________________________________________________________________________").grid(row=13, columnspan=5)
    
    tk.Button(config_popup, text='Cancel', font=("Helvetica", 9, "bold"), command=config_popup_close, width=20, heigh=2).grid(row=14, column=0, pady=8, padx=20, sticky=W, columnspan=3)
    tk.Button(config_popup, text='Save', font=("Helvetica", 9, "bold"), command=config_save, width=20, heigh=2).grid(row=14, column=2, pady=8, padx=20, sticky=E, columnspan=3)
    
    tk.mainloop()


def change_mappings_configs():
    """Function which shows Pop-Up window with question about JIRA credentials, if not entered"""
    global JIRA_BASE_URL_OLD, project_old, JIRA_BASE_URL_NEW, project_new, template_project, new_project_name
    
    def config_save():
        global JIRA_BASE_URL_OLD, project_old, JIRA_BASE_URL_NEW, project_new, template_project, new_project_name
        global mapping_file, protection_password, excel_locked
        
        validation_error = 0
        
        JIRA_BASE_URL_OLD = source_jira.get()
        JIRA_BASE_URL_NEW = target_jira.get()
        project_old = source_project.get()
        project_new = target_project.get()
        template_project = template_proj.get()
        new_project_name = name_proj.get()
        mapping_file = file.get().split('.xls')[0] + '.xlsx'
        protection_password = protect.get().strip()
        config_mapping_popup.destroy()
        
        if protection_password == '':
            excel_locked = 0
        
        try:
            JIRA_BASE_URL_OLD = str(JIRA_BASE_URL_OLD).strip('/').strip()
        except:
            if JIRA_BASE_URL_OLD == '':
                print("[ERROR] Source JIRA URL is empty.")
            else:
                print("[ERROR] Source JIRA URL is invalid.")
            validation_error = 1
        if JIRA_BASE_URL_OLD == '':
            print("[ERROR] Source JIRA URL is empty.")
            validation_error = 1
        
        try:
            JIRA_BASE_URL_NEW = str(JIRA_BASE_URL_NEW).strip('/').strip()
        except:
            if JIRA_BASE_URL_NEW == '':
                print("[ERROR] Target JIRA URL is empty.")
            else:
                print("[ERROR] Target JIRA URL is invalid.")
            validation_error = 1
        if JIRA_BASE_URL_NEW == '':
            print("[ERROR] Target JIRA URL is empty.")
            validation_error = 1
        
        try:
            project_old = str(project_old).strip()
        except:
            if project_old == '':
                print("[ERROR] Source JIRA Project Key is empty.")
            else:
                print("[ERROR] Source JIRA Project Key is invalid.")
            validation_error = 1
        if project_old == '':
            print("[ERROR] Source JIRA Project Key is empty.")
        
        try:
            project_new = str(project_new).strip()
        except:
            print("[ERROR] Target JIRA Project Key is invalid.")
            validation_error = 1
        if project_new == '':
            print("[WARNING] Target JIRA Project Key is empty. Would be used same as Sourse.")
            project_new = project_old
        
        try:
            template_project = str(template_project).strip()
        except:
            template_project = ''
        
        try:
            new_project_name = str(new_project_name).strip()
        except:
            new_project_name = ''
        
        if validation_error == 1:
            print("[WARNING] Mandatory Config data is invalid or empty. Please check the Config data again.")
        save_config()
        config_mapping_popup.quit()
    
    def config_mapping_popup_close():
        config_mapping_popup.destroy()
        config_mapping_popup.quit()
        exit()
    
    def check_similar(field, value):
        """ This function required for fixing same valu duplication issue for second Tk window """
        global JIRA_BASE_URL_OLD, project_old, JIRA_BASE_URL_NEW, project_new, template_project, new_project_name
        
        fields = {"JIRA_BASE_URL_OLD": JIRA_BASE_URL_OLD,
                  "project_old": project_old,
                  "JIRA_BASE_URL_NEW": JIRA_BASE_URL_NEW,
                  "project_new": project_new,
                  "template_project": template_project,
                  "new_project_name": new_project_name,
                  "protection_password": protection_password,
                  }
        for f, v in fields.items():
            if str(value) == str(v) and field != f:
                return check_similar(field, ' ' + str(value))
        else:
            return value
    
    config_mapping_popup = tk.Tk()
    config_mapping_popup.title("JIRA Migration Tool - Configuration for Mapping Generation.")
    
    JIRA_BASE_URL_OLD = check_similar("JIRA_BASE_URL_OLD", JIRA_BASE_URL_OLD)
    
    tk.Label(config_mapping_popup, text="Please enter Source and Target project for migration (Target Project Key will be same, if empty):", foreground="black", font=("Helvetica", 11, "italic", "underline"), pady=10).grid(row=0, column=0, columnspan=4, rowspan=1, sticky=W, padx=60)
    
    tk.Label(config_mapping_popup, text="Source JIRA URL:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=150).grid(row=1, column=0, rowspan=1)
    source_jira = tk.Entry(config_mapping_popup, width=60, textvariable=JIRA_BASE_URL_OLD)
    source_jira.insert(END, JIRA_BASE_URL_OLD)
    source_jira.grid(row=1, column=1, padx=8)
    
    project_old = check_similar("project_old", project_old)
    
    tk.Label(config_mapping_popup, text="Source Project Key:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=150).grid(row=1, column=2, rowspan=1, stick=W)
    source_project = tk.Entry(config_mapping_popup, width=20, textvariable=project_old)
    source_project.insert(END, project_old)
    source_project.grid(row=1, column=3, padx=7, stick=E)
    
    JIRA_BASE_URL_NEW = check_similar("JIRA_BASE_URL_NEW", JIRA_BASE_URL_NEW)
    
    tk.Label(config_mapping_popup, text="Target JIRA URL:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=150).grid(row=2, column=0, rowspan=1)
    target_jira = tk.Entry(config_mapping_popup, width=60, textvariable=JIRA_BASE_URL_NEW)
    target_jira.insert(END, JIRA_BASE_URL_NEW)
    target_jira.grid(row=2, column=1, padx=8)
    
    project_new = check_similar("project_new", project_new)
    
    tk.Label(config_mapping_popup, text="Target Project Key:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=150).grid(row=2, column=2, rowspan=1, stick=W)
    target_project = tk.Entry(config_mapping_popup, width=20, textvariable=project_new)
    target_project.insert(END, project_new)
    target_project.grid(row=2, column=3, padx=7, stick=E)

    protection_password = check_similar("protection_password", protection_password)
    
    tk.Label(config_mapping_popup, text="Excel Lock Password:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=150).grid(row=3, column=0, rowspan=1)
    protect = tk.Entry(config_mapping_popup, width=60, textvariable=protection_password)
    protect.insert(END, protection_password)
    protect.grid(row=3, column=1, padx=8)
    
    mapping_file = 'Migration Template for {} project.xlsx'.format(project_old.strip(), project_new.strip())
    
    tk.Label(config_mapping_popup, text="Template File Name:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=150).grid(row=4, column=0, rowspan=1, sticky=W)
    file = tk.Entry(config_mapping_popup, width=83, textvariable=mapping_file)
    file.insert(END, mapping_file)
    file.grid(row=4, column=1, columnspan=2, padx=0)
    tk.Button(config_mapping_popup, text='Browse', command=load_file, width=15).grid(row=4, column=3, pady=3, padx=8)
    
    tk.Label(config_mapping_popup, text="____________________________________________________________________________________________________________").grid(row=5, columnspan=4)
    
    template_project = check_similar("template_project", template_project)
    
    tk.Label(config_mapping_popup, text="If Source Project doesn't exist, it could be created as copy of Template Project Key:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=550).grid(row=6, column=1, columnspan=2, stick=W)
    template_proj = tk.Entry(config_mapping_popup, width=20, textvariable=template_project)
    template_proj.insert(END, template_project)
    template_proj.grid(row=6, column=3, padx=7, stick=E)
    
    tk.Label(config_mapping_popup, text="and Target Project Name would be:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=550).grid(row=7, column=1, columnspan=2, stick=E, padx=120)
    name_proj = tk.Entry(config_mapping_popup, width=40, textvariable=new_project_name)
    name_proj.insert(END, new_project_name)
    name_proj.grid(row=7, column=2, columnspan=2, padx=7, stick=E)
    
    tk.Label(config_mapping_popup, text="____________________________________________________________________________________________________________").grid(row=8, columnspan=4)
    
    tk.Button(config_mapping_popup, text='Exit', font=("Helvetica", 9, "bold"), command=config_mapping_popup_close, width=20, heigh=2).grid(row=10, column=0, pady=8, padx=100, sticky=W, columnspan=4)
    tk.Button(config_mapping_popup, text='Save', font=("Helvetica", 9, "bold"), command=config_save, width=20, heigh=2).grid(row=10, column=0, pady=8, padx=100, sticky=E, columnspan=4)
    
    tk.mainloop()


def jira_authorization_popup():
    """Function which shows Pop-Up window with question about JIRA credentials, if not entered"""
    global auth, username, password, jira_old, jira_new, atlassian_jira_old, JIRA_BASE_URL_OLD, JIRA_BASE_URL_NEW
    
    def jira_save():
        global auth, username, password, jira_old, jira_new, atlassian_jira_old, JIRA_BASE_URL_OLD, JIRA_BASE_URL_NEW
        
        username = user.get()
        password = passwd.get()
        if len(username) < 3 or len(password) < 3:
            print("Invalid JIRA credentials were entered!")
            os.system("pause")
            exit()
        auth = (username, password)
        jira_popup.destroy()
        
        if verbose_logging == 1:
            print("[INFO] A connection attempt to JIRA server is started.")
        get_jira_connection()
        jira_popup.quit()
    
    def jira_cancel():
        jira_popup.destroy()
        jira_popup.quit()
        print("[ERROR] Invalid JIRA credentials were entered! Program exits...")
        os.system("pause")
        exit()
    
    jira_popup = tk.Tk()
    jira_popup.title("[AUTHORIZATION] JIRA credentials required")
    
    tk.Label(jira_popup, text="To Migrate issues please enter your Username / Password for JIRA access.", foreground="black", font=("Helvetica", 9), padx=10, wraplength=210).grid(row=1, column=0, rowspan=2)
    tk.Label(jira_popup, text="Username").grid(row=1, column=1, pady=5)
    tk.Label(jira_popup, text="Password").grid(row=2, column=1, pady=5)
    
    user = tk.Entry(jira_popup)
    user.grid(row=1, column=2, pady=5)
    passwd = tk.Entry(jira_popup, width=20, show="*")
    passwd.grid(row=2, column=2, pady=5)
    
    tk.Button(jira_popup, text='Cancel', font=("Helvetica", 9, "bold"), command=jira_cancel, width=20, heigh=2).grid(row=4, column=0, pady=8, padx=20, sticky=W, columnspan=2)
    tk.Button(jira_popup, text='OK', font=("Helvetica", 9, "bold"), command=jira_save, width=20, heigh=2).grid(row=4, column=1, pady=8, padx=20, sticky=E, columnspan=2)
    
    tk.mainloop()


def change_migrate_fixversions(*args):
    global migrate_fixversions_check
    migrate_fixversions_check = process_fixversions.get()


def change_migrate_components(*args):
    global migrate_components_check
    migrate_components_check = process_components.get()


def change_migrate_sprints(*args):
    global migrate_sprints_check
    migrate_sprints_check = process_sprints.get()


def change_migrate_attachments(*args):
    global migrate_attachments_check
    migrate_attachments_check = process_attachments.get()


def change_migrate_metadata(*args):
    global migrate_metadata_check
    migrate_metadata_check = process_metadata.get()


def change_migrate_comments(*args):
    global migrate_comments_check
    migrate_comments_check = process_comments.get()


def change_migrate_links(*args):
    global migrate_links_check
    migrate_links_check = process_links.get()


def change_migrate_statuses(*args):
    global migrate_statuses_check
    migrate_statuses_check = process_statuses.get()


def change_migrate_teams(*args):
    global migrate_teams_check
    migrate_teams_check = process_teams.get()


def change_logging(*args):
    global verbose_logging
    verbose_logging = process_logging.get()


def change_linking(*args):
    global create_remote_link_for_old_issue
    create_remote_link_for_old_issue = process_old_linkage.get()


def change_dummy(*args):
    global delete_dummy_flag
    delete_dummy_flag = process_dummy_del.get()


def change_migrated(*args):
    global skip_migrated_flag, merge_projects_flag, merge_projects_start_flag, process_only_last_updated_date
    global refresh_already_migrated_flag
    
    skip_migrated_flag = process_non_migrated.get()
    if skip_migrated_flag == 1 and merge_projects_flag == 1:
        merge_projects_flag = 0
        merge_projects.set(merge_projects_flag)
        merge_projects_start_flag = 1
        merge_projects_start.set(merge_projects_start_flag)
    if skip_migrated_flag == 1:
        process_only_last_updated_date_flag = 0
        process_only_last_updated_date.set(process_only_last_updated_date_flag)
        refresh_already_migrated_flag = 0
        refresh_already_migrated.set(refresh_already_migrated_flag)


def change_process_last_updated(*args):
    global last_updated_days_check, including_dependencies_flag, process_only_last_updated_date
    global refresh_already_migrated_flag
    
    last_updated_days_check = process_last_updated.get()
    if last_updated_days_check == 1:
        including_dependencies_flag = 1
        process_dependencies.set(including_dependencies_flag)
        process_only_last_updated_date_flag = 0
        process_only_last_updated_date.set(process_only_last_updated_date_flag)
        refresh_already_migrated_flag = 0
        refresh_already_migrated.set(refresh_already_migrated_flag)


def change_dependencies(*args):
    global including_dependencies_flag, process_only_last_updated_date, refresh_already_migrated_flag
    global last_updated_days_check
    
    including_dependencies_flag = process_dependencies.get()
    if including_dependencies_flag == 1:
        process_only_last_updated_date_flag = 0
        process_only_last_updated_date.set(process_only_last_updated_date_flag)
        refresh_already_migrated_flag = 0
        refresh_already_migrated.set(refresh_already_migrated_flag)
        last_updated_days_check = 1
        process_last_updated.set(last_updated_days_check)


def change_process_reconciliation(*args):
    global process_reconciliation_flag, process_reconciliation_excel_flag
    
    process_reconciliation_flag = process_reconciliation.get()
    if process_reconciliation_flag == 1:
        process_reconciliation_excel_flag = 1
        process_reconciliation_excel.set(process_reconciliation_excel_flag)
    else:
        process_reconciliation_excel_flag = 0
        process_reconciliation_excel.set(process_reconciliation_excel_flag)


def change_process_reconciliation_excel(*args):
    global process_reconciliation_flag, process_reconciliation_excel_flag
    
    process_reconciliation_excel_flag = process_reconciliation_excel.get()
    if process_reconciliation_excel_flag == 1:
        process_reconciliation_flag = 1
        process_reconciliation.set(process_reconciliation_flag)


def change_read_only(*args):
    global set_source_project_read_only
    set_source_project_read_only = set_read_only.get()


def change_credentials_saved(*args):
    global credentials_saved_flag
    credentials_saved_flag = credentials_saved.get()


def change_bulk_processing(*args):
    global bulk_processing_flag
    bulk_processing_flag = bulk_processing.get()


def change_jsons(*args):
    global multiple_json_data_processing, json_importer_flag, force_update_flag, process_only_last_updated_date_flag
    global json_files_autoupload
    
    multiple_json_data_processing = process_jsons.get()
    if multiple_json_data_processing == 1:
        json_importer_flag = 0
    else:
        json_files_autoupload = 0
        process_jsons_auto.set(json_files_autoupload)
    if json_importer_flag == 0 and (force_update_flag == 1 or including_users_flag == 1 or replace_complete_statuses_flag == 1):
        json_importer_flag = 1
        process_change_history.set(json_importer_flag)
    if multiple_json_data_processing == 1 and process_only_last_updated_date_flag == 1:
        force_update_flag = 0
        force_update.set(force_update_flag)
    elif multiple_json_data_processing == 0 and process_only_last_updated_date_flag == 1:
        force_update_flag = 1
        force_update.set(force_update_flag)


def change_jsons_auto(*args):
    global multiple_json_data_processing, json_importer_flag, force_update_flag, json_files_autoupload
    
    json_files_autoupload = process_jsons_auto.get()
    if json_files_autoupload == 1:
        multiple_json_data_processing = 1
        process_jsons.set(multiple_json_data_processing)
        json_importer_flag = 0


def change_override_template(*args):
    global override_template_flag
    override_template_flag = override_template.get()


def change_control_logic(*args):
    global control_logic_flag
    control_logic_flag = control_logic.get()


def change_retry_logic(*args):
    global retry_logic_flag
    retry_logic_flag = retry_logic.get()


def change_merge_project(*args):
    global merge_projects_flag, merge_projects_start_flag, skip_migrated_flag
    merge_projects_flag = merge_projects.get()
    if merge_projects_start_flag == 1 and merge_projects_flag == 1:
        merge_projects_start_flag = 0
        merge_projects_start.set(merge_projects_start_flag)
    if skip_migrated_flag == 1 and merge_projects_flag == 1:
        skip_migrated_flag = 0
        process_non_migrated.set(skip_migrated_flag)


def change_merge_project_start(*args):
    global merge_projects_start_flag, merge_projects_flag
    merge_projects_start_flag = merge_projects_start.get()
    if merge_projects_flag == 1 and merge_projects_start_flag == 1:
        merge_projects_flag = 0
        merge_projects.set(merge_projects_flag)


def change_migrate_history(*args):
    global json_importer_flag, including_users_flag, replace_complete_statuses_flag, force_update_flag
    global multiple_json_data_processing
    
    json_importer_flag = process_change_history.get()
    if json_importer_flag == 0 and multiple_json_data_processing == 0:
        including_users_flag = 0
        process_users.set(including_users_flag)
        replace_complete_statuses_flag = 0
        process_change_history_statuses.set(replace_complete_statuses_flag)
        force_update_flag = 0
        force_update.set(force_update_flag)


def change_migrate_history_statuses(*args):
    global replace_complete_statuses_flag, json_importer_flag
    replace_complete_statuses_flag = process_change_history_statuses.get()
    if replace_complete_statuses_flag == 1:
        json_importer_flag = 1
        process_change_history.set(json_importer_flag)


def change_users(*args):
    global including_users_flag, json_importer_flag
    including_users_flag = process_users.get()
    if including_users_flag == 1:
        json_importer_flag = 1
        process_change_history.set(json_importer_flag)


def change_force_update(*args):
    global force_update_flag, json_importer_flag, force_sprints_update_flag
    force_update_flag = force_update.get()
    if force_update_flag == 1:
        json_importer_flag = 1
        process_change_history.set(json_importer_flag)
    else:
        force_sprints_update_flag = 0
        force_sprints_update.set(force_sprints_update_flag)


def change_force_sprints_update(*args):
    global force_sprints_update_flag, force_update_flag
    force_sprints_update_flag = force_sprints_update.get()
    if force_sprints_update_flag == 1:
        force_update_flag = 1
        force_update.set(force_update_flag)


def change_validate_template(*args):
    global check_template_flag, skip_existing_issuetypes_validation_flag
    check_template_flag = check_validate_template.get()
    if check_template_flag == 0:
        skip_existing_issuetypes_validation_flag = 0
        skip_existing_issuetypes_validation.set(skip_existing_issuetypes_validation_flag)
    else:
        skip_existing_issuetypes_validation_flag = 1
        skip_existing_issuetypes_validation.set(skip_existing_issuetypes_validation_flag)


def change_skip_existing_issuetypes_validation(*args):
    global skip_existing_issuetypes_validation_flag, check_template_flag
    skip_existing_issuetypes_validation_flag = skip_existing_issuetypes_validation.get()
    if skip_existing_issuetypes_validation_flag == 1:
        check_template_flag = 1
        check_validate_template.set(check_template_flag)


def change_refresh_already_migrated(*args):
    global refresh_already_migrated_flag, last_updated_days_check, skip_migrated_flag, including_dependencies_flag
    global force_sprints_update_flag, process_only_last_updated_date_flag
    
    refresh_already_migrated_flag = refresh_already_migrated.get()
    if refresh_already_migrated_flag == 1:
        last_updated_days_check = 0
        process_last_updated.set(last_updated_days_check)
        skip_migrated_flag = 0
        process_non_migrated.set(skip_migrated_flag)
        including_dependencies_flag = 0
        process_dependencies.set(including_dependencies_flag)
        force_sprints_update_flag = 0
        force_sprints_update.set(force_sprints_update_flag)
        process_only_last_updated_date_flag = 0
        process_only_last_updated_date.set(process_only_last_updated_date_flag)
        process_only_last_updated_date_flag = 0


def change_process_last_updated_date(*args):
    global process_only_last_updated_date_flag, last_updated_days_check, skip_migrated_flag, including_dependencies_flag
    global force_update_flag, control_logic_flag, multiple_json_data_processing, refresh_already_migrated_flag
    
    process_only_last_updated_date_flag = process_only_last_updated_date.get()
    if process_only_last_updated_date_flag == 1:
        last_updated_days_check = 0
        process_last_updated.set(last_updated_days_check)
        skip_migrated_flag = 0
        process_non_migrated.set(skip_migrated_flag)
        including_dependencies_flag = 0
        process_dependencies.set(including_dependencies_flag)
        force_update_flag = 1
        force_update.set(force_update_flag)
        force_sprints_update_flag = 0
        force_sprints_update.set(force_sprints_update_flag)
        control_logic_flag = 1
        control_logic.set(control_logic_flag)
        refresh_already_migrated_flag = 0
        refresh_already_migrated.set(refresh_already_migrated_flag)
    else:
        force_update_flag = 0
        force_update.set(force_update_flag)
        control_logic_flag = 0
        control_logic.set(control_logic_flag)
    if multiple_json_data_processing == 1 and process_only_last_updated_date_flag == 1:
        force_update_flag = 0
        force_update.set(force_update_flag)
    elif multiple_json_data_processing == 0 and process_only_last_updated_date_flag == 1:
        force_update_flag = 1
        force_update.set(force_update_flag)


def change_clear_additional_configuration(*args):
    global process_only_last_updated_date_flag, last_updated_days_check, skip_migrated_flag, including_dependencies_flag
    global merge_projects_flag, set_source_project_read_only, check_template_flag, create_remote_link_for_old_issue
    global clear_additional_configuration_flag, merge_projects_start_flag, multiple_json_data_processing
    global delete_dummy_flag, process_reconciliation_flag, process_reconciliation_excel_flag, verbose_logging
    global refresh_already_migrated_flag, retry_logic_flag, skip_existing_issuetypes_validation_flag
    
    clear_additional_configuration_flag = clear_additional_configuration.get()
    if clear_additional_configuration_flag == 1:
        last_updated_days_check = 0
        process_last_updated.set(last_updated_days_check)
        skip_migrated_flag = 0
        process_non_migrated.set(skip_migrated_flag)
        including_dependencies_flag = 0
        process_dependencies.set(including_dependencies_flag)
        merge_projects_flag = 0
        merge_projects.set(merge_projects_flag)
        process_only_last_updated_date_flag = 0
        process_only_last_updated_date.set(process_only_last_updated_date_flag)
        merge_projects_start_flag = 0
        merge_projects_start.set(merge_projects_start_flag)
        set_source_project_read_only = 0
        set_read_only.set(set_source_project_read_only)
        check_template_flag = 0
        check_validate_template.set(check_template_flag)
        skip_existing_issuetypes_validation_flag = 0
        skip_existing_issuetypes_validation.set(skip_existing_issuetypes_validation_flag)
        clear_additional_configuration_flag = 0
        clear_additional_configuration.set(clear_additional_configuration_flag)
        verbose_logging = 0
        process_logging.set(verbose_logging)
        process_reconciliation_flag = 0
        process_reconciliation.set(process_reconciliation_flag)
        process_reconciliation_excel_flag = 0
        process_reconciliation_excel.set(process_reconciliation_excel_flag)
        create_remote_link_for_old_issue = 0
        process_old_linkage.set(create_remote_link_for_old_issue)
        delete_dummy_flag = 0
        process_dummy_del.set(delete_dummy_flag)
        multiple_json_data_processing = 0
        process_jsons.set(multiple_json_data_processing)
        refresh_already_migrated_flag = 0
        refresh_already_migrated.set(refresh_already_migrated_flag)
        retry_logic_flag = 0
        retry_logic.set(retry_logic_flag)


def change_change_all_configuration(*args):
    global change_configuration_flag, migrate_fixversions_check, migrate_components_check, migrate_sprints_check
    global migrate_attachments_check, migrate_metadata_check, migrate_comments_check, migrate_links_check
    global migrate_statuses_check, migrate_teams_check, json_importer_flag, force_update_flag, including_users_flag
    global replace_complete_statuses_flag, retry_logic_flag, control_logic_flag
    
    change_configuration_flag = change_all_configuration.get()
    if change_configuration_flag == 0:
        migrate_fixversions_check = 0
        process_fixversions.set(migrate_fixversions_check)
        migrate_components_check = 0
        process_components.set(migrate_components_check)
        migrate_sprints_check = 0
        process_sprints.set(migrate_sprints_check)
        migrate_attachments_check = 0
        process_attachments.set(migrate_attachments_check)
        migrate_metadata_check = 0
        process_metadata.set(migrate_metadata_check)
        migrate_comments_check = 0
        process_comments.set(migrate_comments_check)
        migrate_links_check = 0
        process_links.set(migrate_links_check)
        migrate_statuses_check = 0
        process_statuses.set(migrate_statuses_check)
        migrate_teams_check = 0
        process_teams.set(migrate_teams_check)
        json_importer_flag = 0
        process_change_history.set(json_importer_flag)
        including_users_flag = 0
        process_users.set(including_users_flag)
        replace_complete_statuses_flag = 0
        process_change_history_statuses.set(replace_complete_statuses_flag)
        control_logic_flag = 0
        control_logic.set(control_logic_flag)
        retry_logic_flag = 0
        retry_logic.set(retry_logic_flag)
    else:
        migrate_fixversions_check = 1
        process_fixversions.set(migrate_fixversions_check)
        migrate_components_check = 1
        process_components.set(migrate_components_check)
        migrate_sprints_check = 1
        process_sprints.set(migrate_sprints_check)
        migrate_attachments_check = 1
        process_attachments.set(migrate_attachments_check)
        migrate_metadata_check = 1
        process_metadata.set(migrate_metadata_check)
        migrate_comments_check = 1
        process_comments.set(migrate_comments_check)
        migrate_links_check = 1
        process_links.set(migrate_links_check)
        migrate_statuses_check = 1
        process_statuses.set(migrate_statuses_check)
        migrate_teams_check = 1
        process_teams.set(migrate_teams_check)
        json_importer_flag = 1
        process_change_history.set(json_importer_flag)
        including_users_flag = 1
        process_users.set(including_users_flag)
        replace_complete_statuses_flag = 1
        process_change_history_statuses.set(replace_complete_statuses_flag)
        control_logic_flag = 1
        control_logic.set(control_logic_flag)
        retry_logic_flag = 1
        retry_logic.set(retry_logic_flag)


def check_similar(field, value):
    """ This function required for fixing same valu duplication issue for second Tk window """
    global shifted_by, shifted_key_val, last_updated_date, read_only_scheme_name, recently_updated_days
    global mapping_file, default_configuration_file, refresh_issuetypes, reconciliation_updated_days
    
    fields = {"shifted_by": shifted_by,
              "shifted_key_val": shifted_key_val,
              "last_updated_date": last_updated_date,
              "read_only_scheme_name": read_only_scheme_name,
              "recently_updated_days": recently_updated_days,
              "reconciliation_updated_days": reconciliation_updated_days,
              "mapping_file": mapping_file,
              "refresh_issuetypes": refresh_issuetypes,
              "default_configuration_file": default_configuration_file,
              }
    for f, v in fields.items():
        if str(value) == str(v) and field != f:
            return check_similar(field, ' ' + str(value))
    else:
        return value


def check_latest_log_file():
    global log_file, logs_folder
    
    file_path = os.path.join(logs_folder, log_file)
    if os.path.exists(file_path):
        try:
            last_log_number = int(log_file.split('.txt')[0].split('__')[1]) + 1
            log_file = log_file.split('.txt')[0].split('__')[0] + '__' + str(last_log_number) + '.txt'
            check_latest_log_file()
        except:
            log_file = log_file.split('.txt')[0] + '__1.txt'
            check_latest_log_file()
    return


# ------------------ MAIN PROGRAM -----------------------------------
if __name__ == "__main__":
    
    create_temp_folder(logs_folder, clean=False)
    check_latest_log_file()
    logs_file_path = os.path.join(logs_folder, log_file)
    logging.basicConfig(level=logging.INFO, filename=logs_file_path)
    old_print = print
    
    def print(string, string2='', string3='', sep=' ', end='\n'):
        if string2 == '':
            old_print(string, sep=sep, end=end)
            logging.info(string)
        elif string3 == '':
            old_print(string, string2, sep=sep, end=end)
            logging.info(string)
            logging.info(string2)
        else:
            old_print(string, string2, string3, sep=sep, end=end)
            logging.info(string)
            logging.info(string2)
            logging.info(string3)
    
    print("[INFO] Program has started. Please DO NOT CLOSE that window.")
    load_config()
    print("[INFO] Please IGNORE any WARNINGS - the connection issues are covered by Retry logic.")
    print("")
    
    main = tk.Tk()
    Title = main.title("JIRA Migration Tool" + " v_" + current_version)
    
    tk.Label(main, text="Mapping Template Generation", foreground="black", font=("Helvetica", 11, "italic", "underline"), pady=10).grid(row=0, column=0, columnspan=2, rowspan=2, sticky=W, padx=80)
    tk.Label(main, text="Step 1", foreground="black", font=("Helvetica", 12, "bold", "underline"), pady=10).grid(row=0, column=0, columnspan=3, rowspan=2, sticky=W, padx=15)
    
    tk.Button(main, text='Generate Template', font=("Helvetica", 9, "bold"), command=generate_template, width=20, heigh=2).grid(row=0, column=3, pady=5, rowspan=2, sticky=N)
    
    tk.Label(main, text="_____________________________________________________________________________________________________________________________").grid(row=2, columnspan=4)
    
    tk.Label(main, text="Step 2", foreground="black", font=("Helvetica", 12, "bold", "underline"), pady=10).grid(row=4, column=0, columnspan=3, rowspan=1, sticky=W, padx=15)
    tk.Label(main, text="Migration Configuration", foreground="black", font=("Helvetica", 11, "italic", "underline"), pady=10).grid(row=4, column=0, columnspan=4, sticky=W, padx=80)
    
    mapping_file = check_similar("mapping_file", mapping_file)
    
    tk.Label(main, text="Mapping Template:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=150).grid(row=3, column=0, rowspan=1, padx=80, sticky=W, columnspan=1)
    file = tk.Entry(main, width=77, textvariable=mapping_file)
    file.insert(END, mapping_file)
    file.grid(row=3, column=0, columnspan=3, sticky=E, padx=0)
    tk.Button(main, text='Browse', command=load_file, width=15).grid(row=3, column=3, pady=3, padx=8)
    
    bulk_processing = IntVar(value=bulk_processing_flag)
    Checkbutton(main, text="process folder for Bulk templates update", foreground="grey", font=("Helvetica", 9, "italic"), variable=bulk_processing).grid(row=3, column=0, sticky=NE, padx=170, columnspan=4, rowspan=2, pady=27)
    bulk_processing.trace('w', change_bulk_processing)
    
    default_configuration_file = check_similar("default_configuration_file", default_configuration_file)
    
    tk.Label(main, text="Default Mapping File:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=150).grid(row=5, column=0, rowspan=1, padx=80, sticky=W, columnspan=1)
    default_file = tk.Entry(main, width=77, textvariable=default_configuration_file)
    default_file.insert(END, default_configuration_file)
    default_file.grid(row=5, column=0, columnspan=3, sticky=E, padx=0)
    tk.Button(main, text='Browse', command=load_default_file, width=15).grid(row=5, column=3, pady=3, padx=8)
    
    override_template = IntVar(value=override_template_flag)
    Checkbutton(main, text="override template", foreground="grey", font=("Helvetica", 9, "italic"), variable=override_template).grid(row=5, column=0, sticky=NE, padx=170, columnspan=4, rowspan=3, pady=27)
    override_template.trace('w', change_override_template)
    
    change_all_configuration = IntVar(value=change_configuration_flag)
    Checkbutton(main, text="select / unselect all", foreground="grey", font=("Helvetica", 9, "italic"), variable=change_all_configuration).grid(row=10, column=0, sticky=NE, padx=210, columnspan=4, rowspan=2, pady=5)
    change_all_configuration.trace('w', change_change_all_configuration)
    
    process_fixversions = IntVar(value=migrate_fixversions_check)
    Checkbutton(main, text="Migrate all fixVersions / Releases from Source JIRA.", font=("Helvetica", 9, "italic"), variable=process_fixversions).grid(row=6, sticky=W, padx=70, column=0, columnspan=3, pady=0)
    process_fixversions.trace('w', change_migrate_fixversions)
    
    process_components = IntVar(value=migrate_components_check)
    Checkbutton(main, text="Migrate all Components from Source JIRA.", font=("Helvetica", 9, "italic"), variable=process_components).grid(row=7, sticky=W, padx=70, column=0, columnspan=3, pady=0)
    process_components.trace('w', change_migrate_components)
    
    process_sprints = IntVar(value=migrate_sprints_check)
    Checkbutton(main, text="Migrate Sprints (specified in Configs) from Source JIRA (Agile Add-on).", font=("Helvetica", 9, "italic"), variable=process_sprints).grid(row=8, sticky=W, padx=70, column=0, columnspan=3, pady=0)
    process_sprints.trace('w', change_migrate_sprints)
    
    process_teams = IntVar(value=migrate_teams_check)
    Checkbutton(main, text="Migrate Teams from Source JIRA (Portfolio Add-on).", font=("Helvetica", 9, "italic"), variable=process_teams).grid(row=9, sticky=W, padx=70, column=0, columnspan=3, pady=0)
    process_teams.trace('w', change_migrate_teams)
    
    process_metadata = IntVar(value=migrate_metadata_check)
    Checkbutton(main, text="Migrate Metadata (field values) for Issues.", font=("Helvetica", 9, "italic"), variable=process_metadata).grid(row=10, sticky=W, padx=70, column=0, columnspan=3, pady=0)
    process_metadata.trace('w', change_migrate_metadata)
    
    process_attachments = IntVar(value=migrate_attachments_check)
    Checkbutton(main, text="Migrate Attachments from Source JIRA issues.", font=("Helvetica", 9, "italic"), variable=process_attachments).grid(row=11, sticky=W, padx=70, column=0, columnspan=3, pady=0)
    process_attachments.trace('w', change_migrate_attachments)
    
    process_comments = IntVar(value=migrate_comments_check)
    Checkbutton(main, text="Migrate Comments from Source JIRA issues.", font=("Helvetica", 9, "italic"), variable=process_comments).grid(row=12, sticky=W, padx=70, column=0, columnspan=3, pady=0)
    process_comments.trace('w', change_migrate_comments)
    
    process_links = IntVar(value=migrate_links_check)
    Checkbutton(main, text="Migrate Links from Source JIRA issues.", font=("Helvetica", 9, "italic"), variable=process_links).grid(row=13, sticky=W, padx=70, column=0, columnspan=3, pady=0)
    process_links.trace('w', change_migrate_links)
    
    process_statuses = IntVar(value=migrate_statuses_check)
    Checkbutton(main, text="Update Statuses / Resolutions from Source JIRA issues (Project Admin access required).", font=("Helvetica", 9, "italic"), variable=process_statuses).grid(row=14, sticky=W, padx=70, column=0, columnspan=3, pady=0)
    process_statuses.trace('w', change_migrate_statuses)
    
    process_change_history = IntVar(value=json_importer_flag)
    Checkbutton(main, text="Update Change History / Worklogs from Source JIRA issues (Global Admin access required)", font=("Helvetica", 9, "italic"), variable=process_change_history).grid(row=15, sticky=W, padx=70, column=0, columnspan=3, pady=0)
    process_change_history.trace('w', change_migrate_history)
    
    process_users = IntVar(value=including_users_flag)
    Checkbutton(main, text="Including Users", font=("Helvetica", 9, "italic"), variable=process_users).grid(row=15, column=1, sticky=E, padx=136, columnspan=4, pady=0)
    process_users.trace('w', change_users)
    
    force_update = IntVar(value=force_update_flag)
    Checkbutton(main, text="Force update", font=("Helvetica", 9, "italic"), variable=force_update).grid(row=15, column=1, sticky=E, padx=40, columnspan=4, pady=0)
    force_update.trace('w', change_force_update)
    
    force_sprints_update = IntVar(value=force_sprints_update_flag)
    Checkbutton(main, text="incl. Sprints", font=("Helvetica", 9, "italic"), variable=force_sprints_update).grid(row=16, column=1, sticky=E, padx=40, columnspan=4, pady=0)
    force_sprints_update.trace('w', change_force_sprints_update)
    
    process_change_history_statuses = IntVar(value=replace_complete_statuses_flag)
    Checkbutton(main, text="Replace Completed statuses in Change history by Target ones (Agile Reporting - Velocity, Sprint Reports)", font=("Helvetica", 9, "italic"), variable=process_change_history_statuses).grid(row=16, sticky=W, padx=110, column=0, columnspan=5, pady=0)
    process_change_history_statuses.trace('w', change_migrate_history_statuses)
    
    tk.Button(main, text='Change Configuration', font=("Helvetica", 9, "bold"), state='active', command=change_configs, width=20, heigh=2).grid(row=8, column=3, pady=4, rowspan=4)
    
    tk.Label(main, text="_____________________________________________________________________________________________________________________________").grid(row=17, columnspan=4)
    
    tk.Label(main, text="Migration Process", foreground="black", font=("Helvetica", 11, "italic", "underline"), pady=10).grid(row=18, column=0, columnspan=3, sticky=W, padx=80)
    
    tk.Label(main, text="Step 3", foreground="black", font=("Helvetica", 12, "bold", "underline"), pady=10).grid(row=18, column=0, columnspan=3, rowspan=1, sticky=W, padx=15)
    
    tk.Label(main, text="For migration process please enter your Username / Password for JIRA(s) access", foreground="black", font=("Helvetica", 10), padx=10, wraplength=260).grid(row=19, column=0, rowspan=2, columnspan=3, sticky=W, padx=80)
    
    credentials_saved = IntVar(value=credentials_saved_flag)
    Checkbutton(main, text="Save credentials", font=("Helvetica", 9, "italic"), variable=credentials_saved).grid(row=20, column=0, sticky=W, padx=280, columnspan=3, rowspan=2, pady=0)
    credentials_saved.trace('w', change_credentials_saved)
    
    tk.Label(main, text="Username", foreground="black", font=("Helvetica", 10)).grid(row=19, column=1, pady=5, columnspan=3, sticky=W, padx=20)
    tk.Label(main, text="Password", foreground="black", font=("Helvetica", 10)).grid(row=20, column=1, pady=5, columnspan=3, sticky=W, padx=20)
    
    user = tk.Entry(main, textvariable=username)
    user.delete(0, END)
    user.insert(END, username)
    user.grid(row=19, column=1, pady=5, sticky=W, columnspan=3, padx=100)
    
    passwd = tk.Entry(main, width=20, show="*", textvariable=password)
    passwd.delete(0, END)
    passwd.insert(END, password)
    passwd.grid(row=20, column=1, pady=5, sticky=W, columnspan=3, padx=100)
    
    tk.Button(main, text='Start JIRA Migration', font=("Helvetica", 9, "bold"), state='active', command=main_program, width=20, heigh=2).grid(row=19, column=3, pady=4, padx=10, rowspan=2)
    
    tk.Label(main, text="_____________________________________________________________________________________________________________________________").grid(row=21, columnspan=4)
    
    tk.Label(main, text="Additional Configuration", foreground="black", font=("Helvetica", 10, "italic", "underline"), pady=10).grid(row=22, column=0, columnspan=4, sticky=W, padx=300)
    
    process_logging = IntVar(value=verbose_logging)
    Checkbutton(main, text="Switch Verbose Logging ON for migration process.", font=("Helvetica", 9, "italic"), variable=process_logging).grid(row=23, column=0, sticky=W, padx=20, columnspan=3, pady=0)
    process_logging.trace('w', change_logging)
    
    process_dummy_del = IntVar(value=delete_dummy_flag)
    Checkbutton(main, text="Skip deletion of dummy issues (for testing purposes).", font=("Helvetica", 9, "italic"), variable=process_dummy_del).grid(row=24, column=0, sticky=W, padx=20, columnspan=3, pady=0)
    process_dummy_del.trace('w', change_dummy)
    
    process_old_linkage = IntVar(value=create_remote_link_for_old_issue)
    Checkbutton(main, text="Add Remote Links to Source Issues.", font=("Helvetica", 9, "italic"), variable=process_old_linkage).grid(row=23, column=1, sticky=W, padx=70, columnspan=3, pady=0)
    process_old_linkage.trace('w', change_linking)
    
    process_jsons = IntVar(value=multiple_json_data_processing)
    Checkbutton(main, text="Create JSON files", font=("Helvetica", 9, "italic"), variable=process_jsons).grid(row=24, column=1, sticky=W, padx=70, columnspan=3, pady=0)
    process_jsons.trace('w', change_jsons)
    
    process_jsons_auto = IntVar(value=json_files_autoupload)
    Checkbutton(main, text="auto-upload", font=("Helvetica", 9, "italic"), variable=process_jsons_auto).grid(row=24, column=1, sticky=E, padx=140, columnspan=3, pady=0)
    process_jsons_auto.trace('w', change_jsons_auto)
    
    process_non_migrated = IntVar(value=skip_migrated_flag)
    Checkbutton(main, text="Skip already migrated issues.", font=("Helvetica", 9, "italic"), variable=process_non_migrated).grid(row=25, column=1, sticky=W, padx=70, columnspan=3, pady=0)
    process_non_migrated.trace('w', change_migrated)
    
    process_only_last_updated_date = IntVar(value=process_only_last_updated_date_flag)
    Checkbutton(main, text="Force Delta processing after date, i.e. 'last updated' >=  :", font=("Helvetica", 9, "italic"), variable=process_only_last_updated_date).grid(row=25, column=0, sticky=W, padx=20, columnspan=4, pady=0)
    process_only_last_updated_date.trace('w', change_process_last_updated_date)
    
    if last_updated_date == '':
        last_updated_date = 'YYYY-MM-DD'
    
    last_updated_date = check_similar("last_updated_date", last_updated_date)
    
    last_updated_main = tk.Entry(main, width=15, textvariable=last_updated_date)
    last_updated_main.delete(0, END)
    last_updated_main.insert(END, last_updated_date)
    last_updated_main.grid(row=25, column=0, columnspan=4, padx=340, stick=W)
    
    control_logic = IntVar(value=control_logic_flag)
    Checkbutton(main, text="Apply Control logic for recently touched issues.", font=("Helvetica", 9, "italic"), variable=control_logic).grid(row=26, column=0, sticky=W, padx=20, columnspan=3, pady=0)
    control_logic.trace('w', change_control_logic)
    
    retry_logic = IntVar(value=retry_logic_flag)
    Checkbutton(main, text="Re-try all failed issues later.", font=("Helvetica", 9, "italic"), variable=retry_logic).grid(row=26, column=1, sticky=W, padx=70, columnspan=3, pady=0)
    retry_logic.trace('w', change_retry_logic)
    
    process_last_updated = IntVar(value=last_updated_days_check)
    Checkbutton(main, text="ONLY migrate issues updated or created within the last number of days:", font=("Helvetica", 9, "italic"), variable=process_last_updated).grid(row=27, column=0, sticky=W, padx=20, columnspan=4, pady=0)
    process_last_updated.trace('w', change_process_last_updated)
    
    recently_updated_days = check_similar("recently_updated_days", recently_updated_days)
    
    days = tk.Entry(main, width=5, textvariable=recently_updated_days)
    days.insert(END, recently_updated_days)
    days.grid(row=27, column=1, pady=0, sticky=W, columnspan=3, padx=24)
    
    process_dependencies = IntVar(value=including_dependencies_flag)
    Checkbutton(main, text="Including dependencies (Parents / Sub-tasks / Links).", font=("Helvetica", 9, "italic"), variable=process_dependencies).grid(row=27, column=1, sticky=W, padx=55, columnspan=3, pady=0)
    process_dependencies.trace('w', change_dependencies)
    
    refresh_already_migrated = IntVar(value=refresh_already_migrated_flag)
    Checkbutton(main, text="Refresh metadata for already migrated items. Process the following Target issuetypes:", font=("Helvetica", 9, "italic"), variable=refresh_already_migrated).grid(row=28, column=0, sticky=W, padx=20, columnspan=4, pady=0)
    refresh_already_migrated.trace('w', change_refresh_already_migrated)
    
    if refresh_issuetypes == '':
        refresh_issuetypes = 'ALL'
    
    refresh_issuetypes = check_similar("refresh_issuetypes", refresh_issuetypes)
    
    refresh_issuetypes_field = tk.Entry(main, width=41, textvariable=refresh_issuetypes)
    refresh_issuetypes_field.delete(0, END)
    refresh_issuetypes_field.insert(END, refresh_issuetypes)
    refresh_issuetypes_field.grid(row=28, column=0, columnspan=4, padx=82, stick=E)
    
    check_validate_template = IntVar(value=check_template_flag)
    Checkbutton(main, text="Validate Template (check correctness of Issuetypes, Fields and Statuses)", font=("Helvetica", 9, "italic"), variable=check_validate_template).grid(row=29, column=0, sticky=W, padx=20, columnspan=4, pady=0)
    check_validate_template.trace('w', change_validate_template)
    
    skip_existing_issuetypes_validation = IntVar(value=skip_existing_issuetypes_validation_flag)
    Checkbutton(main, text="Skip validation for non-migrated issuetypes.", font=("Helvetica", 9, "italic"), variable=skip_existing_issuetypes_validation).grid(row=29, column=0, sticky=E, padx=140, columnspan=4, pady=0)
    skip_existing_issuetypes_validation.trace('w', change_skip_existing_issuetypes_validation)
    
    process_reconciliation = IntVar(value=process_reconciliation_flag)
    Checkbutton(main, text="Apply Reconciliation logic (diff for Source vs Target) within the last days:", font=("Helvetica", 9, "italic"), variable=process_reconciliation).grid(row=30, column=0, sticky=W, padx=20, columnspan=4, pady=0)
    process_reconciliation.trace('w', change_process_reconciliation)
    
    reconciliation_updated_days = check_similar("reconciliation_updated_days", reconciliation_updated_days)
    
    recon_days = tk.Entry(main, width=5, textvariable=reconciliation_updated_days)
    recon_days.insert(END, reconciliation_updated_days)
    recon_days.grid(row=30, column=1, pady=0, sticky=W, columnspan=3, padx=24)
    
    process_reconciliation_excel = IntVar(value=process_reconciliation_excel_flag)
    Checkbutton(main, text="Generate Diff Excel once migration complete.", font=("Helvetica", 9, "italic"), variable=process_reconciliation_excel).grid(row=30, column=1, sticky=W, padx=55, columnspan=3, pady=0)
    process_reconciliation_excel.trace('w', change_process_reconciliation_excel)
    
    merge_projects_start = IntVar(value=merge_projects_start_flag)
    Checkbutton(main, text="Starting Key in Target Project (i.e. first issue Key):", font=("Helvetica", 9, "italic"), variable=merge_projects_start).grid(row=31, column=0, sticky=W, padx=20, columnspan=4, pady=0)
    merge_projects_start.trace('w', change_merge_project_start)
    
    tk.Label(main, text="OR", font=("Helvetica", 9, "italic")).grid(row=31, column=0, columnspan=4, sticky=W, padx=370)
    
    shifted_by = check_similar("shifted_by", shifted_by)
    
    start_num = tk.Entry(main, width=7, textvariable=shifted_by)
    start_num.insert(END, shifted_by)
    start_num.grid(row=31, column=0, pady=0, sticky=W, columnspan=4, padx=312)
    
    merge_projects = IntVar(value=merge_projects_flag)
    Checkbutton(main, text="Shifting Starting Key from max in Target Project by:", font=("Helvetica", 9, "italic"), variable=merge_projects).grid(row=31, column=0, sticky=E, padx=150, columnspan=4, pady=0)
    merge_projects.trace('w', change_merge_project)
    
    shifted_key_val = check_similar("shifted_key_val", shifted_key_val)
    
    shift_num = tk.Entry(main, width=10, textvariable=shifted_key_val)
    shift_num.insert(END, shifted_key_val)
    shift_num.grid(row=31, column=2, pady=0, sticky=E, columnspan=3, padx=82)
    
    set_read_only = IntVar(value=set_source_project_read_only)
    Checkbutton(main, text="Set Source Project as Read-Only after migration, by updating Permission Scheme to (containing):", font=("Helvetica", 9, "italic"), variable=set_read_only).grid(row=32, column=0, sticky=W, padx=20, columnspan=4, pady=0)
    set_read_only.trace('w', change_read_only)
    
    read_only_scheme_name = check_similar("read_only_scheme_name", read_only_scheme_name)
    
    permission_scheme = tk.Entry(main, width=30, textvariable=read_only_scheme_name)
    permission_scheme.insert(END, read_only_scheme_name)
    permission_scheme.grid(row=32, column=2, pady=0, sticky=W, columnspan=3, padx=35)
    
    tk.Button(main, text='Quit', font=("Helvetica", 9, "bold"), command=main.quit, width=20, heigh=2).grid(row=33, column=0, pady=8, columnspan=4, rowspan=2)
    
    clear_additional_configuration = IntVar(value=clear_additional_configuration_flag)
    Checkbutton(main, text="unselect all", foreground="grey", font=("Helvetica", 9, "italic"), variable=clear_additional_configuration).grid(row=33, column=0, sticky=NW, padx=40, columnspan=4, rowspan=2, pady=5)
    clear_additional_configuration.trace('w', change_clear_additional_configuration)
    
    # The license details could be found here: https://github.com/delsakov/JIRA_Tools/
    # Please do not change line below with copyright
    tk.Label(main, text="Author: Dmitry Elsakov", foreground="grey", font=("Helvetica", 8, "italic"), pady=0).grid(row=34, column=1, sticky=SE, padx=20, columnspan=3)
    
    tk.mainloop()
