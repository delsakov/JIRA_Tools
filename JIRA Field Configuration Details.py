# This Tool has been created by Dmitry Elsakov
# The main source code has been created over weekends and distributed over GPL-3.0 License
# The license details could be found here: https://github.com/delsakov/JIRA_Tools/
# Please do not change notice above and copyright

from jira import JIRA
from tkinter.filedialog import asksaveasfilename
import os
import tkinter as tk
from tkinter import *
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from time import sleep
from sys import exit
import urllib3

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
urllib3.disable_warnings(urllib3.exceptions.HTTPWarning)
urllib3.disable_warnings(urllib3.exceptions.ConnectionError)
urllib3.disable_warnings()

current_version = '0.3'
JIRA_BASE_URL = ''
project = ''
output_file = ''
zoom_scale = 90
verify = True

header_font = Font(color='00000000', bold=True)
header_fill = PatternFill(fill_type="solid", fgColor="8db5e2")
hyperlink = Font(underline='single', color='0563C1')
wb = Workbook()


def get_str_from_lst(lst, sep=',', spacing=' '):
    """This function returns list as comma separated string - for exporting in excel"""
    if lst is None or type(lst) != list:
        return lst
    st = ''
    for i in lst:
        if i != '':
            st += str(i).strip() + sep + spacing
    if spacing == ' ':
        st = st[0:-2]
    else:
        st = st[0:-1]
    return st


def get_fields_list_by_project(jira, project):
    auth_jira = jira
    allfields = auth_jira.fields()
    
    def retrieve_custom_field(field_id):
        for field in allfields:
            if field['id'] == field_id:
                return field['custom']
    
    proj = auth_jira.project(project)
    project_fields = auth_jira.createmeta(projectKeys=proj, expand='projects.issuetypes.fields')
    is_types = project_fields['projects'][0]['issuetypes']
    issuetype_fields = {}
    for issuetype in is_types:
        issuetype_name = issuetype['name']
        issuetype_fields[issuetype_name] = {}
        for field_id in issuetype['fields']:
            field_name = issuetype['fields'][field_id]['name']
            allowed_values = []
            if 'allowedValues' in issuetype['fields'][field_id]:
                for i in issuetype['fields'][field_id]['allowedValues']:
                    if 'children' in i:
                        for ch in i['children']:
                            if 'value' in ch.keys():
                                allowed_values.append([i['value'], ch['value']])
                    elif 'name' in i:
                        allowed_values.append(i['name'])
                    elif 'value' in i:
                        allowed_values.append(i['value'])
                        default_val = None
            
            if issuetype['fields'][field_id]['hasDefaultValue'] is not False:
                if 'name' in issuetype['fields'][field_id]['defaultValue']:
                    default_val = issuetype['fields'][field_id]['defaultValue']['name']
                elif type(issuetype['fields'][field_id]['defaultValue']) == dict:
                    default_val = issuetype['fields'][field_id]['defaultValue']['value']
                elif type(issuetype['fields'][field_id]['defaultValue']) == list:
                    default_val = issuetype['fields'][field_id]['defaultValue'][0]['value']
                else:
                    default_val = issuetype['fields'][field_id]['defaultValue']
            
            field_attributes = {'id': field_id, 'required': issuetype['fields'][field_id]['required'],
                                'custom': retrieve_custom_field(field_id),
                                'type': issuetype['fields'][field_id]['schema']['type'],
                                'custom type': None if 'custom' not in issuetype['fields'][field_id]['schema'] else issuetype['fields'][field_id]['schema']['custom'].replace('com.atlassian.jira.plugin.system.customfieldtypes:', ''),
                                'allowed values': None if allowed_values == [] else allowed_values,
                                'default value': default_val,
                                'validated': True if 'allowedValues' in issuetype['fields'][field_id] else False}
            issuetype_fields[issuetype_name][field_name] = field_attributes
    return issuetype_fields


def create_excel_sheet(sheet_data, title):
    global JIRA_BASE_URL, output_excel
    wb.create_sheet(title)
    ws = wb.get_sheet_by_name(title)
    
    start_column = 1
    start_row = 1
    
    # Creating Excel sheet based on data
    for i in range(len(sheet_data)):
        for y in range(len(sheet_data[i])):
            try:
                ws.cell(row=start_row, column=start_column+y).value = sheet_data[i][y]
            except:
                converted_value = ''
                for letter in sheet_data[i][y]:
                    if letter.isalpha() or letter.isnumeric() or letter in [',', '.', ';', ':', '&', '"', "'", ' ']:
                        converted_value += letter
                    else:
                        converted_value += '?'
                ws.cell(row=start_row, column=start_column+y).value = converted_value
        start_row += 1
    
    for y in range(1, ws.max_column+1):
        ws.cell(row=1, column=y).fill = header_fill
        ws.cell(row=1, column=y).font = header_font
    
    # Column width formatting
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        if length > 80:
            ws.column_dimensions[column_cells[0].column_letter].width = 80
        else:
            ws.column_dimensions[column_cells[0].column_letter].width = length + 4
        
    ws.title = title
    
    sheet_names = wb.sheetnames
    for s in sheet_names:
        ws = wb.get_sheet_by_name(s)
        if ws.dimensions == 'A1:A1':
            wb.remove_sheet(wb[s])

    save_excel()

    
def save_excel():
    global zoom_scale, output_file, project
    try:
        if output_file.split('.xlsx')[0] == '':
            output_file = "Field configuration '{}'.xlsx".format(project)
        
        if os.path.exists(output_file) is True:
            overwrite_popup()
        
        for ws in wb.worksheets:
            ws.sheet_view.zoomScale = zoom_scale
            ws.auto_filter.ref = ws.dimensions
        wb.save(output_file)
        print("[END] File '{}' successfully generated.".format(output_file))
        print()
        sleep(2)
        exit()
    except Exception as e:
        print()
        print("[ERROR] ", e)
        os.system("pause")
        exit()


def select_file():
    global output_file, issues, header
    dir_name = os.getcwd()
    output_file = asksaveasfilename(initialdir=dir_name, title="Select file", filetypes=(("Migration JIRA Template", ".xlsx .xls"), ("all files", "*.*")))
    file.delete(0, END)
    file.insert(0, output_file)


def overwrite_popup():
    """Function which shows Pop-Up window with question about overriding Excel file, if it already exists"""
    global output_file
    
    def create_new():
        global output_file
        popup.destroy()
        popup.quit()
        time_format = "%Y-%m-%dT%H:%M:%S"
        now = datetime.datetime.strptime(str(datetime.datetime.utcnow().isoformat()).split(".", 1)[0], time_format)
        output_file = output_file.split('.xls')[0] + '_' + str(now).replace(':', '-').replace(' ', '_') + '_UTC.xlsx'
    
    def override():
        popup.destroy()
        popup.quit()
    
    popup = tk.Tk()
    popup.title("Override File?")
    
    l1 = tk.Label(popup, text="File '{}' already exist.".format(output_file), foreground="black", font=("Helvetica", 10), pady=4, padx=8)
    l1.grid(row=0, column=0, columnspan=2)
    l2 = tk.Label(popup, text="Do you want to override existing file OR create a new one?", foreground="black", font=("Helvetica", 10), pady=4, padx=8)
    l2.grid(row=1, column=0, columnspan=2)
    
    b1 = tk.Button(popup, text="Override", font=("Helvetica", 9, "bold"), command=override, width=20, heigh=2)
    b1.grid(row=2, column=0, pady=10, padx=8)
    b2 = tk.Button(popup, text="Create New", font=("Helvetica", 9, "bold"), command=create_new, width=20, heigh=2)
    b2.grid(row=2, column=1, pady=10, padx=8)
    
    tk.mainloop()


def jira_authorization_popup():
    """Function which shows Pop-Up window with question about JIRA credentials, if not entered"""
    global auth, username, password, jira, JIRA_BASE_URL, verify
    
    def jira_save():
        global auth, username, password, jira, JIRA_BASE_URL
        
        username = user.get()
        password = passwd.get()
        if len(username) < 3 or len(password) < 3:
            print("Invalid JIRA credentials were entered!")
            os.system("pause")
            exit()
        auth = (username, password)
        jira_popup.destroy()

        try:
            jira1 = JIRA(JIRA_BASE_URL, max_retries=0)
        except:
            jira1 = JIRA(JIRA_BASE_URL, max_retries=0, options={'verify': False})
            verify = False
            print("", "[WARNING] SSL verification failed. Further processing would be with skipping SSL verification -> insecure connection processing.", "", sep='\n')

        try:
            jira = JIRA(JIRA_BASE_URL, auth=auth, max_retries=0, options={'verify': verify})
        except Exception as e:
            print("[ERROR] Login to JIRA failed. Check your Username and Password. Exception: '{}'".format(e))
            os.system("pause")
            exit()
        jira = JIRA(JIRA_BASE_URL, auth=auth, max_retries=3, options={'verify': verify})
        jira_popup.quit()
    
    def jira_cancel():
        jira_popup.destroy()
        jira_popup.quit()
        print("[ERROR] Invalid JIRA credentials were entered!")
        os.system("pause")
        exit()
    
    jira_popup = tk.Tk()
    jira_popup.title("[AUTHORIZATION] JIRA credentials required")
    
    tk.Label(jira_popup, text="To Export Fields configuration please enter your Username / Password for JIRA access.", foreground="black", font=("Helvetica", 9), padx=10, wraplength=210).grid(row=1, column=0, rowspan=2)
    tk.Label(jira_popup, text="Username").grid(row=1, column=1, pady=5)
    tk.Label(jira_popup, text="Password").grid(row=2, column=1, pady=5)
    
    user = tk.Entry(jira_popup)
    user.grid(row=1, column=2, pady=5)
    passwd = tk.Entry(jira_popup, width=20, show="*")
    passwd.grid(row=2, column=2, pady=5)
    
    tk.Button(jira_popup, text='Cancel', font=("Helvetica", 9, "bold"), command=jira_cancel, width=20, heigh=2).grid(row=4, column=0, pady=8, padx=20, sticky=W, columnspan=2)
    tk.Button(jira_popup, text='OK', font=("Helvetica", 9, "bold"), command=jira_save, width=20, heigh=2).grid(row=4, column=1, pady=8, padx=20, sticky=E, columnspan=2)
    
    tk.mainloop()


def main_program():
    global jira, auth, username, password, project,output_file, JIRA_BASE_URL, issue_details, verify
    
    username = user.get().strip()
    password = passwd.get().strip()
    output_file = file.get().split('.xls')[0] + '.xlsx'
    project = source_project.get().strip()
    JIRA_BASE_URL = source_jira.get().strip()
    
    main.destroy()
    if len(username) < 3 or len(password) < 3:
        print('[ERROR] JIRA credentials are required. Please enter them on new window.')
        jira_authorization_popup()
    else:
        auth = (username, password)
        try:
            jira1 = JIRA(JIRA_BASE_URL, max_retries=0)
        except:
            jira1 = JIRA(JIRA_BASE_URL, max_retries=0, options={'verify': False})
            verify = False
            print("", "[WARNING] SSL verification failed. Further processing would be with skipping SSL verification -> insecure connection processing.", "", sep='\n')

        try:
            jira = JIRA(JIRA_BASE_URL, auth=auth, max_retries=0, options={'verify': verify})
        except Exception as e:
            print("[ERROR] Login to JIRA failed. Check your Username and Password. Exception: '{}'".format(e))
            os.system("pause")
            exit()
        jira = JIRA(JIRA_BASE_URL, auth=auth, max_retries=3, options={'verify': False})
    
    issue_details = {}

    if project == '':
        print("[START] Fields configuration downloading from ALL projects")
        print("[INFO] Only projects with correct configuration will be dounloaded.")
        
        projects = jira.projects()
        for project in projects:
            try:
                issue_details[project.key] = get_fields_list_by_project(jira, project.key)
            except:
                print("[WARNING] No access for '{}' project. Skipping...".format(project.key))
        
    else:
        print("[START] Fields configuration downloading from '{}' project".format(project))
        try:
            issue_details[project] = get_fields_list_by_project(jira, project)
        except:
            print("[WARNING] No access for '{}' project. Skipping...".format(project))
        print("[END] Fields configuration successfully processed.", '', sep='\n')

    if len(issue_details) > 0:
        
        sheet_data = [['Project Key', 'IssueType', 'Field Name', 'Field JIRA id', 'is Custom?', 'is Mandatory?', 'Allowed Values', 'Default Value', 'Field Type', 'Field Custom Type']]
        
        for k, v in issue_details.items():
            for issuetype, fields in v.items():
                for field, values in fields.items():
                    sheet_data.append([k, issuetype, field, values['id'], values['custom'], values['required'],
                                       '' if not values['allowed values'] else get_str_from_lst(values['allowed values']),
                                       '' if not values['default value'] else values['default value'],
                                       '' if not values['type'] else values['type'],
                                       '' if not values['custom type'] else values['custom type']])
        create_excel_sheet(sheet_data, 'Fields Configuration')


# ------------------ MAIN PROGRAM -----------------------------------
print("[INFO] Program has started. Please DO NOT CLOSE that window.")
print("[INFO] Please IGNORE any WARNINGS - the connection issues are covered by Retry logic.")
print()

main = tk.Tk()
Title = main.title("JIRA Field Configuration" + " v_" + current_version)

tk.Label(main, text="Please specify JIRA instance URL and Project Key (if empty, ALL project will be processed).", foreground="black", font=("Helvetica", 11, "italic"), padx=10, wraplength=500).grid(row=0, column=0, columnspan=5)

tk.Label(main, text="JIRA Instance URL:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=150).grid(row=1, column=0, rowspan=1)
source_jira = tk.Entry(main, width=50, textvariable=JIRA_BASE_URL)
source_jira.insert(END, JIRA_BASE_URL)
source_jira.grid(row=1, column=1, columnspan=2, padx=8)

tk.Label(main, text="Project Key:", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=150).grid(row=1, column=2, rowspan=1, columnspan=2, sticky=E, padx=80)
source_project = tk.Entry(main, width=10, textvariable=project)
source_project.insert(END, project)
source_project.grid(row=1, column=3, columnspan=1, sticky=E, padx=10)

tk.Label(main, text="____________________________________________________________________________________________________________").grid(row=3, columnspan=4)

tk.Label(main, text="For retrieval fields configuration please enter your Username / Password for JIRA access", foreground="black", font=("Helvetica", 10), padx=10, wraplength=250).grid(row=4, column=0, rowspan=2, columnspan=3, sticky=W, padx=10)
tk.Label(main, text="Username", foreground="black", font=("Helvetica", 10)).grid(row=4, column=1, pady=5, columnspan=2, sticky=W, padx=120)
tk.Label(main, text="Password", foreground="black", font=("Helvetica", 10)).grid(row=5, column=1, pady=5, columnspan=2, sticky=W, padx=120)
user = tk.Entry(main)
user.grid(row=4, column=2, pady=5, sticky=W, columnspan=2, padx=30)
passwd = tk.Entry(main, width=20, show="*")
passwd.grid(row=5, column=2, pady=5, sticky=W, columnspan=2, padx=30)

tk.Button(main, text='Download Fields Data', font=("Helvetica", 9, "bold"), state='active', command=main_program, width=20, heigh=2).grid(row=4, column=3, pady=4, padx=10, rowspan=2)

tk.Label(main, text="____________________________________________________________________________________________________________").grid(row=6, columnspan=4)

tk.Label(main, text="Export File Name", foreground="black", font=("Helvetica", 10), pady=7, padx=5, wraplength=200).grid(row=7, column=0, rowspan=1, padx=20, sticky=W)
file = tk.Entry(main, width=50, textvariable=output_file)
file.insert(END, output_file)
file.grid(row=7, column=1, columnspan=2, padx=8)
tk.Button(main, text='Browse', command=select_file, width=15).grid(row=7, column=3, pady=3, padx=8)

tk.Label(main, text="____________________________________________________________________________________________________________").grid(row=8, columnspan=4)

tk.Button(main, text='Quit', font=("Helvetica", 9, "bold"), command=main.quit, width=20, heigh=2).grid(row=9, column=0, pady=8, columnspan=4, rowspan=2)
tk.Label(main, text="Author: Dmitry Elsakov", foreground="grey", font=("Helvetica", 8, "italic"), pady=10).grid(row=10, column=3, sticky=E, padx=10)

tk.mainloop()
